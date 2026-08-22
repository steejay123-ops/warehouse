import os
import sys

try:
    sys.stdout.reconfigure(encoding='utf-8')
except Exception:
    pass

backend_dir = r"e:\warehouse project\warehouse-backend"
if backend_dir not in sys.path:
    sys.path.insert(0, backend_dir)

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')

import django
django.setup()

from decimal import Decimal
from inventory.models import Item, CountTask, CountTaskHistory

def verify_phase5():
    print("=" * 60)
    print("🛡️ ارزیابی مستقل ایجنت نگهبان فاز ۵ (Guardian Agent Verification)")
    print("=" * 60)

    errors = []

    # 1. بررسی سناریو ۹ (E2E-ITEM-09 - Draft persistence & Sync)
    item9 = Item.objects.filter(fa_unic_code="E2E-ITEM-09").first()
    if not item9:
        errors.append("کالای E2E-ITEM-09 در دیتابیس یافت نشد.")
    else:
        if item9.field_status != 'done':
            errors.append(f"وضعیت میدانی E2E-ITEM-09 برابر {item9.field_status} است، انتظار: done")
        if item9.inventory != Decimal('50.000'):
            errors.append(f"موجودی نهایی E2E-ITEM-09 برابر {item9.inventory} است، انتظار: 50.000")

    task9 = CountTask.objects.filter(item=item9).order_by('-id').first()
    if not task9:
        errors.append("تسک شمارش برای E2E-ITEM-09 یافت نشد.")
    else:
        if task9.status != 'FINAL_APPROVED':
            errors.append(f"وضعیت تسک E2E-ITEM-09 برابر {task9.status} است، انتظار: FINAL_APPROVED")

    # 2. بررسی سناریو ۱۰ (فایل خروجی اکسل)
    excel_path = r"e:\warehouse project\Documents\Browser_E2E_Comprehensive_Testing\inventory_e2e_export.xlsx"
    if not os.path.exists(excel_path) or os.path.getsize(excel_path) < 1000:
        errors.append("فایل خروجی اکسل تولید نشده یا حجم نامعتبر دارد.")
    else:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_path)
            sheet = wb.active
            if sheet.max_row < 2:
                errors.append(f"شیت اکسل فاقد داده است (تعداد ردیف: {sheet.max_row})")
        except Exception as e:
            errors.append(f"خطا در اعتبارسنجی ساختار فایل اکسل: {e}")

    # 3. بررسی فایل‌های اسکرین‌شات فاز ۵
    screenshot_dir = r"e:\warehouse project\Documents\Browser_E2E_Comprehensive_Testing\screenshots"
    expected_screens = [
        'phase5_scenario9_01_draft_saved.png',
        'phase5_scenario9_02_draft_persisted_after_refresh.png',
        'phase5_scenario9_03_synced_manager_approved.png',
        'phase5_scenario10_01_advanced_filters.png',
        'phase5_scenario10_02_task_history_audit_trail.png',
        'phase5_scenario10_03_excel_export_completed.png'
    ]
    for s in expected_screens:
        p = os.path.join(screenshot_dir, s)
        if not os.path.exists(p) or os.path.getsize(p) < 1000:
            errors.append(f"اسکرین‌شات {s} ذخیره نشده یا حجم نامعتبر دارد.")

    if errors:
        print("❌ ایجنت نگهبان فاز ۵ را رد کرد (REJECTED):")
        for err in errors:
            print(f"   - {err}")
        return False
    else:
        print("✅ ایجنت نگهبان فاز ۵ را به طور ۱۰۰٪ تایید کرد (PASSED & VERIFIED).")
        print("   - پایداری پیش‌نویس، همگام‌سازی، تاریخچه رویدادها و ساختار باینری فایل اکسل کاملاً معتبر و صحیح هستند.")
        return True

if __name__ == '__main__':
    ok = verify_phase5()
    sys.exit(0 if ok else 1)
