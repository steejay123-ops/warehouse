"""
ماژول ابزارهای اکسل هوشمند — کاربران (Users)
تولید فایل اکسل خروجی، فایل قالب نمونه، و خوانش/اعتبارسنجی پیشرفته فایل آپلودی
"""
import io
import re
from openpyxl import Workbook, load_workbook
from common.excel_utils import (
    styled_cell, apply_header_styles_to_row, set_column_widths, 
    freeze_header_panes, find_data_start_and_mapping
)
from django.http import HttpResponse
from django.contrib.auth.models import Group

from .models import CustomUser, CustomRole
from warehouses.models import Warehouse


# ── Column Definitions ──────────────────────────────────────────────
USERS_COLUMNS = [
    {'label': 'نام', 'key': 'first_name', 'width': 20, 'type': 'text'},
    {'label': 'نام خانوادگی', 'key': 'last_name', 'width': 20, 'type': 'text'},
    {'label': 'شناسه ورود', 'key': 'username', 'width': 20, 'type': 'text'},
    {'label': 'کد ملی', 'key': 'national_code', 'width': 18, 'type': 'text'},
    {'label': 'تلفن', 'key': 'phone_number', 'width': 18, 'type': 'text'},
    {'label': 'منطقه عملیاتی', 'key': 'operational_zone', 'width': 22, 'type': 'text'},
    {'label': 'شرکت متبوع', 'key': 'company', 'width': 22, 'type': 'text'},
    {'label': 'نقش‌ها', 'key': 'roles', 'width': 30, 'type': 'text'},
    {'label': 'انبارها', 'key': 'warehouses', 'width': 30, 'type': 'text'},
    {'label': 'فعال', 'key': 'is_active', 'width': 12, 'type': 'text'},
]

MAX_IMPORT_ROWS = 500


# ── Smart Data Cleaners & Normalizers ────────────────────────────────

PERSIAN_ARABIC_DIGITS = str.maketrans('۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩', '01234567890123456789')

def normalize_digits(text):
    """تبدیل ارقام فارسی و عربی به ارقام استاندارد انگلیسی"""
    if text is None:
        return ''
    return str(text).translate(PERSIAN_ARABIC_DIGITS).strip()

def normalize_phone(phone_str):
    """پاکسازی و استانداردسازی شماره تماس به قالب 09xxxxxxxxx"""
    if not phone_str:
        return None
    raw = normalize_digits(phone_str)
    digits = re.sub(r'[^\d+]', '', raw)
    if digits.startswith('+98'):
        digits = '0' + digits[3:]
    elif digits.startswith('0098'):
        digits = '0' + digits[4:]
    elif digits.startswith('98') and len(digits) == 12:
        digits = '0' + digits[2:]
    elif digits.startswith('9') and len(digits) == 10:
        digits = '0' + digits
    return digits if digits else None

def normalize_national_code(nid_str):
    """پاکسازی و افزودن صفرهای سمت چپ به کد ملی ۱۰ رقمی"""
    if not nid_str:
        return None
    raw = normalize_digits(nid_str)
    digits = re.sub(r'\D', '', raw)
    if not digits:
        return None
    if len(digits) < 10:
        digits = digits.zfill(10)
    return digits

def normalize_boolean(val_str, default=True):
    """تشخیص مقدار بولی از روی کلمات فارسی، انگلیسی و اعداد"""
    if val_str is None:
        return default
    s = normalize_digits(str(val_str)).lower().strip()
    if s in ('خیر', 'غیرفعال', 'نه', '0', 'no', 'false', 'f', 'off'):
        return False
    if s in ('بله', 'فعال', 'آره', '1', 'yes', 'true', 't', 'on', 'ok'):
        return True
    return default

def split_items(text):
    """تفکیک رشته با جداکننده‌های مختلف (ویرگول فارسی، انگلیسی، سمی‌کالن، اینتر)"""
    if not text:
        return []
    parts = re.split(r'[,،;\n|]+', str(text))
    return [p.strip() for p in parts if p.strip()]


# ── Excel Export & Template ──────────────────────────────────────────

def generate_users_excel(queryset):
    """
    تولید فایل اکسل استاندارد دو سطری از لیست کاربران
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'کاربران'
    ws.sheet_view.rightToLeft = True

    set_column_widths(ws, USERS_COLUMNS)
    freeze_header_panes(ws, row=3)

    # سطر اول: عناوین فارسی
    header_row = []
    for c in USERS_COLUMNS:
        header_row.append(styled_cell(ws, c['label']))
    apply_header_styles_to_row(header_row, is_key_row=False)
    ws.append(header_row)

    # سطر دوم: کلیدهای دیتابیسی
    key_row = []
    for c in USERS_COLUMNS:
        key_row.append(styled_cell(ws, c['key']))
    apply_header_styles_to_row(key_row, is_key_row=True)
    ws.append(key_row)

    # سطر سوم به بعد: داده‌های کاربران
    for row_idx, user in enumerate(queryset, 3):
        role_names = []
        for g in user.groups.all():
            try:
                role_names.append(g.customrole.title or g.name)
            except Exception:
                role_names.append(g.name)
        roles_str = '، '.join(role_names)

        wh_items = []
        for w in user.assigned_warehouses.all():
            wh_items.append(f"{w.name} ({w.code})" if w.code else w.name)
        wh_str = '، '.join(wh_items)

        row_data = [
            user.first_name or '',
            user.last_name or '',
            user.username or '',
            user.national_code or '',
            user.phone_number or '',
            user.operational_zone or '',
            user.company or '',
            roles_str,
            wh_str,
            'بله' if user.is_active else 'خیر',
        ]
        ws.append(row_data)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="users_export.xlsx"'
    return response


def generate_users_template():
    """
    تولید فایل قالب نمونه دو سطری با ۲ سطر داده تستی
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'قالب کاربران'
    ws.sheet_view.rightToLeft = True

    set_column_widths(ws, USERS_COLUMNS)
    freeze_header_panes(ws, row=3)

    # سطر اول: عناوین فارسی
    header_row = []
    for c in USERS_COLUMNS:
        header_row.append(styled_cell(ws, c['label']))
    apply_header_styles_to_row(header_row, is_key_row=False)
    ws.append(header_row)

    # سطر دوم: کلیدهای دیتابیسی
    key_row = []
    for c in USERS_COLUMNS:
        key_row.append(styled_cell(ws, c['key']))
    apply_header_styles_to_row(key_row, is_key_row=True)
    ws.append(key_row)

    sample_data = [
        ['علی', 'محمدی', 'ali.mohammadi', '1234567890', '09121234567', 'منطقه ۱ - جنوب', 'نفت و گاز', 'سرپرست انبار', 'WH-1، WH-2', 'بله'],
        ['فاطمه', 'احمدی', 'fatemeh.ahmadi', '0987654321', '09351234567', 'منطقه ۲ - مرکز', 'پیمانکار فارس عالیش', 'انبارگردان', 'WH-3', 'بله'],
    ]

    for row in sample_data:
        ws.append(row)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="users_template.xlsx"'
    return response


# ── Smart Excel Parser ───────────────────────────────────────────────

def parse_users_excel(file, update_existing=False):
    """
    پارس و اعتبارسنجی هوشمند فایل اکسل کاربران
    - پشتیبانی از تطبیق دوگانه نقش‌ها (عنوان فارسی و نام انگلیسی)
    - پشتیبانی از تطبیق دوگانه انبارها (کد و نام انبار)
    - پاکسازی خودکار ارقام و کد ملی و تلفن
    - پشتیبانی از پارامتر update_existing
    """
    try:
        wb = load_workbook(file, read_only=True, data_only=True)
    except Exception:
        return {'valid_rows': [], 'errors': [{'row': 0, 'field': 'file', 'message': 'فایل اکسل معتبر نیست یا قابل خواندن نمی‌باشد.'}]}

    ws = wb.active
    data_start_row, col_mapping = find_data_start_and_mapping(ws, USERS_COLUMNS)
    rows = list(ws.iter_rows(min_row=data_start_row, values_only=True))

    if len(rows) > MAX_IMPORT_ROWS:
        return {'valid_rows': [], 'errors': [{'row': 0, 'field': 'file', 'message': f'حداکثر {MAX_IMPORT_ROWS} سطر قابل پردازش است. فایل شما {len(rows)} سطر دارد.'}]}

    # ساخت مپ‌های جستجوی هوشمند
    existing_users_by_username = {u.username.lower(): u for u in CustomUser.objects.all()}
    existing_users_by_nid = {
        u.national_code: u for u in CustomUser.objects.exclude(national_code__isnull=True).exclude(national_code='')
    }

    # نقشه نقش‌ها: هم بر اساس name (کد انگلیسی) و هم بر اساس title (عنوان فارسی)
    roles_lookup = {}
    for group in Group.objects.all():
        roles_lookup[group.name.lower().strip()] = group
        try:
            cr = group.customrole
            if cr.title:
                roles_lookup[cr.title.lower().strip()] = group
        except Exception:
            pass

    # نقشه انبارها: هم بر اساس code و هم بر اساس name
    warehouses_lookup = {}
    for wh in Warehouse.objects.all():
        if wh.code:
            warehouses_lookup[wh.code.lower().strip()] = wh
        if wh.name:
            warehouses_lookup[wh.name.lower().strip()] = wh
            # حالت ترکیبی نام و کد مثل "انبار مرکزی (WH-1)"
            if wh.code:
                warehouses_lookup[f"{wh.name.lower().strip()} ({wh.code.lower().strip()})"] = wh

    valid_rows = []
    errors = []
    seen_usernames = set()
    seen_national_codes = set()

    for idx, row in enumerate(rows):
        row_num = idx + data_start_row

        if not row or all(cell is None or str(cell).strip() == '' for cell in row):
            continue

        row = list(row) + [None] * (len(USERS_COLUMNS) - len(row))

        def get_val(key, default_idx, default=''):
            idx_pos = col_mapping.get(key, default_idx)
            if idx_pos < len(row) and row[idx_pos] is not None:
                return str(row[idx_pos]).strip()
            return default

        first_name = get_val('first_name', 0)
        last_name = get_val('last_name', 1)
        username = normalize_digits(get_val('username', 2)).strip()
        raw_nid = get_val('national_code', 3)
        national_code = normalize_national_code(raw_nid)
        phone_number = normalize_phone(get_val('phone_number', 4))
        operational_zone = get_val('operational_zone', 5)
        company = get_val('company', 6)
        roles_str = get_val('roles', 7)
        warehouses_str = get_val('warehouses', 8)
        is_active_str = get_val('is_active', 9, 'بله')
        is_active = normalize_boolean(is_active_str, default=True)

        row_errors = []

        # ── فیلدهای اجباری ──
        if not first_name:
            row_errors.append({'row': row_num, 'field': 'first_name', 'message': 'نام الزامی است.'})
        if not last_name:
            row_errors.append({'row': row_num, 'field': 'last_name', 'message': 'نام خانوادگی الزامی است.'})
        if not username:
            row_errors.append({'row': row_num, 'field': 'username', 'message': 'شناسه ورود الزامی است.'})

        # ── یکتایی و اعتبارسنجی شناسه ورود ──
        is_update_record = False
        target_user_id = None

        if username:
            if ' ' in username:
                row_errors.append({'row': row_num, 'field': 'username', 'message': 'شناسه ورود نباید شامل فاصله باشد.'})
            elif username in seen_usernames:
                row_errors.append({'row': row_num, 'field': 'username', 'message': 'این شناسه ورود در همین فایل تکرار شده است.'})
            else:
                existing_user = existing_users_by_username.get(username.lower())
                if existing_user:
                    if update_existing:
                        is_update_record = True
                        target_user_id = existing_user.id
                    else:
                        row_errors.append({'row': row_num, 'field': 'username', 'message': 'این شناسه ورود قبلاً در سیستم ثبت شده است.'})
                seen_usernames.add(username)

        # ── اعتبارسنجی کد ملی ──
        if national_code:
            if not national_code.isdigit() or len(national_code) != 10:
                row_errors.append({'row': row_num, 'field': 'national_code', 'message': 'کد ملی باید ۱۰ رقم عددی باشد.'})
            elif national_code in seen_national_codes:
                row_errors.append({'row': row_num, 'field': 'national_code', 'message': 'این کد ملی در همین فایل تکرار شده است.'})
            else:
                nid_user = existing_users_by_nid.get(national_code)
                if nid_user:
                    if not update_existing or (is_update_record and nid_user.id != target_user_id):
                        row_errors.append({'row': row_num, 'field': 'national_code', 'message': 'این کد ملی به کاربر دیگری تخصیص دارد.'})
                seen_national_codes.add(national_code)

        # ── تطبیق هوشمند نقش‌ها ──
        resolved_roles = []
        if roles_str:
            role_items = split_items(roles_str)
            for rname in role_items:
                matched_role = roles_lookup.get(rname.lower())
                if not matched_role:
                    row_errors.append({'row': row_num, 'field': 'roles', 'message': f'نقش «{rname}» در سیستم یافت نشد.'})
                elif matched_role not in resolved_roles:
                    resolved_roles.append(matched_role)

        # ── تطبیق هوشمند انبارها ──
        resolved_warehouses = []
        if warehouses_str:
            wh_items = split_items(warehouses_str)
            for witem in wh_items:
                matched_wh = warehouses_lookup.get(witem.lower())
                if not matched_wh:
                    # تلاش با استخراج کد داخل پرانتز اگر وجود داشته باشد
                    match = re.search(r'\((.*?)\)', witem)
                    if match:
                        code_inside = match.group(1).strip().lower()
                        matched_wh = warehouses_lookup.get(code_inside)
                
                if not matched_wh:
                    row_errors.append({'row': row_num, 'field': 'warehouses', 'message': f'انبار «{witem}» در سیستم یافت نشد.'})
                elif matched_wh not in resolved_warehouses:
                    resolved_warehouses.append(matched_wh)

        if row_errors:
            errors.extend(row_errors)
        else:
            valid_rows.append({
                'is_update': is_update_record,
                'user_id': target_user_id,
                'first_name': first_name,
                'last_name': last_name,
                'username': username,
                'national_code': national_code or None,
                'phone_number': phone_number or None,
                'operational_zone': operational_zone or None,
                'company': company or None,
                'is_active': is_active,
                'roles': resolved_roles,
                'warehouses': resolved_warehouses,
            })

    wb.close()
    return {'valid_rows': valid_rows, 'errors': errors}
