"""
ماژول ابزارهای اکسل هوشمند — نقش‌ها (Roles)
تولید فایل اکسل خروجی، فایل قالب نمونه، و خوانش/اعتبارسنجی پیشرفته فایل آپلودی نقش‌ها
"""
import io
import re
from openpyxl import Workbook, load_workbook
from common.excel_utils import (
    styled_cell, apply_header_styles_to_row, set_column_widths, 
    freeze_header_panes, find_data_start_and_mapping
)
from django.http import HttpResponse
from django.contrib.auth.models import Permission

from .models import CustomRole


# ── Column Definitions ──────────────────────────────────────────────
ROLES_COLUMNS = [
    {'label': 'نام یکتا (name)', 'key': 'name', 'width': 22, 'type': 'text'},
    {'label': 'عنوان فارسی (title)', 'key': 'title', 'width': 25, 'type': 'text'},
    {'label': 'رنگ سازمانی (color)', 'key': 'color', 'width': 18, 'type': 'text'},
    {'label': 'نقش والد (parent)', 'key': 'parent', 'width': 25, 'type': 'text'},
    {'label': 'مجوزها (permissions)', 'key': 'permissions', 'width': 50, 'type': 'text'},
]

MAX_IMPORT_ROWS = 500


def split_items(text):
    """تفکیک رشته با انواع جداکننده‌ها"""
    if not text:
        return []
    parts = re.split(r'[,،;\n|]+', str(text))
    return [p.strip() for p in parts if p.strip()]


def generate_roles_excel(queryset):
    """
    تولید فایل اکسل استاندارد دو سطری از لیست نقش‌های سازمانی (۵ ستونه)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'نقش‌ها'
    ws.sheet_view.rightToLeft = True

    set_column_widths(ws, ROLES_COLUMNS)
    freeze_header_panes(ws, row=3)

    # سطر اول: عناوین فارسی
    header_row = []
    for c in ROLES_COLUMNS:
        header_row.append(styled_cell(ws, c['label']))
    apply_header_styles_to_row(header_row, is_key_row=False)
    ws.append(header_row)

    # سطر دوم: کلیدهای دیتابیسی
    key_row = []
    for c in ROLES_COLUMNS:
        key_row.append(styled_cell(ws, c['key']))
    apply_header_styles_to_row(key_row, is_key_row=True)
    ws.append(key_row)

    # سطر سوم به بعد: داده‌ها
    for row_idx, role in enumerate(queryset, 3):
        perms = []
        for p in role.permissions.all():
            perms.append(p.codename)
        perms_str = '، '.join(perms)

        parent_title = ''
        if role.parent:
            parent_title = role.parent.title or role.parent.name

        row_data = [
            role.name or '',
            role.title or role.name or '',
            role.color or '#94a3b8',
            parent_title,
            perms_str,
        ]
        ws.append(row_data)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="roles_export.xlsx"'
    return response


def generate_roles_template():
    """
    تولید فایل قالب نمونه دو سطری با نمونه‌های معتبر ۵ ستونه
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'قالب نقش‌ها'
    ws.sheet_view.rightToLeft = True

    set_column_widths(ws, ROLES_COLUMNS)
    freeze_header_panes(ws, row=3)

    # سطر اول: عناوین فارسی
    header_row = []
    for c in ROLES_COLUMNS:
        header_row.append(styled_cell(ws, c['label']))
    apply_header_styles_to_row(header_row, is_key_row=False)
    ws.append(header_row)

    # سطر دوم: کلیدهای دیتابیسی
    key_row = []
    for c in ROLES_COLUMNS:
        key_row.append(styled_cell(ws, c['key']))
    apply_header_styles_to_row(key_row, is_key_row=True)
    ws.append(key_row)

    sample_data = [
        ['supervisor', 'سرپرست انبار', '#4f46e5', '', 'view_sys_dashboard، view_sys_users، view_wh_docs'],
        ['counter', 'انبارگردان میدانی', '#10b981', 'سرپرست انبار', 'view_sys_counter، view_wh_docs'],
    ]

    for row in sample_data:
        ws.append(row)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="roles_template.xlsx"'
    return response


def parse_roles_excel(file, update_existing=False):
    """
    پارس و اعتبارسنجی هوشمند فایل اکسل نقش‌ها
    - پشتیبانی از تطبیق دسترسی‌ها با نام فارسی یا codename انگلیسی
    - تطبیق نقش والد بر اساس عنوان فارسی یا نام انگلیسی
    - پشتیبانی از پارامتر update_existing
    """
    try:
        wb = load_workbook(file, read_only=True, data_only=True)
    except Exception:
        return {'valid_rows': [], 'errors': [{'row': 0, 'field': 'file', 'message': 'فایل اکسل معتبر نیست یا قابل خواندن نمی‌باشد.'}]}

    ws = wb.active
    data_start_row, col_mapping = find_data_start_and_mapping(ws, ROLES_COLUMNS)
    rows = list(ws.iter_rows(min_row=data_start_row, values_only=True))

    if len(rows) > MAX_IMPORT_ROWS:
        return {'valid_rows': [], 'errors': [{'row': 0, 'field': 'file', 'message': f'حداکثر {MAX_IMPORT_ROWS} سطر قابل پردازش است. فایل شما {len(rows)} سطر دارد.'}]}

    # مپ نقش‌های موجود برای بررسی یکتایی و اتصال والد
    existing_roles_by_name = {r.name.lower().strip(): r for r in CustomRole.objects.all()}
    existing_roles_by_title = {r.title.lower().strip(): r for r in CustomRole.objects.filter(title__isnull=False)}

    # مپ تمام مجوزهای سیستم (هم codename و هم نام فارسی)
    perms_lookup = {}
    for p in Permission.objects.all():
        perms_lookup[p.codename.lower().strip()] = p
        if p.name:
            perms_lookup[p.name.lower().strip()] = p

    valid_rows = []
    errors = []
    seen_names = set()
    new_role_identifiers_in_file = set()

    # فاز اول: جمع‌آوری تمام شناسه‌ها و اسامی موجود در فایل
    for row in rows:
        if row and any(row):
            idx_name = col_mapping.get('name', 0)
            if idx_name < len(row) and row[idx_name]:
                n = str(row[idx_name]).strip().lower()
                new_role_identifiers_in_file.add(n)
            idx_title = col_mapping.get('title', 1)
            if idx_title < len(row) and row[idx_title]:
                t = str(row[idx_title]).strip().lower()
                new_role_identifiers_in_file.add(t)

    for idx, row in enumerate(rows):
        row_num = idx + data_start_row

        if not row or all(cell is None or str(cell).strip() == '' for cell in row):
            continue

        row = list(row) + [None] * (len(ROLES_COLUMNS) - len(row))

        def get_val(key, default_idx, default=''):
            idx_pos = col_mapping.get(key, default_idx)
            if idx_pos < len(row) and row[idx_pos] is not None:
                return str(row[idx_pos]).strip()
            return default

        name = get_val('name', 0)
        title = get_val('title', 1)
        color = get_val('color', 2, '#94a3b8')
        parent_raw = get_val('parent', 3)
        perms_str = get_val('permissions', 4)

        row_errors = []

        # ── فیلدهای اجباری ──
        if not name:
            row_errors.append({'row': row_num, 'field': 'name', 'message': 'نام یکتای سیستمی نقش الزامی است.'})
        if not title:
            row_errors.append({'row': row_num, 'field': 'title', 'message': 'عنوان فارسی نقش الزامی است.'})

        # ── اعتبارسنجی نام یکتا ──
        is_update_record = False
        target_role_id = None

        if name:
            if ' ' in name:
                row_errors.append({'row': row_num, 'field': 'name', 'message': 'نام سیستمی نباید شامل فاصله باشد.'})
            elif name.lower() in seen_names:
                row_errors.append({'row': row_num, 'field': 'name', 'message': 'این نام سیستمی در همین فایل تکرار شده است.'})
            else:
                existing_role = existing_roles_by_name.get(name.lower())
                if existing_role:
                    if update_existing:
                        is_update_record = True
                        target_role_id = existing_role.id
                    else:
                        row_errors.append({'row': row_num, 'field': 'name', 'message': 'این نام سیستمی قبلاً در سیستم ثبت شده است.'})
                seen_names.add(name.lower())

        # ── پاکسازی رنگ سازمانی ──
        if color and not color.startswith('#'):
            color = '#' + color
        if not color or len(color) not in (4, 7):
            color = '#94a3b8'

        # ── اعتبارسنجی نقش والد ──
        resolved_parent_name = None
        if parent_raw:
            p_clean = parent_raw.lower().strip()
            if p_clean in existing_roles_by_name:
                resolved_parent_name = existing_roles_by_name[p_clean].name
            elif p_clean in existing_roles_by_title:
                resolved_parent_name = existing_roles_by_title[p_clean].name
            elif p_clean in new_role_identifiers_in_file:
                resolved_parent_name = parent_raw.strip()
            else:
                row_errors.append({'row': row_num, 'field': 'parent', 'message': f'نقش والد «{parent_raw}» در سیستم یافت نشد.'})

        # ── اعتبارسنجی مجوزها ──
        resolved_permissions = []
        if perms_str:
            perm_items = split_items(perms_str)
            for pitem in perm_items:
                matched_perm = perms_lookup.get(pitem.lower())
                if not matched_perm:
                    row_errors.append({'row': row_num, 'field': 'permissions', 'message': f'دسترسی «{pitem}» در سیستم یافت نشد.'})
                elif matched_perm not in resolved_permissions:
                    resolved_permissions.append(matched_perm)

        if row_errors:
            errors.extend(row_errors)
        else:
            valid_rows.append({
                'is_update': is_update_record,
                'role_id': target_role_id,
                'name': name,
                'title': title,
                'color': color,
                'parent_name': resolved_parent_name,
                'permissions': resolved_permissions,
            })

    wb.close()
    return {'valid_rows': valid_rows, 'errors': errors}
