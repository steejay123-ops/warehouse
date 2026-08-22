import io
import openpyxl
from django.test import TestCase
from django.contrib.auth import get_user_model
from rest_framework.test import APIClient
from rest_framework import status
from axes.models import AccessAttempt
from accounts.models import AuditLog, UserLoginLog

User = get_user_model()

class AuditExportAndUnlockTestCase(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.admin_user = User.objects.create_superuser(
            username='admin_test',
            password='Password123!',
            email='admin@test.com'
        )
        self.normal_user = User.objects.create_user(
            username='user_test',
            password='Password123!',
            email='user@test.com'
        )
        self.client.force_authenticate(user=self.admin_user)

        # Create sample AuditLog records
        AuditLog.objects.create(
            user=self.admin_user,
            actor_username='admin_test',
            module='users',
            action='CREATE',
            severity='info',
            target_repr='تست کاربر جدید',
            target_object_id='100',
            ip_address='192.168.1.50',
            before_state={'username': 'old'},
            after_state={'username': 'new'}
        )
        AuditLog.objects.create(
            user=self.admin_user,
            actor_username='admin_test',
            module='docs',
            action='UPDATE',
            severity='warning',
            target_repr='ویرایش سند ۵۰۰',
            target_object_id='500',
            ip_address='192.168.1.50',
            before_state={'title': 'A'},
            after_state={'title': 'B'}
        )

        # Create sample UserLoginLog records
        UserLoginLog.objects.create(
            user=self.admin_user,
            username_attempted='admin_test',
            status='SUCCESS',
            ip_address='192.168.1.50',
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/120.0.0.0 Safari/537.36'
        )
        UserLoginLog.objects.create(
            user=None,
            username_attempted='locked_user',
            status='FAILED_LOCKED',
            failure_reason='بیش از حد مجاز تلاش ناموفق',
            ip_address='192.168.1.100',
            user_agent='Mozilla/5.0'
        )

        # Create sample axes AccessAttempt
        AccessAttempt.objects.create(
            username='locked_user',
            ip_address='192.168.1.100',
            failures_since_start=5
        )

    def test_audit_export_xlsx(self):
        """تست خروجی اکسل ممیزی با دو ردیف هدر (فارسی + نام دیتابیسی) و زمان شمسی"""
        response = self.client.get('/api/auth/audit-logs/export_excel/', {
            'file_format': 'xlsx',
            'module': 'users',
            'columns': 'index,id,user,module,action,severity,created_at'
        })
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        # بررسی ساختار شیت اکسل
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        self.assertEqual(ws.views.sheetView[0].rightToLeft, True)
        
        # ردیف ۱: عناوین فارسی
        row1 = [cell.value for cell in ws[1]]
        self.assertIn('شناسه', row1)
        self.assertIn('ماژول', row1)
        self.assertIn('تاریخ و زمان', row1)

        # ردیف ۲: نام‌های دیتابیسی
        row2 = [cell.value for cell in ws[2]]
        self.assertIn('id', row2)
        self.assertIn('module', row2)
        self.assertIn('created_at', row2)

        # ردیف ۳: سطرهای داده با تاریخ شمسی
        row3 = [cell.value for cell in ws[3]]
        created_at_idx = row2.index('created_at')
        shamsi_date = row3[created_at_idx]
        self.assertIsInstance(shamsi_date, str)
        self.assertRegex(shamsi_date, r'^\d{4}/\d{2}/\d{2} \d{2}:\d{2}:\d{2}$')

    def test_audit_export_csv(self):
        """تست خروجی متنی CSV ممیزی با هدر دو ردیفه و تاریخ شمسی"""
        response = self.client.get('/api/auth/audit-logs/export_excel/', {
            'file_format': 'csv',
            'severity': 'info',
            'columns': 'id,user,created_at'
        })
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue('text/csv' in response['Content-Type'])
        
        csv_text = response.content.decode('utf-8-sig')
        lines = [line.strip() for line in csv_text.splitlines() if line.strip()]
        self.assertTrue(len(lines) >= 3)
        self.assertIn('شناسه', lines[0])
        self.assertEqual(lines[1], 'id,user,created_at')

    def test_login_export_xlsx(self):
        """تست خروجی اکسل تاریخچه ورود با هدر نام دیتابیسی و زمان شمسی"""
        response = self.client.get('/api/auth/login-logs/export_excel/', {
            'file_format': 'xlsx',
            'status': 'FAILED_LOCKED',
            'columns': 'id,username_attempted,status,created_at'
        })
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        self.assertEqual(ws.views.sheetView[0].rightToLeft, True)
        
        # ردیف ۲ باید شامل نام‌های دیتابیسی باشد
        row2 = [cell.value for cell in ws[2]]
        self.assertEqual(row2, ['id', 'username_attempted', 'status', 'created_at'])

        # ردیف ۳ باید دارای زمان شمسی باشد
        row3 = [cell.value for cell in ws[3]]
        shamsi_val = row3[3]
        self.assertRegex(shamsi_val, r'^\d{4}/\d{2}/\d{2} \d{2}:\d{2}:\d{2}$')

    def test_login_export_csv(self):
        """تست خروجی CSV تاریخچه ورود با دو ردیف هدر"""
        response = self.client.get('/api/auth/login-logs/export_excel/', {
            'file_format': 'csv',
            'search': 'locked_user',
            'columns': 'id,username_attempted,status'
        })
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue('text/csv' in response['Content-Type'])
        
        csv_text = response.content.decode('utf-8-sig')
        lines = [line.strip() for line in csv_text.splitlines() if line.strip()]
        self.assertEqual(lines[1], 'id,username_attempted,status')

    def test_locked_users_endpoint(self):
        """تست استعلام کاربران قفل شده"""
        response = self.client.get('/api/auth/login-logs/locked_users/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        data = response.json()
        self.assertIn('locked_users', data)
        usernames = [u['username'] for u in data['locked_users']]
        self.assertIn('locked_user', usernames)

    def test_reset_lockout_endpoint(self):
        """تست بازنشانی قفل کاربر مسدود شده و پاکسازی رکوردهای axes"""
        self.assertTrue(AccessAttempt.objects.filter(username='locked_user').exists())

        response = self.client.post('/api/auth/login-logs/reset_lockout/', {
            'username': 'locked_user',
            'ip_address': '192.168.1.100'
        }, format='json')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        data = response.json()
        self.assertTrue(data.get('success'))

        # Verify AccessAttempt is removed
        self.assertFalse(AccessAttempt.objects.filter(username='locked_user').exists())

        # Verify an AuditLog is recorded for the unlock action
        unlock_log = AuditLog.objects.filter(action='UPDATE', target_repr__contains='locked_user').first()
        self.assertIsNotNone(unlock_log)
