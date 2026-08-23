from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework_simplejwt.views import TokenObtainPairView
from django.contrib.auth.models import Group, Permission
from .models import CustomUser, CustomRole
from .serializers import UserSerializer, CustomTokenObtainPairSerializer, CustomRoleSerializer, PermissionSerializer

from common.mixins import DeleteImpactMixin

class CustomTokenObtainPairView(TokenObtainPairView):
    serializer_class = CustomTokenObtainPairSerializer

class UserViewSet(DeleteImpactMixin, viewsets.ModelViewSet):
    queryset = CustomUser.objects.all()
    serializer_class = UserSerializer
    pagination_class = None

    def get_queryset(self):
        qs = CustomUser.objects.all()
        user = self.request.user
        
        has_perm = self.request.query_params.get('has_perm')
        if has_perm:
            from django.db.models import Q
            qs = qs.filter(
                Q(user_permissions__codename=has_perm) | 
                Q(groups__permissions__codename=has_perm)
            ).distinct()

        if user and user.is_authenticated:
            # Hide superusers from non-superuser users
            if not user.is_superuser:
                qs = qs.exclude(is_superuser=True)
        return qs

    def get_permissions(self):
        from rest_framework.permissions import IsAuthenticated, AllowAny
        from .permissions import HasMenuAccess
        
        if self.action == 'verify_card':
            permission_classes = [AllowAny()]
        elif self.action in ['change_password', 'update_preferences', 'my_avatar']:
            permission_classes = [IsAuthenticated()]
        elif self.action in ['list', 'retrieve', 'export_excel', 'download_template']:
            permission_classes = [HasMenuAccess('view_sys_users')]
        elif self.action in ['create', 'import_excel']:
            permission_classes = [HasMenuAccess('perm_usr_add')]
        else: # update, partial_update, destroy, toggle_status, user_avatar
            permission_classes = [HasMenuAccess('perm_usr_edit')]
            
        return permission_classes

    def _validate_user_deactivation(self, request, user, action_name="حذف"):
        from rest_framework.exceptions import ValidationError
        from django.db.models import Q
        from django.contrib.auth.models import Permission
        
        # 1. Self-deletion guard
        if request.user.id == user.id:
            raise ValidationError(f"شما نمی‌توانید حساب کاربری خود را {action_name} کنید.")
            
        # 2. Protect last active superuser
        if user.is_superuser:
            active_superusers = CustomUser.objects.filter(
                is_active=True, is_superuser=True
            ).exclude(id=user.id).count()
            
            if active_superusers == 0:
                raise ValidationError(f"امکان {action_name} تنها مدیر (Admin) فعال سیستم وجود ندارد.")
                
        # 3. Protect last key roles
        KEY_PERMISSIONS = {
            'can_act_as_counter': 'انبارگردان',
            'can_act_as_supervisor': 'سرپرست انبار',
            'can_act_as_manager': 'مدیر انبار',
            'can_act_as_doc_worker': 'کارشناس اسناد',
            'can_act_as_doc_supervisor': 'سرپرست اسناد'
        }
        
        user_perms = set(
            user.user_permissions.filter(codename__in=KEY_PERMISSIONS.keys()).values_list('codename', flat=True)
        ) | set(
            Permission.objects.filter(group__user=user, codename__in=KEY_PERMISSIONS.keys()).values_list('codename', flat=True)
        )

        for perm in user_perms:
            other_active_users = CustomUser.objects.filter(
                is_active=True
            ).exclude(id=user.id).filter(
                Q(user_permissions__codename=perm) | 
                Q(groups__permissions__codename=perm)
            ).count()
            
            if other_active_users == 0:
                role_name = KEY_PERMISSIONS.get(perm, perm)
                raise ValidationError(f"امکان {action_name} این کاربر وجود ندارد. سیستم باید حداقل یک کاربر فعال با دسترسی «{role_name}» داشته باشد.")

    def perform_create(self, serializer):
        from .audit_utils import log_audit_event
        from rest_framework.exceptions import PermissionDenied
        user_permissions = serializer.validated_data.get('user_permissions', [])
        if user_permissions and not (self.request.user and self.request.user.is_superuser):
            from .models import SENSITIVE_PERMISSION_CODENAMES
            assigned_codenames = set(
                Permission.objects.filter(id__in=[p.id if hasattr(p, 'id') else p for p in user_permissions])
                .values_list('codename', flat=True)
            )
            if assigned_codenames.intersection(SENSITIVE_PERMISSION_CODENAMES):
                raise PermissionDenied("تنها مدیر ارشد سامانه (Superuser) مجاز به اعطا یا تغییر دسترسی‌های حساس است.")

        instance = serializer.save(created_by=self.request.user if self.request.user.is_authenticated else None)
        try:
            log_audit_event(
                module='users',
                action='CREATE',
                severity='warning',
                target_model='CustomUser',
                target_object_id=instance.id,
                target_repr=f"{instance.username} ({instance.first_name} {instance.last_name})",
                user=self.request.user if self.request.user.is_authenticated else None,
                after_state={'username': instance.username, 'first_name': instance.first_name, 'last_name': instance.last_name, 'is_active': instance.is_active}
            )
        except Exception:
            pass

    def perform_update(self, serializer):
        from .audit_utils import log_audit_event, calculate_model_diff
        from rest_framework.exceptions import PermissionDenied
        user_permissions = serializer.validated_data.get('user_permissions', None)
        if user_permissions is not None and not (self.request.user and self.request.user.is_superuser):
            from .models import SENSITIVE_PERMISSION_CODENAMES
            assigned_codenames = set(
                Permission.objects.filter(id__in=[p.id if hasattr(p, 'id') else p for p in user_permissions])
                .values_list('codename', flat=True)
            )
            if assigned_codenames.intersection(SENSITIVE_PERMISSION_CODENAMES):
                raise PermissionDenied("تنها مدیر ارشد سامانه (Superuser) مجاز به اعطا یا تغییر دسترسی‌های حساس است.")

        old_instance = self.get_object()
        before_state = {'username': old_instance.username, 'first_name': old_instance.first_name, 'last_name': old_instance.last_name, 'is_active': old_instance.is_active}
        instance = serializer.save(modified_by=self.request.user if self.request.user.is_authenticated else None)
        after_state = {'username': instance.username, 'first_name': instance.first_name, 'last_name': instance.last_name, 'is_active': instance.is_active}
        diff_b, diff_a = calculate_model_diff(before_state, after_state)
        if diff_b or diff_a:
            try:
                log_audit_event(
                    module='users',
                    action='UPDATE',
                    severity='warning',
                    target_model='CustomUser',
                    target_object_id=instance.id,
                    target_repr=f"{instance.username} ({instance.first_name} {instance.last_name})",
                    user=self.request.user if self.request.user.is_authenticated else None,
                    before_state=diff_b,
                    after_state=diff_a
                )
            except Exception:
                pass

    def destroy(self, request, *args, **kwargs):
        user = self.get_object()
        self._validate_user_deactivation(request, user, action_name="حذف")
        from .audit_utils import log_audit_event
        try:
            log_audit_event(
                module='users',
                action='DELETE',
                severity='critical',
                target_model='CustomUser',
                target_object_id=user.id,
                target_repr=f"{user.username} ({user.first_name} {user.last_name})",
                user=request.user if request.user.is_authenticated else None,
                before_state={'username': user.username, 'first_name': user.first_name, 'last_name': user.last_name, 'is_active': user.is_active}
            )
        except Exception:
            pass
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=['patch'])
    def toggle_status(self, request, pk=None):
        user = self.get_object()
        from rest_framework.exceptions import ValidationError
        from .audit_utils import log_audit_event
        
        # If user is currently active and is being deactivated
        if user.is_active:
            self._validate_user_deactivation(request, user, action_name="غیرفعال‌سازی")
                    
        old_active = user.is_active
        user.is_active = not user.is_active
        user.save()

        try:
            log_audit_event(
                module='users',
                action='UPDATE',
                severity='critical' if not user.is_active else 'warning',
                target_model='CustomUser',
                target_object_id=user.id,
                target_repr=f"{user.username} ({'فعال‌سازی' if user.is_active else 'غیرفعال‌سازی'})",
                user=request.user if request.user.is_authenticated else None,
                before_state={'is_active': old_active},
                after_state={'is_active': user.is_active}
            )
        except Exception:
            pass

        return Response({'status': 'success', 'is_active': user.is_active})

    @action(detail=False, methods=['post'])
    def change_password(self, request):
        user = request.user
        old_password = request.data.get('old_password')
        new_password = request.data.get('new_password')
        
        if not user.check_password(old_password):
            return Response({'error': 'رمز عبور فعلی نادرست است.'}, status=400)
            
        if new_password == '123456':
            return Response({'error': 'استفاده از رمز عبور پیش‌فرض (123456) مجاز نیست.'}, status=400)
            
        from django.utils import timezone
        user.set_password(new_password)
        user.requires_password_change = False
        user.password_changed_at = timezone.now()
        user.save()

        from .audit_utils import log_audit_event
        log_audit_event(
            user=user,
            module='users',
            action='UPDATE',
            severity='warning',
            target_model='CustomUser',
            target_object_id=user.id,
            target_repr=f"تغییر رمز عبور کاربر {user.username}",
            details={'reason': 'تغییر کلمه عبور توسط خود کاربر'},
            ip_address=getattr(request, 'META', {}).get('REMOTE_ADDR')
        )
        
        return Response({'success': True, 'message': 'رمز عبور با موفقیت تغییر یافت.'})

    @action(detail=True, methods=['post'])
    def admin_reset_password(self, request, pk=None):
        from django.utils import timezone
        user = self.get_object()
        user.set_password('123456')
        user.requires_password_change = True
        user.password_changed_at = timezone.now()
        user.save()

        from .audit_utils import log_audit_event
        log_audit_event(
            user=request.user,
            module='users',
            action='UPDATE',
            severity='warning',
            target_model='CustomUser',
            target_object_id=user.id,
            target_repr=f"بازنشانی رمز عبور کاربر {user.username} توسط مدیر",
            details={'target_user_id': user.id, 'target_username': user.username},
            ip_address=getattr(request, 'META', {}).get('REMOTE_ADDR')
        )
            
        return Response({'success': True, 'message': 'رمز عبور با موفقیت به مقدار پیش‌فرض تغییر یافت و کاربر باید دوباره لاگین کند.'})

    @action(detail=False, methods=['post'])
    def update_preferences(self, request):
        user = request.user
        prefs = request.data.get('preferences', {})
        if isinstance(prefs, dict):
            # Update specific keys instead of overriding completely
            if not isinstance(user.ui_preferences, dict):
                user.ui_preferences = {}
            user.ui_preferences.update(prefs)
            user.save()
            return Response({'status': 'success', 'preferences': user.ui_preferences})
        return Response({'error': 'Invalid preferences format'}, status=400)

    @action(detail=False, methods=['post', 'delete'], url_path='me/avatar')
    def my_avatar(self, request):
        """Update or delete the authenticated user's own avatar."""
        from .avatar_utils import process_and_optimize_avatar, delete_user_avatar_file
        user = request.user
        
        if request.method == 'DELETE':
            delete_user_avatar_file(user)
            user.avatar = None
            user.save()
            return Response({'success': True, 'avatar': None, 'message': 'تصویر پروفایل با موفقیت حذف شد.'})
            
        avatar_file = request.FILES.get('avatar')
        if not avatar_file:
            return Response({'error': 'فایل تصویری ارسال نشده است.'}, status=400)
            
        try:
            optimized_file = process_and_optimize_avatar(avatar_file)
            delete_user_avatar_file(user)
            user.avatar = optimized_file
            user.save()
            return Response({
                'success': True,
                'avatar': user.avatar.url,
                'message': 'تصویر پروفایل با موفقیت بروزرسانی شد.'
            })
        except Exception as e:
            return Response({'error': str(e)}, status=400)

    @action(detail=True, methods=['post', 'delete'], url_path='avatar')
    def user_avatar(self, request, pk=None):
        """Update or delete avatar for a specific user (admin action)."""
        from .avatar_utils import process_and_optimize_avatar, delete_user_avatar_file
        user = self.get_object()
        
        if request.method == 'DELETE':
            delete_user_avatar_file(user)
            user.avatar = None
            user.save()
            return Response({'success': True, 'avatar': None, 'message': 'تصویر کاربر با موفقیت حذف شد.'})
            
        avatar_file = request.FILES.get('avatar')
        if not avatar_file:
            return Response({'error': 'فایل تصویری ارسال نشده است.'}, status=400)
            
        try:
            optimized_file = process_and_optimize_avatar(avatar_file)
            delete_user_avatar_file(user)
            user.avatar = optimized_file
            user.save()
            return Response({
                'success': True,
                'avatar': user.avatar.url,
                'message': 'تصویر کاربر با موفقیت بروزرسانی شد.'
            })
        except Exception as e:
            return Response({'error': str(e)}, status=400)

    # ── Excel Import/Export Actions ──────────────────────────────────
    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Download all users as an Excel file."""
        from .excel_utils import generate_users_excel
        queryset = self.get_queryset()
        return generate_users_excel(queryset)

    @action(detail=False, methods=['get'])
    def download_template(self, request):
        """Download an empty Excel template with sample data."""
        from .excel_utils import generate_users_template
        return generate_users_template()

    @action(detail=False, methods=['post'])
    def import_excel(self, request):
        """Upload an Excel file and bulk-create or update users."""
        from .excel_utils import parse_users_excel
        file = request.FILES.get('file')
        if not file:
            return Response({'success': False, 'errors': [{'row': 0, 'field': 'file', 'message': 'فایلی انتخاب نشده است.'}]}, status=400)

        if not file.name.endswith('.xlsx'):
            return Response({'success': False, 'errors': [{'row': 0, 'field': 'file', 'message': 'فقط فایل‌های با فرمت xlsx پشتیبانی می‌شوند.'}]}, status=400)

        update_existing = request.POST.get('update_existing') in ('true', 'True', True, '1', 1) or request.data.get('update_existing') in ('true', 'True', True, '1', 1)
        result = parse_users_excel(file, update_existing=update_existing)

        # If there are file-level errors (row=0), return immediately
        file_errors = [e for e in result['errors'] if e['row'] == 0]
        if file_errors:
            return Response({'success': False, 'summary': {'total_rows': 0, 'created': 0, 'updated': 0, 'skipped': 0}, 'errors': file_errors}, status=400)

        created_count = 0
        updated_count = 0
        for row_data in result['valid_rows']:
            is_update = row_data.pop('is_update', False)
            user_id = row_data.pop('user_id', None)
            roles = row_data.pop('roles', [])
            warehouses = row_data.pop('warehouses', [])

            if is_update and user_id:
                user = CustomUser.objects.filter(id=user_id).first()
                if user:
                    for k, v in row_data.items():
                        setattr(user, k, v)
                    if request.user and request.user.is_authenticated:
                        user.modified_by = request.user
                    user.save()
                    if roles:
                        user.groups.set(roles)
                    if warehouses:
                        user.assigned_warehouses.set(warehouses)
                    updated_count += 1
                    continue

            password = row_data.get('national_code') or '123456'
            user = CustomUser(**row_data)
            user.set_password(password)
            user.requires_password_change = True
            if request.user and request.user.is_authenticated:
                user.created_by = request.user
            user.save()

            if roles:
                user.groups.set(roles)
            if warehouses:
                user.assigned_warehouses.set(warehouses)
            created_count += 1

        total_rows = created_count + updated_count + len(result['errors'])
        from .audit_utils import log_audit_event
        log_audit_event(
            user=request.user,
            module='users',
            action='IMPORT',
            severity='info',
            target_model='CustomUser',
            target_repr=f"بارگذاری اکسل پرسنل ({created_count} ایجاد، {updated_count} ویرایش)",
            details={'total_rows': total_rows, 'created': created_count, 'updated': updated_count, 'skipped': len(result['errors'])},
            ip_address=getattr(request, 'META', {}).get('REMOTE_ADDR')
        )

        return Response({
            'success': True,
            'summary': {
                'total_rows': total_rows,
                'created': created_count,
                'updated': updated_count,
                'skipped': len(result['errors'])
            },
            'errors': result['errors']
        })

    @action(detail=False, methods=['get'], url_path='verify_card')
    def verify_card(self, request):
        """Public verification endpoint for personnel ID card scanning."""
        code = request.query_params.get('code', '').strip()
        user_id = None
        if code.upper().startswith('EMP-'):
            try:
                user_id = int(code[4:]) - 1000
            except ValueError:
                pass
        elif code.isdigit():
            user_id = int(code)
            
        user = None
        if user_id is not None and user_id > 0:
            user = CustomUser.objects.filter(id=user_id).first()
        if not user and code:
            user = CustomUser.objects.filter(national_code=code).first() or CustomUser.objects.filter(username=code).first()
            
        if not user:
            return Response({'valid': False, 'message': 'پرسنل با این شناسه در سامانه یافت نشد.'}, status=404)
            
        roles = []
        for g in user.groups.all():
            try:
                cr = g.customrole
                roles.append({'id': cr.id, 'title': cr.title or cr.name, 'color': cr.color or '#4f46e5'})
            except Exception:
                roles.append({'id': g.id, 'title': g.name, 'color': '#4f46e5'})
                
        wh_names = list(user.assigned_warehouses.values_list('name', flat=True))
        
        avatar_url = user.avatar.url if user.avatar else None
        
        return Response({
            'valid': True,
            'is_active': user.is_active,
            'id': user.id,
            'personnel_code': f"EMP-{1000 + user.id}",
            'first_name': user.first_name,
            'last_name': user.last_name,
            'national_code': user.national_code or '---',
            'phone_number': user.phone_number or '---',
            'operational_zone': user.operational_zone or 'انبار مرکزی',
            'company': user.company or 'فارس عــالیش',
            'avatar': avatar_url,
            'blood_type': user.blood_type or 'O+',
            'emergency_contact': user.emergency_contact or user.phone_number or '۰۲۱-۸۸۹۹۰۰۱۱',
            'roles': roles if roles else [{'id': 0, 'title': 'پرسنل عملیات انبار', 'color': '#4f46e5'}],
            'assigned_warehouses': wh_names if wh_names else ['انبار مرکزی']
        })

class CustomRoleViewSet(DeleteImpactMixin, viewsets.ModelViewSet):
    queryset = CustomRole.objects.all()
    serializer_class = CustomRoleSerializer
    pagination_class = None

    def get_permissions(self):
        from .permissions import HasMenuAccess
        return [HasMenuAccess('perm_usr_role')]

    def get_queryset(self):
        return CustomRole.objects.all()

    def _check_sensitive_permissions(self, permissions):
        user = self.request.user
        if not user or not user.is_superuser:
            from .models import SENSITIVE_PERMISSION_CODENAMES
            if permissions:
                if hasattr(permissions, 'values_list'):
                    assigned_codenames = set(permissions.values_list('codename', flat=True))
                else:
                    assigned_codenames = set(
                        Permission.objects.filter(id__in=[p.id if hasattr(p, 'id') else p for p in permissions])
                        .values_list('codename', flat=True)
                    )
                if assigned_codenames.intersection(SENSITIVE_PERMISSION_CODENAMES):
                    from rest_framework.exceptions import PermissionDenied
                    raise PermissionDenied("تنها مدیر ارشد سامانه (Superuser) مجاز به اعطا یا تغییر دسترسی‌های حساس و بحرانی است.")

    def perform_create(self, serializer):
        permissions = serializer.validated_data.get('permissions', [])
        self._check_sensitive_permissions(permissions)
        instance = serializer.save()
        from .audit_utils import log_audit_event
        log_audit_event(
            user=self.request.user,
            module='users',
            action='CREATE',
            severity='info',
            target_model='CustomRole',
            target_object_id=instance.id,
            target_repr=f"ایجاد نقش {instance.title or instance.name}",
            details={'role_name': instance.name, 'title': instance.title, 'permissions_count': len(permissions)},
            ip_address=getattr(self.request, 'META', {}).get('REMOTE_ADDR')
        )

    def perform_update(self, serializer):
        permissions = serializer.validated_data.get('permissions', None)
        if permissions is not None:
            self._check_sensitive_permissions(permissions)
        instance = serializer.save()
        from .audit_utils import log_audit_event
        log_audit_event(
            user=self.request.user,
            module='users',
            action='UPDATE',
            severity='info',
            target_model='CustomRole',
            target_object_id=instance.id,
            target_repr=f"ویرایش نقش {instance.title or instance.name}",
            details={'role_name': instance.name, 'title': instance.title},
            ip_address=getattr(self.request, 'META', {}).get('REMOTE_ADDR')
        )

    def perform_destroy(self, instance):
        role_id = instance.id
        role_repr = f"حذف نقش {instance.title or instance.name}"
        super().perform_destroy(instance)
        from .audit_utils import log_audit_event
        log_audit_event(
            user=self.request.user,
            module='users',
            action='DELETE',
            severity='warning',
            target_model='CustomRole',
            target_object_id=role_id,
            target_repr=role_repr,
            details={'role_id': role_id},
            ip_address=getattr(self.request, 'META', {}).get('REMOTE_ADDR')
        )

    # ── Excel Import/Export Actions ──────────────────────────────────
    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Download all roles as an Excel file."""
        from .roles_excel_utils import generate_roles_excel
        queryset = self.get_queryset()
        return generate_roles_excel(queryset)

    @action(detail=False, methods=['get'])
    def download_template(self, request):
        """Download an empty Excel template with sample data."""
        from .roles_excel_utils import generate_roles_template
        return generate_roles_template()

    @action(detail=False, methods=['post'])
    def import_excel(self, request):
        """Upload an Excel file and bulk-create or update roles."""
        from .roles_excel_utils import parse_roles_excel
        file = request.FILES.get('file')
        if not file:
            return Response({'success': False, 'errors': [{'row': 0, 'field': 'file', 'message': 'فایلی انتخاب نشده است.'}]}, status=400)

        if not file.name.endswith('.xlsx'):
            return Response({'success': False, 'errors': [{'row': 0, 'field': 'file', 'message': 'فقط فایل‌های با فرمت xlsx پشتیبانی می‌شوند.'}]}, status=400)

        update_existing = request.POST.get('update_existing') in ('true', 'True', True, '1', 1) or request.data.get('update_existing') in ('true', 'True', True, '1', 1)
        result = parse_roles_excel(file, update_existing=update_existing)

        file_errors = [e for e in result['errors'] if e['row'] == 0]
        if file_errors:
            return Response({'success': False, 'summary': {'total_rows': 0, 'created': 0, 'updated': 0, 'skipped': 0}, 'errors': file_errors}, status=400)

        created_count = 0
        updated_count = 0
        # Two-pass creation: first create/update roles without parents, then set parents
        processed_roles = {}
        for row_data in result['valid_rows']:
            is_update = row_data.pop('is_update', False)
            role_id = row_data.pop('role_id', None)
            permissions = row_data.pop('permissions', [])
            parent_name = row_data.pop('parent_name', None)

            if is_update and role_id:
                role = CustomRole.objects.filter(id=role_id).first()
                if role:
                    role.title = row_data['title']
                    role.color = row_data['color']
                    role.save()
                    if permissions:
                        role.permissions.set(permissions)
                    processed_roles[row_data['name']] = {'role': role, 'parent_name': parent_name}
                    updated_count += 1
                    continue

            role = CustomRole(name=row_data['name'], title=row_data['title'], color=row_data['color'])
            role.save()
            if permissions:
                role.permissions.set(permissions)
            processed_roles[row_data['name']] = {'role': role, 'parent_name': parent_name}
            created_count += 1

        # Second pass: resolve parents
        for name, info in processed_roles.items():
            if info['parent_name']:
                parent = CustomRole.objects.filter(name__iexact=info['parent_name']).first()
                if not parent:
                    parent = CustomRole.objects.filter(title__iexact=info['parent_name']).first()
                if parent and parent.id != info['role'].id:
                    info['role'].parent = parent
                    info['role'].save()

        total_rows = created_count + updated_count + len(result['errors'])
        from .audit_utils import log_audit_event
        log_audit_event(
            user=request.user,
            module='users',
            action='IMPORT',
            severity='info',
            target_model='CustomRole',
            target_repr=f"بارگذاری اکسل نقش‌ها ({created_count} ایجاد، {updated_count} ویرایش)",
            details={'total_rows': total_rows, 'created': created_count, 'updated': updated_count, 'skipped': len(result['errors'])},
            ip_address=getattr(request, 'META', {}).get('REMOTE_ADDR')
        )

        return Response({
            'success': True,
            'summary': {
                'total_rows': total_rows,
                'created': created_count,
                'updated': updated_count,
                'skipped': len(result['errors'])
            },
            'errors': result['errors']
        })



class PermissionViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Permission.objects.filter(content_type__model__in=['customuser', 'group', 'warehouse', 'record'])
    serializer_class = PermissionSerializer
    pagination_class = None

    def get_permissions(self):
        from .permissions import HasMenuAccess
        return [HasMenuAccess('perm_usr_role')]


from rest_framework.permissions import IsAuthenticated
from .models import UserTableViewState
from .serializers import UserTableViewStateSerializer

class UserTableViewStateViewSet(viewsets.ModelViewSet):
    serializer_class = UserTableViewStateSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = None

    def get_queryset(self):
        # Only return views for the current user
        qs = UserTableViewState.objects.filter(user=self.request.user)
        table_name = self.request.query_params.get('table_name')
        if table_name:
            qs = qs.filter(table_name=table_name)
        return qs

    @action(detail=True, methods=['post'])
    def set_last_selected(self, request, pk=None):
        view_state = self.get_object()
        
        # Reset others for the same table
        UserTableViewState.objects.filter(
            user=request.user, 
            table_name=view_state.table_name
        ).update(is_last_selected=False)
        
        # Set this one
        view_state.is_last_selected = True
        view_state.save()
        
        return Response({'status': 'success', 'message': 'نمای انتخاب شده با موفقیت ذخیره شد.'})


import csv
from io import StringIO
from django.http import HttpResponse
from django.utils.timezone import now, timedelta
from rest_framework.views import APIView
from rest_framework.pagination import PageNumberPagination
from .models import UserLoginLog, AuditLog
from .serializers import UserLoginLogSerializer, AuditLogSerializer, AuditLogListSerializer
from .audit_utils import log_login_event

from django.utils import timezone
from django.utils.dateparse import parse_datetime, parse_date
from datetime import datetime

def _parse_query_date(date_str, is_end_of_day=False):
    if not date_str:
        return None
    try:
        s = str(date_str).strip()
        if len(s) == 10 and '-' in s:
            d = parse_date(s)
            if d:
                time_val = datetime.max.time() if is_end_of_day else datetime.min.time()
                dt = datetime.combine(d, time_val)
                return timezone.make_aware(dt) if timezone.is_naive(dt) else dt
        dt = parse_datetime(s)
        if not dt and 'Z' in s:
            try:
                dt = datetime.fromisoformat(s.replace('Z', '+00:00'))
            except Exception:
                dt = None
        if not dt:
            d = parse_date(s[:10])
            if d:
                time_val = datetime.max.time() if is_end_of_day else datetime.min.time()
                dt = datetime.combine(d, time_val)
        if dt:
            if timezone.is_naive(dt):
                dt = timezone.make_aware(dt)
            return dt
    except Exception:
        pass
    return None

class StandardLogPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 500

class LogoutView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        user = request.user
        log_login_event(
            username=user.username,
            status='LOGOUT',
            user=user
        )
        return Response({'detail': 'خروج با موفقیت ثبت شد.'})

def get_db_total_storage_info():
    """
    استعلام حجم فیزیکی کل دیتابیس فعلی بر روی دیسک با PostgreSQL
    """
    from django.db import connection
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT pg_database_size(current_database())")
            row = cursor.fetchone()
            if row and row[0] is not None:
                return int(row[0])
    except Exception:
        pass
    return None

def get_table_storage_info(table_name):
    """
    استعلام حجم فیزیکی دیسک برای یک جدول با پشتیبانی از PostgreSQL
    """
    from django.db import connection
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT pg_total_relation_size(%s)", [table_name])
            row = cursor.fetchone()
            if row and row[0] is not None:
                return int(row[0])
    except Exception:
        pass
    return None

def format_bytes_to_human(size_bytes):
    """
    تبدیل بایت به فرمت خوانا (B, KB, MB, GB)
    """
    if size_bytes is None or size_bytes < 0:
        return '—'
    if size_bytes < 1024:
        return f"{size_bytes} B"
    elif size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.1f} KB"
    elif size_bytes < 1024 * 1024 * 1024:
        return f"{size_bytes / (1024 * 1024):.2f} MB"
    else:
        return f"{size_bytes / (1024 * 1024 * 1024):.2f} GB"

def build_log_storage_payload(audit_count=0, login_count=0):
    """
    تولید ساختار متادیتای حجم کل پایگاه داده و تفکیک درصدی لاگ‌ها
    """
    db_total_bytes = get_db_total_storage_info()
    audit_table_size = get_table_storage_info('accounts_auditlog')
    login_table_size = get_table_storage_info('accounts_userloginlog')

    # Fallback در صورت دسترسی نداشتن به توابع سیستمی
    if not audit_table_size and audit_count > 0:
        audit_table_size = audit_count * 2560
    elif not audit_table_size:
        audit_table_size = 0

    if not login_table_size and login_count > 0:
        login_table_size = login_count * 1536
    elif not login_table_size:
        login_table_size = 0

    total_logs_size = audit_table_size + login_table_size

    if not db_total_bytes or db_total_bytes <= 0:
        db_total_bytes = max(total_logs_size * 5, 15 * 1024 * 1024)

    audit_percent = round((audit_table_size / db_total_bytes) * 100, 2) if db_total_bytes > 0 else 0.0
    login_percent = round((login_table_size / db_total_bytes) * 100, 2) if db_total_bytes > 0 else 0.0
    total_logs_percent = round((total_logs_size / db_total_bytes) * 100, 2) if db_total_bytes > 0 else 0.0

    return {
        'db_total_bytes': db_total_bytes,
        'db_total_formatted': format_bytes_to_human(db_total_bytes),
        'audit_bytes': audit_table_size,
        'audit_formatted': format_bytes_to_human(audit_table_size),
        'audit_percent': audit_percent,
        'login_bytes': login_table_size,
        'login_formatted': format_bytes_to_human(login_table_size),
        'login_percent': login_percent,
        'total_logs_bytes': total_logs_size,
        'total_logs_formatted': format_bytes_to_human(total_logs_size),
        'total_logs_percent': total_logs_percent,
        'avg_row_size_kb': round((audit_table_size / (audit_count * 1024)), 2) if (audit_table_size and audit_count > 0) else 2.5
    }

def format_shamsi_datetime(dt):
    """
    تبدیل شیء تاریخ/زمان به رشته فرمت‌بندی‌شده شمسی (YYYY/MM/DD HH:MM:SS)
    """
    if not dt:
        return "—"
    try:
        import jdatetime
        from django.utils import timezone
        if timezone.is_aware(dt):
            dt = timezone.localtime(dt)
        jdt = jdatetime.datetime.fromgregorian(datetime=dt)
        return jdt.strftime('%Y/%m/%d %H:%M:%S')
    except Exception:
        return dt.strftime('%Y-%m-%d %H:%M:%S') if hasattr(dt, 'strftime') else str(dt)


class UserLoginLogViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = UserLoginLog.objects.all().select_related('user').prefetch_related('user__groups__customrole').order_by('-created_at')
    serializer_class = UserLoginLogSerializer
    pagination_class = StandardLogPagination

    def get_permissions(self):
        from .permissions import HasMenuAccess
        return [HasMenuAccess('perm_sys_logs') | HasMenuAccess('perm_sys_audit_export') | HasMenuAccess('perm_sys_purge_logs')]

    def get_queryset(self):
        req = getattr(self, 'request', None)
        qs = super().get_queryset()
        params = getattr(req, 'query_params', getattr(req, 'GET', {})) if req else {}


        username = params.get('username')
        if username:
            qs = qs.filter(username_attempted__icontains=username)

        status = params.get('status')
        if status:
            s_clean = str(status).strip().upper()
            if s_clean == 'FAILED':
                qs = qs.filter(status__startswith='FAILED')
            else:
                qs = qs.filter(status=s_clean)

        ip_address = params.get('ip_address')
        if ip_address:
            qs = qs.filter(ip_address__icontains=ip_address)

        user_id = params.get('user')
        if user_id:
            qs = qs.filter(user_id=user_id)

        search = params.get('search')
        if search:
            from django.db.models import Q
            qs = qs.filter(
                Q(username_attempted__icontains=search) |
                Q(user__first_name__icontains=search) |
                Q(user__last_name__icontains=search) |
                Q(ip_address__icontains=search) |
                Q(failure_reason__icontains=search)
            )

        from_date = _parse_query_date(params.get('from_date'), is_end_of_day=False)
        if from_date:
            qs = qs.filter(created_at__gte=from_date)

        to_date = _parse_query_date(params.get('to_date'), is_end_of_day=True)
        if to_date:
            qs = qs.filter(created_at__lte=to_date)

        return qs

    @action(detail=False, methods=['get'])
    def stats(self, request):
        from django.db.models import Count
        now_time = now()
        today_start = now_time.replace(hour=0, minute=0, second=0, microsecond=0)
        last_24h = now_time - timedelta(hours=24)

        base_qs = self.filter_queryset(self.get_queryset())
        total_logins = base_qs.count()
        logins_24h = base_qs.filter(created_at__gte=last_24h).count()
        failed_24h = base_qs.filter(created_at__gte=last_24h).exclude(status__in=['SUCCESS', 'LOGOUT']).count()
        success_24h = base_qs.filter(created_at__gte=last_24h, status='SUCCESS').count()

        status_breakdown = dict(
            base_qs.filter(created_at__gte=last_24h)
            .values_list('status')
            .annotate(c=Count('id'))
        )

        audit_total = AuditLog.objects.count()
        storage_info = build_log_storage_payload(audit_count=audit_total, login_count=total_logins)

        return Response({
            'total_all_time': total_logins,
            'logins_24h': logins_24h,
            'success_24h': success_24h,
            'failed_24h': failed_24h,
            'status_breakdown': status_breakdown,
            'storage': storage_info
        })

    @action(detail=False, methods=['post'], permission_classes=[permissions.IsAuthenticated])
    def heartbeat(self, request):
        """
        ثبت حضور فعال روزانه کاربر (یک‌بار در هر روز تقویمی)
        """
        user = request.user
        if not user or not user.is_authenticated:
            return Response({'detail': 'احراز هویت نشده'}, status=status.HTTP_401_UNAUTHORIZED)
        
        now_time = now()
        today_start = now_time.replace(hour=0, minute=0, second=0, microsecond=0)
        
        # بررسی اینکه آیا امروز قبلاً رکوردی برای کاربر ثبت شده است یا خیر
        already_logged_today = UserLoginLog.objects.filter(
            user=user,
            created_at__gte=today_start
        ).exists()
        
        if not already_logged_today:
            from .middleware import get_client_ip
            device_model = request.data.get('device_model') or request.headers.get('Sec-Ch-Ua-Model')
            log_login_event(
                username=user.username,
                status='DAILY_ACTIVE',
                user=user,
                ip_address=get_client_ip(request),
                user_agent=request.META.get('HTTP_USER_AGENT', ''),
                device_model=device_model,
                failure_reason=None
            )
            return Response({'status': 'logged', 'message': 'لاگ حضور روزانه با موفقیت ثبت شد.'})
        
        return Response({'status': 'already_recorded', 'message': 'حضور کاربر در تاریخ امروز قبلاً ثبت شده است.'})

    def _check_export_permission(self, request):
        user = getattr(request, 'user', None)
        if not (user and user.is_authenticated and (
            user.is_superuser or
            user.has_perm('accounts.perm_sys_audit_export') or
            user.has_perm('accounts.perm_sys_purge_logs') or
            user.has_perm('accounts.perm_sys_logs')
        )):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("تنها کاربران دارای مجوز خروجی ممیزی (perm_sys_audit_export) یا مدیر سیستم مجاز به دریافت گزارش ورودها هستند.")

    @action(detail=False, methods=['get', 'post'])
    def export_excel(self, request):
        self._check_export_permission(request)
        data = request.data if request.method == 'POST' and request.data else request.query_params

        qs = self.get_queryset()
        if request.method == 'POST' and request.data:
            username = data.get('username')
            if username:
                qs = qs.filter(username_attempted__icontains=username)
            status_val = data.get('status')
            if status_val:
                s_clean = str(status_val).strip().upper()
                if s_clean == 'FAILED':
                    qs = qs.filter(status__startswith='FAILED')
                else:
                    qs = qs.filter(status=s_clean)
            ip_address = data.get('ip_address')
            if ip_address:
                qs = qs.filter(ip_address__icontains=ip_address)
            search = data.get('search')
            if search:
                from django.db.models import Q
                qs = qs.filter(
                    Q(username_attempted__icontains=search) |
                    Q(user__first_name__icontains=search) |
                    Q(user__last_name__icontains=search) |
                    Q(ip_address__icontains=search) |
                    Q(failure_reason__icontains=search)
                )
            from_date = _parse_query_date(data.get('from_date'), is_end_of_day=False)
            if from_date:
                qs = qs.filter(created_at__gte=from_date)
            to_date = _parse_query_date(data.get('to_date'), is_end_of_day=True)
            if to_date:
                qs = qs.filter(created_at__lte=to_date)

        export_limit = 10000
        qs = qs[:export_limit]

        all_cols_def = [
            ('index', 'ردیف', 8),
            ('id', 'شناسه', 10),
            ('username_attempted', 'نام کاربری ورودی', 20),
            ('user_name', 'نام و نام خانوادگی', 22),
            ('status', 'وضعیت ورود', 20),
            ('ip_address', 'آدرس آی‌پی', 18),
            ('device_model', 'مدل دستگاه', 20),
            ('user_agent', 'مرورگر / دستگاه', 30),
            ('failure_reason', 'علت شکست', 25),
            ('created_at', 'تاریخ و زمان', 22),
        ]

        req_cols = data.get('columns')
        if req_cols:
            if isinstance(req_cols, str):
                req_cols = [c.strip() for c in req_cols.split(',') if c.strip()]
            selected_cols = [c for c in all_cols_def if c[0] in req_cols]
            if not selected_cols:
                selected_cols = all_cols_def
        else:
            selected_cols = all_cols_def

        format_type = str(data.get('file_format') or data.get('export_format') or data.get('format', 'xlsx')).lower().strip()

        if format_type == 'csv':
            output = StringIO()
            writer = csv.writer(output)
            writer.writerow([c[1] for c in selected_cols])  # ردیف ۱: عنوان فارسی
            writer.writerow([c[0] for c in selected_cols])  # ردیف ۲: نام فیلد دیتابیس
            for idx, item in enumerate(qs, 1):
                user_name = f"{item.user.first_name} {item.user.last_name}".strip() if item.user else "—"
                row_vals = []
                for col_key, _, _ in selected_cols:
                    if col_key == 'index':
                        row_vals.append(idx)
                    elif col_key == 'id':
                        row_vals.append(item.id)
                    elif col_key == 'username_attempted':
                        row_vals.append(item.username_attempted)
                    elif col_key == 'user_name':
                        row_vals.append(user_name)
                    elif col_key == 'status':
                        row_vals.append(item.get_status_display())
                    elif col_key == 'ip_address':
                        row_vals.append(item.ip_address or "—")
                    elif col_key == 'device_model':
                        row_vals.append(item.device_model or "—")
                    elif col_key == 'user_agent':
                        row_vals.append(item.user_agent or "—")
                    elif col_key == 'failure_reason':
                        row_vals.append(item.failure_reason or "—")
                    elif col_key == 'created_at':
                        row_vals.append(format_shamsi_datetime(item.created_at))
                writer.writerow(row_vals)

            response = HttpResponse(output.getvalue().encode('utf-8-sig'), content_type='text/csv; charset=utf-8')
            response['Content-Disposition'] = 'attachment; filename="login_history_report.csv"'
            return response

        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "تاریخچه ورود کاربران"
        ws.views.sheetView[0].rightToLeft = True

        header_font = Font(name='Tahoma', size=10, bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='4F46E5')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        key_font = Font(name='Consolas', size=9, bold=True, color='475569')
        key_fill = PatternFill('solid', fgColor='F1F5F9')
        key_align = Alignment(horizontal='center', vertical='center')

        data_font = Font(name='Tahoma', size=9)
        data_align_center = Alignment(horizontal='center', vertical='center')
        data_align_right = Alignment(horizontal='right', vertical='center')
        zebra_fill = PatternFill('solid', fgColor='F8FAFC')
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )

        # ردیف ۱: عناوین فارسی
        header_titles = [c[1] for c in selected_cols]
        ws.append(header_titles)
        ws.row_dimensions[1].height = 28
        for col_idx in range(1, len(selected_cols) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # ردیف ۲: نام‌های دیتابیسی ستون‌ها
        db_keys = [c[0] for c in selected_cols]
        ws.append(db_keys)
        ws.row_dimensions[2].height = 20
        for col_idx in range(1, len(selected_cols) + 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.font = key_font
            cell.fill = key_fill
            cell.alignment = key_align
            cell.border = thin_border

        # سطرهای داده از ردیف ۳
        for idx, item in enumerate(qs, 1):
            user_name = f"{item.user.first_name} {item.user.last_name}".strip() if item.user else "—"
            row_vals = []
            for col_key, _, _ in selected_cols:
                if col_key == 'index':
                    row_vals.append(idx)
                elif col_key == 'id':
                    row_vals.append(item.id)
                elif col_key == 'username_attempted':
                    row_vals.append(item.username_attempted)
                elif col_key == 'user_name':
                    row_vals.append(user_name)
                elif col_key == 'status':
                    row_vals.append(item.get_status_display())
                elif col_key == 'ip_address':
                    row_vals.append(item.ip_address or "—")
                elif col_key == 'device_model':
                    row_vals.append(item.device_model or "—")
                elif col_key == 'user_agent':
                    row_vals.append((item.user_agent or "—")[:120])
                elif col_key == 'failure_reason':
                    row_vals.append(item.failure_reason or "—")
                elif col_key == 'created_at':
                    row_vals.append(format_shamsi_datetime(item.created_at))

            ws.append(row_vals)
            row_num = idx + 2
            ws.row_dimensions[row_num].height = 22
            is_zebra = (idx % 2 == 0)

            for col_idx, (col_key, _, _) in enumerate(selected_cols, 1):
                c = ws.cell(row=row_num, column=col_idx)
                c.font = data_font
                c.border = thin_border
                if is_zebra:
                    c.fill = zebra_fill
                if col_key in ('user_name', 'user_agent', 'failure_reason'):
                    c.alignment = data_align_right
                else:
                    c.alignment = data_align_center

        for col_idx, (_, _, width_val) in enumerate(selected_cols, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width_val

        ws.freeze_panes = 'A3'

        import io
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="login_history_report.xlsx"'
        return response

    @action(detail=False, methods=['get'])
    def export_csv(self, request):
        return self.export_excel(request)

    @action(detail=False, methods=['post'])
    def purge(self, request):
        """
        پاکسازی تاریخچه ورود کاربران توسط مدیر سیستم بر اساس فیلترهای زمانی و وضعیت
        """
        user = getattr(request, 'user', None)
        if not (user and user.is_authenticated and (
            user.is_superuser or
            user.has_perm('accounts.perm_sys_purge_logs')
        )):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("تنها مدیر ارشد سامانه یا کاربران دارای مجوز اختصاصی پاکسازی لاگ‌ها (perm_sys_purge_logs) مجاز به حذف لاگ‌های ورود هستند.")

        data = request.data or {}
        dry_run = data.get('dry_run') in (True, 'true', '1', 1)
        
        qs = UserLoginLog.objects.all()

        from_date = _parse_query_date(data.get('from_date'), is_end_of_day=False)
        if from_date:
            qs = qs.filter(created_at__gte=from_date)

        to_date = _parse_query_date(data.get('to_date'), is_end_of_day=True)
        if to_date:
            qs = qs.filter(created_at__lte=to_date)

        status_val = data.get('status')
        if status_val:
            qs = qs.filter(status=status_val)

        username = data.get('username')
        if username:
            qs = qs.filter(username_attempted__icontains=username)

        days = data.get('days')
        if days and str(days).isdigit():
            cutoff = now() - timedelta(days=int(days))
            qs = qs.filter(created_at__lt=cutoff)

        count = qs.count()

        if dry_run:
            return Response({
                'success': True,
                'dry_run': True,
                'count': count,
                'message': f"{count} رکورد لاگ ورود کاربران منطبق با شرایط انتخابی برای پاکسازی است."
            })

        confirm_text = str(data.get('confirm_text', '')).strip()
        if confirm_text != 'PURGE_LOGIN_LOGS_CONFIRM':
            return Response({
                'success': False,
                'error': 'جهت پاکسازی قطعی تاریخچه ورود، عبارت تاییدیه امنیتی PURGE_LOGIN_LOGS_CONFIRM باید به صورت دقیق وارد شود.'
            }, status=status.HTTP_400_BAD_REQUEST)

        if count == 0:
            return Response({
                'success': True,
                'purged_count': 0,
                'message': 'هیچ رکوردی منطبق با فیلترهای انتخابی یافت نشد.'
            })

        batch_size = 5000
        while True:
            batch_ids = list(qs.values_list('id', flat=True)[:batch_size])
            if not batch_ids:
                break
            UserLoginLog.objects.filter(id__in=batch_ids).delete()
            if len(batch_ids) < batch_size:
                break

        from .audit_utils import log_audit_event
        desc_parts = []
        if from_date:
            desc_parts.append(f"از: {from_date.strftime('%Y-%m-%d')}")
        if to_date:
            desc_parts.append(f"تا: {to_date.strftime('%Y-%m-%d')}")
        if status_val:
            desc_parts.append(f"وضعیت: {status_val}")
        if username:
            desc_parts.append(f"کاربر: {username}")
        if days:
            desc_parts.append(f"قدیمی‌تر از {days} روز")
        criteria_str = ' | '.join(desc_parts) if desc_parts else 'کلیه لاگ‌های ورود'

        log_audit_event(
            user=user,
            module='system',
            action='DELETE',
            severity='critical',
            details={
                'description': f"پاکسازی قطعی {count} رکورد از لاگ‌های ورود کاربران ({criteria_str})",
                'purged_count': count,
                'filters': {
                    'from_date': str(from_date) if from_date else None,
                    'to_date': str(to_date) if to_date else None,
                    'status': status_val,
                    'username': username,
                    'days': days
                }
            },
            ip_address=getattr(request, 'META', {}).get('REMOTE_ADDR')
        )

        return Response({
            'success': True,
            'purged_count': count,
            'message': f"{count} رکورد تاریخچه ورود با موفقیت پاکسازی شد."
        })

    @action(detail=False, methods=['get'])
    def locked_users(self, request):
        user = getattr(request, 'user', None)
        if not (user and user.is_authenticated and (
            user.is_superuser or
            user.has_perm('accounts.perm_sys_logs') or
            user.has_perm('accounts.view_wh_audit') or
            user.has_perm('accounts.perm_usr_edit')
        )):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("عدم دسترسی به مشاهده وضعیت کاربران مسدود شده.")

        from axes.models import AccessAttempt
        from django.conf import settings
        failure_limit = getattr(settings, 'AXES_FAILURE_LIMIT', 5)

        attempts = AccessAttempt.objects.all().order_by('-attempt_time')
        results = []
        for att in attempts:
            results.append({
                'username': att.username,
                'ip_address': att.ip_address,
                'failures': att.failures_since_start,
                'is_locked': att.failures_since_start >= failure_limit,
                'attempt_time': att.attempt_time.isoformat() if att.attempt_time else None,
                'user_agent': att.user_agent or '',
            })

        return Response({
            'locked_users': results,
            'total_locked': len([r for r in results if r['is_locked']])
        })

    @action(detail=False, methods=['post'])
    def reset_lockout(self, request):
        user = getattr(request, 'user', None)
        if not (user and user.is_authenticated and (
            user.is_superuser or
            user.has_perm('accounts.perm_usr_edit') or
            user.has_perm('accounts.perm_sys_logs')
        )):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("شما دسترسی لازم برای رفع مسدودیت کاربران را ندارید.")

        data = request.data or {}
        username = str(data.get('username', '')).strip()
        ip_address = str(data.get('ip_address', '')).strip()

        if not username and not ip_address:
            return Response({'error': 'نام کاربری یا آدرس آی‌پی جهت رفع مسدودیت الزامی است.'}, status=status.HTTP_400_BAD_REQUEST)

        from axes.utils import reset
        from axes.models import AccessAttempt

        if username:
            try:
                reset(username=username)
            except Exception:
                pass
            AccessAttempt.objects.filter(username=username).delete()

        if ip_address:
            try:
                reset(ip=ip_address)
            except Exception:
                pass
            AccessAttempt.objects.filter(ip_address=ip_address).delete()

        # Log security audit event
        from .audit_utils import log_audit_event
        from .middleware import get_client_ip
        target_name = f"کاربر «{username}»" if username else f"آی‌پی «{ip_address}»"
        log_audit_event(
            user=user,
            module='users',
            action='UPDATE',
            target_repr=target_name,
            details={'description': f"رفع مسدودیت امنیتی {target_name} و بازنشانی شمارنده تلاش‌های ناموفق سیستم ضد نفوذ (Axes)"},
            severity='warning',
            ip_address=get_client_ip(request)
        )

        return Response({
            'success': True,
            'username': username,
            'ip_address': ip_address,
            'message': f"قفل امنیتی {target_name} با موفقیت بازنشانی شد و کاربر مجدداً مجاز به ورود است."
        })


class AuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = AuditLog.objects.all().select_related('user', 'warehouse').prefetch_related('user__groups__customrole').order_by('-created_at')
    serializer_class = AuditLogSerializer
    pagination_class = StandardLogPagination

    def get_serializer_class(self):
        if self.action == 'list':
            return AuditLogListSerializer
        return AuditLogSerializer

    def get_permissions(self):
        from .permissions import HasMenuAccess
        return [HasMenuAccess('perm_sys_logs') | HasMenuAccess('view_wh_audit') | HasMenuAccess('perm_sys_audit_export') | HasMenuAccess('perm_sys_purge_logs')]

    def get_queryset(self):
        req = getattr(self, 'request', None)
        user = getattr(req, 'user', None) if req else None
        qs = super().get_queryset()
        params = getattr(req, 'query_params', getattr(req, 'GET', {})) if req else {}

        # Warehouse restriction: non-superusers without perm_sys_logs only see assigned warehouses
        if user and user.is_authenticated and not (user.is_superuser or user.has_perm('accounts.perm_sys_logs')):
            assigned_wh_ids = list(user.assigned_warehouses.values_list('id', flat=True))
            qs = qs.filter(warehouse_id__in=assigned_wh_ids)


        module = params.get('module')
        if module:
            qs = qs.filter(module=module)

        action_param = params.get('action_type') or params.get('action')
        if action_param:
            qs = qs.filter(action=action_param)

        severity = params.get('severity')
        if severity:
            qs = qs.filter(severity=severity)

        warehouse_id = params.get('warehouse')
        if warehouse_id and str(warehouse_id).strip().upper() != 'ALL':
            qs = qs.filter(warehouse_id=warehouse_id)

        user_id = params.get('user')
        if user_id:
            qs = qs.filter(user_id=user_id)

        search = params.get('search')
        if search:
            from django.db.models import Q
            qs = qs.filter(
                Q(target_repr__icontains=search) |
                Q(target_object_id__icontains=search) |
                Q(user__username__icontains=search) |
                Q(user__first_name__icontains=search) |
                Q(user__last_name__icontains=search) |
                Q(actor_username__icontains=search) |
                Q(actor_name__icontains=search)
            )

        from_date = _parse_query_date(params.get('from_date'), is_end_of_day=False)
        if from_date:
            qs = qs.filter(created_at__gte=from_date)

        to_date = _parse_query_date(params.get('to_date'), is_end_of_day=True)
        if to_date:
            qs = qs.filter(created_at__lte=to_date)

        return qs

    @action(detail=False, methods=['get'])
    def stats(self, request):
        from django.db.models import Count, Q
        now_time = now()
        today_start = now_time.replace(hour=0, minute=0, second=0, microsecond=0)
        last_24h = now_time - timedelta(hours=24)

        base_qs = self.filter_queryset(self.get_queryset())
        total_logs = base_qs.count()
        logs_24h = base_qs.filter(created_at__gte=last_24h).count()
        critical_24h = base_qs.filter(created_at__gte=last_24h, severity='critical').count()
        rollbacks_24h = base_qs.filter(created_at__gte=last_24h, action='ROLLBACK').count()

        aggregates = base_qs.aggregate(
            total_all_time=Count('id'),
            critical_all_time=Count('id', filter=Q(severity='critical')),
            warning_all_time=Count('id', filter=Q(severity='warning')),
            rollbacks_all_time=Count('id', filter=Q(action='ROLLBACK'))
        )

        warning_24h = base_qs.filter(created_at__gte=last_24h, severity='warning').count()

        module_breakdown = dict(
            base_qs.filter(created_at__gte=last_24h)
            .values_list('module')
            .annotate(c=Count('id'))
        )

        login_total = UserLoginLog.objects.count()
        audit_total = aggregates['total_all_time'] or 0
        storage_info = build_log_storage_payload(audit_count=audit_total, login_count=login_total)

        return Response({
            'total_all_time': aggregates['total_all_time'] or 0,
            'critical_all_time': aggregates['critical_all_time'] or 0,
            'warning_all_time': aggregates['warning_all_time'] or 0,
            'rollbacks_all_time': aggregates['rollbacks_all_time'] or 0,
            'logs_24h': logs_24h,
            'audits_24h': logs_24h,
            'critical_24h': critical_24h,
            'critical_count': critical_24h,
            'warning_24h': warning_24h,
            'warning_count': warning_24h,
            'rollbacks_24h': rollbacks_24h,
            'module_breakdown': module_breakdown,
            'storage': storage_info
        })

    def _check_export_permission(self, request):
        user = getattr(request, 'user', None)
        if not (user and user.is_authenticated and (
            user.is_superuser or
            user.has_perm('accounts.perm_sys_audit_export') or
            user.has_perm('accounts.perm_sys_purge_logs') or
            user.has_perm('accounts.perm_sys_logs')
        )):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("تنها کاربران دارای مجوز خروجی ممیزی (perm_sys_audit_export) یا مدیر سیستم مجاز به دریافت گزارش لاگ‌ها هستند.")

    @action(detail=False, methods=['get', 'post'])
    def export_excel(self, request):
        self._check_export_permission(request)
        data = request.data if request.method == 'POST' and request.data else request.query_params

        qs = self.get_queryset()
        if request.method == 'POST' and request.data:
            module = data.get('module')
            if module:
                qs = qs.filter(module=module)
            action_param = data.get('action_type') or data.get('action')
            if action_param:
                qs = qs.filter(action=action_param)
            severity = data.get('severity')
            if severity:
                qs = qs.filter(severity=severity)
            warehouse_id = data.get('warehouse') or data.get('warehouse_id')
            if warehouse_id and str(warehouse_id) != 'ALL':
                qs = qs.filter(warehouse_id=warehouse_id)
            user_id = data.get('user')
            if user_id:
                qs = qs.filter(user_id=user_id)
            search = data.get('search')
            if search:
                from django.db.models import Q
                qs = qs.filter(
                    Q(target_repr__icontains=search) |
                    Q(target_object_id__icontains=search) |
                    Q(user__username__icontains=search) |
                    Q(user__first_name__icontains=search) |
                    Q(user__last_name__icontains=search) |
                    Q(actor_username__icontains=search) |
                    Q(actor_name__icontains=search)
                )
            from_date = _parse_query_date(data.get('from_date'), is_end_of_day=False)
            if from_date:
                qs = qs.filter(created_at__gte=from_date)
            to_date = _parse_query_date(data.get('to_date'), is_end_of_day=True)
            if to_date:
                qs = qs.filter(created_at__lte=to_date)

        export_limit = 10000
        qs = qs[:export_limit]

        all_cols_def = [
            ('index', 'ردیف', 8),
            ('id', 'شناسه', 10),
            ('user', 'کاربر عامل', 22),
            ('warehouse', 'انبار', 18),
            ('module', 'ماژول', 18),
            ('action', 'نوع عملیات', 16),
            ('severity', 'سطح اهمیت', 14),
            ('target_repr', 'موجودیت هدف', 25),
            ('target_object_id', 'شناسه سند', 14),
            ('changes_summary', 'خلاصه تغییرات', 35),
            ('ip_address', 'آدرس آی‌پی', 18),
            ('created_at', 'تاریخ و زمان', 22),
        ]

        col_alias_map = {
            'user_display': 'user',
            'actor_username': 'user',
            'warehouse_name': 'warehouse',
            'module_display': 'module',
            'action_display': 'action',
            'target': 'target_repr',
            'object_id': 'target_object_id',
            'changes': 'changes_summary',
            'ip': 'ip_address',
            'date': 'created_at',
            'timestamp': 'created_at'
        }

        req_cols = data.get('columns') or request.query_params.get('columns')
        if req_cols:
            if isinstance(req_cols, str):
                req_cols = [c.strip() for c in req_cols.split(',') if c.strip()]
            normalized_cols = set([col_alias_map.get(c, c) for c in req_cols])
            selected_cols = [c for c in all_cols_def if c[0] in normalized_cols or (c[0] == 'index' and ('index' in normalized_cols or 'radif' in normalized_cols))]
            if not selected_cols:
                selected_cols = all_cols_def
        else:
            selected_cols = all_cols_def

        def _get_changes_summary(item):
            if not item.before_state and not item.after_state:
                return "—"
            diffs = []
            before = item.before_state or {}
            after = item.after_state or {}
            all_keys = set(before.keys()) | set(after.keys())
            for k in sorted(all_keys):
                old_v = before.get(k)
                new_v = after.get(k)
                if old_v != new_v:
                    diffs.append(f"{k}: {old_v} ➔ {new_v}")
            return ' | '.join(diffs) if diffs else "بدون تغییر داده‌ای"

        # اگر درخواست CSV باشد
        is_csv = (
            request.query_params.get('format') == 'csv' or
            request.query_params.get('file_format') == 'csv' or
            data.get('file_format') == 'csv' or
            data.get('format') == 'csv' or
            request.path.endswith('export_csv')
        )
        if is_csv:
            import csv
            import io
            output = io.StringIO()
            writer = csv.writer(output)
            # سطر ۱: عناوین فارسی
            writer.writerow([c[1] for c in selected_cols])
            # سطر ۲: نام‌های دیتابیسی
            writer.writerow([c[0] for c in selected_cols])

            for idx, item in enumerate(qs, 1):
                if item.user:
                    u_full = f"{item.user.first_name} {item.user.last_name}".strip()
                    user_display = u_full if u_full else item.user.username
                elif item.actor_name:
                    user_display = f"{item.actor_name} (سابق)"
                elif item.actor_username:
                    user_display = f"{item.actor_username} (سابق)"
                else:
                    user_display = "سیستم"

                wh_name = item.warehouse.name if item.warehouse else "عمومی / سیستم"
                row_vals = []
                for col_key, _, _ in selected_cols:
                    if col_key == 'index':
                        row_vals.append(idx)
                    elif col_key == 'id':
                        row_vals.append(item.id)
                    elif col_key in ('user', 'user_display'):
                        row_vals.append(user_display)
                    elif col_key in ('warehouse', 'warehouse_name'):
                        row_vals.append(wh_name)
                    elif col_key in ('module', 'module_display'):
                        row_vals.append(item.get_module_display() if hasattr(item, 'get_module_display') else item.module)
                    elif col_key in ('action', 'action_display'):
                        row_vals.append(item.get_action_display() if hasattr(item, 'get_action_display') else item.action)
                    elif col_key == 'severity':
                        row_vals.append(item.get_severity_display() if hasattr(item, 'get_severity_display') else item.severity)
                    elif col_key == 'target_repr':
                        row_vals.append(item.target_repr or "—")
                    elif col_key == 'target_object_id':
                        row_vals.append(item.target_object_id or "—")
                    elif col_key == 'changes_summary':
                        row_vals.append(_get_changes_summary(item))
                    elif col_key == 'ip_address':
                        row_vals.append(item.ip_address or "—")
                    elif col_key == 'created_at':
                        row_vals.append(format_shamsi_datetime(item.created_at))
                writer.writerow(row_vals)

            response = HttpResponse(output.getvalue().encode('utf-8-sig'), content_type='text/csv; charset=utf-8')
            response['Content-Disposition'] = 'attachment; filename="audit_logs_report.csv"'
            return response

        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "گزارش رهگیری تغییرات"
        ws.views.sheetView[0].rightToLeft = True

        header_font = Font(name='Tahoma', size=10, bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='4F46E5')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        key_font = Font(name='Consolas', size=9, bold=True, color='475569')
        key_fill = PatternFill('solid', fgColor='F1F5F9')
        key_align = Alignment(horizontal='center', vertical='center')

        data_font = Font(name='Tahoma', size=9)
        data_align_center = Alignment(horizontal='center', vertical='center')
        data_align_right = Alignment(horizontal='right', vertical='center')
        zebra_fill = PatternFill('solid', fgColor='F8FAFC')
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )

        # سطر ۱: عناوین نمایشی فارسی
        ws.row_dimensions[1].height = 28
        for col_idx, (_, persian_title, _) in enumerate(selected_cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=persian_title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # سطر ۲: کلیدهای سیستمی لاتین
        ws.row_dimensions[2].height = 20
        for col_idx, (col_key, _, _) in enumerate(selected_cols, 1):
            cell = ws.cell(row=2, column=col_idx, value=col_key)
            cell.font = key_font
            cell.fill = key_fill
            cell.alignment = key_align
            cell.border = thin_border

        # سطرهای داده از ردیف ۳
        for idx, item in enumerate(qs, 1):
            if item.user:
                u_full = f"{item.user.first_name} {item.user.last_name}".strip()
                user_display = u_full if u_full else item.user.username
            elif item.actor_name:
                user_display = f"{item.actor_name} (سابق)"
            elif item.actor_username:
                user_display = f"{item.actor_username} (سابق)"
            else:
                user_display = "سیستم"

            wh_name = item.warehouse.name if item.warehouse else "عمومی / سیستم"
            row_vals = []
            for col_key, _, _ in selected_cols:
                if col_key == 'index':
                    row_vals.append(idx)
                elif col_key == 'id':
                    row_vals.append(item.id)
                elif col_key in ('user', 'user_display'):
                    row_vals.append(user_display)
                elif col_key in ('warehouse', 'warehouse_name'):
                    row_vals.append(wh_name)
                elif col_key in ('module', 'module_display'):
                    row_vals.append(item.get_module_display() if hasattr(item, 'get_module_display') else item.module)
                elif col_key in ('action', 'action_display'):
                    row_vals.append(item.get_action_display() if hasattr(item, 'get_action_display') else item.action)
                elif col_key == 'severity':
                    row_vals.append(item.get_severity_display() if hasattr(item, 'get_severity_display') else item.severity)
                elif col_key == 'target_repr':
                    row_vals.append(item.target_repr or "—")
                elif col_key == 'target_object_id':
                    row_vals.append(item.target_object_id or "—")
                elif col_key == 'changes_summary':
                    summary_txt = _get_changes_summary(item)
                    row_vals.append(summary_txt[:1000] if len(summary_txt) > 1000 else summary_txt)
                elif col_key == 'ip_address':
                    row_vals.append(item.ip_address or "—")
                elif col_key == 'created_at':
                    row_vals.append(format_shamsi_datetime(item.created_at))

            ws.append(row_vals)
            row_num = idx + 2
            ws.row_dimensions[row_num].height = 22
            is_zebra = (idx % 2 == 0)

            for col_idx, (col_key, _, _) in enumerate(selected_cols, 1):
                c = ws.cell(row=row_num, column=col_idx)
                c.font = data_font
                c.border = thin_border
                if is_zebra:
                    c.fill = zebra_fill
                if col_key in ('user_display', 'warehouse_name', 'target_repr', 'changes_summary'):
                    c.alignment = data_align_right
                else:
                    c.alignment = data_align_center

        for col_idx, (_, _, width_val) in enumerate(selected_cols, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width_val

        ws.freeze_panes = 'A3'

        import io
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="audit_logs_report.xlsx"'
        return response

    @action(detail=False, methods=['get'])
    def export_csv(self, request):
        return self.export_excel(request)

    @action(detail=False, methods=['post'])
    def purge(self, request):
        user = getattr(request, 'user', None)
        if not (user and user.is_authenticated and (
            user.is_superuser or
            user.has_perm('accounts.perm_sys_purge_logs')
        )):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("تنها مدیر ارشد سامانه یا کاربران دارای مجوز اختصاصی پاکسازی لاگ‌ها (perm_sys_purge_logs) مجاز به حذف لاگ‌های ممیزی هستند.")

        data = request.data or {}
        dry_run = data.get('dry_run') in (True, 'true', '1', 1)
        
        qs = AuditLog.objects.all()

        from_date = _parse_query_date(data.get('from_date'), is_end_of_day=False)
        if from_date:
            qs = qs.filter(created_at__gte=from_date)

        to_date = _parse_query_date(data.get('to_date'), is_end_of_day=True)
        if to_date:
            qs = qs.filter(created_at__lte=to_date)

        warehouse_id = data.get('warehouse') or data.get('warehouse_id')
        if warehouse_id and str(warehouse_id) != 'ALL':
            qs = qs.filter(warehouse_id=warehouse_id)

        module = data.get('module')
        if module:
            qs = qs.filter(module=module)

        days = data.get('days')
        if days and str(days).isdigit():
            cutoff = now() - timedelta(days=int(days))
            qs = qs.filter(created_at__lt=cutoff)

        count = qs.count()

        if dry_run:
            return Response({
                'success': True,
                'dry_run': True,
                'count': count,
                'message': f"{count} رکورد لاگ ممیزی منطبق با شرایط انتخابی برای پاکسازی است."
            })

        confirm_text = str(data.get('confirm_text', '')).strip()
        if confirm_text != 'PURGE_AUDIT_LOGS_CONFIRM':
            return Response({
                'success': False,
                'error': 'جهت پاکسازی قطعی لاگ‌های ممیزی، عبارت تاییدیه امنیتی PURGE_AUDIT_LOGS_CONFIRM باید به صورت دقیق وارد شود.'
            }, status=status.HTTP_400_BAD_REQUEST)

        if count == 0:
            return Response({
                'success': True,
                'purged_count': 0,
                'message': 'هیچ رکوردی منطبق با فیلترهای انتخابی یافت نشد.'
            })

        batch_size = 5000
        while True:
            batch_ids = list(qs.values_list('id', flat=True)[:batch_size])
            if not batch_ids:
                break
            AuditLog.objects.filter(id__in=batch_ids).delete()
            if len(batch_ids) < batch_size:
                break

        from .audit_utils import log_audit_event
        desc_parts = []
        if from_date:
            desc_parts.append(f"از: {from_date.strftime('%Y-%m-%d')}")
        if to_date:
            desc_parts.append(f"تا: {to_date.strftime('%Y-%m-%d')}")
        if warehouse_id:
            desc_parts.append(f"انبار: {warehouse_id}")
        if module:
            desc_parts.append(f"ماژول: {module}")
        if days:
            desc_parts.append(f"قدیمی‌تر از {days} روز")
        criteria_str = ' | '.join(desc_parts) if desc_parts else 'کلیه لاگ‌ها'

        log_audit_event(
            user=user,
            module='system',
            action='DELETE',
            severity='critical',
            details={
                'description': f"پاکسازی قطعی {count} رکورد از لاگ‌های ممیزی ({criteria_str})",
                'purged_count': count,
                'filters': {
                    'from_date': str(from_date) if from_date else None,
                    'to_date': str(to_date) if to_date else None,
                    'warehouse': warehouse_id,
                    'module': module,
                    'days': days
                }
            },
            ip_address=getattr(request, 'META', {}).get('REMOTE_ADDR')
        )

        return Response({
            'success': True,
            'purged_count': count,
            'message': f"{count} رکورد لاگ ممیزی با موفقیت پاکسازی شد."
        })

    @action(detail=True, methods=['get'])
    def preview_revert(self, request, pk=None):
        user = getattr(request, 'user', None)
        if not (user and user.is_authenticated and (
            user.is_superuser or
            user.has_perm('accounts.perm_rollback_single') or
            user.has_perm('accounts.perm_rollback_bulk') or
            user.has_perm('accounts.perm_rollback_data') or
            user.has_perm('accounts.perm_restore_deleted')
        )):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("تنها کاربران مجاز (دارای مجوز بازگردانی یا مدیر سیستم) مجاز به مشاهده پیش‌نمایش بازگردانی هستند.")
        
        log_entry = self.get_object()
        from .rollback_service import get_revert_preview
        result = get_revert_preview(log_entry)
        return Response(result)

    @action(detail=True, methods=['post'])
    def revert(self, request, pk=None):
        user = getattr(request, 'user', None)
        if not (user and user.is_authenticated and (
            user.is_superuser or
            user.has_perm('accounts.perm_rollback_single') or
            user.has_perm('accounts.perm_rollback_bulk') or
            user.has_perm('accounts.perm_rollback_data') or
            user.has_perm('accounts.perm_restore_deleted')
        )):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("تنها کاربران مجاز (دارای مجوز بازگردانی perm_rollback_single یا مدیر سیستم) مجاز به بازگردانی این رکورد هستند.")
        
        log_entry = self.get_object()
        from .rollback_service import revert_log_entry
        from .middleware import get_current_ip
        
        reason = request.data.get('reason')
        ip_address = get_current_ip() or getattr(request, 'META', {}).get('REMOTE_ADDR')
        
        res = revert_log_entry(log_entry, user=request.user, reason=reason, ip_address=ip_address)
        if not res.get('success'):
            return Response(res, status=status.HTTP_400_BAD_REQUEST)
        return Response(res)

    @action(detail=False, methods=['post'])
    def bulk_revert(self, request):
        user = getattr(request, 'user', None)
        if not (user and user.is_authenticated and (
            user.is_superuser or
            user.has_perm('accounts.perm_rollback_bulk') or
            user.has_perm('accounts.perm_rollback_data')
        )):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("تنها مدیر ارشد سامانه یا کاربران دارای مجوز اختصاصی بازگردانی گروهی (perm_rollback_bulk) مجاز به بازگردانی دسته‌ای رکوردها هستند.")
        
        log_ids = request.data.get('log_ids', [])
        if not log_ids or not isinstance(log_ids, list):
            return Response({'success': False, 'error': 'لیست شناسه‌های لاگ الزامی است.'}, status=status.HTTP_400_BAD_REQUEST)
            
        reason = request.data.get('reason')
        from .rollback_service import revert_log_entry
        from .middleware import get_current_ip
        ip_address = get_current_ip() or getattr(request, 'META', {}).get('REMOTE_ADDR')
        
        success_count = 0
        errors = []
        
        allowed_qs = self.get_queryset()
        for lid in log_ids:
            log_item = allowed_qs.filter(id=lid).first()
            if not log_item:
                errors.append(f"لاگ #{lid} یافت نشد یا در قلمرو انبار مجاز شما قرار ندارد.")
                continue
            res = revert_log_entry(log_item, user=request.user, reason=reason, ip_address=ip_address)
            if res.get('success'):
                success_count += 1
            else:
                errors.append(f"لاگ #{lid}: {res.get('error')}")
                
        return Response({
            'success': success_count > 0,
            'success_count': success_count,
            'errors': errors,
            'message': f"{success_count} مورد با موفقیت بازگردانی شد." + (f" ({len(errors)} خطا)" if errors else "")
        })

    @action(detail=False, methods=['post'])
    def preview_point_in_time_rollback(self, request):
        user = getattr(request, 'user', None)
        if not (user and user.is_authenticated and (
            user.is_superuser or
            user.has_perm('accounts.perm_rollback_bulk') or
            user.has_perm('accounts.perm_rollback_data')
        )):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("تنها کاربران دارای مجوز بازگردانی گروهی (perm_rollback_bulk) یا مدیر سیستم مجاز به شبیه‌سازی بازگردانی به تاریخ هستند.")

        data = request.data or {}
        target_date_raw = data.get('target_datetime') or data.get('target_date')
        if not target_date_raw:
            return Response({'can_rollback': False, 'message': 'تاریخ و زمان مقصد جهت بازگردانی الزامی است.'}, status=status.HTTP_400_BAD_REQUEST)

        from .rollback_service import preview_point_in_time_rollback
        target_datetime = _parse_query_date(target_date_raw, is_end_of_day=False)
        if not target_datetime:
            return Response({'can_rollback': False, 'message': 'فرمت تاریخ و زمان نامعتبر است.'}, status=status.HTTP_400_BAD_REQUEST)

        warehouse_id = data.get('warehouse') or data.get('warehouse_id')
        module = data.get('module')
        target_model = data.get('target_model')

        res = preview_point_in_time_rollback(
            target_datetime=target_datetime,
            warehouse_id=warehouse_id,
            module=module,
            target_model=target_model
        )
        return Response(res)

    @action(detail=False, methods=['post'])
    def execute_point_in_time_rollback(self, request):
        user = getattr(request, 'user', None)
        if not (user and user.is_authenticated and (
            user.is_superuser or
            user.has_perm('accounts.perm_rollback_bulk') or
            user.has_perm('accounts.perm_rollback_data')
        )):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("تنها مدیر ارشد سامانه یا کاربران دارای مجوز اختصاصی بازگردانی گروهی (perm_rollback_bulk) مجاز به بازگردانی داده‌ها به یک تاریخ مشخص هستند.")

        data = request.data or {}
        target_date_raw = data.get('target_datetime') or data.get('target_date')
        if not target_date_raw:
            return Response({'success': False, 'error': 'تاریخ و زمان مقصد جهت بازگردانی الزامی است.'}, status=status.HTTP_400_BAD_REQUEST)

        from .rollback_service import execute_point_in_time_rollback
        from .middleware import get_current_ip
        target_datetime = _parse_query_date(target_date_raw, is_end_of_day=False)
        if not target_datetime:
            return Response({'success': False, 'error': 'فرمت تاریخ و زمان نامعتبر است.'}, status=status.HTTP_400_BAD_REQUEST)

        warehouse_id = data.get('warehouse') or data.get('warehouse_id')
        module = data.get('module')
        target_model = data.get('target_model')
        reason = data.get('reason')
        ip_address = get_current_ip() or getattr(request, 'META', {}).get('REMOTE_ADDR')

        res = execute_point_in_time_rollback(
            target_datetime=target_datetime,
            user=user,
            warehouse_id=warehouse_id,
            module=module,
            target_model=target_model,
            reason=reason,
            ip_address=ip_address
        )
        if not res.get('success'):
            return Response(res, status=status.HTTP_400_BAD_REQUEST)
        return Response(res)


class DatabaseBackupViewSet(viewsets.ViewSet):
    """
    مدیریت نسخه‌های پشتیبان و بازیابی پایگاه‌داده (Database Backup & Restore)
    - عملیات مدیریت/ایجاد/دانلود: perm_sys_backup_manage یا perm_sys_backup_restore یا Superuser
    - عملیات بازیابی دیتابیس: انحصاری با perm_sys_backup_restore یا Superuser
    """
    def _check_manage_permission(self, request):
        user = getattr(request, 'user', None)
        if not (user and user.is_authenticated and (
            user.is_superuser or
            user.has_perm('accounts.perm_sys_backup_manage') or
            user.has_perm('accounts.perm_sys_backup_restore')
        )):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("تنها مدیر ارشد سامانه یا کاربران دارای مجوز مدیریت نسخه پشتیبان (perm_sys_backup_manage) به این بخش دسترسی دارند.")

    def _check_restore_permission(self, request):
        user = getattr(request, 'user', None)
        if not (user and user.is_authenticated and (
            user.is_superuser or
            user.has_perm('accounts.perm_sys_backup_restore')
        )):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("تنها مدیر ارشد سامانه یا کاربران دارای مجوز صریح بازیابی پایگاه داده (perm_sys_backup_restore) به این بخش دسترسی دارند.")

    def list(self, request):
        self._check_manage_permission(request)
        from .backup_service import get_backup_list
        backups = get_backup_list()
        return Response(backups)

    def create(self, request):
        self._check_manage_permission(request)
        from .backup_service import create_database_backup
        description = request.data.get('description', 'پشتیبان‌گیری دستی توسط کاربر')
        try:
            meta = create_database_backup(user=request.user, description=description)
            return Response({'success': True, 'backup': meta}, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({'success': False, 'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'])
    def verify(self, request):
        self._check_manage_permission(request)
        filename = request.data.get('filename')
        if not filename:
            return Response({'is_valid': False, 'error': 'نام فایل پشتیبان الزامی است.'}, status=status.HTTP_400_BAD_REQUEST)
        from .backup_service import verify_backup_integrity
        result = verify_backup_integrity(filename)
        return Response(result)

    @action(detail=False, methods=['post'])
    def restore(self, request):
        self._check_restore_permission(request)
        filename = request.data.get('filename')
        confirm_text = request.data.get('confirm_text', '')

        # الزامی بودن تاییدیه متنی صریح
        if confirm_text.strip() != 'RESTORE_DATABASE_CONFIRM':
            return Response({
                'success': False,
                'error': 'جهت بازیابی پایگاه داده، عبارت تاییدیه امنیتی RESTORE_DATABASE_CONFIRM باید به طور دقیق وارد شود.'
            }, status=status.HTTP_400_BAD_REQUEST)

        if not filename:
            return Response({'success': False, 'error': 'نام فایل پشتیبان الزامی است.'}, status=status.HTTP_400_BAD_REQUEST)

        from .backup_service import restore_database_backup
        from .middleware import get_current_ip
        ip_address = get_current_ip() or getattr(request, 'META', {}).get('REMOTE_ADDR')

        try:
            res = restore_database_backup(filename, user=request.user, ip_address=ip_address)
            return Response(res)
        except Exception as e:
            return Response({'success': False, 'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'])
    def download(self, request):
        self._check_manage_permission(request)
        filename = request.query_params.get('filename')
        if not filename or '/' in filename or '\\' in filename:
            return Response({'error': 'نام فایل نامعتبر است.'}, status=status.HTTP_400_BAD_REQUEST)

        from django.conf import settings
        file_path = os.path.join(settings.BASE_DIR, 'backups', 'db', filename)
        if not os.path.exists(file_path):
            return Response({'error': 'فایل یافت نشد.'}, status=status.HTTP_404_NOT_FOUND)

        from django.http import FileResponse
        response = FileResponse(open(file_path, 'rb'), as_attachment=True, filename=filename)
        return response



