import datetime as _dt
from decimal import Decimal

from openpyxl.cell import WriteOnlyCell, Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from django.utils import timezone

# ─── استایل‌ها (هم‌رنگ پالت indigo/slate رابط کاربری) ───
HEADER_FILL = PatternFill('solid', fgColor='4F46E5')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
KEY_FONT = Font(italic=True, color='94A3B8', size=9)
TITLE_FONT = Font(bold=True, size=13, color='1E293B')
ZEBRA_FILL = PatternFill('solid', fgColor='F1F5F9')
THIN = Side(style='thin', color='CBD5E1')
HEADER_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal='center', vertical='center')

NUMBER_FORMATS = {
    'number': '#,##0.###',
    'date': 'yyyy/mm/dd',
    'datetime': 'yyyy/mm/dd hh:mm',
}

COL_WIDTHS = {
    'boolean': 10,
    'number': 14,
    'date': 14,
    'datetime': 19,
    'choice': 16,
    'text': 25,
}

def jalali_now_str():
    now = timezone.localtime()
    try:
        import jdatetime
        return jdatetime.datetime.fromgregorian(datetime=now).strftime('%Y/%m/%d %H:%M')
    except ImportError:
        return now.strftime('%Y-%m-%d %H:%M')

def get_cell_value(v):
    if v is None:
        return ''
    if isinstance(v, bool):
        return 'بله' if v else 'خیر'
    if isinstance(v, (int, float, Decimal)):
        return v
    if isinstance(v, _dt.datetime):
        if timezone.is_aware(v):
            v = timezone.localtime(v).replace(tzinfo=None)
        return v
    if isinstance(v, _dt.date):
        return v
    if hasattr(v, 'isoformat'):
        return v.isoformat()
    return str(v)

def styled_cell(ws, value, number_format=None, fill=None, is_write_only=False):
    """
    یاری‌رسان برای تولید سلول استایل‌دار.
    اگر ورک‌بوک write_only باشد، باید از WriteOnlyCell استفاده شود.
    """
    if is_write_only:
        c = WriteOnlyCell(ws, value=value)
    else:
        c = Cell(ws, value=value)
        
    if number_format and not isinstance(value, str):
        c.number_format = number_format
    if fill is not None:
        c.fill = fill
    return c

def apply_header_styles_to_row(row_cells, is_key_row=False):
    """
    اعمال استایل هدر (آبی/سفید) یا کلید (طوسی/ایتالیک) روی یک لیست از سلول‌ها.
    """
    for cell in row_cells:
        if is_key_row:
            cell.font = KEY_FONT
            cell.alignment = CENTER
        else:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = HEADER_BORDER
            cell.alignment = CENTER

def set_column_widths(ws, columns, is_write_only=False):
    """
    تنظیم عرض ستون‌ها بر اساس نوع فیلد
    """
    from openpyxl.utils import get_column_letter
    for i, c in enumerate(columns, start=1):
        # 'c' can be a dict (from reports) or just a string key
        col_type = c.get('type') if isinstance(c, dict) else 'text'
        ws.column_dimensions[get_column_letter(i)].width = (
            COL_WIDTHS.get(col_type, 25)
        )

def freeze_header_panes(ws, row=3, is_write_only=False):
    """
    فریز کردن پنل هدر.
    در حالت معمولی از ws.freeze_panes استفاده می‌شود.
    در حالت write_only از ws.sheet_view.pane.
    """
    try:
        if is_write_only:
            from openpyxl.worksheet.views import Pane
            ws.sheet_view.pane = Pane(
                topLeftCell=f'A{row}', ySplit=row-1, state='frozen', activePane='bottomLeft',
            )
        else:
            ws.freeze_panes = f'A{row}'
    except Exception:
        pass


def find_data_start_and_mapping(ws, expected_keys, max_scan_rows=5):
    """
    اسکن چند سطر اول برای پیدا کردن هدر (انگلیسی یا فارسی) و استخراج ردیف شروع داده‌ها.
    برمی‌گرداند: (data_start_row_idx, col_mapping)
    col_mapping: دیکشنری {key: column_index_0_based}
    """
    # Create a reverse mapping for finding english keys from persian labels if needed
    # expected_keys is assumed to be a list of dicts like: [{'key': 'username', 'label': 'شناسه ورود'}, ...]
    
    key_set = {k['key'] for k in expected_keys}
    label_to_key = {k['label'].strip(): k['key'] for k in expected_keys if 'label' in k}
    
    # We will read up to max_scan_rows
    rows = list(ws.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True))
    
    best_row_idx = -1
    best_mapping = {}
    best_score = 0
    
    for idx, row in enumerate(rows):
        if not row:
            continue
            
        current_mapping = {}
        score = 0
        
        for col_idx, cell in enumerate(row):
            if not cell:
                continue
            val = str(cell).strip()
            
            # Direct english key match
            if val in key_set:
                current_mapping[val] = col_idx
                score += 2 # Give higher weight to english key matches
            # Persian label match
            elif val in label_to_key:
                current_mapping[label_to_key[val]] = col_idx
                score += 1
                
        if score > best_score:
            best_score = score
            best_mapping = current_mapping
            best_row_idx = idx
            
    if best_row_idx == -1:
        # Fallback to row 1 as header, row 2 as data (if completely unknown)
        return (2, {})
        
    return (best_row_idx + 2, best_mapping) # +1 to 1-based row index, +1 to point to data row
