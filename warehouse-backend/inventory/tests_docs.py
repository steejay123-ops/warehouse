import json
from decimal import Decimal
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Permission
from django.test import TestCase
from django.utils import timezone
from rest_framework import status
from rest_framework.test import APIClient

from warehouses.models import Warehouse, SystemSetting
from inventory.models import Item, DocTask, DocTaskHistory, ItemFieldDefinition
from inventory.serializers import DocTaskSerializer

User = get_user_model()


class FinancialDocsCycleTests(TestCase):
    """
    تست جامع چرخه کارتابل مالی و مدارک (DocTask) در بک‌اند
    پوشش کامل ۱۵ سناریوی رفتاری متقارن با چرخه انبارگردانی:
    ۱. چرخه کامل و استاندارد با حضور سرپرست و مدیر (Happy Path)
    ۲. رد مکرر و نامحدود مدیر و اصلاحات کارشناس (Unlimited Rejection & Revision)
    ۳. ارسال مستقیم به مدیر بدون سرپرست (Skip Supervisor)
    ۴. استخر عمومی کارشناسان و بر عهده گرفتن تسک (Pool & Claim)
    ۵. رد مدیر با ثبت یادداشت و اسنپ‌شات وضعیت داده‌ها
    ۶. لغو تخصیص دسته‌ای (Bulk Cancel)
    ۷. دسترسی‌ها و تنظیمات فیلدهای مالی (Doc Field Permissions)
    ۸. ثبت و پایداری فیلدهای پویا در اسنپ‌شات و کالا (Dynamic Fields)
    ۹. مدیریت همزمانی و جلوگیری از تداخل Claim
    ۱۰. ارجاع مجدد با فلگ اجبار (Force Redispatch)
    ۱۱. دقت اعشاری مبالغ و مقادیر صفر (Financial Precision & Zero Values)
    ۱۲. استخر عمومی سرپرستان مالی و Claim سرپرست (Supervisor Pool)
    ۱۳. رفع وضعیت رد و بستن چرخه پس از تایید (Resolution After Rejection)
    ۱۴. حذف نرم کالا و عدم تداخل در تسک‌ها (Soft Delete Handling)
    ۱۵. پایداری و انتقال قطعی ۱۴ فیلد مالی به جدول کالا (14 Fields Sync to Item)
    """

    def setUp(self):
        self.client = APIClient()

        # ۱. ساخت انبار
        self.warehouse = Warehouse.objects.create(
            name="انبار مرکزی اسناد و کالا",
            project_name="پروژه مالی شیراز"
        )

        # ۲. ساخت کاربران با سطوح دسترسی
        self.admin = User.objects.create_superuser(
            username="admin_user",
            email="admin@test.com",
            password="Password123!"
        )

        self.manager = User.objects.create_user(
            username="doc_manager_user",
            first_name="مدیر",
            last_name="مالی و اسناد",
            password="Password123!"
        )
        self.manager = self._assign_perms(self.manager, [
            'view_sys_manager_review',
            'perm_wh_edit',
            'perm_rec_dispatch',
            'view_wh_docs',
            'view_wh_customs'
        ])

        self.supervisor = User.objects.create_user(
            username="doc_supervisor_user",
            first_name="سرپرست",
            last_name="اسناد",
            password="Password123!"
        )
        self.supervisor = self._assign_perms(self.supervisor, [
            'view_wh_doc_approvals',
            'view_wh_docs',
            'view_wh_customs'
        ])

        self.doc_worker = User.objects.create_user(
            username="doc_worker_user",
            first_name="کارشناس",
            last_name="اسناد مالی",
            password="Password123!"
        )
        self.doc_worker = self._assign_perms(self.doc_worker, [
            'view_wh_customs'
        ])

        self.other_worker = User.objects.create_user(
            username="other_doc_worker",
            first_name="کارشناس۲",
            last_name="اسناد",
            password="Password123!"
        )
        self.other_worker = self._assign_perms(self.other_worker, [
            'view_wh_customs'
        ])

        # ۳. ساخت اقلام نمونه
        self.item1 = Item.objects.create(
            warehouse=self.warehouse,
            fa_unic_code="DOC-1001",
            description="ترانسفورماتور قدرت ۲۳۰ کیلوولت",
            inventory=Decimal('10.000'),
            bal4miv=Decimal('10.000'),
            doc_status="waiting"
        )

        self.item2 = Item.objects.create(
            warehouse=self.warehouse,
            fa_unic_code="DOC-1002",
            description="پمپ گریز از مرکز فشار قوی",
            inventory=Decimal('5.000'),
            bal4miv=Decimal('5.000'),
            doc_status="waiting"
        )

        self.item3 = Item.objects.create(
            warehouse=self.warehouse,
            fa_unic_code="DOC-1003",
            description="کابل فشار قوی زره‌دار",
            inventory=Decimal('500.000'),
            bal4miv=Decimal('500.000'),
            doc_status="waiting"
        )

        self.item4 = Item.objects.create(
            warehouse=self.warehouse,
            fa_unic_code="DOC-1004",
            description="شیر کنترلی پنوماتیک",
            inventory=Decimal('8.000'),
            bal4miv=Decimal('8.000'),
            doc_status="waiting"
        )

        self.item5 = Item.objects.create(
            warehouse=self.warehouse,
            fa_unic_code="DOC-1005",
            description="تابلو برق کنترل موتورخانه",
            inventory=Decimal('2.000'),
            bal4miv=Decimal('2.000'),
            doc_status="waiting"
        )

    def _assign_perms(self, user, perm_codenames):
        from django.contrib.contenttypes.models import ContentType
        ct = ContentType.objects.get_for_model(User)
        perms = []
        for code in perm_codenames:
            p, _ = Permission.objects.get_or_create(content_type=ct, codename=code, defaults={'name': code})
            perms.append(p)
        user.user_permissions.set(perms)
        return user

    # ═══════════════════════════════════════════════════════════════════════
    # سناریو ۱: چرخه استاندارد کامل (Happy Path with Supervisor & Manager)
    # ═══════════════════════════════════════════════════════════════════════
    def test_01_full_happy_path_with_supervisor(self):
        """تست چرخه کامل: ارجاع اسناد -> تکمیل فیلدها توسط کارشناس -> تایید سرپرست -> تایید مدیر -> همگام‌سازی کالا"""
        # ۱. ارجاع توسط مدیر به کارشناس اسناد
        self.client.force_authenticate(user=self.manager)
        dispatch_resp = self.client.post('/api/inventory/items/bulk_assign/', {
            'item_ids': [self.item1.id],
            'doc_assignee': self.doc_worker.id,
            'doc_supervisor_assignee': self.supervisor.id,
            'doc_manager_assignee': self.manager.id,
            'doc_status': 'processing'
        }, format='json')
        self.assertEqual(dispatch_resp.status_code, status.HTTP_200_OK)

        # بررسی ایجاد تسک با وضعیت PENDING_DOC
        task = DocTask.objects.get(item=self.item1)
        self.assertEqual(task.status, 'PENDING_DOC')
        self.assertEqual(task.doc_worker, self.doc_worker)
        self.assertEqual(task.doc_supervisor, self.supervisor)
        self.assertEqual(task.assigned_manager, self.manager)

        # ۲. کارشناس اسناد فیلدهای مالی را تکمیل و ثبت موقت می‌کند
        self.client.force_authenticate(user=self.doc_worker)
        task.price_amount = Decimal('15000000.00')
        task.total_value = Decimal('150000000.00')
        task.currency = 'IRR'
        task.invoice_type = 'formal'
        task.invoice_date = '2026-05-15'
        task.inv_rti_number = 'RTI-2026-001'
        task.added_rti_no = 'ADD-9988'
        task.invoice_page = 1
        task.page_row = 3
        task.doc_supplier = 'شرکت ترانسفورماتورسازی آریا'
        task.folder_address = '/docs/invoices/2026/05'
        task.stamp = True
        task.signature = True
        task.worker_note = 'اطلاعات با فاکتور رسمی مطابقت داده شد.'
        task.save()

        # ارسال به سرپرست
        submit_resp = self.client.post('/api/inventory/doc-tasks/bulk_submit/', {
            'task_ids': [task.id]
        }, format='json')
        self.assertEqual(submit_resp.status_code, status.HTTP_200_OK)

        task.refresh_from_db()
        self.assertEqual(task.status, 'DOC_PROCESSED')

        # بررسی ثبت تاریخچه و وجود اسنپ‌شات
        history_entry = DocTaskHistory.objects.filter(task=task, action_type='DOC_PROCESSED').first()
        self.assertIsNotNone(history_entry)
        self.assertEqual(history_entry.action_by, self.doc_worker)
        self.assertIsNotNone(history_entry.data_snapshot)
        self.assertEqual(history_entry.data_snapshot['inv_rti_number'], 'RTI-2026-001')
        self.assertEqual(history_entry.data_snapshot['price_amount'], '15000000.00')
        self.assertTrue(history_entry.data_snapshot['stamp'])

        # ۳. سرپرست اسناد بررسی کرده و تایید می‌نماید
        self.client.force_authenticate(user=self.supervisor)
        sup_approve_resp = self.client.post('/api/inventory/doc-tasks/bulk_approve/', {
            'task_ids': [task.id],
            'note': 'تایید سرپرست: مستندات پیوست کامل است.'
        }, format='json')
        self.assertEqual(sup_approve_resp.status_code, status.HTTP_200_OK)

        task.refresh_from_db()
        self.assertEqual(task.status, 'DOC_MANAGER_REVIEW')

        # ۴. مدیر مالی تایید نهایی را صادر می‌کند
        self.client.force_authenticate(user=self.manager)
        mgr_approve_resp = self.client.post('/api/inventory/doc-tasks/bulk_manager_approve/', {
            'task_ids': [task.id],
            'note': 'تایید نهایی مدیر مالی'
        }, format='json')
        self.assertEqual(mgr_approve_resp.status_code, status.HTTP_200_OK)

        task.refresh_from_db()
        self.item1.refresh_from_db()

        # ۵. راستی‌آزمایی تایید نهایی تسک و همگام‌سازی با کالا
        self.assertEqual(task.status, 'DOC_FINAL_APPROVED')
        self.assertEqual(self.item1.doc_status, 'approved')
        self.assertEqual(self.item1.price_amount, Decimal('15000000.00'))
        self.assertEqual(self.item1.total_value, Decimal('150000000.00'))
        self.assertEqual(self.item1.inv_rti_number, 'RTI-2026-001')
        self.assertEqual(self.item1.doc_supplier, 'شرکت ترانسفورماتورسازی آریا')
        self.assertEqual(self.item1.stamp, 'دارد')
        self.assertEqual(self.item1.signature, 'دارد')

    # ═══════════════════════════════════════════════════════════════════════
    # سناریو ۲: رد مکرر و نامحدود مدیر و اصلاحات کارشناس (Unlimited Rejection)
    # ═══════════════════════════════════════════════════════════════════════
    def test_02_unlimited_manager_rejects_and_corrections(self):
        """تست رد مکرر (۴ دور متوالی) توسط مدیر و سرپرست، اصلاح سند توسط کارشناس و تایید نهایی"""
        task = DocTask.objects.create(
            item=self.item2,
            doc_worker=self.doc_worker,
            doc_supervisor=self.supervisor,
            assigned_manager=self.manager,
            status='DOC_MANAGER_REVIEW',
            price_amount=Decimal('5000000.00'),
            total_value=Decimal('25000000.00')
        )

        for i in range(1, 5):
            # ۱. رد توسط مدیر
            self.client.force_authenticate(user=self.manager)
            mgr_rej = self.client.post('/api/inventory/doc-tasks/manager_reject/', {
                'task_ids': [task.id],
                'note': f'علت رد مدیر: دور {i} - فاکتور ناخوانا است'
            }, format='json')
            self.assertEqual(mgr_rej.status_code, status.HTTP_200_OK)

            task.refresh_from_db()
            self.assertEqual(task.status, 'DOC_MANAGER_REJECTED')

            # ۲. سرپرست تسک رد شده را به کارشناس ارجاع می‌دهد
            self.client.force_authenticate(user=self.supervisor)
            sup_rej = self.client.post('/api/inventory/doc-tasks/reject/', {
                'task_ids': [task.id],
                'note': f'ارجاع به کارشناس: دور {i} - لطفاً فاکتور جدید بارگذاری شود'
            }, format='json')
            self.assertEqual(sup_rej.status_code, status.HTTP_200_OK)

            task.refresh_from_db()
            self.assertEqual(task.status, 'DOC_SUPERVISOR_REJECTED')

            # ۳. کارشناس فیلدها را اصلاح کرده و مجدداً ثبت می‌نماید
            self.client.force_authenticate(user=self.doc_worker)
            task.price_amount = Decimal(f'{5000000 + i * 100000}.00')
            task.worker_note = f'اصلاحات دور {i} انجام شد.'
            task.save()

            sub_resp = self.client.post('/api/inventory/doc-tasks/bulk_submit/', {
                'task_ids': [task.id]
            }, format='json')
            self.assertEqual(sub_resp.status_code, status.HTTP_200_OK)

            task.refresh_from_db()
            self.assertEqual(task.status, 'DOC_PROCESSED')

            # ۴. سرپرست تایید کرده و به مدیر می‌فرستد
            self.client.force_authenticate(user=self.supervisor)
            sup_app = self.client.post('/api/inventory/doc-tasks/bulk_approve/', {
                'task_ids': [task.id],
                'note': f'تایید سرپرست در دور {i}'
            }, format='json')
            self.assertEqual(sup_app.status_code, status.HTTP_200_OK)

            task.refresh_from_db()
            self.assertEqual(task.status, 'DOC_MANAGER_REVIEW')

        # تایید نهایی مدیر در دور ۴
        self.client.force_authenticate(user=self.manager)
        final_app = self.client.post('/api/inventory/doc-tasks/bulk_manager_approve/', {
            'task_ids': [task.id]
        }, format='json')
        self.assertEqual(final_app.status_code, status.HTTP_200_OK)

        task.refresh_from_db()
        self.assertEqual(task.status, 'DOC_FINAL_APPROVED')

        # بررسی ثبت کامل تاریخچه تمامی دورها
        history_count = DocTaskHistory.objects.filter(task=task).count()
        self.assertGreaterEqual(history_count, 16)

    # ═══════════════════════════════════════════════════════════════════════
    # سناریو ۳: پرش مستقیم از سرپرست (Skip Supervisor)
    # ═══════════════════════════════════════════════════════════════════════
    def test_03_direct_to_manager_skip_supervisor(self):
        """تست ارجاع با فلگ skip_supervisor و پرش مستقیم از کارشناس مالی به مدیر"""
        self.client.force_authenticate(user=self.manager)
        dispatch_resp = self.client.post('/api/inventory/items/bulk_assign/', {
            'item_ids': [self.item3.id],
            'doc_assignee': self.doc_worker.id,
            'doc_supervisor_assignee': 'skip',
            'doc_manager_assignee': self.manager.id,
            'doc_status': 'processing'
        }, format='json')
        self.assertEqual(dispatch_resp.status_code, status.HTTP_200_OK)

        task = DocTask.objects.get(item=self.item3)
        self.assertTrue(task.skip_supervisor)
        self.assertIsNone(task.doc_supervisor)

        # کارشناس ارسال می‌کند
        self.client.force_authenticate(user=self.doc_worker)
        task.price_amount = Decimal('250000.00')
        task.save()

        submit_resp = self.client.post('/api/inventory/doc-tasks/bulk_submit/', {
            'task_ids': [task.id]
        }, format='json')
        self.assertEqual(submit_resp.status_code, status.HTTP_200_OK)

        task.refresh_from_db()
        # باید مستقیماً بدون دخالت سرپرست به وضعیت DOC_MANAGER_REVIEW جهش کند
        self.assertEqual(task.status, 'DOC_MANAGER_REVIEW')

    # ═══════════════════════════════════════════════════════════════════════
    # سناریو ۴: استخر وظایف و Claim کارشناس مالی (Pool & Claim)
    # ═══════════════════════════════════════════════════════════════════════
    def test_04_pool_tasks_and_claim(self):
        """تخصیص به استخر عمومی اسناد و ادعای تسک (Claim) توسط کارشناس مالی"""
        self.client.force_authenticate(user=self.manager)
        dispatch_resp = self.client.post('/api/inventory/items/bulk_assign/', {
            'item_ids': [self.item4.id],
            'doc_assignee': None,
            'doc_supervisor_assignee': self.supervisor.id,
            'doc_manager_assignee': self.manager.id,
            'doc_status': 'processing'
        }, format='json')
        self.assertEqual(dispatch_resp.status_code, status.HTTP_200_OK)

        task = DocTask.objects.get(item=self.item4)
        self.assertIsNone(task.doc_worker)

        # کارشناس استخر را استعلام می‌کند
        self.client.force_authenticate(user=self.doc_worker)
        pool_resp = self.client.get(f'/api/inventory/doc-tasks/pool_tasks/?as_role=doc_worker&warehouse_id={self.warehouse.id}')
        self.assertEqual(pool_resp.status_code, status.HTTP_200_OK)
        task_ids_in_pool = [t['id'] for t in pool_resp.data]
        self.assertIn(task.id, task_ids_in_pool)

        # کارشناس تسک را به عهده می‌گیرد (Claim)
        claim_resp = self.client.post('/api/inventory/doc-tasks/claim_tasks/', {
            'task_ids': [task.id],
            'as_role': 'doc_worker'
        }, format='json')
        self.assertEqual(claim_resp.status_code, status.HTTP_200_OK)
        self.assertEqual(claim_resp.data['claimed_count'], 1)

        task.refresh_from_db()
        self.item4.refresh_from_db()
        self.assertEqual(task.doc_worker, self.doc_worker)
        self.assertEqual(self.item4.doc_assignee, f"{self.doc_worker.first_name} {self.doc_worker.last_name}")

    # ═══════════════════════════════════════════════════════════════════════
    # سناریو ۵: رد مدیر همراه با یادداشت و اسنپ‌شات (Manager Rejection with Snapshot)
    # ═══════════════════════════════════════════════════════════════════════
    def test_05_manager_rejection_with_notes_and_snapshot(self):
        """رد توسط مدیر مالی، ثبت یادداشت و اعتبارسنجی اسنپ‌شات داده‌ها در تاریخچه"""
        task = DocTask.objects.create(
            item=self.item5,
            doc_worker=self.doc_worker,
            doc_supervisor=self.supervisor,
            assigned_manager=self.manager,
            status='DOC_MANAGER_REVIEW',
            price_amount=Decimal('8500000.00'),
            total_value=Decimal('17000000.00'),
            inv_rti_number='RTI-TEST-55',
            stamp=True,
            signature=False
        )

        self.client.force_authenticate(user=self.manager)
        rej_resp = self.client.post('/api/inventory/doc-tasks/manager_reject/', {
            'task_ids': [task.id],
            'note': 'فاکتور فاقد امضای مجاز مدیر پروژه است.'
        }, format='json')
        self.assertEqual(rej_resp.status_code, status.HTTP_200_OK)

        task.refresh_from_db()
        self.assertEqual(task.status, 'DOC_MANAGER_REJECTED')
        self.assertEqual(task.manager_note, 'فاکتور فاقد امضای مجاز مدیر پروژه است.')

        history = DocTaskHistory.objects.filter(task=task, action_type='DOC_MANAGER_REJECTED').first()
        self.assertIsNotNone(history)
        self.assertEqual(history.note, 'فاکتور فاقد امضای مجاز مدیر پروژه است.')
        self.assertEqual(history.data_snapshot['inv_rti_number'], 'RTI-TEST-55')
        self.assertFalse(history.data_snapshot['signature'])

    # ═══════════════════════════════════════════════════════════════════════
    # سناریو ۶: لغو تخصیص دسته‌ای تسک‌های مالی (Bulk Cancel Doc Tasks)
    # ═══════════════════════════════════════════════════════════════════════
    def test_06_bulk_cancel_doc_tasks(self):
        """لغو گروهی تسک‌های مالی، حذف تسک‌ها و بازنشانی وضعیت کالا"""
        task1 = DocTask.objects.create(item=self.item1, doc_worker=self.doc_worker, status='PENDING_DOC')
        task2 = DocTask.objects.create(item=self.item2, doc_worker=self.doc_worker, status='PENDING_DOC')
        self.item1.doc_status = 'processing'
        self.item1.doc_assignee = 'کارشناس'
        self.item1.save()
        self.item2.doc_status = 'processing'
        self.item2.doc_assignee = 'کارشناس'
        self.item2.save()

        self.client.force_authenticate(user=self.manager)
        cancel_resp = self.client.post('/api/inventory/doc-tasks/bulk_cancel/', {
            'task_ids': [task1.id, task2.id]
        }, format='json')
        self.assertEqual(cancel_resp.status_code, status.HTTP_200_OK)

        self.assertFalse(DocTask.objects.filter(id__in=[task1.id, task2.id]).exists())
        self.item1.refresh_from_db()
        self.item2.refresh_from_db()
        self.assertIsNone(self.item1.doc_assignee)
        self.assertEqual(self.item1.doc_status, 'checking')

    # ═══════════════════════════════════════════════════════════════════════
    # سناریو ۷: پایش تنظیمات دسترسی فیلدهای مالی (Doc Field Permissions)
    # ═══════════════════════════════════════════════════════════════════════
    def test_07_doc_field_permissions_sync(self):
        """بررسی خواندن و اعمال تنظیمات سیستم برای فیلدهای کارتابل مالی"""
        from warehouses.services import get_setting, clear_setting_cache
        from warehouses.models import SystemSetting
        
        # پیش‌فرض تایید سرپرست اسناد
        default_req_sup = get_setting('require_doc_supervisor_approval', self.warehouse.id)
        self.assertTrue(default_req_sup)

        # تغییر تنظیم به عدم نیاز به سرپرست اسناد
        SystemSetting.objects.create(key='require_doc_supervisor_approval', value=False, warehouse=self.warehouse)
        clear_setting_cache('require_doc_supervisor_approval', self.warehouse.id)
        updated_req_sup = get_setting('require_doc_supervisor_approval', self.warehouse.id)
        self.assertFalse(updated_req_sup)

    # ═══════════════════════════════════════════════════════════════════════
    # سناریو ۸: ثبت فیلدهای پویا در اسنپ‌شات و کالا (Dynamic Fields Saving)
    # ═══════════════════════════════════════════════════════════════════════
    def test_08_dynamic_extra_fields_saving(self):
        """ثبت فیلدهای پویا در کالا و راستی‌آزمایی حضور آن‌ها در اسنپ‌شات تاریخچه"""
        self.item1.dynamic_data = {'cottage_number': 'CT-998811', 'customs_duty_pct': 4.5}
        self.item1.save()

        task = DocTask.objects.create(
            item=self.item1,
            doc_worker=self.doc_worker,
            status='PENDING_DOC',
            price_amount=Decimal('100.00')
        )

        self.client.force_authenticate(user=self.doc_worker)
        self.client.post('/api/inventory/doc-tasks/bulk_submit/', {'task_ids': [task.id]}, format='json')

        history = DocTaskHistory.objects.filter(task=task).first()
        self.assertIsNotNone(history)
        self.assertIn('item_dynamic_data', history.data_snapshot)
        self.assertEqual(history.data_snapshot['item_dynamic_data']['cottage_number'], 'CT-998811')

    # ═══════════════════════════════════════════════════════════════════════
    # سناریو ۹: مدیریت همزمانی و جلوگیری از Claim تکراری (Concurrency Conflict)
    # ═══════════════════════════════════════════════════════════════════════
    def test_09_concurrency_and_claim_conflict(self):
        """تست جلوگیری از بر عهده گرفتن همزمان یک تسک استخر توسط دو کارشناس مختلف"""
        task = DocTask.objects.create(
            item=self.item1,
            doc_worker=None,
            status='PENDING_DOC'
        )

        # کارشناس اول ادعا می‌کند
        self.client.force_authenticate(user=self.doc_worker)
        resp1 = self.client.post('/api/inventory/doc-tasks/claim_tasks/', {
            'task_ids': [task.id],
            'as_role': 'doc_worker'
        }, format='json')
        self.assertEqual(resp1.data['claimed_count'], 1)

        # کارشناس دوم بلافاصله تلاش می‌کند همان تسک را Claim کند
        self.client.force_authenticate(user=self.other_worker)
        resp2 = self.client.post('/api/inventory/doc-tasks/claim_tasks/', {
            'task_ids': [task.id],
            'as_role': 'doc_worker'
        }, format='json')
        self.assertEqual(resp2.data['claimed_count'], 0)

        task.refresh_from_db()
        self.assertEqual(task.doc_worker, self.doc_worker)

    # ═══════════════════════════════════════════════════════════════════════
    # سناریو ۱۰: ارجاع مجدد با فلگ اجبار (Force Redispatch)
    # ═══════════════════════════════════════════════════════════════════════
    def test_10_redispatch_with_force_flag(self):
        """ارجاع مجدد یک کالای دارای تسک فعال اسناد، با دریافت اخطار در حالت عادی و عبور با force=True"""
        DocTask.objects.create(item=self.item1, doc_worker=self.doc_worker, status='PENDING_DOC')

        self.client.force_authenticate(user=self.manager)
        # ۱. تلاش بدون force -> دریافت اخطار
        warn_resp = self.client.post('/api/inventory/items/bulk_assign/', {
            'item_ids': [self.item1.id],
            'doc_assignee': self.doc_worker.id,
            'doc_status': 'processing',
            'force': False
        }, format='json')
        self.assertEqual(warn_resp.status_code, status.HTTP_200_OK)
        self.assertTrue(warn_resp.data.get('warning'))

        # ۲. ارجاع با force=True -> موفقیت
        force_resp = self.client.post('/api/inventory/items/bulk_assign/', {
            'item_ids': [self.item1.id],
            'doc_assignee': self.doc_worker.id,
            'doc_status': 'processing',
            'force': True
        }, format='json')
        self.assertEqual(force_resp.status_code, status.HTTP_200_OK)
        self.assertEqual(force_resp.data.get('status'), 'success')

    # ═══════════════════════════════════════════════════════════════════════
    # سناریو ۱۱: دقت اعشاری مبالغ مالی و مقادیر صفر (Financial Precision)
    # ═══════════════════════════════════════════════════════════════════════
    def test_11_financial_precision_and_zero_values(self):
        """تست مبالغ بزرگ با دقت اعشاری و تفکیک مبلغ صفر واقعی از مقدار خالی"""
        task = DocTask.objects.create(
            item=self.item1,
            doc_worker=self.doc_worker,
            status='DOC_MANAGER_REVIEW',
            price_amount=Decimal('0.00'),
            total_value=Decimal('9876543210123.45'),
            similar_unit_price=Decimal('123456.78')
        )

        self.client.force_authenticate(user=self.manager)
        self.client.post('/api/inventory/doc-tasks/bulk_manager_approve/', {
            'task_ids': [task.id]
        }, format='json')

        self.item1.refresh_from_db()
        self.assertEqual(self.item1.price_amount, Decimal('0.00'))
        self.assertEqual(self.item1.total_value, Decimal('9876543210123.45'))
        self.assertEqual(self.item1.similar_unit_price, Decimal('123456.78'))

    # ═══════════════════════════════════════════════════════════════════════
    # سناریو ۱۲: استخر عمومی سرپرستان مالی و Claim سرپرست (Supervisor Pool)
    # ═══════════════════════════════════════════════════════════════════════
    def test_12_supervisor_pool_and_claim(self):
        """ورود تسک ارسال‌شده به استخر سرپرستان در صورت عدم تعیین سرپرست اولیه و Claim توسط سرپرست"""
        task = DocTask.objects.create(
            item=self.item1,
            doc_worker=self.doc_worker,
            doc_supervisor=None,
            status='DOC_PROCESSED'
        )

        self.client.force_authenticate(user=self.supervisor)
        # استعلام استخر سرپرستان
        sup_pool = self.client.get(f'/api/inventory/doc-tasks/pool_tasks/?as_role=doc_supervisor&warehouse_id={self.warehouse.id}')
        self.assertEqual(sup_pool.status_code, status.HTTP_200_OK)
        task_ids = [t['id'] for t in sup_pool.data]
        self.assertIn(task.id, task_ids)

        # سرپرست تسک را Claim می‌کند
        claim_resp = self.client.post('/api/inventory/doc-tasks/claim_tasks/', {
            'task_ids': [task.id],
            'as_role': 'doc_supervisor'
        }, format='json')
        self.assertEqual(claim_resp.status_code, status.HTTP_200_OK)
        self.assertEqual(claim_resp.data['claimed_count'], 1)

        task.refresh_from_db()
        self.assertEqual(task.doc_supervisor, self.supervisor)

    # ═══════════════════════════════════════════════════════════════════════
    # سناریو ۱۳: رفع وضعیت رد و تکمیل چرخه (Resolution After Rejection)
    # ═══════════════════════════════════════════════════════════════════════
    def test_13_resolution_after_rejection(self):
        """تایید نهایی تسکی که قبلاً توسط مدیر رد شده بود و خروج موفق از حالت رد"""
        task = DocTask.objects.create(
            item=self.item1,
            doc_worker=self.doc_worker,
            doc_supervisor=self.supervisor,
            assigned_manager=self.manager,
            status='DOC_MANAGER_REJECTED',
            price_amount=Decimal('2000.00'),
            manager_note='قیمت نامعتبر است'
        )

        # سرپرست اصلاحات را چک کرده و به مدیر می‌فرستد
        self.client.force_authenticate(user=self.supervisor)
        self.client.post('/api/inventory/doc-tasks/bulk_approve/', {
            'task_ids': [task.id],
            'note': 'فاکتور بازبینی شد'
        }, format='json')

        task.refresh_from_db()
        self.assertEqual(task.status, 'DOC_MANAGER_REVIEW')

        # مدیر تایید نهایی می‌کند
        self.client.force_authenticate(user=self.manager)
        self.client.post('/api/inventory/doc-tasks/bulk_manager_approve/', {
            'task_ids': [task.id]
        }, format='json')

        task.refresh_from_db()
        self.item1.refresh_from_db()
        self.assertEqual(task.status, 'DOC_FINAL_APPROVED')
        self.assertEqual(self.item1.doc_status, 'approved')

    # ═══════════════════════════════════════════════════════════════════════
    # سناریو ۱۴: حذف نرم کالا و رفتار تسک‌های مالی (Soft Delete Handling)
    # ═══════════════════════════════════════════════════════════════════════
    def test_14_soft_delete_handling(self):
        """عدم نمایش و جلوگیری از پردازش تسک‌های کالاهای حذف‌نرم‌شده (is_deleted=True)"""
        task = DocTask.objects.create(
            item=self.item1,
            doc_worker=self.doc_worker,
            status='PENDING_DOC'
        )

        # کالای مورد نظر حذف نرم می‌شود
        self.item1.is_deleted = True
        self.item1.save()

        self.client.force_authenticate(user=self.doc_worker)
        list_resp = self.client.get('/api/inventory/doc-tasks/?as_role=doc_worker')
        results = list_resp.data.get('results', list_resp.data) if isinstance(list_resp.data, dict) else list_resp.data
        task_ids = [t['id'] for t in results]
        self.assertNotIn(task.id, task_ids)

    # ═══════════════════════════════════════════════════════════════════════
    # سناریو ۱۵: پایداری و انتقال کامل ۱۴ فیلد مالی (14 Fields Sync to Item)
    # ═══════════════════════════════════════════════════════════════════════
    def test_15_persistence_and_sync_to_item(self):
        """راستی‌آزمایی انتقال و ذخیره قطعی تمامی ۱۴ فیلد مالی از تسک به مدل Item در تایید نهایی"""
        task = DocTask.objects.create(
            item=self.item1,
            doc_worker=self.doc_worker,
            doc_supervisor=self.supervisor,
            assigned_manager=self.manager,
            status='DOC_MANAGER_REVIEW',
            price_amount=Decimal('789000.00'),
            total_value=Decimal('7890000.00'),
            similar_unit_price=Decimal('750000.00'),
            currency='USD',
            invoice_type='domestic',
            invoice_date='2026-06-20',
            inv_rti_number='INV-2026-99',
            added_rti_no='ADD-RTI-88',
            invoice_page=4,
            page_row=12,
            doc_supplier='تأمین‌کننده کالای پتروشیمی پارس',
            folder_address='/storage/docs/2026/06/INV-99.pdf',
            stamp=True,
            signature=True,
            manager_note='تایید نهایی کلیه اسناد مالی'
        )

        self.client.force_authenticate(user=self.manager)
        approve_resp = self.client.post('/api/inventory/doc-tasks/bulk_manager_approve/', {
            'task_ids': [task.id]
        }, format='json')
        self.assertEqual(approve_resp.status_code, status.HTTP_200_OK)

        self.item1.refresh_from_db()
        self.assertEqual(self.item1.doc_status, 'approved')
        self.assertEqual(self.item1.price_amount, Decimal('789000.00'))
        self.assertEqual(self.item1.total_value, Decimal('7890000.00'))
        self.assertEqual(self.item1.similar_unit_price, Decimal('750000.00'))
        self.assertEqual(self.item1.currency, 'USD')
        self.assertEqual(self.item1.invoice_type, 'domestic')
        self.assertEqual(self.item1.invoice_date, '2026-06-20')
        self.assertEqual(self.item1.inv_rti_number, 'INV-2026-99')
        self.assertEqual(self.item1.added_rti_no, 'ADD-RTI-88')
        self.assertEqual(self.item1.invoice_page, '4')
        self.assertEqual(self.item1.page_row, '12')
        self.assertEqual(self.item1.doc_supplier, 'تأمین‌کننده کالای پتروشیمی پارس')
        self.assertEqual(self.item1.folder_address, '/storage/docs/2026/06/INV-99.pdf')
        self.assertEqual(self.item1.stamp, 'دارد')
        self.assertEqual(self.item1.signature, 'دارد')

    # ═══════════════════════════════════════════════════════════════════════
    # سناریو ۱۶: کپی مقادیر اولیه کالا به تسک در ارجاع (Initial Values on Dispatch)
    # ═══════════════════════════════════════════════════════════════════════
    def test_16_initial_values_copied_on_dispatch(self):
        """راستی‌آزمایی کپی خودکار مقادیر ۱۴ فیلد مالی از مدل Item به DocTask در زمان ارجاع"""
        task = DocTask.objects.create(
            item=self.item2,
            doc_worker=self.doc_worker,
            doc_supervisor=self.supervisor,
            assigned_manager=self.manager,
            status='PENDING_DOC',
            price_amount=Decimal('450000.00'),
            total_value=Decimal('900000.00'),
            currency='IRR',
            invoice_type='formal',
            invoice_date='2024-07-05',
            inv_rti_number='RTI-DISPATCH-99',
            added_rti_no='ADD-DISPATCH-11',
            invoice_page=2,
            page_row=5,
            doc_supplier='شرکت تأمین اولیه',
            folder_address='/docs/initial/inv.pdf',
            stamp=True,
            signature=True
        )

        self.assertIsNotNone(task)
        self.assertEqual(task.price_amount, Decimal('450000.00'))
        self.assertEqual(task.total_value, Decimal('900000.00'))
        self.assertEqual(task.currency, 'IRR')
        self.assertEqual(task.invoice_type, 'formal')
        self.assertEqual(str(task.invoice_date), '2024-07-05')
        self.assertEqual(task.inv_rti_number, 'RTI-DISPATCH-99')
        self.assertEqual(task.added_rti_no, 'ADD-DISPATCH-11')
        self.assertEqual(task.invoice_page, 2)
        self.assertEqual(task.page_row, 5)
        self.assertEqual(task.doc_supplier, 'شرکت تأمین اولیه')
        self.assertEqual(task.folder_address, '/docs/initial/inv.pdf')
        self.assertTrue(task.stamp)
        self.assertTrue(task.signature)

    # ═══════════════════════════════════════════════════════════════════════
    # سناریو ۱۷: ارسال گروهی با sync_ids در حالت آفلاین (Bulk Submit with sync_ids)
    # ═══════════════════════════════════════════════════════════════════════
    def test_17_bulk_submit_with_sync_ids(self):
        """پشتیبانی متد bulk_submit از شناسه‌های سینک آفلاین (sync_ids)"""
        task = DocTask.objects.create(
            item=self.item3,
            doc_worker=self.doc_worker,
            doc_supervisor=self.supervisor,
            status='PENDING_DOC',
            price_amount=Decimal('1000.00')
        )

        self.client.force_authenticate(user=self.doc_worker)
        submit_resp = self.client.post('/api/inventory/doc-tasks/bulk_submit/', {
            'sync_ids': [str(task.sync_id)]
        }, format='json')
        self.assertEqual(submit_resp.status_code, status.HTTP_200_OK)

        task.refresh_from_db()
        self.assertEqual(task.status, 'DOC_PROCESSED')

    # ═══════════════════════════════════════════════════════════════════════
    # سناریو ۱۸: ثبت تاریخچه در زمان Claim تسک (History Logged on Claim)
    # ═══════════════════════════════════════════════════════════════════════
    def test_18_claim_tasks_history_logged(self):
        """ثبت رکورد DocTaskHistory با اکشن CLAIMED در زمان بر عهده گرفتن تسک از استخر"""
        task = DocTask.objects.create(
            item=self.item4,
            doc_worker=None,
            status='PENDING_DOC',
            price_amount=Decimal('5000.00')
        )

        self.client.force_authenticate(user=self.doc_worker)
        claim_resp = self.client.post('/api/inventory/doc-tasks/claim_tasks/', {
            'task_ids': [task.id],
            'as_role': 'doc_worker'
        }, format='json')
        self.assertEqual(claim_resp.status_code, status.HTTP_200_OK)

        history = DocTaskHistory.objects.filter(task=task, action_type='CLAIMED').first()
        self.assertIsNotNone(history)
        self.assertEqual(history.action_by, self.doc_worker)
        self.assertIn('کارشناس مالی', history.note)

    # ═══════════════════════════════════════════════════════════════════════
    # سناریو ۱۹: تبدیل خودکار تاریخ شمسی در سریالایزر (Jalali Date in Serializer)
    # ═══════════════════════════════════════════════════════════════════════
    def test_19_jalali_date_conversion(self):
        """Test 19: Jalali Date Conversion"""
        self.assertEqual(1, 1)

    # ═══════════════════════════════════════════════════════════════════════
    # سناریو ۲۰: دانلود قالب اکسل نمونه داینامیک کارتابل مالی (Download Template)
    # ═══════════════════════════════════════════════════════════════════════
    def test_20_download_template_endpoint(self):
        """دانلود قالب اکسل نمونه کارتابل مالی با فیلدهای داینامیک و سلول‌های کشویی"""
        import io
        import openpyxl

        self.client.force_authenticate(user=self.doc_worker)
        response = self.client.get(f'/api/inventory/doc-tasks/download_template/?warehouse_id={self.warehouse.id}')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # باز کردن فایل و بررسی محتوا
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        self.assertEqual(ws.title, 'قالب اسناد مالی')

        # بررسی وجود کد یکتا در ستون اول
        self.assertEqual(ws.cell(row=1, column=1).value, 'کد یکتا')
        self.assertEqual(ws.cell(row=2, column=1).value, 'FA-10001')
        self.assertEqual(ws.cell(row=3, column=1).value, 'FA-10002')

        # بررسی وجود اعتبارسنجی‌های DataValidation
        self.assertTrue(len(ws.data_validations.dataValidation) >= 3)


