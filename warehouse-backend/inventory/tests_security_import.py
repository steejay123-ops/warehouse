import os
import tempfile
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Permission
from django.test import TransactionTestCase
from rest_framework import status
from rest_framework.test import APIClient

from warehouses.models import Warehouse
from inventory.models import ImportLog

User = get_user_model()


class DownloadImportLogSecurityTests(TransactionTestCase):
    def setUp(self):
        self.client = APIClient()

        self.warehouse = Warehouse.objects.create(name="انبار تست ۱")
        self.other_warehouse = Warehouse.objects.create(name="انبار تست ۲")

        # کاربران
        self.user_with_perm = User.objects.create_user(
            username="importer_user",
            password="Password123!"
        )
        from django.contrib.contenttypes.models import ContentType
        ct = ContentType.objects.get_for_model(User)
        perm, _ = Permission.objects.get_or_create(content_type=ct, codename="perm_rec_import", defaults={'name': "Import Excel"})
        self.user_with_perm.user_permissions.add(perm)
        self.user_with_perm.assigned_warehouses.add(self.warehouse)

        self.user_no_perm = User.objects.create_user(
            username="unauthorized_user",
            password="Password123!"
        )

        self.import_id = "testimport123"
        self.log_file_path = os.path.join(tempfile.gettempdir(), f"import_log_{self.import_id}.xlsx")
        with open(self.log_file_path, "wb") as f:
            f.write(b"dummy excel content")

        self.import_log = ImportLog.objects.create(
            import_id=self.import_id,
            warehouse=self.warehouse,
            imported_by=self.user_with_perm,
            file_name="test.xlsx"
        )

    def tearDown(self):
        if os.path.exists(self.log_file_path):
            try:
                os.remove(self.log_file_path)
            except OSError:
                pass

    def test_unauthenticated_access_is_blocked(self):
        """درخواست بدون احراز هویت باید مسدود شود (۴۰۱ یا ۴۰۳)"""
        url = f"/api/inventory/items/download_import_log/?import_id={self.import_id}"
        response = self.client.get(url)
        self.assertIn(response.status_code, [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN])

    def test_user_without_permission_is_blocked(self):
        """کاربر لاگین‌شده بدون پرمیشن perm_rec_import باید ۴۰۳ شود"""
        self.client.force_authenticate(user=self.user_no_perm)
        url = f"/api/inventory/items/download_import_log/?import_id={self.import_id}"
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_path_traversal_attempt_is_rejected(self):
        """تلاش برای Path Traversal باید ارور ۴۰۰ برگرداند"""
        self.client.force_authenticate(user=self.user_with_perm)
        bad_ids = [
            "../../etc/passwd",
            "..\\..\\windows\\win.ini",
            "/etc/passwd",
            "import_id/../secret"
        ]
        for bad_id in bad_ids:
            url = f"/api/inventory/items/download_import_log/?import_id={bad_id}"
            response = self.client.get(url)
            self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_user_from_other_warehouse_cannot_download_others_log(self):
        """کاربر دارای پرمیشن ایمپورت در انبار دیگر نباید لاگ انبار ۱ را ببیند (۴۰۳)"""
        other_user = User.objects.create_user(
            username="other_wh_user",
            password="Password123!"
        )
        from django.contrib.contenttypes.models import ContentType
        ct = ContentType.objects.get_for_model(User)
        perm = Permission.objects.get(codename="perm_rec_import", content_type=ct)
        other_user.user_permissions.add(perm)
        other_user.assigned_warehouses.add(self.other_warehouse)

        self.client.force_authenticate(user=other_user)
        url = f"/api/inventory/items/download_import_log/?import_id={self.import_id}"
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_authorized_user_can_download(self):
        """کاربر مجاز باید بتواند فایل لاگ را دانلود کند (۲۰۰)"""
        self.client.force_authenticate(user=self.user_with_perm)
        url = f"/api/inventory/items/download_import_log/?import_id={self.import_id}"
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.content, b"dummy excel content")

    def test_clear_warehouse_data_blocks_unauthorized_warehouse(self):
        """کاربر انبار ۱ نباید بتواند داده‌های انبار ۲ را پاک کند (۴۰۳)"""
        self.client.force_authenticate(user=self.user_with_perm)
        url = "/api/inventory/items/clear_warehouse_data/"
        response = self.client.post(url, {'warehouse_id': self.other_warehouse.id}, format='json')
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_latest_import_blocks_unauthorized_warehouse(self):
        """کاربر انبار ۱ نباید بتواند آخرین ایمپورت‌های انبار ۲ را مشاهده کند (۴۰۳)"""
        self.client.force_authenticate(user=self.user_with_perm)
        url = f"/api/inventory/items/latest_import/?warehouse_id={self.other_warehouse.id}"
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_revert_import_blocks_unauthorized_warehouse(self):
        """کاربر انبار ۱ نباید بتواند ایمپورت متعلق به انبار ۲ را بازگردانی کند (۴۰۳)"""
        other_import = ImportLog.objects.create(
            import_id="otherimport999",
            warehouse=self.other_warehouse,
            file_name="other.xlsx"
        )
        self.client.force_authenticate(user=self.user_with_perm)
        url = "/api/inventory/items/revert_import/"
        response = self.client.post(url, {'import_id': other_import.import_id}, format='json')
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_import_excel_blocks_unauthorized_warehouse(self):
        """کاربر انبار ۱ نباید بتواند در انبار ۲ اکسل بارگذاری کند (۴۰۳)"""
        import io
        from django.core.files.uploadedfile import SimpleUploadedFile
        fake_excel = SimpleUploadedFile("items.xlsx", b"dummy content", content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.client.force_authenticate(user=self.user_with_perm)
        url = "/api/inventory/items/import_excel/"
        response = self.client.post(url, {'warehouse_id': self.other_warehouse.id, 'file': fake_excel}, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_delete_from_excel_blocks_unauthorized_warehouse(self):
        """کاربر انبار ۱ نباید بتواند از طریق اکسل در انبار ۲ حذف انجام دهد (۴۰۳)"""
        from django.core.files.uploadedfile import SimpleUploadedFile
        fake_excel = SimpleUploadedFile("delete.xlsx", b"dummy content", content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.client.force_authenticate(user=self.user_with_perm)
        url = "/api/inventory/items/delete_from_excel/"
        response = self.client.post(url, {'warehouse_id': self.other_warehouse.id, 'file': fake_excel}, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_import_excel_with_foreign_item_id_fails_to_steal_item(self):
        """تلاش برای جابجایی یا ویرایش کالای انبار دیگر با ستون id باید رد شده و انبار کالا تغییر نکند"""
        import io
        import openpyxl
        from django.core.files.uploadedfile import SimpleUploadedFile
        from inventory.models import Item

        foreign_item = Item.objects.create(
            warehouse=self.other_warehouse,
            fa_unic_code="FA-FOREIGN-1",
            description="کالای انبار دوم"
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['id', 'fa_unic_code', 'description'])
        ws.append([foreign_item.id, 'FA-FOREIGN-1', 'تلاش برای سرقت کالا'])
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        excel_file = SimpleUploadedFile("steal.xlsx", out.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        self.client.force_authenticate(user=self.user_with_perm)
        url = "/api/inventory/items/import_excel/"
        response = self.client.post(url, {
            'warehouse_id': self.warehouse.id,
            'conflict_strategy': 'replace',
            'file': excel_file
        }, format='multipart')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        async def _consume(gen):
            res = []
            async for chunk in gen:
                res.append(chunk.decode('utf-8') if isinstance(chunk, bytes) else chunk)
            return "".join(res)

        import asyncio
        import json
        content = asyncio.run(_consume(response.streaming_content))
        lines = [json.loads(l) for l in content.strip().split('\n') if l.strip()]
        msgs = [l.get('msg', '') for l in lines]
        self.assertTrue(any("مغایرت انبار" in m for m in msgs), f"Expected 'مغایرت انبار' in messages: {msgs}")

        foreign_item.refresh_from_db()
        self.assertEqual(foreign_item.warehouse_id, self.other_warehouse.id)
        self.assertEqual(foreign_item.description, "کالای انبار دوم")

    def test_import_excel_with_conflicting_warehouse_column_fails(self):
        """ردیف با نام انبار مغایر با انبار انتخابی فرآیند باید رد شود"""
        import io
        import openpyxl
        from django.core.files.uploadedfile import SimpleUploadedFile

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['fa_unic_code', 'description', 'warehouse'])
        ws.append(['FA-NEW-99', 'تست مغایرت انبار', self.other_warehouse.name])
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        excel_file = SimpleUploadedFile("conflict_wh.xlsx", out.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        self.client.force_authenticate(user=self.user_with_perm)
        url = "/api/inventory/items/import_excel/"
        response = self.client.post(url, {
            'warehouse_id': self.warehouse.id,
            'file': excel_file
        }, format='multipart')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        async def _consume(gen):
            res = []
            async for chunk in gen:
                res.append(chunk.decode('utf-8') if isinstance(chunk, bytes) else chunk)
            return "".join(res)

        import asyncio
        import json
        content = asyncio.run(_consume(response.streaming_content))
        lines = [json.loads(l) for l in content.strip().split('\n') if l.strip()]
        msgs = [l.get('msg', '') for l in lines]
        self.assertTrue(any("مغایرت انبار" in m for m in msgs), f"Expected 'مغایرت انبار' in messages: {msgs}")

    def test_broken_row_does_not_rollback_valid_rows_and_emits_summary(self):
        """یک ردیف خراب نباید کل تراکنش و ردیف‌های معتبر قبل و بعد از خود را رول‌بک کند"""
        import io
        import openpyxl
        from django.core.files.uploadedfile import SimpleUploadedFile
        from inventory.models import Item

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['fa_unic_code', 'description', 'total_value'])
        ws.append(['FA-GOOD-1', 'کالای معتبر ۱', 1000])
        # ردیف خراب: مقدار نامعتبر متنی برای فیلد اعشاری total_value
        ws.append(['FA-BAD-2', 'کالای خراب ۲', 'INVALID_NUMERIC_VALUE_XYZ'])
        ws.append(['FA-GOOD-3', 'کالای معتبر ۳', 3000])
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        excel_file = SimpleUploadedFile("mixed.xlsx", out.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        self.client.force_authenticate(user=self.user_with_perm)
        url = "/api/inventory/items/import_excel/"
        response = self.client.post(url, {
            'warehouse_id': self.warehouse.id,
            'file': excel_file
        }, format='multipart')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        async def _consume(gen):
            res = []
            async for chunk in gen:
                res.append(chunk.decode('utf-8') if isinstance(chunk, bytes) else chunk)
            return "".join(res)

        import asyncio
        import json
        content = asyncio.run(_consume(response.streaming_content))
        lines = [json.loads(l) for l in content.strip().split('\n') if l.strip()]

        summary = next((l for l in lines if l.get('type') == 'summary'), None)
        self.assertIsNotNone(summary, "پیام خلاصه نهایی summary باید ارسال شود")
        self.assertEqual(summary.get('status'), 'success')
        self.assertEqual(summary.get('created'), 2)
        self.assertEqual(summary.get('failed'), 1)

        self.assertTrue(Item.objects.filter(fa_unic_code='FA-GOOD-1', warehouse=self.warehouse).exists())
        self.assertTrue(Item.objects.filter(fa_unic_code='FA-GOOD-3', warehouse=self.warehouse).exists())
        self.assertFalse(Item.objects.filter(fa_unic_code='FA-BAD-2').exists())

    def test_user_with_only_import_perm_can_access_parse_headers_and_export_columns(self):
        """کاربری که فقط پرمیشن ایمپورت دارد ولی perm_wh_edit ندارد، باید بتواند هدرها و ستون‌ها را بخواند"""
        import io
        import openpyxl
        from django.core.files.uploadedfile import SimpleUploadedFile

        self.client.force_authenticate(user=self.user_with_perm)

        # ۱. ستون‌های قابل شناسایی
        res_cols = self.client.get('/api/inventory/items/export_columns/', {'warehouse_id': self.warehouse.id})
        self.assertEqual(res_cols.status_code, status.HTTP_200_OK)
        self.assertTrue(any(c.get('key') == 'fa_unic_code' for c in res_cols.data))

        # ۲. خواندن هدرهای فایل اکسل
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['fa_unic_code', 'description'])
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        excel_file = SimpleUploadedFile("headers.xlsx", out.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        res_parse = self.client.post('/api/inventory/items/parse_headers/', {
            'file': excel_file,
            'warehouse_id': self.warehouse.id
        }, format='multipart')
        self.assertEqual(res_parse.status_code, status.HTTP_200_OK)
        self.assertIn('fa_unic_code', res_parse.data.get('found_fields', []))

        # ۳. دریافت آخرین وضعیت ایمپورت
        res_latest = self.client.get('/api/inventory/items/latest_import/', {'warehouse_id': self.warehouse.id})
        self.assertEqual(res_latest.status_code, status.HTTP_200_OK)

    def test_user_without_perm_cannot_access_parse_headers(self):
        """کاربر بدون پرمیشن perm_rec_import نباید به parse_headers دسترسی داشته باشد (۴۰۳)"""
        import io
        import openpyxl
        from django.core.files.uploadedfile import SimpleUploadedFile

        self.client.force_authenticate(user=self.user_no_perm)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['fa_unic_code', 'description'])
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        excel_file = SimpleUploadedFile("headers.xlsx", out.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        res = self.client.post('/api/inventory/items/parse_headers/', {
            'file': excel_file,
            'warehouse_id': self.warehouse.id
        }, format='multipart')
        self.assertEqual(res.status_code, status.HTTP_403_FORBIDDEN)

    def test_update_empty_does_not_overwrite_zero_or_false(self):
        """در حالت تکمیل نواقص، موجودی صفر و مقادیر False نباید با مقادیر اکسل بازنویسی شوند"""
        import io
        import openpyxl
        from django.core.files.uploadedfile import SimpleUploadedFile
        from inventory.models import Item

        item = Item.objects.create(
            warehouse=self.warehouse,
            fa_unic_code="FA-ZERO-TEST",
            inventory=0,
            has_conflict=False,
            description=""
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['fa_unic_code', 'inventory', 'has_conflict', 'description'])
        ws.append(['FA-ZERO-TEST', 999, True, 'شرح اضافه شده'])
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        excel_file = SimpleUploadedFile("update_empty.xlsx", out.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        self.client.force_authenticate(user=self.user_with_perm)
        url = "/api/inventory/items/import_excel/"
        response = self.client.post(url, {
            'warehouse_id': self.warehouse.id,
            'conflict_strategy': 'update_empty',
            'file': excel_file
        }, format='multipart')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        async def _consume(gen):
            res = []
            async for chunk in gen:
                res.append(chunk.decode('utf-8') if isinstance(chunk, bytes) else chunk)
            return "".join(res)

        import asyncio
        asyncio.run(_consume(response.streaming_content))

        item.refresh_from_db()
        # شرح خالی پر شد
        self.assertEqual(item.description, 'شرح اضافه شده')
        # موجودی صفر دست‌نخورده باقی ماند
        self.assertEqual(item.inventory, 0)
        # وضعیت تداخل دست‌نخورده باقی ماند
        self.assertEqual(item.has_conflict, False)

    def test_delete_from_excel_with_float_ids_succeeds(self):
        """حذف از اکسل با شناسه‌های اعشاری مثل 101.0 نباید خطای دیتابیسی بدهد"""
        import io
        import openpyxl
        from django.core.files.uploadedfile import SimpleUploadedFile
        from inventory.models import Item

        item = Item.objects.create(
            warehouse=self.warehouse,
            fa_unic_code="FA-DEL-FLOAT",
            description="کالای حذف شونده"
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['id', 'fa_unic_code'])
        # اکسل عدد را اعشاری لود می‌کند مثلاً 101.0
        ws.append([f"{item.id}.0", 'FA-DEL-FLOAT'])
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        excel_file = SimpleUploadedFile("delete_float.xlsx", out.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        self.client.force_authenticate(user=self.user_with_perm)
        url = "/api/inventory/items/delete_from_excel/"
        response = self.client.post(url, {
            'warehouse_id': self.warehouse.id,
            'file': excel_file
        }, format='multipart')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data.get('status'), 'success')

        # کالا با موفقیت حذف شده است
        self.assertFalse(Item.objects.filter(id=item.id).exists())

    def test_tag_separator_deduplication_persian_and_english_commas(self):
        """تگ‌های دارای کامای فارسی و انگلیسی باید درست تفکیک و بدون تکرار ذخیره شوند"""
        import io
        import openpyxl
        from django.core.files.uploadedfile import SimpleUploadedFile
        from inventory.models import Item

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['fa_unic_code', 'description', 'my_tag'])
        ws.append(['FA-TAG-101', 'تست تفکیک تگ', 'پارت۱،پارت۲'])
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        excel_file = SimpleUploadedFile("test_tags.xlsx", out.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        self.client.force_authenticate(user=self.user_with_perm)
        url = "/api/inventory/items/import_excel/"
        response = self.client.post(url, {
            'warehouse_id': self.warehouse.id,
            'file': excel_file,
            'import_tag': 'پارت۲,پارت۳',
            'conflict_action': 'replace'
        }, format='multipart')

        self.assertEqual(response.status_code, status.HTTP_200_OK)

        async def _consume(streaming_content):
            res = []
            async for chunk in streaming_content:
                res.append(chunk.decode('utf-8') if isinstance(chunk, bytes) else chunk)
            return "".join(res)

        import asyncio
        asyncio.run(_consume(response.streaming_content))

        created_item = Item.objects.filter(fa_unic_code='FA-TAG-101', warehouse=self.warehouse).first()
        self.assertIsNotNone(created_item)
        tags = created_item.my_tag.split('،')
        # تگ 'پارت۲' نباید تکراری باشد
        self.assertEqual(len(tags), 3)
        self.assertEqual(set(tags), {'پارت۱', 'پارت۲', 'پارت۳'})

    def test_download_template_uppercase_dynamic_field_no_duplicates(self):
        """فیلدهای پویای دارای حروف بزرگ نباید ستون تکراری با حروف کوچک در قالب تولید کنند"""
        import io
        import openpyxl
        from inventory.models import ItemFieldDefinition

        ItemFieldDefinition.objects.create(
            warehouse=self.warehouse,
            name="BatchNo",
            label="شماره بچ",
            field_type="text"
        )

        self.client.force_authenticate(user=self.user_with_perm)
        url = f"/api/inventory/items/download_template/?warehouse_id={self.warehouse.id}"
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)

        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        first_row = [cell.value for cell in ws[1]]

        # BatchNo باید دقیقاً یک بار در هدر باشد و batchno نباید وجود داشته باشد
        self.assertEqual(first_row.count("BatchNo"), 1)
        self.assertNotIn("batchno", first_row)

    def test_export_columns_checks_warehouse_access(self):
        """دسترسی به export_columns انبار غیرمجاز باید ۴۰۳ بدهد"""
        self.client.force_authenticate(user=self.user_with_perm)
        url = f"/api/inventory/items/export_columns/?warehouse_id={self.other_warehouse.id}"
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_import_pre_counted_items_success_and_consistency(self):
        """تست ورود اقلام از قبل شمرده‌شده: تسک شمارش تایید نهایی، وضعیت done و مغایرت خودکار"""
        import io
        import openpyxl
        from django.core.files.uploadedfile import SimpleUploadedFile
        from inventory.models import Item, CountTask, CountTaskHistory

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['fa_unic_code', 'description', 'inventory', 'bal4miv'])
        # ردیف ۱: بدون مغایرت (50 == 50)
        ws.append(['PRE-COUNT-01', 'کالای بدون مغایرت', 50, 50])
        # ردیف ۲: دارای مغایرت (30 != 35)
        ws.append(['PRE-COUNT-02', 'کالای دارای مغایرت', 30, 35])
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        excel_file = SimpleUploadedFile("pre_counted.xlsx", out.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        from django.contrib.contenttypes.models import ContentType
        ct = ContentType.objects.get_for_model(User)
        perm_finalize, _ = Permission.objects.get_or_create(content_type=ct, codename="perm_inventory_finalize", defaults={'name': "Finalize Inventory"})
        self.user_with_perm.user_permissions.add(perm_finalize)

        self.client.force_authenticate(user=self.user_with_perm)
        url = "/api/inventory/items/import_excel/"
        response = self.client.post(url, {
            'warehouse_id': self.warehouse.id,
            'file': excel_file,
            'is_pre_counted': 'true',
            'conflict_strategy': 'replace'
        }, format='multipart')

        self.assertEqual(response.status_code, status.HTTP_200_OK)

        async def _consume(streaming_content):
            res = []
            async for chunk in streaming_content:
                res.append(chunk.decode('utf-8') if isinstance(chunk, bytes) else chunk)
            return "".join(res)

        import asyncio
        asyncio.run(_consume(response.streaming_content))

        # بررسی کالاها در دیتابیس
        item1 = Item.objects.filter(fa_unic_code='PRE-COUNT-01', warehouse=self.warehouse).first()
        self.assertIsNotNone(item1)
        self.assertEqual(item1.field_status, 'done')
        self.assertEqual(item1.doc_status, 'waiting') # گردش مالی دست‌نخورده مانده است
        self.assertFalse(item1.has_conflict) # 50 == 50

        item2 = Item.objects.filter(fa_unic_code='PRE-COUNT-02', warehouse=self.warehouse).first()
        self.assertIsNotNone(item2)
        self.assertEqual(item2.field_status, 'done')
        self.assertEqual(item2.doc_status, 'waiting')
        self.assertTrue(item2.has_conflict) # 30 != 35

        # بررسی تسک‌های شمارش (CountTask)
        task1 = CountTask.objects.filter(item=item1).first()
        self.assertIsNotNone(task1)
        self.assertEqual(task1.status, 'FINAL_APPROVED')
        self.assertEqual(float(task1.counted_balance), 50.0)

        history1 = CountTaskHistory.objects.filter(task=task1).first()
        self.assertIsNotNone(history1)
        self.assertEqual(history1.action_type, 'FINAL_APPROVED')

        task2 = CountTask.objects.filter(item=item2).first()
        self.assertIsNotNone(task2)
        self.assertEqual(task2.status, 'FINAL_APPROVED')
        self.assertEqual(float(task2.counted_balance), 30.0)

    def test_import_pre_counted_permission_denied_for_unauthorized_user(self):
        """کاربری که دسترسی ورود فایل عادی دارد ولی perm_inventory_finalize ندارد باید با ۴۰۳ مسدود شود"""
        import io
        import openpyxl
        from django.core.files.uploadedfile import SimpleUploadedFile
        from django.contrib.contenttypes.models import ContentType

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['fa_unic_code', 'description', 'inventory'])
        ws.append(['PRE-COUNT-DENIED', 'کالای غیرمجاز', 10])
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        excel_file = SimpleUploadedFile("denied.xlsx", out.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        # ساخت کاربری که پرمیشن ورودی کالا (perm_rec_import) دارد تا از گیت منو رد شود
        normal_import_user = User.objects.create_user(
            username="normal_importer",
            password="Password123!"
        )
        ct = ContentType.objects.get_for_model(User)
        perm_import = Permission.objects.get(codename="perm_rec_import", content_type=ct)
        normal_import_user.user_permissions.add(perm_import)
        normal_import_user.assigned_warehouses.add(self.warehouse)

        # ارسال با is_pre_counted=true توسط این کاربر باید به طور صریح با ۴۰۳ و پیام مربوطه رد شود
        self.client.force_authenticate(user=normal_import_user)
        url = "/api/inventory/items/import_excel/"
        response = self.client.post(url, {
            'warehouse_id': self.warehouse.id,
            'file': excel_file,
            'is_pre_counted': 'true',
        }, format='multipart')

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
        self.assertIn('شما مجوز ثبت مستقیم اقلام در وضعیت از قبل شمرده‌شده را ندارید', response.json().get('error', ''))

    def test_import_pre_counted_resurrect_tombstone_fresh_inventory(self):
        """احیای رکورد حذف‌شده با فایل شمرده‌شده باید موجودی جدید اکسل را در کالا و تسک شمارش بنشاند"""
        import io
        import openpyxl
        from django.core.files.uploadedfile import SimpleUploadedFile
        from django.contrib.contenttypes.models import ContentType
        from inventory.models import Item, CountTask

        # ایجاد کالای حذف‌شده در دیتابیس با موجودی قبلی 10
        tombstone = Item.all_objects.create(
            warehouse=self.warehouse,
            fa_unic_code="TOMB-RESURRECT-01",
            description="کالای حذف‌شده قدیمی",
            inventory=10,
            bal4miv=25,
            is_deleted=True,
            field_status='waiting'
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['fa_unic_code', 'description', 'inventory', 'bal4miv'])
        # در اکسل موجودی شمرده‌شده جدید 25 است (برابر با دفتری 25 -> بدون مغایرت)
        ws.append(['TOMB-RESURRECT-01', 'کالای احیاشده', 25, 25])
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        excel_file = SimpleUploadedFile("resurrect.xlsx", out.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        ct = ContentType.objects.get_for_model(User)
        perm_finalize, _ = Permission.objects.get_or_create(content_type=ct, codename="perm_inventory_finalize", defaults={'name': "Finalize Inventory"})
        self.user_with_perm.user_permissions.add(perm_finalize)

        self.client.force_authenticate(user=self.user_with_perm)
        url = "/api/inventory/items/import_excel/"
        response = self.client.post(url, {
            'warehouse_id': self.warehouse.id,
            'file': excel_file,
            'is_pre_counted': 'true',
            'conflict_strategy': 'replace'
        }, format='multipart')

        self.assertEqual(response.status_code, status.HTTP_200_OK)

        async def _consume(streaming_content):
            res = []
            async for chunk in streaming_content:
                res.append(chunk.decode('utf-8') if isinstance(chunk, bytes) else chunk)
            return "".join(res)

        import asyncio
        asyncio.run(_consume(response.streaming_content))

        # کالا باید از حالت حذف خارج شده باشد و موجودی آن ۲۵ باشد (نه مقدار کهنه ۱۰)
        tombstone.refresh_from_db()
        self.assertFalse(tombstone.is_deleted)
        self.assertEqual(float(tombstone.inventory), 25.0)
        self.assertEqual(tombstone.field_status, 'done')
        self.assertFalse(tombstone.has_conflict) # 25 == 25

        task = CountTask.objects.filter(item=tombstone).first()
        self.assertIsNotNone(task)
        self.assertEqual(task.status, 'FINAL_APPROVED')
        # مقدار تسک شمارش نیز باید دقیقاً مقدار جدید ۲۵ باشد نه مقدار کهنه ۱۰
        self.assertEqual(float(task.counted_balance), 25.0)

    def test_import_pre_counted_update_empty_consistent_values(self):
        """در استراتژی update_empty مقادیر Item.inventory، has_conflict و CountTask کاملاً هم‌خوان باشند"""
        import io
        import openpyxl
        from django.core.files.uploadedfile import SimpleUploadedFile
        from django.contrib.contenttypes.models import ContentType
        from inventory.models import Item, CountTask

        # ایجاد کالای موجود با موجودی 10 و دفتری 10 (بدون مغایرت اولیه)
        existing = Item.objects.create(
            warehouse=self.warehouse,
            fa_unic_code="EXISTING-EMPTY-TEST",
            description="کالای موجود با موجودی پر",
            inventory=10,
            bal4miv=10,
            field_status='waiting',
            has_conflict=False
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['fa_unic_code', 'inventory', 'bal4miv'])
        # در اکسل موجودی 99 ارسال می‌شود اما چون استراتژی update_empty است و موجودی کالا خالی نیست، بازنویسی نمی‌شود
        ws.append(['EXISTING-EMPTY-TEST', 99, 10])
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        excel_file = SimpleUploadedFile("update_empty.xlsx", out.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        ct = ContentType.objects.get_for_model(User)
        perm_finalize, _ = Permission.objects.get_or_create(content_type=ct, codename="perm_inventory_finalize", defaults={'name': "Finalize Inventory"})
        self.user_with_perm.user_permissions.add(perm_finalize)

        self.client.force_authenticate(user=self.user_with_perm)
        url = "/api/inventory/items/import_excel/"
        response = self.client.post(url, {
            'warehouse_id': self.warehouse.id,
            'file': excel_file,
            'is_pre_counted': 'true',
            'conflict_strategy': 'update_empty'
        }, format='multipart')

        self.assertEqual(response.status_code, status.HTTP_200_OK)

        async def _consume(streaming_content):
            res = []
            async for chunk in streaming_content:
                res.append(chunk.decode('utf-8') if isinstance(chunk, bytes) else chunk)
            return "".join(res)

        import asyncio
        asyncio.run(_consume(response.streaming_content))

        existing.refresh_from_db()
        # موجودی در update_empty تغییر نکرده و همان 10 باقی مانده است
        self.assertEqual(float(existing.inventory), 10.0)
        # مغایرت باید با موجودی نهایی ماندگار (10) سنجیده شده باشد نه عدد 99 اکسل! پس 10 == 10 یعنی False
        self.assertFalse(existing.has_conflict)
        self.assertEqual(existing.field_status, 'done')

        task = CountTask.objects.filter(item=existing).first()
        self.assertIsNotNone(task)
        self.assertEqual(task.status, 'FINAL_APPROVED')
        # مقدار شمارش تسک هم با موجودی واقعی کالا هم‌خوان است (10)
        self.assertEqual(float(task.counted_balance), 10.0)

    def test_import_doc_pre_approved_permission_denied_without_perm(self):
        """کاربر بدون مجوز تایید اسناد باید در صورت ارسال is_doc_pre_approved خطای ۴۰۳ دریافت کند"""
        import io
        import openpyxl
        from django.core.files.uploadedfile import SimpleUploadedFile

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['fa_unic_code', 'description', 'price_amount'])
        ws.append(['DOC-DENIED-01', 'کالای رد مجوز اسناد', 150000])
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        excel_file = SimpleUploadedFile("test_doc_denied.xlsx", out.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        # user_with_perm دارای perm_rec_import است اما فاقد perm_doc_approve_action است
        self.client.force_authenticate(user=self.user_with_perm)
        url = "/api/inventory/items/import_excel/"
        response = self.client.post(url, {
            'warehouse_id': self.warehouse.id,
            'file': excel_file,
            'is_doc_pre_approved': 'true'
        }, format='multipart')

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
        self.assertIn('شما مجوز ثبت مستقیم اقلام در وضعیت اسناد تاییدشده را ندارید', response.json().get('error', ''))

    def test_import_doc_pre_approved_success_with_perm_and_financial_fields(self):
        """کاربر مجاز با ارسال is_doc_pre_approved باید آیتم تاییدشده و تسک مالی DOC_FINAL_APPROVED ایجاد کند"""
        import io
        import openpyxl
        from django.core.files.uploadedfile import SimpleUploadedFile
        from django.contrib.contenttypes.models import ContentType
        from inventory.models import Item, DocTask, DocTaskHistory

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['fa_unic_code', 'description', 'price_amount', 'invoice_type', 'doc_supplier', 'stamp', 'signature'])
        ws.append(['DOC-SUCCESS-01', 'کالای اسناد کامل', 450000, 'رسمی', 'شرکت تامین پیشرو', 'دارد', 'دارد'])
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        excel_file = SimpleUploadedFile("test_doc_success.xlsx", out.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        ct = ContentType.objects.get_for_model(User)
        perm_doc, _ = Permission.objects.get_or_create(content_type=ct, codename="perm_doc_approve_action", defaults={'name': "Approve Docs"})
        self.user_with_perm.user_permissions.add(perm_doc)

        self.client.force_authenticate(user=self.user_with_perm)
        url = "/api/inventory/items/import_excel/"
        response = self.client.post(url, {
            'warehouse_id': self.warehouse.id,
            'file': excel_file,
            'is_doc_pre_approved': 'true'
        }, format='multipart')

        self.assertEqual(response.status_code, status.HTTP_200_OK)

        async def _consume(streaming_content):
            res = []
            async for chunk in streaming_content:
                res.append(chunk.decode('utf-8') if isinstance(chunk, bytes) else chunk)
            return "".join(res)

        import asyncio
        asyncio.run(_consume(response.streaming_content))

        # بررسی ایجاد آیتم
        item = Item.objects.filter(fa_unic_code='DOC-SUCCESS-01').first()
        self.assertIsNotNone(item)
        self.assertEqual(item.doc_status, 'done')
        self.assertEqual(float(item.price_amount), 450000.0)

        # بررسی تشکیل DocTask
        doc_task = DocTask.objects.filter(item=item).first()
        self.assertIsNotNone(doc_task)
        self.assertEqual(doc_task.status, 'DOC_FINAL_APPROVED')
        self.assertEqual(float(doc_task.price_amount), 450000.0)
        self.assertEqual(doc_task.invoice_type, 'رسمی')
        self.assertEqual(doc_task.doc_supplier, 'شرکت تامین پیشرو')
        self.assertTrue(doc_task.stamp)
        self.assertTrue(doc_task.signature)

        # بررسی ثبت تاریخچه DocTaskHistory
        history = DocTaskHistory.objects.filter(task=doc_task).first()
        self.assertIsNotNone(history)
        self.assertEqual(history.action_type, 'DOC_FINAL_APPROVED')
        self.assertIsNotNone(history.data_snapshot)
        self.assertEqual(history.data_snapshot.get('invoice_type'), 'رسمی')

    def test_import_both_pre_counted_and_doc_pre_approved_together(self):
        """ارسال هم‌زمان هر دو تاگل: شمارش فیزیکی و اسناد هر دو ۱۰۰٪ تایید نهایی شوند"""
        import io
        import openpyxl
        from django.core.files.uploadedfile import SimpleUploadedFile
        from django.contrib.contenttypes.models import ContentType
        from inventory.models import Item, CountTask, DocTask

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['fa_unic_code', 'description', 'inventory', 'bal4miv', 'price_amount'])
        ws.append(['BOTH-APPROVED-01', 'کالای کامل هر دو فاز', 50, 50, 200000])
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        excel_file = SimpleUploadedFile("test_both.xlsx", out.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        ct = ContentType.objects.get_for_model(User)
        perm_finalize, _ = Permission.objects.get_or_create(content_type=ct, codename="perm_inventory_finalize", defaults={'name': "Finalize Inventory"})
        perm_doc, _ = Permission.objects.get_or_create(content_type=ct, codename="perm_doc_approve_action", defaults={'name': "Approve Docs"})
        self.user_with_perm.user_permissions.add(perm_finalize, perm_doc)

        self.client.force_authenticate(user=self.user_with_perm)
        url = "/api/inventory/items/import_excel/"
        response = self.client.post(url, {
            'warehouse_id': self.warehouse.id,
            'file': excel_file,
            'is_pre_counted': 'true',
            'is_doc_pre_approved': 'true'
        }, format='multipart')

        self.assertEqual(response.status_code, status.HTTP_200_OK)

        async def _consume(streaming_content):
            res = []
            async for chunk in streaming_content:
                res.append(chunk.decode('utf-8') if isinstance(chunk, bytes) else chunk)
            return "".join(res)

        import asyncio
        asyncio.run(_consume(response.streaming_content))

        item = Item.objects.filter(fa_unic_code='BOTH-APPROVED-01').first()
        self.assertIsNotNone(item)
        self.assertEqual(item.field_status, 'done')
        self.assertEqual(item.doc_status, 'done')

        c_task = CountTask.objects.filter(item=item).first()
        self.assertIsNotNone(c_task)
        self.assertEqual(c_task.status, 'FINAL_APPROVED')

        d_task = DocTask.objects.filter(item=item).first()
        self.assertIsNotNone(d_task)
        self.assertEqual(d_task.status, 'DOC_FINAL_APPROVED')

    def test_import_replace_empty_inventory_preserves_existing_balance(self):
        """در استراتژی replace، اگر سلول موجودی اکسل خالی باشد، موجودی قبلی کالا صفر نشود"""
        import io
        import openpyxl
        from django.core.files.uploadedfile import SimpleUploadedFile
        from django.contrib.contenttypes.models import ContentType
        from inventory.models import Item, CountTask

        # ایجاد کالای موجود با موجودی 15.0
        existing = Item.objects.create(
            warehouse=self.warehouse,
            fa_unic_code="PRESERVE-INV-TEST",
            description="کالای با موجودی 15",
            inventory=15.0,
            bal4miv=15.0,
            field_status='waiting'
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['fa_unic_code', 'description', 'inventory'])
        # سلول موجودی را خالی (None) ارسال می‌کنیم
        ws.append(['PRESERVE-INV-TEST', 'شرح ویرایش‌شده جدید', None])
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        excel_file = SimpleUploadedFile("test_preserve.xlsx", out.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        ct = ContentType.objects.get_for_model(User)
        perm_finalize, _ = Permission.objects.get_or_create(content_type=ct, codename="perm_inventory_finalize", defaults={'name': "Finalize Inventory"})
        self.user_with_perm.user_permissions.add(perm_finalize)

        self.client.force_authenticate(user=self.user_with_perm)
        url = "/api/inventory/items/import_excel/"
        response = self.client.post(url, {
            'warehouse_id': self.warehouse.id,
            'file': excel_file,
            'is_pre_counted': 'true',
            'conflict_strategy': 'replace'
        }, format='multipart')

        self.assertEqual(response.status_code, status.HTTP_200_OK)

        async def _consume(streaming_content):
            res = []
            async for chunk in streaming_content:
                res.append(chunk.decode('utf-8') if isinstance(chunk, bytes) else chunk)
            return "".join(res)

        import asyncio
        asyncio.run(_consume(response.streaming_content))

        existing.refresh_from_db()
        # شرح تغییر کرده است
        self.assertEqual(existing.description, 'شرح ویرایش‌شده جدید')
        # موجودی نباید به 0.0 پاک شده باشد؛ باید همان 15.0 حفظ شده باشد!
        self.assertEqual(float(existing.inventory), 15.0)
        self.assertEqual(existing.field_status, 'done')

        task = CountTask.objects.filter(item=existing).first()
        self.assertIsNotNone(task)
        self.assertEqual(task.status, 'FINAL_APPROVED')
        self.assertEqual(float(task.counted_balance), 15.0)




