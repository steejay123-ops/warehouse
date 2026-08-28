from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.http import StreamingHttpResponse, HttpResponse
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters
from rest_framework.parsers import MultiPartParser, FormParser
from .signals import broadcast_count_task_update, broadcast_doc_task_update
from .models import Item, ImportLog, ImportHistory, ItemFieldDefinition, ItemPhoto
from .models import CountTaskHistory
from common.sync_models import soft_delete_queryset
from common.warehouse_scope import scope_queryset, can_access_warehouse
from .serializers import ItemSerializer, CountTaskSerializer, DocTaskSerializer, ItemFieldDefinitionSerializer, ItemPhotoSerializer, photo_prefetch
from django.forms.models import model_to_dict
from warehouses.models import Warehouse
from common.mixins import DeleteImpactMixin
from .models import CountTask, DocTask, DocTaskHistory
from django.db.models import Q
from django.utils import timezone
import openpyxl
import uuid
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import PatternFill
import json
import tempfile
import os
import traceback
import queue
import threading
import asyncio
from datetime import datetime
from django.core.cache import cache
from django.db import transaction
import logging
import re
from decimal import Decimal
from django.utils import timezone
import time

logger = logging.getLogger(__name__)

from decimal import Decimal, InvalidOperation

def _parse_decimal(val):
    if val is None or str(val).strip() == '':
        return None
    try:
        return Decimal(str(val).strip().replace(',', '.'))
    except (InvalidOperation, ValueError, TypeError):
        return None

def _create_doc_task_snapshot(task):
    """ایجاد اسنپ‌شات از تمام مقادیر ۱۴ فیلد مالی و فیلدهای پویای تسک در لحظه عملیات"""
    return {
        'added_rti_no': task.added_rti_no,
        'inv_rti_number': task.inv_rti_number,
        'invoice_type': task.invoice_type,
        'invoice_date': str(task.invoice_date) if task.invoice_date else None,
        'invoice_page': task.invoice_page,
        'page_row': task.page_row,
        'doc_supplier': task.doc_supplier,
        'total_value': str(task.total_value) if task.total_value is not None else None,
        'price_amount': str(task.price_amount) if task.price_amount is not None else None,
        'similar_unit_price': str(task.similar_unit_price) if task.similar_unit_price is not None else None,
        'currency': task.currency,
        'folder_address': task.folder_address,
        'stamp': task.stamp,
        'signature': task.signature,
        'worker_note': task.worker_note,
        'supervisor_note': task.supervisor_note,
        'manager_note': task.manager_note,
        'status': task.status,
        'item_dynamic_data': task.item.dynamic_data if (task.item and getattr(task.item, 'dynamic_data', None)) else {}
    }

from rest_framework.pagination import PageNumberPagination
from .filters import ItemFilter, ItemFieldDefinitionFilter


def _soft_delete_items_cascade(items_qs):
    """
    حذف نرم گروهی آیتم‌ها + Cascade دستی به مدل‌های سینک‌شونده وابسته
    (CountTask، CountTaskHistory و ItemPhoto) تا کلاینت آفلاین tombstone همه را بگیرد.

    ItemPhoto اضافه شد چون بدون آن، عکس‌های کالای حذف‌شده در دیتابیس زنده
    می‌ماندند و کلاینت آفلاینی که کالا را حذف‌شده دیده هیچ‌وقت خبردار نمی‌شد که
    عکس‌ها هم رفته‌اند. is_primary هم پاک می‌شود تا رکورد tombstone با
    is_primary=True به کلاینت نرسد. مسیر معکوس (احیای کالا) در
    `_restore_item_photos_cascade` است و باید همیشه با این تابع جفت بماند.

    DocTask/ImportHistory عمداً دست نمی‌خورند (خارج از دامنه سینک فاز ۰؛
    نمایششان با فیلتر item__is_deleted=False کنترل می‌شود).
    خروجی: تعداد آیتم‌های حذف‌نرم‌شده.
    """
    item_ids = list(items_qs.values_list('id', flat=True))
    if not item_ids:
        return 0
    ItemPhoto.objects.filter(item_id__in=item_ids).update(
        is_deleted=True, is_primary=False, updated_at=timezone.now()
    )
    soft_delete_queryset(CountTaskHistory.objects.filter(task__item_id__in=item_ids))
    soft_delete_queryset(CountTask.objects.filter(item_id__in=item_ids))
    return soft_delete_queryset(Item.objects.filter(id__in=item_ids))


def _restore_item_photos_cascade(item_ids):
    """
    برگرداندن عکس‌های کالاهایی که از حالت حذف‌نرم احیا شده‌اند.

    چرا لازم است: حذف کالا عکس‌هایش را هم tombstone می‌کند. اگر احیا این کار را
    برنگرداند، تصویری که کاربر گرفته برای همیشه نامرئی می‌ماند — فایل روی دیسک
    هست، ردیف با is_deleted=True مانده و هیچ مسیر دیگری آن را زنده نمی‌کند.
    یعنی از دست رفتن داده کاربر.

    is_primary در حذف پاک شده بود، پس برای هر کالا که پس از احیا عکس شاخص ندارد
    دوباره یکی انتخاب می‌شود؛ وگرنه کالا عکس دارد ولی هیچ‌کدام شاخص نیست و
    بند انگشتی (thumbnail) کالا خالی می‌ماند.

    خروجی: تعداد عکس‌های احیاشده.
    """
    item_ids = list(item_ids)
    if not item_ids:
        return 0

    now = timezone.now()
    # updated_at صریح ست می‌شود چون update() سیگنال auto_now را رد می‌کند و
    # بدون آن، احیا در دلتای Pull به کلاینت آفلاین نمی‌رسد.
    restored = ItemPhoto.all_objects.filter(item_id__in=item_ids, is_deleted=True).update(
        is_deleted=False, updated_at=now
    )
    if not restored:
        return 0

    for item_id in item_ids:
        live = ItemPhoto.objects.filter(item_id=item_id)
        if live.filter(is_primary=True).exists():
            continue
        successor = live.order_by('display_order', '-created_at', '-id').first()
        if successor:
            ItemPhoto.objects.filter(id=successor.id).update(is_primary=True, updated_at=now)

    return restored


def _cleanup_old_import_logs(max_age_hours=24):
    """پاکسازی فایل‌های گزارش اکسل قدیمی از پوشه موقت سرور"""
    try:
        temp_dir = tempfile.gettempdir()
        now = time.time()
        for fname in os.listdir(temp_dir):
            if fname.startswith("import_log_") and fname.endswith(".xlsx"):
                fpath = os.path.join(temp_dir, fname)
                if os.path.isfile(fpath):
                    try:
                        if (now - os.path.getmtime(fpath)) > (max_age_hours * 3600):
                            os.remove(fpath)
                    except OSError:
                        pass
    except Exception:
        pass


class ItemPagination(PageNumberPagination):
    page_size = 100
    page_size_query_param = 'page_size'
    max_page_size = 500

from django.db.models import Case, When, Value, IntegerField, Q

class PriorityOrderingFilter(filters.OrderingFilter):
    def filter_queryset(self, request, queryset, view):
        ordering = self.get_ordering(request, queryset, view)
        
        search_value = request.query_params.get('search', '').strip()
        search_fields = getattr(view, 'search_fields', [])
        
        if search_value and search_fields:
            # Create OR conditions for all search fields starting with the search value
            q_objects = Q()
            for field in search_fields:
                kwargs = {f"{field}__istartswith": search_value}
                q_objects |= Q(**kwargs)
                
            queryset = queryset.annotate(
                match_priority=Case(
                    When(q_objects, then=Value(0)),
                    default=Value(1),
                    output_field=IntegerField()
                )
            )
            if ordering:
                return queryset.order_by(*ordering, 'match_priority')
            else:
                return queryset.order_by('match_priority')
        
        if ordering:
            return queryset.order_by(*ordering)
        return queryset

class ItemFieldDefinitionViewSet(viewsets.ModelViewSet):
    queryset = ItemFieldDefinition.objects.all()
    serializer_class = ItemFieldDefinitionSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_class = ItemFieldDefinitionFilter
    search_fields = ['name', 'label']

    def get_queryset(self):
        return scope_queryset(super().get_queryset(), self.request.user)

    def perform_create(self, serializer):
        # اگر رکورد حذف‌نرم با همان (انبار، نام) وجود دارد، احیا می‌شود؛
        # وگرنه INSERT به قید unique_together دیتابیس می‌خورد (اعتبارسنجی DRF
        # فقط رکوردهای زنده را می‌بیند).
        tombstone = ItemFieldDefinition.all_objects.filter(
            warehouse=serializer.validated_data.get('warehouse'),
            name=serializer.validated_data.get('name'),
            is_deleted=True,
        ).first()
        if tombstone:
            serializer.instance = tombstone
            instance = serializer.save(created_by=self.request.user, is_deleted=False)
        else:
            instance = serializer.save(created_by=self.request.user)

        from accounts.audit_utils import log_audit_event
        log_audit_event(
            user=self.request.user,
            warehouse=instance.warehouse,
            module='warehouses',
            action='CREATE',
            severity='info',
            target_model='ItemFieldDefinition',
            target_object_id=instance.id,
            target_repr=f"تعریف فیلد پویا «{instance.label or instance.name}» در انبار {instance.warehouse.name if instance.warehouse else '—'}",
            details={'name': instance.name, 'label': instance.label, 'field_type': instance.field_type},
            ip_address=getattr(self.request, 'META', {}).get('REMOTE_ADDR')
        )

    def perform_destroy(self, instance):
        from accounts.audit_utils import log_audit_event
        wh = instance.warehouse
        f_repr = f"حذف فیلد پویا «{instance.label or instance.name}»"
        f_id = instance.id
        instance.soft_delete()
        log_audit_event(
            user=self.request.user,
            warehouse=wh,
            module='warehouses',
            action='DELETE',
            severity='warning',
            target_model='ItemFieldDefinition',
            target_object_id=f_id,
            target_repr=f_repr,
            details={'field_id': f_id},
            ip_address=getattr(self.request, 'META', {}).get('REMOTE_ADDR')
        )

    def perform_update(self, serializer):
        old_instance = self.get_object()
        old_name = old_instance.name
        new_instance = serializer.save(modified_by=self.request.user)
        new_name = new_instance.name
        
        # If the system name changed, migrate all item data in this warehouse
        if old_name != new_name:
            items_to_update = []
            items = Item.objects.filter(warehouse=new_instance.warehouse)
            for item in items:
                if item.dynamic_data and old_name in item.dynamic_data:
                    item.dynamic_data[new_name] = item.dynamic_data.pop(old_name)
                    items_to_update.append(item)
            
            if items_to_update:
                now = timezone.now()
                for item in items_to_update:
                    item.updated_at = now  # bulk_update سیگنال auto_now را رد می‌کند
                Item.objects.bulk_update(items_to_update, ['dynamic_data', 'updated_at'])

        from accounts.audit_utils import log_audit_event
        log_audit_event(
            user=self.request.user,
            warehouse=new_instance.warehouse,
            module='warehouses',
            action='UPDATE',
            severity='info',
            target_model='ItemFieldDefinition',
            target_object_id=new_instance.id,
            target_repr=f"ویرایش فیلد پویا «{new_instance.label or new_instance.name}»",
            details={'name': new_instance.name, 'label': new_instance.label, 'old_name': old_name},
            ip_address=getattr(self.request, 'META', {}).get('REMOTE_ADDR')
        )

    @action(detail=False, methods=['post'])
    def copy_from_warehouse(self, request):
        source_warehouse_id = request.data.get('source_warehouse_id')
        target_warehouse_id = request.data.get('target_warehouse_id')
        
        if not source_warehouse_id or not target_warehouse_id:
            return Response({'error': 'شناسه انبار مبدأ و مقصد الزامی است.'}, status=400)
            
        try:
            with transaction.atomic():
                source_fields = ItemFieldDefinition.objects.filter(warehouse_id=source_warehouse_id)
                target_existing_names = set(ItemFieldDefinition.objects.filter(warehouse_id=target_warehouse_id).values_list('name', flat=True))
                # فیلدهای حذف‌نرم مقصد: به‌جای ساخت (که به unique_together می‌خورد) احیا می‌شوند
                target_tombstones = {
                    fd.name: fd
                    for fd in ItemFieldDefinition.all_objects.filter(warehouse_id=target_warehouse_id, is_deleted=True)
                }

                new_fields = []
                resurrected = 0
                for field in source_fields:
                    if field.name in target_existing_names:
                        continue
                    tombstone = target_tombstones.get(field.name)
                    if tombstone:
                        tombstone.label = field.label
                        tombstone.field_type = field.field_type
                        tombstone.is_required = field.is_required
                        tombstone.default_value = field.default_value
                        tombstone.is_active = field.is_active
                        tombstone.is_deleted = False
                        tombstone.save()
                        resurrected += 1
                    else:
                        new_field = ItemFieldDefinition(
                            warehouse_id=target_warehouse_id,
                            name=field.name,
                            label=field.label,
                            field_type=field.field_type,
                            is_required=field.is_required,
                            default_value=field.default_value,
                            is_active=field.is_active,
                            created_by=request.user
                        )
                        new_fields.append(new_field)

                if new_fields:
                    ItemFieldDefinition.objects.bulk_create(new_fields)

                from accounts.audit_utils import log_audit_event
                log_audit_event(
                    user=request.user,
                    warehouse=Warehouse.objects.filter(id=target_warehouse_id).first(),
                    module='warehouses',
                    action='CREATE',
                    severity='info',
                    target_model='ItemFieldDefinition',
                    target_repr=f"کپی {len(new_fields) + resurrected} فیلد پویا از انبار #{source_warehouse_id} به #{target_warehouse_id}",
                    details={'source_warehouse_id': source_warehouse_id, 'target_warehouse_id': target_warehouse_id, 'copied_count': len(new_fields) + resurrected},
                    ip_address=getattr(request, 'META', {}).get('REMOTE_ADDR')
                )

                return Response({'message': f'{len(new_fields) + resurrected} فیلد با موفقیت کپی شد.'}, status=200)
        except Exception as e:
            return Response({'error': str(e)}, status=500)

class ItemViewSet(DeleteImpactMixin, viewsets.ModelViewSet):
    queryset = Item.objects.all()
    serializer_class = ItemSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, PriorityOrderingFilter]
    filterset_class = ItemFilter
    pagination_class = ItemPagination
    search_fields = ['fa_unic_code', 'description', 'po', 'pl', 'pk_number', 'my_tag']
    ordering_fields = '__all__'
    parser_classes = (MultiPartParser, FormParser, *viewsets.ModelViewSet.parser_classes)

    def get_queryset(self):
        # photo_prefetch جلوی N+1 عکس‌ها را می‌گیرد: بدون آن هر ردیف یک COUNT و
        # یک SELECT جدا برای بندانگشتی می‌زد (صفحه ۱۰۰ ردیفی = ۲۰۰ کوئری اضافه).
        queryset = super().get_queryset().prefetch_related(photo_prefetch())
        return scope_queryset(queryset, self.request.user)

    def perform_create(self, serializer):
        from accounts.audit_utils import log_audit_event
        # احیای رکورد حذف‌نرم با همان (انبار، کد یکتا) به‌جای INSERT تکراری
        tombstone = Item.all_objects.filter(
            warehouse=serializer.validated_data.get('warehouse'),
            fa_unic_code=serializer.validated_data.get('fa_unic_code'),
            is_deleted=True,
        ).first()
        if tombstone:
            serializer.instance = tombstone
            instance = serializer.save(is_deleted=False)
            # احیای کالا باید عکس‌هایش را هم برگرداند؛ حذف آن‌ها را tombstone کرده بود.
            _restore_item_photos_cascade([instance.id])
        else:
            instance = serializer.save()

        try:
            log_audit_event(
                module='docs',
                action='CREATE',
                target_model='Item',
                target_object_id=instance.id,
                target_repr=f"{instance.fa_unic_code or ''} - {instance.description or ''}"[:255],
                warehouse=instance.warehouse,
                user=self.request.user if self.request.user.is_authenticated else None,
                after_state=model_to_dict(instance, exclude=['photo'])
            )
        except Exception:
            pass

    def perform_update(self, serializer):
        from accounts.audit_utils import log_audit_event, calculate_model_diff
        old_instance = self.get_object()
        before_state = model_to_dict(old_instance, exclude=['photo'])
        instance = serializer.save()
        after_state = model_to_dict(instance, exclude=['photo'])
        diff_b, diff_a = calculate_model_diff(before_state, after_state)
        if diff_b or diff_a:
            try:
                log_audit_event(
                    module='docs',
                    action='UPDATE',
                    target_model='Item',
                    target_object_id=instance.id,
                    target_repr=f"{instance.fa_unic_code or ''} - {instance.description or ''}"[:255],
                    warehouse=instance.warehouse,
                    user=self.request.user if self.request.user.is_authenticated else None,
                    before_state=diff_b,
                    after_state=diff_a
                )
            except Exception:
                pass

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        
        user = request.user
        if not (user.is_superuser or user.has_perm('accounts.perm_wh_edit')):
            instance = self.get_object()
            from warehouses.services import get_setting
            
            if user.has_perm('accounts.view_sys_counter'):
                perms = get_setting('field_permissions_counter', instance.warehouse_id) or {}
            elif user.has_perm('accounts.view_sys_supervisor') or user.has_perm('accounts.view_sys_financial'):
                perms = get_setting('field_permissions_doc', instance.warehouse_id) or {}
            else:
                perms = {}
                
            editable_fields = {k for k, v in perms.items() if isinstance(v, dict) and v.get('editable')}
            
            invalid_fields = []
            for field in request.data.keys():
                if field == 'dynamic_data':
                    dyn_data = request.data.get('dynamic_data')
                    if isinstance(dyn_data, str):
                        try:
                            import json
                            dyn_data = json.loads(dyn_data)
                        except:
                            dyn_data = {}
                    if isinstance(dyn_data, dict):
                        for dyn_k in dyn_data.keys():
                            if dyn_k not in editable_fields:
                                invalid_fields.append(dyn_k)
                elif field not in editable_fields and field not in ['id', 'warehouse']:
                    invalid_fields.append(field)
            
            if invalid_fields:
                from accounts.audit_utils import log_audit_event
                log_audit_event(
                    user=user,
                    module='docs',
                    action='UPDATE',
                    severity='warning',
                    target_model='Item',
                    target_object_id=instance.id,
                    target_repr=f"تلاش غیرمجاز برای ویرایش فیلدها",
                    warehouse=instance.warehouse,
                    details={'invalid_fields': invalid_fields},
                    ip_address=getattr(request, 'META', {}).get('REMOTE_ADDR')
                )
                return Response({
                    'error': 'شما مجوز ویرایش این فیلدها را ندارید.',
                    'invalid_fields': invalid_fields
                }, status=400)
                
        # Re-inject partial since kwargs is passed to super
        kwargs['partial'] = partial
        return super().update(request, *args, **kwargs)

    def perform_destroy(self, instance):
        from accounts.audit_utils import log_audit_event
        try:
            log_audit_event(
                module='docs',
                action='DELETE',
                severity='critical',
                target_model='Item',
                target_object_id=instance.id,
                target_repr=f"{instance.fa_unic_code or ''} - {instance.description or ''}"[:255],
                warehouse=instance.warehouse,
                user=self.request.user if self.request.user.is_authenticated else None,
                before_state=model_to_dict(instance, exclude=['photo'])
            )
        except Exception:
            pass

        with transaction.atomic():
            _soft_delete_items_cascade(Item.objects.filter(id=instance.id))

    def get_permissions(self):
        from accounts.permissions import HasMenuAccess, CanManageItemPhotos
        from rest_framework.permissions import AllowAny, IsAuthenticated

        if self.action == 'photos':
            # خواندن عکس برای هر کاربر لاگین‌شده، ولی آپلود فقط برای نقش‌هایی که
            # با عکس کالا کار می‌کنند. پیش از این POST هم IsAuthenticated بود.
            if self.request.method in ('GET', 'HEAD', 'OPTIONS'):
                permission_classes = [IsAuthenticated()]
            else:
                permission_classes = [IsAuthenticated(), CanManageItemPhotos()]
        elif self.action in ['list', 'retrieve', 'dashboard_stats', 'export_columns', 'download_template']:
            permission_classes = [IsAuthenticated()]
        elif self.action == 'bulk_assign':
            permission_classes = [HasMenuAccess('perm_rec_dispatch')]
        elif self.action in ['export_excel', 'export_excel_mt']: # I'll just secure export here in case it's added
            permission_classes = [HasMenuAccess('view_sys_export')]
        elif self.action in ['import_excel', 'cancel_import', 'revert_import', 'download_import_log', 'delete_from_excel', 'clear_warehouse_data', 'latest_import', 'parse_headers']:
            permission_classes = [HasMenuAccess('perm_rec_import')]
        elif self.action in ['reject', 'manager_reject']:
            permission_classes = [HasMenuAccess('perm_rec_recount')]
        elif self.action in ['update', 'partial_update', 'bulk_update', 'bulk_tag']:
            permission_classes = [HasMenuAccess('perm_wh_edit') | HasMenuAccess('view_sys_counter') | HasMenuAccess('view_sys_supervisor')]
        else: # create, destroy, etc
            permission_classes = [HasMenuAccess('perm_wh_edit')]

        return permission_classes

    @action(detail=True, methods=['get', 'post'])
    def photos(self, request, pk=None):
        from .views_photos import handle_item_photos_upload, list_item_photos
        item = self.get_object()
        if request.method == 'GET':
            return list_item_photos(request, item)
        return handle_item_photos_upload(request, item)


    @action(detail=False, methods=['post'])
    def export_excel(self, request):
        import openpyxl
        import io
        import zipfile
        from django.http import HttpResponse
        
        data_scope = request.data.get('data_scope', 'all')
        columns_scope = request.data.get('columns_scope', 'all_db')
        columns_list = request.data.get('columns_list', [])
        warehouse_id = request.data.get('warehouse_id') or request.query_params.get('warehouse_id')
        
        from django.http import QueryDict
        original_query_params = request._request.GET
        try:
            q = QueryDict(mutable=True)
            for k, v in request.data.items():
                if isinstance(v, list):
                    q.setlist(k, [str(x) for x in v])
                else:
                    q[k] = str(v)
            request._request.GET = q
            
            if data_scope == 'selected':
                selected_ids = request.data.get('selected_ids', [])
                queryset = self.get_queryset().filter(id__in=selected_ids)
                queryset = self.filter_queryset(queryset)
            else:
                queryset = self.filter_queryset(self.get_queryset())
        finally:
            request._request.GET = original_query_params
            
        expected_fields_dict = self.get_expected_fields(warehouse_id)
        valid_fields = {f.name: f for f in Item._meta.fields if f.name != 'dynamic_data'}
        
        canonical_headers = list(dict.fromkeys(expected_fields_dict.values()))
        if columns_scope == 'all_db':
            headers = canonical_headers
        elif columns_scope in ['visible', 'custom']:
            headers = list(dict.fromkeys([expected_fields_dict[c] for c in columns_list if c in expected_fields_dict]))
            if not headers:
                headers = canonical_headers
        else:
            headers = canonical_headers
            
        header_labels = [(valid_fields[h].verbose_name if hasattr(valid_fields[h], 'verbose_name') and valid_fields[h].verbose_name else h) if h in valid_fields else f"{h} (داینامیک)" for h in headers]
        header_keys = [h for h in headers]
        
        CHUNK_SIZE = 100_000
        total_count = queryset.count()

        def _format_row(item):
            row = []
            for h in headers:
                if h in valid_fields:
                    val = getattr(item, h, '')
                else:
                    val = item.dynamic_data.get(h, '') if item.dynamic_data else ''
                    
                if val is None:
                    val = ''
                elif hasattr(val, 'username'):
                    val = f"{val.first_name} {val.last_name}".strip() or val.username
                elif val.__class__.__name__ == 'Warehouse':
                    val = str(val)
                elif hasattr(val, 'project_name') and getattr(val, 'project_name'):
                    val = getattr(val, 'project_name')
                else:
                    val = str(val)
                row.append(val)
            return row

        if total_count <= CHUNK_SIZE:
            wb = openpyxl.Workbook(write_only=True)
            ws = wb.create_sheet(title="Export")
            ws.append(header_labels)
            ws.append(header_keys)
            
            for item in queryset.iterator(chunk_size=5000):
                ws.append(_format_row(item))
                
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="export_items.xlsx"'
            response['X-Export-Type'] = 'single'
            response['X-Total-Records'] = str(total_count)
            wb.save(response)
            return response
        else:
            # Multi-part ZIP output for very large datasets
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                part_num = 1
                current_wb = openpyxl.Workbook(write_only=True)
                current_ws = current_wb.create_sheet(title=f"Part_{part_num}")
                current_ws.append(header_labels)
                current_ws.append(header_keys)
                
                rows_in_current_part = 0
                for item in queryset.iterator(chunk_size=5000):
                    current_ws.append(_format_row(item))
                    rows_in_current_part += 1
                    
                    if rows_in_current_part >= CHUNK_SIZE:
                        part_buffer = io.BytesIO()
                        current_wb.save(part_buffer)
                        part_buffer.seek(0)
                        zip_file.writestr(f"export_part_{part_num}.xlsx", part_buffer.getvalue())
                        part_buffer.close()
                        
                        part_num += 1
                        current_wb = openpyxl.Workbook(write_only=True)
                        current_ws = current_wb.create_sheet(title=f"Part_{part_num}")
                        current_ws.append(header_labels)
                        current_ws.append(header_keys)
                        rows_in_current_part = 0
                
                if rows_in_current_part > 0:
                    part_buffer = io.BytesIO()
                    current_wb.save(part_buffer)
                    part_buffer.seek(0)
                    zip_file.writestr(f"export_part_{part_num}.xlsx", part_buffer.getvalue())
                    part_buffer.close()
            
            zip_buffer.seek(0)
            response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
            response['Content-Disposition'] = 'attachment; filename="export_items.zip"'
            response['X-Export-Type'] = 'zip'
            response['X-Total-Records'] = str(total_count)
            response['X-Total-Parts'] = str(part_num)
            return response

    @action(detail=False, methods=['get'])
    def export_columns(self, request):
        warehouse_id = request.query_params.get('warehouse_id')
        if warehouse_id:
            try:
                wh_id_val = int(warehouse_id)
            except (ValueError, TypeError):
                return Response({'error': 'شناسه انبار نامعتبر است.'}, status=400)
            if not can_access_warehouse(request.user, wh_id_val):
                return Response({'error': 'شما به این انبار دسترسی ندارید.'}, status=403)
            warehouse_id = wh_id_val

        valid_fields = Item._meta.fields
        columns = []
        for f in valid_fields:
            if f.name == 'dynamic_data': continue
            label = getattr(f, 'verbose_name', '') or f.name
            columns.append({"key": f.name, "label": str(label)})
            
        if warehouse_id:
            dynamic_defs = ItemFieldDefinition.objects.filter(warehouse_id=warehouse_id, is_active=True)
            for d in dynamic_defs:
                columns.append({"key": d.name, "label": f"{d.name} (داینامیک)"})
                
        return Response(columns)

    @action(detail=False, methods=['get'])
    def download_template(self, request):
        import openpyxl
        from openpyxl.styles import PatternFill
        from django.http import HttpResponse

        warehouse_id = request.query_params.get('warehouse_id')
        if warehouse_id:
            try:
                wh_id_val = int(warehouse_id)
            except (ValueError, TypeError):
                return Response({'error': 'شناسه انبار نامعتبر است.'}, status=400)
            if not can_access_warehouse(request.user, wh_id_val):
                return Response({'error': 'شما به این انبار دسترسی ندارید.'}, status=403)
            warehouse_id = wh_id_val

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Template"

        expected_fields = self.get_expected_fields(warehouse_id)
        EXCLUDED_TEMPLATE_FIELDS = {'sync_id', 'is_deleted', 'created_at', 'updated_at', 'created_by', 'modified_by'}
        headers = [h for h in dict.fromkeys(expected_fields.values()) if h not in EXCLUDED_TEMPLATE_FIELDS]

        # Add dynamic fields if warehouse_id is provided
        dynamic_fields = []
        if warehouse_id:
            dynamic_defs = ItemFieldDefinition.objects.filter(warehouse_id=warehouse_id, is_active=True)
            for d in dynamic_defs:
                dynamic_fields.append(d)

        # Write headers
        ws.append(headers)

        # Color system fields in the header
        from warehouses.services import get_setting
        restricted_fields = get_setting('SENSITIVE_EXCEL_FIELDS', warehouse_id) or ['doc_status', 'field_status', 'tag_status']
        
        system_fields = ['id', 'created_at', 'updated_at', 'created_by', 'modified_by'] + restricted_fields
        sys_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid") # Warning color (yellow/orange)
        for col_idx, h in enumerate(headers, 1):
            if h in system_fields:
                ws.cell(row=1, column=col_idx).fill = sys_fill

        # Write sample data
        sample_row_1 = []
        sample_row_2 = []
        valid_fields_dict = {f.name: f for f in Item._meta.fields}

        for h in headers:
            field = valid_fields_dict.get(h)
            
            # System fields mock data
            if h in system_fields:
                if h == 'id':
                    sample_row_1.append(101)
                    sample_row_2.append(102)
                elif h in ['created_at', 'updated_at']:
                    sample_row_1.append('2023-10-01 12:00')
                    sample_row_2.append('2023-10-01 12:05')
                elif h in ['created_by', 'modified_by']:
                    sample_row_1.append('سیستم / کاربر ۱')
                    sample_row_2.append('سیستم / کاربر ۲')
                else:
                    sample_row_1.append('مقدار سیستمی')
                    sample_row_2.append('مقدار سیستمی')
                continue
                
            # Dynamic fields mock data
            dyn_field = next((df for df in dynamic_fields if df.name == h), None)
            if dyn_field:
                if dyn_field.field_type == 'number':
                    sample_row_1.append(10)
                    sample_row_2.append(25)
                elif dyn_field.field_type == 'boolean':
                    sample_row_1.append(True)
                    sample_row_2.append(False)
                elif dyn_field.field_type == 'date':
                    sample_row_1.append('1402/05/10')
                    sample_row_2.append('1402/06/15')
                else:
                    sample_row_1.append('مقدار تست ۱')
                    sample_row_2.append('مقدار تست ۲')
                continue

            if h == 'fa_unic_code':
                sample_row_1.append('FA-10001')
                sample_row_2.append('FA-10002')
            elif h == 'warehouse':
                sample_row_1.append('انبار مرکزی')
                sample_row_2.append('انبار مرکزی')
            elif h in ['inventory', 'bal4miv']:
                sample_row_1.append(100.0)
                sample_row_2.append(50.5)
            elif field and field.get_internal_type() == 'BooleanField':
                sample_row_1.append(False)
                sample_row_2.append(True)
            elif field and field.get_internal_type() == 'DateField':
                sample_row_1.append('2023-01-01')
                sample_row_2.append('2023-02-01')
            else:
                sample_row_1.append('نمونه داده')
                sample_row_2.append('نمونه داده ۲')

        ws.append(sample_row_1)
        ws.append(sample_row_2)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Warehouse_Template.xlsx"'
        wb.save(response)

        return response

    @action(detail=False, methods=['post'])
    def bulk_update(self, request):
        try:
            records = request.data
            if not isinstance(records, list):
                return Response({"error": "Data must be a list of records"}, status=400)
            
            record_dict = {str(r['id']): r for r in records if 'id' in r}
            items = self.get_queryset().filter(id__in=record_dict.keys())
            
            # اعتبارسنجی امنیتی: جلوگیری از ویرایش رکوردهای خارج از انبار مجاز کاربر
            if len(items) != len(record_dict):
                missing_ids = set(record_dict.keys()) - set(str(item.id) for item in items)
                if Item.objects.filter(id__in=missing_ids).exists():
                    return Response({"error": "دسترسی غیرمجاز: یک یا چند قلم کالا خارج از انبار مجاز شماست."}, status=403)
            
            user = request.user
            check_field_perms = not (user.is_superuser or user.has_perm('accounts.perm_wh_edit'))
            warehouse_perms_cache = {}
            from warehouses.services import get_setting

            items_to_update = []
            update_fields = set()
            valid_fields = {f.name: f for f in Item._meta.fields}

            for item in items:
                record = record_dict[str(item.id)]

                # بررسی ماتریس دسترسی فیلدها برای کاربران دارای محدودیت نقش (مانند انبارگردان/سرپرست)
                if check_field_perms:
                    wh_id = item.warehouse_id
                    if wh_id not in warehouse_perms_cache:
                        if user.has_perm('accounts.view_sys_counter'):
                            perms = get_setting('field_permissions_counter', wh_id) or {}
                        elif user.has_perm('accounts.view_sys_supervisor') or user.has_perm('accounts.view_sys_financial'):
                            perms = get_setting('field_permissions_doc', wh_id) or {}
                        else:
                            perms = {}
                        warehouse_perms_cache[wh_id] = {k for k, v in perms.items() if isinstance(v, dict) and v.get('editable')}

                    editable_fields = warehouse_perms_cache[wh_id]
                    invalid_fields = []
                    for key in record.keys():
                        if key in ['id', 'warehouse', 'warehouse_id', 'created_at', 'updated_at', 'created_by', 'modified_by']:
                            continue
                        if key == 'dynamic_data':
                            dyn_data = record.get('dynamic_data')
                            if isinstance(dyn_data, str):
                                try:
                                    import json
                                    dyn_data = json.loads(dyn_data)
                                except Exception:
                                    dyn_data = {}
                            if isinstance(dyn_data, dict):
                                for dyn_k in dyn_data.keys():
                                    if dyn_k not in editable_fields:
                                        invalid_fields.append(dyn_k)
                        elif key not in editable_fields:
                            invalid_fields.append(key)

                    if invalid_fields:
                        from accounts.audit_utils import log_audit_event
                        log_audit_event(
                            user=user,
                            module='docs',
                            action='UPDATE',
                            severity='warning',
                            target_model='Item',
                            target_object_id=item.id,
                            target_repr=f"تلاش غیرمجاز برای ویرایش فیلدهای قفل‌شده در ویرایش دسته‌ای کالا {item.fa_unic_code}",
                            warehouse=item.warehouse,
                            details={'invalid_fields': invalid_fields, 'item_id': item.id},
                            ip_address=getattr(request, 'META', {}).get('REMOTE_ADDR')
                        )
                        return Response({
                            'error': f'شما مجوز ویرایش فیلدهای ({", ".join(invalid_fields)}) را در کالا {item.fa_unic_code} ندارید.',
                            'invalid_fields': invalid_fields
                        }, status=400)

                for key, value in record.items():
                    if key in valid_fields and key != 'id':
                        field = valid_fields[key]
                        if field.is_relation and field.many_to_one:
                            setattr(item, field.attname, value)
                            update_fields.add(field.name)
                        else:
                            setattr(item, key, value)
                            update_fields.add(key)
                items_to_update.append(item)
            
            if items_to_update and update_fields:
                for item in items_to_update:
                    item.updated_at = timezone.now()
                    item.modified_by = request.user
                update_fields.update(['updated_at', 'modified_by'])
                Item.objects.bulk_update(items_to_update, list(update_fields))

                from accounts.audit_utils import log_audit_event
                first_item = items_to_update[0] if items_to_update else None
                log_audit_event(
                    user=request.user,
                    warehouse=getattr(first_item, 'warehouse', None) if first_item else None,
                    module='docs',
                    action='BULK_UPDATE',
                    severity='info',
                    target_model='Item',
                    target_repr=f"ویرایش دسته‌ای {len(items_to_update)} کالا",
                    details={
                        'updated_count': len(items_to_update),
                        'fields_modified': list(update_fields),
                        'item_ids': [item.id for item in items_to_update[:50]]
                    },
                    ip_address=getattr(request, 'META', {}).get('REMOTE_ADDR')
                )
            
            return Response({"success": f"Updated {len(items_to_update)} items"})
        except Exception as e:
            logger.exception("خطا در اجرای bulk_update: %s", e)
            return Response({"error": "خطایی در ذخیره‌سازی دسته‌ای کالاها رخ داد. لطفاً داده‌های ارسالی را بررسی فرمایید."}, status=400)

    @action(detail=False, methods=['post'])
    def bulk_assign(self, request):
        ids = request.data.get('item_ids', request.data.get('ids', []))
        
        # Field Counting Assignments
        field_assignee_id = request.data.get('field_assignee')
        supervisor_id = request.data.get('supervisor_assignee')
        manager_id = request.data.get('manager_assignee')
        
        # Document Phase Assignments
        doc_assignee_id = request.data.get('doc_assignee') or request.data.get('doc_assignee_id')
        doc_supervisor_id = request.data.get('doc_supervisor_assignee') or request.data.get('doc_supervisor_id')
        doc_manager_id = request.data.get('doc_manager_assignee') or request.data.get('doc_manager_id')

        
        field_status = request.data.get('field_status')
        doc_status = request.data.get('doc_status')
        force = request.data.get('force', False)
        select_all = request.data.get('select_all', False)
        filters_dict = request.data.get('filters', {})
        
        if select_all:
            from .filters import ItemFilter
            items = ItemFilter(filters_dict or {}, queryset=self.get_queryset()).qs
            if filters_dict and 'search' in filters_dict and filters_dict['search']:
                s = filters_dict['search']
                items = items.filter(
                    Q(fa_unic_code__icontains=s) |
                    Q(description__icontains=s) |
                    Q(po__icontains=s) |
                    Q(pl__icontains=s) |
                    Q(pk_number__icontains=s) |
                    Q(my_tag__icontains=s)
                )
            if filters_dict and 'warehouse' in filters_dict and filters_dict['warehouse']:
                items = items.filter(warehouse_id=filters_dict['warehouse'])
        elif ids:
            items = self.get_queryset().filter(id__in=ids)
            ids_set = set(str(i) for i in ids)
            if len(items) != len(ids_set):
                missing_ids = ids_set - set(str(item.id) for item in items)
                if Item.objects.filter(id__in=missing_ids).exists():
                    return Response({"error": "دسترسی غیرمجاز: یک یا چند قلم کالا خارج از انبار مجاز شماست."}, status=403)
        else:
            items = self.get_queryset().none()
        
        # بررسی مجوز و اعتبارسنجی در صورت درخواست بازشماری
        if field_status == 'recount':
            if not request.user.is_superuser and not request.user.has_perm('accounts.perm_rec_recount'):
                return Response({'error': 'شما مجوز ثبت درخواست بازشماری (perm_rec_recount) را ندارید.'}, status=403)
                
            invalid_uncounted_count = items.filter(field_status__in=['waiting', 'counting', 'در انتظار شمارش']).count()
            if invalid_uncounted_count > 0:
                return Response({
                    'error': f'تعداد {invalid_uncounted_count} مورد از کالاهای انتخابی هنوز شمارش اولیه نشده‌اند و امکان درخواست بازشماری برای آن‌ها وجود ندارد.'
                }, status=400)
        
        # هشدار برای ارسال مجدد کالایی که CountTask فعال دارد
        if field_status == 'counting' and not force:
            from .models import CountTask
            active_count_items = CountTask.objects.filter(
                item__in=items
            ).exclude(status='FINAL_APPROVED').values('item_id').distinct().count()
            if active_count_items > 0:
                return Response({
                    'warning': True,
                    'message': f'تعداد {active_count_items} قلم از کالاهای انتخاب شده در حال حاضر دارای فرآیند شمارش فعال هستند. آیا از ارجاع مجدد اطمینان دارید؟'
                }, status=200)
                
        # هشدار برای ارسال مجدد کالایی که DocTask فعال دارد
        if doc_status in ['checking', 'processing'] and not force:
            from .models import DocTask
            active_doc_items = DocTask.objects.filter(
                item__in=items
            ).exclude(status='DOC_FINAL_APPROVED').values('item_id').distinct().count()
            if active_doc_items > 0:
                return Response({
                    'warning': True,
                    'message': f'تعداد {active_doc_items} قلم از کالاهای انتخاب شده در حال حاضر دارای فرآیند بررسی اسناد فعال هستند. آیا از ارجاع مجدد اطمینان دارید؟'
                }, status=200)
        
        from django.contrib.auth import get_user_model
        User = get_user_model()
        
        def _get_user_or_none(uid):
            if uid:
                try:
                    return User.objects.get(id=uid)
                except (User.DoesNotExist, ValueError):
                    return None
            return None

        counter_user = _get_user_or_none(field_assignee_id)
        manager_user = _get_user_or_none(manager_id)
        
        doc_worker_user = _get_user_or_none(doc_assignee_id)
        doc_manager_user = _get_user_or_none(doc_manager_id)
        
        supervisor_user = None
        skip_supervisor = False
        if supervisor_id:
            if str(supervisor_id) == 'skip':
                skip_supervisor = True
            else:
                supervisor_user = _get_user_or_none(supervisor_id)
                
        doc_supervisor_user = None
        doc_skip_supervisor = False
        if doc_supervisor_id:
            if str(doc_supervisor_id) == 'skip':
                doc_skip_supervisor = True
            else:
                doc_supervisor_user = _get_user_or_none(doc_supervisor_id)
        
        update_data = {}
        if 'field_assignee' in request.data:
            if counter_user:
                update_data['field_assignee'] = f"{counter_user.first_name} {counter_user.last_name}".strip() or counter_user.username
            else:
                update_data['field_assignee'] = 'استخر عمومی'
                
        if 'doc_assignee' in request.data:
            if doc_worker_user:
                update_data['doc_assignee'] = f"{doc_worker_user.first_name} {doc_worker_user.last_name}".strip() or doc_worker_user.username
            else:
                update_data['doc_assignee'] = 'استخر عمومی'

        if field_status is not None:
            update_data['field_status'] = field_status
            if field_status == 'recount':
                update_data['has_conflict'] = True
        if doc_status is not None:
            update_data['doc_status'] = doc_status
            
        created_count_tasks = False
        created_doc_tasks = False

        with transaction.atomic():
            items_list = list(items.select_for_update())
            if not items_list:
                return Response({'status': 'success', 'updated': 0})

            if update_data:
                update_data['updated_at'] = timezone.now()
                update_data['modified_by'] = request.user
                items.update(**update_data)
                
            from warehouses.services import get_setting
            first_item = items_list[0] if items_list else None
            wh_id = first_item.warehouse_id if first_item else None
            
            # Create CountTasks if it's a field dispatch
            if field_status == 'counting':
                from .models import CountTask
                from common.sync_models import soft_delete_queryset
                
                # باطل‌سازی تسک‌های فعال قبلی این کالاها برای جلوگیری از ایجاد تسک‌های تکراری در کارتابل
                soft_delete_queryset(CountTask.objects.filter(item__in=items_list).exclude(status='FINAL_APPROVED'))
                
                # بررسی تنظیم تایید سرپرست
                req_supervisor = get_setting('require_supervisor_approval', wh_id)
                
                tasks_to_create = []
                for item in items_list:
                    tasks_to_create.append(CountTask(
                        item=item,
                        counter=counter_user,
                        supervisor=supervisor_user if (req_supervisor and not skip_supervisor) else None,
                        skip_supervisor=skip_supervisor,
                        assigned_manager=manager_user,
                        status='PENDING_COUNT',
                        created_by=request.user,
                        modified_by=request.user
                    ))
                if tasks_to_create:
                    CountTask.objects.bulk_create(tasks_to_create)
                    created_count_tasks = True

            # Create or update CountTasks if it's a recount request
            if field_status == 'recount':
                from .models import CountTask, CountTaskHistory
                from warehouses.services import get_setting
                
                for item in items_list:
                    latest_task = CountTask.objects.filter(item=item).order_by('-created_at').first()
                    req_supervisor = get_setting('require_supervisor_approval', item.warehouse_id)
                    target_status = 'MANAGER_REJECTED' if (req_supervisor and (latest_task and latest_task.supervisor)) else 'PENDING_COUNT'
                    
                    if latest_task:
                        prev_counted = latest_task.counted_balance
                        latest_task.status = target_status
                        latest_task.manager_note = 'درخواست بازشماری از تخصیص کالا (ثبت مغایرت)'
                        if target_status == 'PENDING_COUNT':
                            latest_task.counted_balance = None
                        latest_task.modified_by = request.user
                        latest_task.updated_at = timezone.now()
                        latest_task.save()
                        
                        CountTaskHistory.objects.create(
                            task=latest_task,
                            action_by=request.user,
                            action_type='MANAGER_REJECTED',
                            counted_balance=prev_counted,
                            note='درخواست بازشماری از صفحه تخصیص'
                        )
                    else:
                        new_task = CountTask.objects.create(
                            item=item,
                            status=target_status,
                            manager_note='درخواست بازشماری از تخصیص کالا (ثبت مغایرت)',
                            created_by=request.user,
                            modified_by=request.user
                        )
                        CountTaskHistory.objects.create(
                            task=new_task,
                            action_by=request.user,
                            action_type='MANAGER_REJECTED',
                            note='ایجاد تسک بازشماری از صفحه تخصیص'
                        )
                created_count_tasks = True

            # Create DocTasks if it's a document dispatch
            if doc_status == 'processing':
                from .models import DocTask
                import re
                from datetime import date as _dt_date
                
                # حذف تسک‌های فعال قبلی این کالاها برای جلوگیری از ایجاد تسک‌های تکراری در کارتابل اسناد
                DocTask.objects.filter(item__in=items_list).exclude(status='DOC_FINAL_APPROVED').delete()
                
                # بررسی تنظیم تایید سرپرست اسناد
                req_doc_supervisor = get_setting('require_doc_supervisor_approval', wh_id)
                
                doc_tasks_to_create = []
                for item in items_list:
                    # تبدیل امن مقادیر اولیه فیلدهای مالی کالا
                    p_amount = item.price_amount
                    s_price = item.similar_unit_price
                    t_value = item.total_value
                    curr = item.currency if item.currency in ['IRR', 'USD', 'EUR', 'OTHER'] else None
                    inv_type = item.invoice_type if item.invoice_type in ['formal', 'domestic', 'foreign', 'consignment'] else None
                    
                    # صفحه و ردیف فاکتور
                    p_row = int(str(item.page_row).strip()) if item.page_row and str(item.page_row).strip().isdigit() else None
                    inv_page = int(str(item.invoice_page).strip()) if item.invoice_page and str(item.invoice_page).strip().isdigit() else None
                    
                    # مهر و امضا
                    stamp_val = bool(item.stamp) if isinstance(item.stamp, bool) else (str(item.stamp).lower() in ['true', '1', 'بله', 'دارد', 'yes'])
                    sign_val = bool(item.signature) if isinstance(item.signature, bool) else (str(item.signature).lower() in ['true', '1', 'بله', 'دارد', 'yes'])
                    
                    # تاریخ فاکتور
                    inv_date = None
                    if item.invoice_date:
                        inv_date_str = str(item.invoice_date).strip()
                        jalali_match = re.match(r'^(\d{4})[/-](\d{1,2})[/-](\d{1,2})$', inv_date_str)
                        if jalali_match:
                            y, m, d = int(jalali_match.group(1)), int(jalali_match.group(2)), int(jalali_match.group(3))
                            if 1300 <= y <= 1500:
                                try:
                                    import jdatetime
                                    inv_date = jdatetime.date(y, m, d).togregorian()
                                except Exception:
                                    pass
                            elif 1900 <= y <= 2100:
                                try:
                                    inv_date = _dt_date(y, m, d)
                                except Exception:
                                    pass
                        elif isinstance(item.invoice_date, _dt_date):
                            inv_date = item.invoice_date

                    doc_tasks_to_create.append(DocTask(
                        item=item,
                        doc_worker=doc_worker_user,
                        doc_supervisor=doc_supervisor_user if (req_doc_supervisor and not doc_skip_supervisor) else None,
                        skip_supervisor=doc_skip_supervisor,
                        assigned_manager=doc_manager_user,
                        status='PENDING_DOC',
                        price_amount=p_amount,
                        similar_unit_price=s_price,
                        total_value=t_value,
                        currency=curr,
                        invoice_type=inv_type,
                        invoice_date=inv_date,
                        inv_rti_number=item.inv_rti_number or None,
                        added_rti_no=item.added_rti_no or None,
                        page_row=p_row,
                        invoice_page=inv_page,
                        doc_supplier=item.doc_supplier or None,
                        folder_address=item.folder_address or None,
                        stamp=stamp_val,
                        signature=sign_val,
                        created_by=request.user,
                        modified_by=request.user,
                        sync_id=uuid.uuid4()
                    ))
                if doc_tasks_to_create:
                    DocTask.objects.bulk_create(doc_tasks_to_create)
                    created_doc_tasks = True

        if created_count_tasks:
            broadcast_count_task_update()
        if created_doc_tasks:
            broadcast_doc_task_update()

        from accounts.audit_utils import log_audit_event
        log_audit_event(
            user=request.user,
            warehouse=getattr(first_item, 'warehouse', None) if first_item else None,
            module='dispatch',
            action='UPDATE',
            severity='info',
            target_model='Item',
            target_repr=f"تخصیص و ارجاع گروهی {len(items_list)} کالا",
            details={
                'items_count': len(items_list),
                'field_status': field_status,
                'doc_status': doc_status,
                'field_assignee': str(counter_user) if counter_user else update_data.get('field_assignee'),
                'supervisor_assignee': str(supervisor_user) if supervisor_user else ('عدم نیاز' if skip_supervisor else None),
                'manager_assignee': str(manager_user) if manager_user else None,
                'doc_assignee': str(doc_worker_user) if doc_worker_user else update_data.get('doc_assignee'),
                'doc_supervisor_assignee': str(doc_supervisor_user) if doc_supervisor_user else ('عدم نیاز' if doc_skip_supervisor else None),
                'doc_manager_assignee': str(doc_manager_user) if doc_manager_user else None,
            },
            ip_address=getattr(request, 'META', {}).get('REMOTE_ADDR')
        )
            
        return Response({'status': 'success', 'updated': len(items_list)})

    @action(detail=False, methods=['post'])
    def bulk_tag(self, request):
        action = request.data.get('action')
        tag_val = (request.data.get('tag') or '').strip()
        item_ids = request.data.get('item_ids', [])
        select_all = request.data.get('select_all', False)
        filters_dict = request.data.get('filters', {})
        
        if action and (item_ids or select_all):
            if select_all:
                from .filters import ItemFilter
                items = ItemFilter(filters_dict or {}, queryset=self.get_queryset()).qs
                if filters_dict and 'search' in filters_dict and filters_dict['search']:
                    s = filters_dict['search']
                    items = items.filter(
                        Q(fa_unic_code__icontains=s) |
                        Q(description__icontains=s) |
                        Q(po__icontains=s) |
                        Q(pl__icontains=s) |
                        Q(pk_number__icontains=s) |
                        Q(my_tag__icontains=s)
                    )
                if filters_dict and 'warehouse' in filters_dict and filters_dict['warehouse']:
                    items = items.filter(warehouse_id=filters_dict['warehouse'])
            else:
                items = self.get_queryset().filter(id__in=item_ids)
                item_ids_set = set(str(i) for i in item_ids)
                if len(items) != len(item_ids_set):
                    missing_ids = item_ids_set - set(str(item.id) for item in items)
                    if Item.objects.filter(id__in=missing_ids).exists():
                        return Response({"error": "دسترسی غیرمجاز: یک یا چند قلم کالا خارج از انبار مجاز شماست."}, status=403)
            updated_count = 0
            with transaction.atomic():
                for item in items:
                    current_tags = [t.strip() for t in (item.my_tag or '').split('،') if t.strip()]
                    if action == 'add' and tag_val:
                        if tag_val not in current_tags:
                            current_tags.append(tag_val)
                    elif action == 'remove' and tag_val:
                        current_tags = [t for t in current_tags if t != tag_val]
                    elif action == 'clear':
                        current_tags = []
                    
                    new_tag_str = '، '.join(current_tags)
                    if item.my_tag != new_tag_str:
                        item.my_tag = new_tag_str
                        item.updated_at = timezone.now()
                        item.modified_by = request.user
                        item.save(update_fields=['my_tag', 'updated_at', 'modified_by'])
                    updated_count += 1
            return Response({'status': 'success', 'updated': updated_count})

        updates = request.data.get('updates', [])
        updated_count = 0
        if updates:
            with transaction.atomic():
                for up in updates:
                    item_id = up.get('id')
                    tag = up.get('my_tag', '')
                    if item_id:
                        self.get_queryset().filter(id=item_id).update(
                            my_tag=tag,
                            updated_at=timezone.now(), 
                            modified_by=request.user
                        )
                        updated_count += 1
        else:
            # Fallback for old bulk_tag format if any
            ids = request.data.get('ids', [])
            tag = request.data.get('tag')
            items = self.get_queryset().filter(id__in=ids)
            if tag == 'conflict':
                items.update(has_conflict=True, updated_at=timezone.now(), modified_by=request.user)
            updated_count = items.count()
            
        return Response({'status': 'success', 'updated': updated_count})

    def get_expected_fields(self, warehouse_id=None):
        fields = {}
        for f in Item._meta.fields:
            name = f.name
            if name == 'dynamic_data': continue
            if name == 'warehouse':
                fields['warehouse'] = 'warehouse'
            elif name in ['created_by', 'modified_by']:
                fields[name] = name
            else:
                fields[name] = name
                
        if warehouse_id:
            from .models import ItemFieldDefinition
            dynamic_defs = ItemFieldDefinition.objects.filter(warehouse_id=warehouse_id, is_active=True)
            for d in dynamic_defs:
                fields[d.name] = d.name
                fields[d.name.lower()] = d.name
                
        return fields

    @action(detail=False, methods=['post'])
    def parse_headers(self, request):
        from warehouses.services import get_setting
        file_obj = request.FILES.get('file')
        warehouse_id = request.data.get('warehouse_id')
        
        if not file_obj:
            return Response({'error': 'هیچ فایلی ارسال نشده است.'}, status=400)

        if warehouse_id and not can_access_warehouse(request.user, warehouse_id):
            return Response({'error': 'شما به این انبار دسترسی ندارید.'}, status=403)
            
        temp_path = None
        try:
            fd, temp_path = tempfile.mkstemp(suffix='.xlsx')
            with os.fdopen(fd, 'wb') as f:
                for chunk in file_obj.chunks():
                    f.write(chunk)
            
            wb = openpyxl.load_workbook(temp_path, read_only=True, data_only=True)
            ws = wb.active
            
            # Read only the first row for headers to maximize speed
            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
            # Convert to lower case to remove case sensitivity
            raw_headers = [str(val).strip().lower() if val is not None else '' for val in first_row]
            wb.close()
            
            expected_fields = self.get_expected_fields(warehouse_id)
            
            found_fields = []
            for h in raw_headers:
                if h in expected_fields:
                    found_fields.append(expected_fields[h])
            
            all_expected = set(expected_fields.values())
            missing_fields = list(all_expected - set(found_fields))
            
            # Check for sensitive fields
            restricted_fields = get_setting('SENSITIVE_EXCEL_FIELDS', warehouse_id) or ['doc_status', 'field_status', 'tag_status']
            has_restricted = any(f in found_fields for f in restricted_fields)
            is_superuser = request.user.is_superuser if request.user and request.user.is_authenticated else False
            
            return Response({
                'status': 'success', 
                'found_fields': list(set(found_fields)), 
                'missing_fields': list(set(missing_fields)),
                'has_restricted_fields': has_restricted,
                'is_superuser': is_superuser
            })
        except Exception as e:
            logger.exception("Error parsing excel headers")
            return Response({'error': 'خطا در خواندن فایل اکسل. لطفاً از معتبر بودن و فرمت استاندارد فایل اطمینان حاصل کنید.'}, status=400)
        finally:
            if temp_path and os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                except Exception:
                    pass

    @action(detail=False, methods=['post'])
    def cancel_import(self, request):
        import_id = request.data.get('import_id')
        if not import_id:
            return Response({'error': 'شناسه فرآیند الزامی است.'}, status=400)
            
        import re
        if not re.match(r'^[a-zA-Z0-9_-]{1,64}$', str(import_id)):
            return Response({'error': 'شناسه فرآیند نامعتبر است.'}, status=400)

        import_log = ImportLog.objects.filter(import_id=import_id).first()
        if import_log and not can_access_warehouse(request.user, import_log.warehouse_id):
            return Response({'error': 'شما دسترسی لازم برای لغو این فرآیند را ندارید.'}, status=403)

        cache.set(f"cancel_import_{import_id}", True, timeout=3600)
        return Response({'status': 'cancelled'})

    @action(detail=False, methods=['get'])
    def download_import_log(self, request):
        import_id = request.query_params.get('import_id')
        if not import_id:
            return Response({'error': 'شناسه فرآیند الزامی است.'}, status=400)

        import re
        if not re.match(r'^[a-zA-Z0-9_-]{1,64}$', str(import_id)):
            return Response({'error': 'شناسه فرآیند نامعتبر است.'}, status=400)

        import_log = ImportLog.objects.filter(import_id=import_id).first()
        if not import_log:
            return Response({'error': 'فایل لاگ یافت نشد یا منقضی شده است.'}, status=404)

        # کنترل دسترسی کاربر به لاگ (سوپریوزر، ثبت‌کننده لاگ یا دسترسی مجاز به انبار)
        is_owner = (import_log.imported_by_id == request.user.id)
        has_wh_access = can_access_warehouse(request.user, import_log.warehouse_id)
        if not (request.user.is_superuser or is_owner or has_wh_access):
            return Response({'error': 'شما دسترسی لازم برای دریافت این فایل لاگ را ندارید.'}, status=403)

        temp_dir = os.path.abspath(tempfile.gettempdir())
        file_path = os.path.abspath(os.path.join(temp_dir, f"import_log_{import_id}.xlsx"))

        # محافظت قطعی در برابر Path Traversal
        try:
            if os.path.commonpath([temp_dir, file_path]) != temp_dir:
                return Response({'error': 'مسیر فایل نامعتبر است.'}, status=400)
        except ValueError:
            return Response({'error': 'مسیر فایل نامعتبر است.'}, status=400)

        if not os.path.exists(file_path):
            return Response({'error': 'فایل لاگ یافت نشد یا منقضی شده است.'}, status=404)

        with open(file_path, 'rb') as f:
            file_data = f.read()

        response = HttpResponse(
            file_data,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="import_log_{import_id}.xlsx"'
        return response

    @action(detail=False, methods=['post'])
    def import_excel(self, request):
        file_obj = request.FILES.get('file')
        warehouse_id = request.data.get('warehouse_id')

        if warehouse_id is not None and str(warehouse_id).strip() != '':
            try:
                warehouse_id_val = int(warehouse_id)
            except (ValueError, TypeError):
                return Response({'error': 'شناسه انبار نامعتبر است.'}, status=400)

            if not can_access_warehouse(request.user, warehouse_id_val):
                return Response({'error': 'شما به این انبار دسترسی ندارید.'}, status=403)
            warehouse_id = warehouse_id_val
        else:
            warehouse_id = None
            if not can_access_warehouse(request.user, None):
                return Response({'error': 'تعیین انبار مجاز برای کاربران با دسترسی محدود الزامی است.'}, status=403)

        from warehouses.services import get_setting
        sys_conflict_strategy = get_setting('default_conflict_strategy', warehouse_id)

        conflict_strategy = request.data.get('conflict_strategy') or sys_conflict_strategy
        import_tag = request.data.get('import_tag', '')
        import_id = request.data.get('import_id', '')
        is_pre_counted_raw = request.data.get('is_pre_counted', False)
        is_pre_counted = str(is_pre_counted_raw).strip().lower() in ['true', '1', 'yes']

        if is_pre_counted:
            can_pre_count = (
                request.user.is_superuser or 
                request.user.has_perm('accounts.perm_inventory_finalize')
            )
            if not can_pre_count:
                return Response({'error': 'شما مجوز ثبت مستقیم اقلام در وضعیت از قبل شمرده‌شده را ندارید.'}, status=403)

        is_doc_pre_approved_raw = request.data.get('is_doc_pre_approved', False)
        is_doc_pre_approved = str(is_doc_pre_approved_raw).strip().lower() in ['true', '1', 'yes']

        if is_doc_pre_approved:
            can_pre_approve_docs = (
                request.user.is_superuser or 
                request.user.has_perm('accounts.perm_doc_approve_action')
            )
            if not can_pre_approve_docs:
                return Response({'error': 'شما مجوز ثبت مستقیم اقلام در وضعیت اسناد تاییدشده را ندارید.'}, status=403)

        if not file_obj:
            return Response({'error': 'هیچ فایلی ارسال نشده است.'}, status=400)
            
        # Save file to temp to allow streaming response generator to read it
        fd, temp_path = tempfile.mkstemp(suffix='.xlsx')
        with os.fdopen(fd, 'wb') as f:
            for chunk in file_obj.chunks():
                f.write(chunk)
                
        user_id = request.user.id if request.user.is_authenticated else None
        original_file_name = file_obj.name

        def worker(q):
            try:
                q.put(json.dumps({"type": "info", "msg": ">> فایل با موفقیت دریافت شد. در حال خواندن محتوا..."}) + "\n")
                wb = openpyxl.load_workbook(temp_path, data_only=True)
                ws = wb.active
                
                raw_headers = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[1]]
                expected_fields = self.get_expected_fields(warehouse_id)
                
                dynamic_fields_keys = []
                if warehouse_id:
                    dynamic_fields_keys = [d.name for d in ItemFieldDefinition.objects.filter(warehouse_id=warehouse_id, is_active=True)]
                
                found_fields = []
                col_indices = {}
                
                for idx, h in enumerate(raw_headers):
                    if h in expected_fields:
                        found_fields.append(expected_fields[h])
                        col_indices[expected_fields[h]] = idx
                
                all_expected_db_fields = set(expected_fields.values())
                missing_fields = list(all_expected_db_fields - set(found_fields))
                
                if 'fa_unic_code' not in found_fields and 'id' not in found_fields:
                    q.put(json.dumps({"type": "err", "msg": ">> خطای حیاتی: ستون 'fa_unic_code' یا 'id' در فایل اکسل یافت نشد. یکی از این ستون‌ها اجباری است."}) + "\n")
                    q.put(json.dumps({
                        "type": "summary",
                        "status": "failed",
                        "created": 0, "updated": 0, "skipped": 0, "failed": 0,
                        "found_fields": list(set(found_fields)),
                        "missing_fields": list(set(missing_fields)),
                        "error_details": [{"row": 0, "code": "HEADER", "error": "ستون fa_unic_code یا id یافت نشد"}]
                    }) + "\n")
                    return
                
                created = 0
                updated = 0
                skipped = 0
                failed = 0
                error_details = []

                def _mark_item_as_counted(item, counted_qty, note_reason):
                    c_task = CountTask.objects.select_for_update().filter(item=item).first()
                    prev_status = c_task.status if c_task else None
                    prev_balance = c_task.counted_balance if c_task else None

                    if c_task:
                        c_task.status = 'FINAL_APPROVED'
                        c_task.counted_balance = counted_qty
                        c_task.manager_note = note_reason
                        c_task.modified_by_id = user_id
                        c_task.updated_at = timezone.now()
                        c_task.save(update_fields=['status', 'counted_balance', 'manager_note', 'modified_by', 'updated_at'])
                    else:
                        c_task = CountTask.objects.create(
                            item=item,
                            status='FINAL_APPROVED',
                            counted_balance=counted_qty,
                            manager_note=note_reason,
                            created_by_id=user_id,
                            modified_by_id=user_id
                        )

                    audit_note = note_reason
                    if prev_status and prev_status != 'FINAL_APPROVED':
                        audit_note = f"{note_reason} (تغییر وضعیت از: {prev_status}، موجودی قبلی: {prev_balance})"

                    CountTaskHistory.objects.create(
                        task=c_task,
                        action_by_id=user_id,
                        action_type='FINAL_APPROVED',
                        counted_balance=counted_qty,
                        note=audit_note
                    )
                    return c_task

                def _mark_item_docs_as_approved(item, financial_data, note_reason):
                    d_task = DocTask.objects.select_for_update().filter(item=item).first()
                    prev_status = d_task.status if d_task else None

                    def _get_val(key):
                        # ۱. اولویت اول: مقدار موجود در فایل اکسل جاری
                        if key in financial_data and financial_data[key] is not None and str(financial_data[key]).strip() != '':
                            return financial_data[key]
                        # ۲. اولویت دوم: حفظ مقدار قبلی ثبت‌شده توسط کاربر در پرونده تسک اسناد
                        if d_task and getattr(d_task, key, None) is not None and str(getattr(d_task, key, None)).strip() != '':
                            return getattr(d_task, key, None)
                        # ۳. اولویت سوم: استفاده از مقدار موجود در مدل کالا
                        return getattr(item, key, None)

                    def _parse_date_flexible(val):
                        if not val:
                            return None
                        if hasattr(val, 'date') and callable(getattr(val, 'date')):
                            return val.date()
                        if hasattr(val, 'strftime') and hasattr(val, 'year'):
                            return val
                        val_str = str(val).strip()
                        if not val_str:
                            return None
                        # تطبیق تاریخ‌های شمسی یا میلادی چهاررقمی (مانند 1403/05/12 یا 2024-05-12)
                        match = re.match(r'^(\d{4})[/-](\d{1,2})[/-](\d{1,2})', val_str)
                        if match:
                            y, m, d = int(match.group(1)), int(match.group(2)), int(match.group(3))
                            if 1300 <= y <= 1500:
                                try:
                                    import jdatetime
                                    return jdatetime.date(y, m, d).togregorian()
                                except Exception:
                                    pass
                            elif 1900 <= y <= 2100:
                                try:
                                    from datetime import date as _dt_date
                                    return _dt_date(y, m, d)
                                except Exception:
                                    pass
                        try:
                            from django.utils.dateparse import parse_date
                            return parse_date(val_str)
                        except Exception:
                            return None

                    def _parse_positive_int(v):
                        if v is None or str(v).strip() == '':
                            return None
                        v_str = str(v).strip()
                        persian_digits = '۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩'
                        english_digits = '01234567890123456789'
                        v_str = v_str.translate(str.maketrans(persian_digits, english_digits))
                        match = re.search(r'\d+', v_str)
                        if match:
                            try:
                                val_int = int(match.group(0))
                                return val_int if val_int >= 0 else None
                            except (ValueError, TypeError):
                                return None
                        return None

                    def _clean_currency(v):
                        if not v:
                            return None
                        v_str = str(v).strip()
                        if not v_str:
                            return None
                        curr_map = {
                            'ریال': 'IRR',
                            'تومان': 'IRR',
                            'دلار': 'USD',
                            'یورو': 'EUR',
                            'درهم': 'OTHER',
                            'سایر': 'OTHER'
                        }
                        for fa_name, code in curr_map.items():
                            if fa_name in v_str:
                                return code
                        return v_str[:10]

                    def _clean_invoice_type(v):
                        if not v:
                            return None
                        v_str = str(v).strip()
                        if not v_str:
                            return None
                        type_map = {
                            'داخلی': 'domestic',
                            'خارجی': 'foreign',
                            'امانی': 'consignment',
                        }
                        for fa_name, code in type_map.items():
                            if fa_name in v_str:
                                return code
                        return v_str[:20]

                    # بررسی هوشمند وضعیت مهر اسناد با حفظ مقدار پیشین تسک
                    raw_stamp = _get_val('stamp')
                    parsed_stamp = None
                    if raw_stamp is not None:
                        if isinstance(raw_stamp, bool):
                            parsed_stamp = raw_stamp
                        elif str(raw_stamp).strip() in ['دارد', '1', 'true', 'True', 'بله']:
                            parsed_stamp = True
                        elif str(raw_stamp).strip() in ['ندارد', '0', 'false', 'False', 'خیر']:
                            parsed_stamp = False

                    final_stamp = parsed_stamp if parsed_stamp is not None else (d_task.stamp if d_task else False)

                    # بررسی هوشمند وضعیت امضای اسناد با حفظ مقدار پیشین تسک
                    raw_sig = _get_val('signature')
                    parsed_sig = None
                    if raw_sig is not None:
                        if isinstance(raw_sig, bool):
                            parsed_sig = raw_sig
                        elif str(raw_sig).strip() in ['دارد', '1', 'true', 'True', 'بله']:
                            parsed_sig = True
                        elif str(raw_sig).strip() in ['ندارد', '0', 'false', 'False', 'خیر']:
                            parsed_sig = False

                    final_sig = parsed_sig if parsed_sig is not None else (d_task.signature if d_task else False)

                    task_fields = {
                        'status': 'DOC_FINAL_APPROVED',
                        'manager_note': note_reason,
                        'modified_by_id': user_id,
                        'updated_at': timezone.now(),
                        'added_rti_no': str(_get_val('added_rti_no'))[:100] if _get_val('added_rti_no') else None,
                        'inv_rti_number': str(_get_val('inv_rti_number'))[:100] if _get_val('inv_rti_number') else None,
                        'invoice_type': _clean_invoice_type(_get_val('invoice_type')),
                        'invoice_date': _parse_date_flexible(_get_val('invoice_date')),
                        'invoice_page': _parse_positive_int(_get_val('invoice_page')),
                        'page_row': _parse_positive_int(_get_val('page_row')),
                        'doc_supplier': str(_get_val('doc_supplier'))[:255] if _get_val('doc_supplier') else None,
                        'total_value': _parse_decimal(_get_val('total_value')),
                        'price_amount': _parse_decimal(_get_val('price_amount')),
                        'similar_unit_price': _parse_decimal(_get_val('similar_unit_price')),
                        'currency': _clean_currency(_get_val('currency')),
                        'folder_address': str(_get_val('folder_address'))[:500] if _get_val('folder_address') else None,
                        'stamp': final_stamp,
                        'signature': final_sig,
                    }

                    if d_task:
                        for k, v in task_fields.items():
                            setattr(d_task, k, v)
                        d_task.save(update_fields=list(task_fields.keys()))
                    else:
                        d_task = DocTask.objects.create(
                            item=item,
                            created_by_id=user_id,
                            **task_fields
                        )

                    audit_doc_note = note_reason
                    if prev_status and prev_status != 'DOC_FINAL_APPROVED':
                        audit_doc_note = f"{note_reason} (تغییر وضعیت از: {prev_status})"

                    DocTaskHistory.objects.create(
                        task=d_task,
                        action_by_id=user_id,
                        action_type='DOC_FINAL_APPROVED',
                        note=audit_doc_note,
                        data_snapshot=_create_doc_task_snapshot(d_task)
                    )
                    return d_task
                
                # Fetch sensitive fields settings and user permissions
                restricted_fields = get_setting('SENSITIVE_EXCEL_FIELDS', warehouse_id) or ['doc_status', 'field_status', 'tag_status']
                is_superuser = request.user.is_superuser if request.user and request.user.is_authenticated else False
                
                # Setup output workbook for log
                out_wb = openpyxl.Workbook(write_only=True)
                out_ws = out_wb.create_sheet('Log')
                log_headers = list(raw_headers) + ['وضعیت پردازش', 'جزئیات پیام']
                out_ws.append(log_headers)
                
                FILL_COLORS = {
                    'created': PatternFill(start_color="C6F6D5", end_color="C6F6D5", fill_type="solid"),
                    'updated': PatternFill(start_color="BFDBFE", end_color="BFDBFE", fill_type="solid"),
                    'warn': PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid"),
                    'err': PatternFill(start_color="FECDD3", end_color="FECDD3", fill_type="solid")
                }
                
                def append_colored_row(row_data_tuple, status_code, message):
                    row_cells = []
                    for val in list(row_data_tuple) + [status_code, message]:
                        c = WriteOnlyCell(out_ws, value=val)
                        if status_code in FILL_COLORS:
                            c.fill = FILL_COLORS[status_code]
                        row_cells.append(c)
                    out_ws.append(row_cells)

                q.put(json.dumps({"type": "info", "msg": ">> پردازش سطرها شروع شد..."}) + "\n")
                
                try:
                    history_records = []
                    valid_fields_dict = {f.name: f for f in Item._meta.fields}
                    
                    with transaction.atomic():
                        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                            # Check cancel flag
                            if import_id and cache.get(f"cancel_import_{import_id}"):
                                raise Exception("CANCELED_BY_USER")

                            row_data = {}
                            for db_field, col_idx in col_indices.items():
                                row_data[db_field] = row[col_idx]
                            
                            raw_fa = row_data.get('fa_unic_code')
                            fa_unic_code = str(raw_fa).strip() if raw_fa is not None and str(raw_fa).strip() != '' else None
                            
                            raw_id = row_data.pop('id', None)
                            item_id = str(raw_id).strip() if raw_id is not None and str(raw_id).strip() != '' else None
                            
                            if not fa_unic_code and not item_id:
                                failed += 1
                                err_msg = "ستون fa_unic_code و id خالی است"
                                error_details.append({"row": row_idx, "code": "N/A", "error": err_msg})
                                append_colored_row(row, 'err', err_msg)
                                q.put(json.dumps({"type": "err", "msg": f"[ردیف {row_idx}] خطا: {err_msg}"}) + "\n")
                                continue
                            
                            # Ignore Excel data for core system fields
                            for ignore_field in ['created_at', 'updated_at', 'created_by', 'modified_by', 'sync_id', 'is_deleted']:
                                row_data.pop(ignore_field, None)
                                
                            # Ignore sensitive fields if not superuser
                            if not is_superuser:
                                for ignore_field in restricted_fields:
                                    row_data.pop(ignore_field, None)
                                
                            if 'hov_date' in row_data:
                                date_val = row_data['hov_date']
                                if isinstance(date_val, datetime):
                                    row_data['hov_date'] = date_val.date()
                                else:
                                    row_data['hov_date'] = None
                                    
                            excel_tag = row_data.pop('my_tag', '')
                            if not excel_tag: excel_tag = ''
                            excel_tag = str(excel_tag).strip().replace(',', '،')
                            
                            final_tags = []
                            if excel_tag:
                                final_tags.extend([t.strip() for t in excel_tag.split('،') if t.strip()])
                            if import_tag:
                                import_tag_clean = import_tag.replace(',', '،')
                                final_tags.extend([t.strip() for t in import_tag_clean.split('،') if t.strip()])
                            
                            unique_tags = list(set(final_tags))
                            if unique_tags:
                                row_data['my_tag'] = '،'.join(unique_tags)
                            else:
                                row_data['my_tag'] = ''

                            warehouse_str = row_data.pop('warehouse', None)
                            target_warehouse_id = int(warehouse_id) if warehouse_id else None
                            
                            if warehouse_str:
                                wh_str = str(warehouse_str).strip()
                                wh = Warehouse.objects.filter(Q(name__iexact=wh_str) | Q(code__iexact=wh_str)).first()
                                if not wh:
                                    failed += 1
                                    err_msg = f"انبار با نام یا کد '{wh_str}' یافت نشد."
                                    error_details.append({"row": row_idx, "code": fa_unic_code, "error": err_msg})
                                    append_colored_row(row, 'err', err_msg)
                                    q.put(json.dumps({"type": "err", "msg": f"[ردیف {row_idx}] خطا در {fa_unic_code}: {err_msg}"}) + "\n")
                                    continue

                                # اگر فرآیند برای انبار مشخصی اجرا شده، ردیف نباید به انبار دیگری اشاره کند
                                if warehouse_id and wh.id != int(warehouse_id):
                                    failed += 1
                                    err_msg = f"مغایرت انبار: فرآیند برای انبار شناسه {warehouse_id} آغاز شده اما این ردیف انبار '{wh.name}' را مشخص کرده است."
                                    error_details.append({"row": row_idx, "code": fa_unic_code, "error": err_msg})
                                    append_colored_row(row, 'err', err_msg)
                                    q.put(json.dumps({"type": "err", "msg": f"[ردیف {row_idx}] خطا در {fa_unic_code}: {err_msg}"}) + "\n")
                                    continue

                                target_warehouse_id = wh.id

                            if target_warehouse_id and not can_access_warehouse(request.user, target_warehouse_id):
                                failed += 1
                                err_msg = "شما دسترسی لازم به انبار مشخص‌شده در این ردیف را ندارید."
                                error_details.append({"row": row_idx, "code": fa_unic_code, "error": err_msg})
                                append_colored_row(row, 'err', err_msg)
                                q.put(json.dumps({"type": "err", "msg": f"[ردیف {row_idx}] خطا در {fa_unic_code}: {err_msg}"}) + "\n")
                                continue

                            # Extract dynamic data
                            dynamic_data_updates = {}
                            if target_warehouse_id:
                                for d_key in dynamic_fields_keys:
                                    if d_key in row_data and row_data[d_key] is not None:
                                        dynamic_data_updates[d_key] = row_data.pop(d_key)

                            if conflict_strategy == 'replace':
                                defaults = {}
                                for k, v in row_data.items():
                                    if k == 'fa_unic_code':
                                        continue
                                    if v is None:
                                        f = valid_fields_dict.get(k)
                                        if f and f.null:
                                            defaults[k] = None
                                        elif f and f.has_default():
                                            defaults[k] = f.default if not callable(f.default) else f.default()
                                        elif f and isinstance(f, (models.CharField, models.TextField)):
                                            defaults[k] = ''
                                    else:
                                        defaults[k] = v
                            else:
                                defaults = {k: v for k, v in row_data.items() if k != 'fa_unic_code' and v is not None}
                            if target_warehouse_id: defaults['warehouse_id'] = target_warehouse_id
                            if user_id: defaults['modified_by_id'] = user_id

                            item_display_code = fa_unic_code or f"ID:{item_id}"
                            existing_item = None
                            if item_id:
                                try:
                                    clean_id = int(float(item_id))
                                except ValueError:
                                    clean_id = item_id

                                existing_item = Item.objects.filter(id=clean_id).first()
                                if existing_item:
                                    # جلوگیری از سرقت یا جابجایی ناخواسته کالا بین انبارها
                                    if target_warehouse_id and existing_item.warehouse_id != target_warehouse_id:
                                        failed += 1
                                        err_msg = f"مغایرت انبار: کالای شناسه {clean_id} متعلق به انبار دیگری است و امکان جابجایی یا ویرایش آن از این انبار وجود ندارد."
                                        error_details.append({"row": row_idx, "code": item_display_code, "error": err_msg})
                                        append_colored_row(row, 'err', err_msg)
                                        q.put(json.dumps({"type": "err", "msg": f"[ردیف {row_idx}] خطا: {err_msg}"}) + "\n")
                                        continue

                                    if not can_access_warehouse(request.user, existing_item.warehouse_id):
                                        failed += 1
                                        err_msg = f"عدم دسترسی: شما به انبار کالای شناسه {clean_id} دسترسی ندارید."
                                        error_details.append({"row": row_idx, "code": item_display_code, "error": err_msg})
                                        append_colored_row(row, 'err', err_msg)
                                        q.put(json.dumps({"type": "err", "msg": f"[ردیف {row_idx}] خطا: {err_msg}"}) + "\n")
                                        continue

                                    # Check if ID points to a different fa_unic_code
                                    if fa_unic_code:
                                        db_code = str(existing_item.fa_unic_code).strip() if existing_item.fa_unic_code else ""
                                        if db_code and db_code != fa_unic_code:
                                            failed += 1
                                            err_msg = f"مغایرت شناسه: کالا با id={clean_id} دارای کد {db_code} است ولی اکسل کد {fa_unic_code} را ارسال کرده است."
                                            error_details.append({"row": row_idx, "code": fa_unic_code, "error": err_msg})
                                            append_colored_row(row, 'err', err_msg)
                                            q.put(json.dumps({"type": "err", "msg": f"[ردیف {row_idx}] خطا: {err_msg}"}) + "\n")
                                            continue

                            # هرگز انبار یک کالای موجود را از طریق فایل اکسل تغییر ندهید (جلوگیری از سرقت یا جابجایی انبار)
                            if existing_item:
                                defaults.pop('warehouse_id', None)
                            
                            if not existing_item and fa_unic_code and target_warehouse_id:
                                existing_item = Item.objects.filter(fa_unic_code=fa_unic_code, warehouse_id=target_warehouse_id).first()
                                # Check if fa_unic_code points to a different ID
                                if existing_item and item_id:
                                    if str(existing_item.id) != str(item_id).split('.')[0]: # split to handle '1.0'
                                        failed += 1
                                        err_msg = f"مغایرت شناسه: کد {fa_unic_code} متعلق به id={existing_item.id} است ولی اکسل id={item_id} را ارسال کرده است."
                                        error_details.append({"row": row_idx, "code": fa_unic_code, "error": err_msg})
                                        append_colored_row(row, 'err', err_msg)
                                        q.put(json.dumps({"type": "err", "msg": f"[ردیف {row_idx}] خطا: {err_msg}"}) + "\n")
                                        continue
                                
                            # رکورد حذف‌نرم با همین کد؟ (unique_together هنوز فعال است)
                            # → باید احیا شود، وگرنه INSERT خطای دیتابیس می‌دهد.
                            resurrect_tombstone = None
                            if not existing_item and fa_unic_code and target_warehouse_id:
                                resurrect_tombstone = Item.all_objects.filter(
                                    fa_unic_code=fa_unic_code,
                                    warehouse_id=target_warehouse_id,
                                    is_deleted=True,
                                ).first()

                            item_display_code = fa_unic_code or f"ID:{item_id}"
                            
                            # Build merged dynamic_data for this item if needed
                            final_dynamic_data = {}
                            if existing_item and existing_item.dynamic_data:
                                final_dynamic_data = existing_item.dynamic_data.copy()
                            if dynamic_data_updates:
                                final_dynamic_data.update(dynamic_data_updates)
                            
                            if final_dynamic_data:
                                defaults['dynamic_data'] = final_dynamic_data

                            if is_pre_counted:
                                defaults['field_status'] = 'done'
                                ref_item = existing_item or resurrect_tombstone
                                raw_inv = row_data.get('inventory')
                                if (raw_inv is None or str(raw_inv).strip() == '') and ref_item:
                                    inv_val = ref_item.inventory
                                    defaults['inventory'] = ref_item.inventory
                                else:
                                    inv_val = defaults.get('inventory')

                                raw_bal = row_data.get('bal4miv')
                                if (raw_bal is None or str(raw_bal).strip() == '') and ref_item:
                                    bal_val = ref_item.bal4miv
                                    defaults['bal4miv'] = ref_item.bal4miv
                                else:
                                    bal_val = defaults.get('bal4miv')

                                c_dec = _parse_decimal(inv_val)
                                s_dec = _parse_decimal(bal_val)
                                if c_dec is not None and s_dec is not None:
                                    defaults['has_conflict'] = (c_dec != s_dec)
                                elif c_dec is not None:
                                    defaults['has_conflict'] = (c_dec != Decimal('0'))
                                else:
                                    defaults['has_conflict'] = False

                            if is_doc_pre_approved:
                                defaults['doc_status'] = 'done'
                                
                            try:
                                with transaction.atomic():
                                    if resurrect_tombstone:
                                        # احیای رکورد حذف‌نرم: از دید کاربر «رکورد جدید» است،
                                        # اما ردیف قبلی (با همان sync_id) به‌روزرسانی و زنده می‌شود.
                                        from django.core.serializers.json import DjangoJSONEncoder
                                        old_state = model_to_dict(resurrect_tombstone)
                                        old_state_json = json.loads(json.dumps(old_state, cls=DjangoJSONEncoder))

                                        if user_id: defaults['created_by_id'] = user_id
                                        if 'tag_status' not in defaults:
                                            defaults['tag_status'] = 'چاپ نشده'
                                        Item.all_objects.filter(id=resurrect_tombstone.id).update(
                                            fa_unic_code=fa_unic_code,
                                            is_deleted=False,
                                            updated_at=timezone.now(),
                                            **defaults,
                                        )
                                        _restore_item_photos_cascade([resurrect_tombstone.id])
                                        history_records.append(ImportHistory(item=resurrect_tombstone, action='update', previous_state=old_state_json))

                                        if is_pre_counted:
                                            counted_qty = defaults.get('inventory', resurrect_tombstone.inventory)
                                            _mark_item_as_counted(resurrect_tombstone, counted_qty, f"ثبت مستقیم اقلام شمرده‌شده از اکسل ({original_file_name})")

                                        if is_doc_pre_approved:
                                            _mark_item_docs_as_approved(resurrect_tombstone, defaults, f"ثبت مستقیم اسناد تاییدشده از اکسل ({original_file_name})")

                                        created += 1
                                        append_colored_row(row, 'created', 'ثبت رکورد جدید (احیای رکورد حذف‌شده)')
                                        q.put(json.dumps({"type": "created", "msg": f"[ردیف {row_idx}] ثبت رکورد جدید (احیا): {fa_unic_code}"}) + "\n")
                                    elif existing_item:
                                        # Append new tags to existing item's tags
                                        if existing_item.my_tag:
                                            existing_tags = [t.strip() for t in existing_item.my_tag.split('،') if t.strip()]
                                            new_tags = [t.strip() for t in defaults.get('my_tag', '').split('،') if t.strip()]
                                            combined_tags = list(set(existing_tags + new_tags))
                                            defaults['my_tag'] = '،'.join(combined_tags) if combined_tags else ''

                                        if conflict_strategy == 'ignore':
                                            skipped += 1
                                            append_colored_row(row, 'warn', 'نادیده گرفتن رکورد تکراری')
                                            q.put(json.dumps({"type": "warn", "msg": f"[ردیف {row_idx}] نادیده گرفتن رکورد تکراری: {item_display_code}"}) + "\n")
                                            continue
                                        elif conflict_strategy == 'log':
                                            error_details.append({"row": row_idx, "code": item_display_code, "error": "تداخل رکورد (ثبت در لاگ)"})
                                            skipped += 1
                                            append_colored_row(row, 'warn', 'تداخل رکورد (ثبت در لاگ)')
                                            q.put(json.dumps({"type": "warn", "msg": f"[ردیف {row_idx}] تداخل رکورد: {item_display_code}"}) + "\n")
                                            continue
                                        elif conflict_strategy == 'update_empty':
                                            def _is_field_empty(val):
                                                return val is None or (isinstance(val, str) and val.strip() == '')

                                            new_defaults = {
                                                k: v for k, v in defaults.items()
                                                if _is_field_empty(getattr(existing_item, k, None)) and not _is_field_empty(v)
                                            }
                                            # Always update fa_unic_code if empty
                                            if fa_unic_code and _is_field_empty(existing_item.fa_unic_code):
                                                new_defaults['fa_unic_code'] = fa_unic_code
                                            # Always update tag since we append them
                                            if defaults.get('my_tag') and defaults.get('my_tag') != existing_item.my_tag:
                                                new_defaults['my_tag'] = defaults['my_tag']

                                            if is_pre_counted:
                                                new_defaults['field_status'] = 'done'
                                                final_inv = new_defaults.get('inventory', existing_item.inventory)
                                                final_bal = new_defaults.get('bal4miv', existing_item.bal4miv)
                                                c_dec = _parse_decimal(final_inv)
                                                s_dec = _parse_decimal(final_bal)
                                                if c_dec is not None and s_dec is not None:
                                                    new_defaults['has_conflict'] = (c_dec != s_dec)
                                                elif c_dec is not None:
                                                    new_defaults['has_conflict'] = (c_dec != Decimal('0'))
                                                else:
                                                    new_defaults['has_conflict'] = False

                                            if is_doc_pre_approved:
                                                new_defaults['doc_status'] = 'done'

                                            if new_defaults:
                                                from django.core.serializers.json import DjangoJSONEncoder
                                                old_state = model_to_dict(existing_item)
                                                old_state_json = json.loads(json.dumps(old_state, cls=DjangoJSONEncoder))

                                                Item.objects.filter(id=existing_item.id).update(**{**new_defaults, 'updated_at': timezone.now()})
                                                history_records.append(ImportHistory(item=existing_item, action='update', previous_state=old_state_json))

                                                if is_pre_counted:
                                                    counted_qty = new_defaults.get('inventory', getattr(existing_item, 'inventory', None))
                                                    _mark_item_as_counted(existing_item, counted_qty, f"به‌روزرسانی مستقیم اقلام شمرده‌شده از اکسل ({original_file_name})")

                                                if is_doc_pre_approved:
                                                    _mark_item_docs_as_approved(existing_item, new_defaults, f"به‌روزرسانی مستقیم اسناد تاییدشده از اکسل ({original_file_name})")

                                                updated += 1
                                                append_colored_row(row, 'updated', 'تکمیل نواقص رکورد')
                                                q.put(json.dumps({"type": "updated", "msg": f"[ردیف {row_idx}] تکمیل نواقص رکورد: {item_display_code}"}) + "\n")
                                            else:
                                                skipped += 1
                                                append_colored_row(row, 'warn', 'بدون نقص، نادیده گرفته شد')
                                                q.put(json.dumps({"type": "warn", "msg": f"[ردیف {row_idx}] بدون نقص، نادیده گرفته شد: {item_display_code}"}) + "\n")
                                        elif conflict_strategy == 'replace':
                                            if fa_unic_code: defaults['fa_unic_code'] = fa_unic_code

                                            from django.core.serializers.json import DjangoJSONEncoder
                                            old_state = model_to_dict(existing_item)
                                            old_state_json = json.loads(json.dumps(old_state, cls=DjangoJSONEncoder))

                                            Item.objects.filter(id=existing_item.id).update(**{**defaults, 'updated_at': timezone.now()})
                                            history_records.append(ImportHistory(item=existing_item, action='update', previous_state=old_state_json))

                                            if is_pre_counted:
                                                counted_qty = defaults.get('inventory', getattr(existing_item, 'inventory', None))
                                                _mark_item_as_counted(existing_item, counted_qty, f"به‌روزرسانی مستقیم اقلام شمرده‌شده از اکسل ({original_file_name})")

                                            if is_doc_pre_approved:
                                                _mark_item_docs_as_approved(existing_item, defaults, f"به‌روزرسانی مستقیم اسناد تاییدشده از اکسل ({original_file_name})")

                                            updated += 1
                                            append_colored_row(row, 'updated', 'بروزرسانی کامل رکورد')
                                            q.put(json.dumps({"type": "updated", "msg": f"[ردیف {row_idx}] بروزرسانی رکورد: {item_display_code}"}) + "\n")
                                    else:
                                        if not target_warehouse_id:
                                            failed += 1
                                            err_msg = "بدون انبار امکان ساخت رکورد جدید نیست"
                                            error_details.append({"row": row_idx, "code": item_display_code, "error": err_msg})
                                            append_colored_row(row, 'err', err_msg)
                                            q.put(json.dumps({"type": "err", "msg": f"[ردیف {row_idx}] خطا در {item_display_code}: {err_msg}"}) + "\n")
                                            continue

                                        if not fa_unic_code:
                                            failed += 1
                                            err_msg = "برای ساخت رکورد جدید، کد یکتا (fa_unic_code) الزامی است"
                                            error_details.append({"row": row_idx, "code": "N/A", "error": err_msg})
                                            append_colored_row(row, 'err', err_msg)
                                            q.put(json.dumps({"type": "err", "msg": f"[ردیف {row_idx}] خطا: {err_msg}"}) + "\n")
                                            continue

                                        if user_id: defaults['created_by_id'] = user_id
                                        if 'tag_status' not in defaults:
                                            defaults['tag_status'] = 'چاپ نشده'
                                        new_item = Item.objects.create(fa_unic_code=fa_unic_code, **defaults)
                                        history_records.append(ImportHistory(item=new_item, action='create'))

                                        if is_pre_counted:
                                            counted_qty = new_item.inventory
                                            _mark_item_as_counted(new_item, counted_qty, f"ثبت مستقیم اقلام شمرده‌شده از فایل اکسل: {original_file_name}")

                                        if is_doc_pre_approved:
                                            _mark_item_docs_as_approved(new_item, defaults, f"ثبت مستقیم اسناد تاییدشده از فایل اکسل: {original_file_name}")

                                        created += 1
                                        append_colored_row(row, 'created', 'ثبت رکورد جدید')
                                        q.put(json.dumps({"type": "created", "msg": f"[ردیف {row_idx}] ثبت رکورد جدید: {fa_unic_code}"}) + "\n")
                            except Exception as row_err:
                                if str(row_err) == "CANCELED_BY_USER":
                                    raise
                                failed += 1
                                from django.db import DatabaseError
                                if isinstance(row_err, DatabaseError):
                                    logger.warning(f"[Import Row {row_idx}] Database error for {item_display_code}: {row_err}")
                                    err_msg = "خطا در قالب داده یا محدودیت‌های پایگاه داده (طول فیلد یا نوع نامعتبر)"
                                else:
                                    err_msg = str(row_err)
                                error_details.append({"row": row_idx, "code": item_display_code, "error": err_msg})
                                append_colored_row(row, 'err', err_msg)
                                q.put(json.dumps({"type": "err", "msg": f"[ردیف {row_idx}] خطا در {item_display_code}: {err_msg}"}) + "\n")

                        log_record = ImportLog.objects.create(
                            import_id=import_id,
                            warehouse_id=target_warehouse_id if 'target_warehouse_id' in locals() and target_warehouse_id else None,
                            imported_by_id=user_id,
                            file_name=original_file_name,
                            records_created=created,
                            records_updated=updated,
                            records_skipped=skipped,
                            records_failed=failed,
                            conflict_strategy=conflict_strategy,
                            error_details=error_details
                        )
                        
                        for hr in history_records:
                            hr.import_log = log_record
                        ImportHistory.objects.bulk_create(history_records)

                    req_wh_id = int(warehouse_id) if (warehouse_id and str(warehouse_id).isdigit()) else None
                    effective_wh_id = req_wh_id or (target_warehouse_id if 'target_warehouse_id' in locals() and target_warehouse_id else None)

                    if is_pre_counted:
                        try:
                            if effective_wh_id:
                                broadcast_count_task_update(warehouse_id=effective_wh_id)
                            else:
                                broadcast_count_task_update()
                        except Exception as bc_err:
                            logger.warning(f"Failed to broadcast count task update after pre-counted import: {bc_err}")

                    if is_doc_pre_approved:
                        try:
                            if effective_wh_id:
                                broadcast_doc_task_update(warehouse_id=effective_wh_id)
                            else:
                                broadcast_doc_task_update()
                        except Exception as bc_err:
                            logger.warning(f"Failed to broadcast doc task update after pre-approved docs import: {bc_err}")

                    # ثبت لاگ کلان ممیزی برای بارگذاری اکسل (خارج از تراکنش جهت جلوگیری از رول‌بک ایمپورت در صورت خطای جانبی)
                    try:
                        from accounts.audit_utils import log_audit_event
                        from django.contrib.auth import get_user_model
                        actor = get_user_model().objects.filter(id=user_id).first() if user_id else None
                        target_wh = Warehouse.objects.filter(id=effective_wh_id).first() if effective_wh_id else None
                        log_audit_event(
                            user=actor,
                            warehouse=target_wh,
                            module='docs',
                            action='IMPORT',
                            severity='info',
                            target_model='Item',
                            target_repr=f"بارگذاری اکسل کالاها ({original_file_name})",
                            details={
                                'import_id': import_id,
                                'file_name': original_file_name,
                                'is_pre_counted': is_pre_counted,
                                'is_doc_pre_approved': is_doc_pre_approved,
                                'records_created': created,
                                'records_updated': updated,
                                'records_skipped': skipped,
                                'records_failed': failed,
                                'conflict_strategy': conflict_strategy
                            }
                        )
                    except Exception as log_err:
                        logger.warning(f"Failed to log audit event for excel import: {log_err}")

                    # Save the colored log workbook
                    if import_id:
                        try:
                            _cleanup_old_import_logs(max_age_hours=24)
                            out_file_path = os.path.join(tempfile.gettempdir(), f"import_log_{import_id}.xlsx")
                            out_wb.save(out_file_path)
                        except Exception as wb_err:
                            logger.warning(f"Failed to save import log workbook: {wb_err}")

                except Exception as ex:
                    if str(ex) == "CANCELED_BY_USER":
                        q.put(json.dumps({
                            "type": "summary",
                            "status": "cancelled",
                            "created": 0, "updated": 0, "skipped": 0, "failed": 0,
                            "found_fields": [], "missing_fields": [],
                            "error_details": []
                        }) + "\n")
                        return
                    else:
                        traceback.print_exc()
                        q.put(json.dumps({"type": "err", "msg": f">> خطای بحرانی در فرآیند: {str(ex)}"}) + "\n")
                        q.put(json.dumps({
                            "type": "summary",
                            "status": "failed",
                            "created": 0,
                            "updated": 0,
                            "skipped": skipped if 'skipped' in locals() else 0,
                            "failed": (failed if 'failed' in locals() else 0) + 1,
                            "found_fields": list(set(found_fields)) if 'found_fields' in locals() else [],
                            "missing_fields": list(set(missing_fields)) if 'missing_fields' in locals() else [],
                            "error_details": (error_details if 'error_details' in locals() else []) + [{"row": 0, "code": "CRASH", "error": f"خطای سیستمی: {str(ex)}"}]
                        }) + "\n")
                        return

                q.put(json.dumps({
                    "type": "summary",
                    "status": "success",
                    "created": created,
                    "updated": updated,
                    "skipped": skipped,
                    "failed": failed,
                    "found_fields": list(set(found_fields)),
                    "missing_fields": list(set(missing_fields)),
                    "error_details": error_details
                }) + "\n")
                
            except Exception as e:
                traceback.print_exc()
                q.put(json.dumps({"type": "err", "msg": f">> خطای سیستمی: {str(e)}"}) + "\n")
                q.put(json.dumps({
                    "type": "summary",
                    "status": "failed",
                    "created": 0,
                    "updated": 0,
                    "skipped": 0,
                    "failed": 1,
                    "found_fields": [],
                    "missing_fields": [],
                    "error_details": [{"row": 0, "code": "FATAL", "error": str(e)}]
                }) + "\n")
            finally:
                q.put(None)
                if import_id:
                    cache.delete(f"cancel_import_{import_id}")
                if os.path.exists(temp_path):
                    try:
                        os.remove(temp_path)
                    except:
                        pass
                try:
                    from django.db import connection
                    connection.close()
                except Exception:
                    pass

        async def stream_logs_async():
            q = queue.Queue()
            thread = threading.Thread(target=worker, args=(q,))
            thread.start()
            
            while True:
                chunk = await asyncio.to_thread(q.get)
                if chunk is None:
                    break
                yield chunk

        response = StreamingHttpResponse(
            stream_logs_async(), 
            content_type='application/x-ndjson'
        )
        response['X-Accel-Buffering'] = 'no'
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response['X-Content-Type-Options'] = 'nosniff'
        return response

    @action(detail=False, methods=['post'])
    def revert_import(self, request):
        import_id = request.data.get('import_id')
        if not import_id:
            return Response({'error': 'شناسه فرآیند الزامی است.'}, status=400)
            
        try:
            import_log = ImportLog.objects.get(import_id=import_id)
        except ImportLog.DoesNotExist:
            return Response({'error': 'فرآیندی با این شناسه یافت نشد.'}, status=404)

        if not can_access_warehouse(request.user, import_log.warehouse_id):
            return Response({'error': 'شما به این انبار دسترسی ندارید.'}, status=403)

        if import_log.is_reverted:
            return Response({'error': 'این فرآیند قبلاً بازگردانی شده است.'}, status=400)
            
        time_elapsed = (timezone.now() - import_log.imported_at).total_seconds()
        if time_elapsed > 900: # 15 minutes
            return Response({'error': 'مهلت ۱۵ دقیقه‌ای برای بازگردانی این فرآیند به پایان رسیده است.'}, status=400)
            
        try:
            with transaction.atomic():
                histories = import_log.histories.all()
                items_to_create = []
                for history in histories:
                    if history.action == 'create':
                        if history.item:
                            # بازگردانی ایجاد = حذف (نرم) رکورد ساخته‌شده به همراه تسک‌های اسناد ایجادشده
                            _soft_delete_items_cascade(Item.objects.filter(id=history.item_id))
                            DocTask.objects.filter(item_id=history.item_id).delete()
                    elif history.action == 'update' or history.action == 'delete':
                        if history.previous_state:
                            state = history.previous_state.copy()

                            # Handle foreign keys correctly for both update and create
                            fk_fields = ['warehouse', 'created_by', 'modified_by']
                            for fk in fk_fields:
                                if fk in state and isinstance(state[fk], int):
                                    state[f'{fk}_id'] = state.pop(fk)

                            if history.action == 'update' and history.item:
                                # بازگردانی مشخصات کالا
                                Item.all_objects.filter(id=history.item.id).update(**{**state, 'updated_at': timezone.now()})

                                # بازگردانی وضعیت تسک شمارش در صورت لزوم
                                prev_field_status = state.get('field_status')
                                if prev_field_status and prev_field_status != 'done':
                                    c_task = CountTask.objects.filter(item=history.item).first()
                                    if c_task:
                                        last_hist = CountTaskHistory.objects.filter(task=c_task).exclude(
                                            action_by_id=import_log.imported_by_id, action_type='FINAL_APPROVED'
                                        ).order_by('-created_at').first()
                                        if last_hist:
                                            c_task.status = last_hist.action_type
                                            c_task.counted_balance = last_hist.counted_balance
                                            c_task.save(update_fields=['status', 'counted_balance', 'updated_at'])
                                        else:
                                            c_task.status = 'PENDING_COUNT'
                                            c_task.counted_balance = None
                                            c_task.save(update_fields=['status', 'counted_balance', 'updated_at'])

                                # بازگردانی وضعیت تسک اسناد در صورت لزوم
                                prev_doc_status = state.get('doc_status')
                                if prev_doc_status and prev_doc_status != 'done':
                                    d_task = DocTask.objects.filter(item=history.item).first()
                                    if d_task:
                                        last_d_hist = DocTaskHistory.objects.filter(task=d_task).exclude(
                                            action_by_id=import_log.imported_by_id, action_type='DOC_FINAL_APPROVED'
                                        ).order_by('-created_at').first()
                                        if last_d_hist:
                                            d_task.status = last_d_hist.action_type
                                            d_task.save(update_fields=['status', 'updated_at'])
                                        else:
                                            d_task.status = 'PENDING_DOC'
                                            d_task.save(update_fields=['status', 'updated_at'])
                            elif history.action == 'delete':
                                # حذف حالا نرم است؛ ردیف tombstone هنوز با همه داده‌ها موجود است
                                # → فقط احیا می‌شود. اگر (به هر دلیل) واقعاً حذف شده بود، بازسازی.
                                tombstone_ids = list(
                                    Item.all_objects.filter(
                                        warehouse_id=state.get('warehouse_id'),
                                        fa_unic_code=state.get('fa_unic_code'),
                                        is_deleted=True,
                                    ).values_list('id', flat=True)
                                )
                                if tombstone_ids:
                                    Item.all_objects.filter(id__in=tombstone_ids).update(
                                        is_deleted=False, updated_at=timezone.now()
                                    )
                                    # عکس‌ها هم با حذف کالا tombstone شده بودند
                                    _restore_item_photos_cascade(tombstone_ids)
                                else:
                                    items_to_create.append(Item(**state))
                
                if items_to_create:
                    Item.objects.bulk_create(items_to_create, ignore_conflicts=True)
                            
                import_log.is_reverted = True
                import_log.save()

                try:
                    wh_id = import_log.warehouse_id
                    if wh_id:
                        broadcast_count_task_update(warehouse_id=wh_id)
                        broadcast_doc_task_update(warehouse_id=wh_id)
                    else:
                        broadcast_count_task_update()
                        broadcast_doc_task_update()
                except Exception as bc_err:
                    logger.warning(f"Failed to broadcast task updates after revert_import: {bc_err}")

                from accounts.audit_utils import log_audit_event
                log_audit_event(
                    user=request.user,
                    warehouse=import_log.warehouse,
                    module='docs',
                    action='ROLLBACK',
                    severity='warning',
                    target_model='Item',
                    target_repr=f"بازگردانی بارگذاری اکسل ({import_log.file_name})",
                    details={
                        'import_id': import_id,
                        'file_name': import_log.file_name,
                        'affected_records': len(histories)
                    },
                    ip_address=getattr(request, 'META', {}).get('REMOTE_ADDR')
                )
                
            return Response({'status': 'success', 'msg': 'فرآیند با موفقیت بازگردانی شد.', 'affected_records': len(histories)})
        except Exception as e:
            logger.exception(f"Error reverting import {import_id}")
            return Response({'error': 'خطا در بازگردانی فرآیند در دیتابیس.'}, status=500)

    @action(detail=False, methods=['post'])
    def clear_warehouse_data(self, request):
        warehouse_id = request.data.get('warehouse_id')
        if not warehouse_id:
            return Response({'error': 'شناسه انبار الزامی است.'}, status=400)

        try:
            warehouse_id_val = int(warehouse_id)
        except (ValueError, TypeError):
            return Response({'error': 'شناسه انبار نامعتبر است.'}, status=400)

        if not can_access_warehouse(request.user, warehouse_id_val):
            return Response({'error': 'شما به این انبار دسترسی ندارید.'}, status=403)
        warehouse_id = warehouse_id_val

        try:
            items_deleted = 0
            user_id = request.user.id if request.user.is_authenticated else None
            import_id = f"clear_{datetime.now().strftime('%Y%m%d%H%M%S')}_{warehouse_id}"
            
            with transaction.atomic():
                import_log = ImportLog.objects.create(
                    import_id=import_id,
                    warehouse_id=warehouse_id,
                    imported_by_id=user_id,
                    file_name="حذف تمامی داده‌های انبار"
                )
                
                items_qs = Item.objects.filter(warehouse_id=warehouse_id)
                
                # We need to bulk read the items to save them
                from django.core.serializers.json import DjangoJSONEncoder
                histories = []
                for item in items_qs.iterator(chunk_size=1000):
                    state = model_to_dict(item)
                    state = json.loads(json.dumps(state, cls=DjangoJSONEncoder))
                    
                    histories.append(ImportHistory(
                        import_log=import_log,
                        item=None, 
                        action='delete',
                        previous_state=state
                    ))
                    if len(histories) >= 500:
                        ImportHistory.objects.bulk_create(histories, batch_size=500)
                        histories.clear()
                
                if histories:
                    ImportHistory.objects.bulk_create(histories, batch_size=500)
                    histories.clear()
                    
                deleted = _soft_delete_items_cascade(items_qs)
                items_deleted = deleted
                
                import_log.records_created = items_deleted # Store count here
                import_log.save()

                from accounts.audit_utils import log_audit_event
                log_audit_event(
                    user=request.user,
                    warehouse_id=warehouse_id,
                    module='docs',
                    action='DELETE',
                    severity='critical',
                    target_model='Warehouse',
                    target_object_id=warehouse_id,
                    target_repr=f"پاکسازی کامل داده‌های انبار #{warehouse_id}",
                    details={
                        'operation': 'CLEAR_WAREHOUSE_DATA',
                        'items_deleted': items_deleted,
                        'import_id': import_id
                    },
                    ip_address=getattr(request, 'META', {}).get('REMOTE_ADDR')
                )
                
            return Response({'status': 'success', 'msg': f'{items_deleted} رکورد با موفقیت حذف شدند.', 'import_id': import_id})
        except Exception as e:
            return Response({'error': f'خطا در حذف داده‌ها: {str(e)}'}, status=500)

    @action(detail=False, methods=['get'])
    def latest_import(self, request):
        warehouse_id = request.query_params.get('warehouse_id')
        if not warehouse_id:
            return Response({'error': 'شناسه انبار الزامی است.'}, status=400)

        try:
            warehouse_id_val = int(warehouse_id)
        except (ValueError, TypeError):
            return Response({'error': 'شناسه انبار نامعتبر است.'}, status=400)

        if not can_access_warehouse(request.user, warehouse_id_val):
            return Response({'error': 'شما به این انبار دسترسی ندارید.'}, status=403)
        warehouse_id = warehouse_id_val

        try:
            logs = ImportLog.objects.filter(warehouse_id=warehouse_id, is_reverted=False).order_by('-imported_at')[:10]
            
            recent_imports = []
            for log in logs:
                time_elapsed = (timezone.now() - log.imported_at).total_seconds()
                if time_elapsed <= 900:
                    recent_imports.append({
                        'import_id': log.import_id,
                        'file_name': log.file_name,
                        'imported_at': log.imported_at.isoformat(),
                        'time_remaining_seconds': 900 - int(time_elapsed),
                        'records_affected': log.records_created + log.records_updated
                    })
                
            return Response({
                'latest_import': recent_imports[0] if recent_imports else None,
                'recent_imports': recent_imports
            })
        except Exception as e:
            return Response({'error': str(e)}, status=500)

    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def delete_from_excel(self, request):
        warehouse_id = request.data.get('warehouse_id')
        file_obj = request.FILES.get('file')

        if not warehouse_id or not file_obj:
            return Response({'error': 'پارامترهای warehouse_id و file الزامی است.'}, status=400)

        try:
            warehouse_id_val = int(warehouse_id)
        except (ValueError, TypeError):
            return Response({'error': 'شناسه انبار نامعتبر است.'}, status=400)

        if not can_access_warehouse(request.user, warehouse_id_val):
            return Response({'error': 'شما به این انبار دسترسی ندارید.'}, status=403)
        warehouse_id = warehouse_id_val

        try:
            wb = openpyxl.load_workbook(file_obj, data_only=True)
            sheet = wb.active
            headers = [str(cell.value).strip() if cell.value else '' for cell in sheet[1]]
            
            id_idx = headers.index('id') if 'id' in headers else None
            fa_unic_idx = headers.index('fa_unic_code') if 'fa_unic_code' in headers else None
            
            if id_idx is None and fa_unic_idx is None:
                return Response({'error': 'ستون fa_unic_code یا id در فایل اکسل یافت نشد.'}, status=400)
                
            ids_to_delete = []
            fa_unics_to_delete = []
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                id_val = row[id_idx] if id_idx is not None else None
                fa_val = row[fa_unic_idx] if fa_unic_idx is not None else None
                
                if id_val is not None and str(id_val).strip():
                    raw_str = str(id_val).strip()
                    try:
                        clean_id = int(float(raw_str))
                        ids_to_delete.append(clean_id)
                    except (ValueError, TypeError):
                        ids_to_delete.append(raw_str)
                elif fa_val is not None and str(fa_val).strip():
                    fa_unics_to_delete.append(str(fa_val).strip())
                    
            if not ids_to_delete and not fa_unics_to_delete:
                return Response({'error': 'هیچ شناسه یا کدی در فایل یافت نشد.'}, status=400)
                
            items_deleted = 0
            user_id = request.user.id if request.user.is_authenticated else None
            import_id = f"del_{datetime.now().strftime('%Y%m%d%H%M%S')}_{warehouse_id}"
            
            with transaction.atomic():
                import_log = ImportLog.objects.create(
                    import_id=import_id,
                    warehouse_id=warehouse_id,
                    imported_by_id=user_id,
                    file_name=f"حذف گروهی (اکسل): {file_obj.name}"
                )
                
                query = Q()
                if ids_to_delete:
                    query |= Q(id__in=ids_to_delete)
                if fa_unics_to_delete:
                    query |= Q(fa_unic_code__in=fa_unics_to_delete)
                    
                items_qs = Item.objects.filter(Q(warehouse_id=warehouse_id) & query)
                
                # We need to bulk read the items to save them
                from django.core.serializers.json import DjangoJSONEncoder
                histories = []
                for item in items_qs.iterator(chunk_size=1000):
                    # For delete action, item foreign key might be nullified when item is deleted. 
                    # We store it anyway.
                    state = model_to_dict(item)
                    state = json.loads(json.dumps(state, cls=DjangoJSONEncoder))
                    
                    histories.append(ImportHistory(
                        import_log=import_log,
                        item=None, # item will be deleted, don't set foreign key
                        action='delete',
                        previous_state=state
                    ))
                    if len(histories) >= 500:
                        ImportHistory.objects.bulk_create(histories, batch_size=500)
                        histories.clear()
                
                if histories:
                    ImportHistory.objects.bulk_create(histories, batch_size=500)
                    histories.clear()
                    
                deleted = _soft_delete_items_cascade(items_qs)
                items_deleted = deleted
                
                import_log.records_created = items_deleted # Store count in this field for history
                import_log.save()

                from accounts.audit_utils import log_audit_event
                log_audit_event(
                    user=request.user,
                    warehouse_id=warehouse_id,
                    module='docs',
                    action='DELETE',
                    severity='critical',
                    target_model='Item',
                    target_repr=f"حذف گروهی {items_deleted} کالا از فایل اکسل ({file_obj.name})",
                    details={
                        'file_name': file_obj.name,
                        'items_deleted': items_deleted,
                        'import_id': import_id
                    },
                    ip_address=getattr(request, 'META', {}).get('REMOTE_ADDR')
                )
                
            return Response({'status': 'success', 'msg': f'{items_deleted} رکورد با موفقیت از انبار حذف شدند.', 'import_id': import_id})
        except Exception as e:
            return Response({'error': f'خطا در پردازش فایل: {str(e)}'}, status=500)

    @action(detail=False, methods=['get'])
    def dashboard_stats(self, request):
        import jdatetime
        from datetime import timedelta
        from django.utils import timezone
        from django.db.models import Q
        
        project_id = request.query_params.get('project_id')
        items = Item.objects.all()
        if project_id and project_id != 'ALL':
            items = items.filter(warehouse_id=project_id)
            
        now = timezone.localtime(timezone.now()) if timezone.is_aware(timezone.now()) else timezone.now()
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        
        # Overall
        total_quantity = items.count()
        total_counted = items.exclude(field_status__in=['waiting', 'counting', 'recount', 'در انتظار شمارش', 'بازشماری']).count()
        printed_tags = items.filter(tag_status__in=['printed', 'reprint', 'چاپ شده', 'چاپ مجدد']).count()
        docs_approved = items.filter(doc_status__in=['done', 'approved']).count()
        conflicts = items.filter(Q(has_conflict=True) | Q(field_status='recount')).count()
        done = items.filter(field_status='done', doc_status__in=['done', 'approved']).count()
        
        # Days stats (last 7 days)
        weekly_data = []
        
        for i in range(6, -1, -1):
            d_start = today_start - timedelta(days=i)
            d_end = d_start + timedelta(days=1)
            
            day_items = items.filter(updated_at__gte=d_start, updated_at__lt=d_end)
            
            c_count = day_items.exclude(field_status__in=['waiting', 'counting', 'recount', 'در انتظار شمارش', 'بازشماری']).count()
            c_docs = day_items.filter(doc_status__in=['done', 'approved']).count()
            c_feed = day_items.filter(field_status='done', doc_status__in=['done', 'approved']).count()
            
            jdt = jdatetime.datetime.fromgregorian(datetime=d_start)
            day_label = 'امروز' if i == 0 else ('دیروز' if i == 1 else jdt.strftime('%A'))
            
            weekly_data.append({
                'day': day_label,
                'count': c_count,
                'docs': c_docs,
                'feed': c_feed
            })
            
        today_stats = weekly_data[-1]
        yesterday_stats = weekly_data[-2]
        
        last_7_days_count = sum(d['count'] for d in weekly_data)
        last_7_days_docs = sum(d['docs'] for d in weekly_data)
        last_7_days_feed = sum(d['feed'] for d in weekly_data)
        
        max_val = max([max(d['count'], d['docs'], d['feed']) for d in weekly_data] + [10])
        
        return Response({
            'overall': {
                'total': total_quantity,
                'counted': total_counted,
                'printed': printed_tags,
                'docs_approved': docs_approved,
                'conflicts': conflicts,
                'done': done,
                'ready_to_feed': max(0, total_counted - done)
            },
            'today': today_stats,
            'yesterday': yesterday_stats,
            'last_week_totals': {
                'count': last_7_days_count,
                'docs': last_7_days_docs,
                'feed': last_7_days_feed
            },
            'weekly_data': weekly_data,
            'overallMax': max_val
        })

from .serializers import CountTaskSerializer
from .models import CountTask, CountTaskHistory
from rest_framework import permissions, viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db import transaction
from common.sync_models import soft_delete_queryset
from .serializers import photo_prefetch
from common.warehouse_scope import can_access_warehouse
from .signals import broadcast_count_task_update

class CountTaskViewSet(viewsets.ModelViewSet):
    serializer_class = CountTaskSerializer

    def get_serializer_context(self):
        context = super().get_serializer_context()
        as_role = self.request.query_params.get('as_role')
        user = self.request.user
        
        # اگر نقش یکی از نقش‌های مدیریتی، نظارتی یا پیگیری باشد، یا اکشن خروجی اکسل باشد
        # شمارش نباید کور باشد و اطلاعات موجودی کامل باید ارسال گردد.
        if as_role in ['manager', 'supervisor', 'count_tracking', 'doc_supervisor', 'doc_worker']:
            context['is_blind'] = False
        elif self.action in ['export_excel', 'get_export_columns', 'bulk_manager_approve', 'bulk_manager_reject', 'bulk_approve', 'bulk_reject']:
            context['is_blind'] = False
        elif as_role == 'counter':
            context['is_blind'] = True
        elif user.is_authenticated and (user.is_superuser or user.has_perm('accounts.view_sys_manager_review') or user.has_perm('accounts.view_sys_supervisor') or user.has_perm('accounts.view_wh_stocktaking')):
            if not user.has_perm('accounts.can_act_as_counter') and not user.has_perm('accounts.view_sys_counter'):
                context['is_blind'] = False
        return context

    def get_permissions(self):
        from accounts.permissions import HasMenuAccess
        from rest_framework.permissions import IsAuthenticated
        
        if self.action in ['list', 'retrieve', 'pool_tasks', 'claim_tasks', 'get_export_columns', 'export_excel']:
            permission_classes = [HasMenuAccess('view_sys_counter') | HasMenuAccess('view_sys_supervisor') | HasMenuAccess('view_sys_manager_review') | HasMenuAccess('view_sys_recounts') | HasMenuAccess('view_wh_stocktaking')]
        elif self.action == 'bulk_submit':
            permission_classes = [HasMenuAccess('view_sys_counter') | HasMenuAccess('can_act_as_counter')]
        elif self.action == 'bulk_approve':
            permission_classes = [HasMenuAccess('view_sys_supervisor') | HasMenuAccess('can_act_as_supervisor') | HasMenuAccess('perm_feed_approve_action')]
        elif self.action in ['reject', 'bulk_reject']:
            permission_classes = [HasMenuAccess('perm_rec_recount') | HasMenuAccess('view_sys_supervisor') | HasMenuAccess('view_sys_manager_review') | HasMenuAccess('can_act_as_supervisor')]
        elif self.action in ['manager_reject', 'bulk_manager_reject']:
            permission_classes = [HasMenuAccess('perm_rec_recount') | HasMenuAccess('view_sys_manager_review') | HasMenuAccess('can_act_as_manager') | HasMenuAccess('perm_inventory_finalize')]
        elif self.action in ['bulk_manager_approve', 'bulk_cancel']:
            permission_classes = [HasMenuAccess('view_sys_manager_review') | HasMenuAccess('can_act_as_manager') | HasMenuAccess('perm_inventory_finalize')]
        else: # create, update, etc
            permission_classes = [IsAuthenticated()]
            
        return permission_classes

    def perform_destroy(self, instance):
        # حذف نرم + tombstone تاریخچه، تا کلاینت‌های آفلاین باخبر شوند
        with transaction.atomic():
            soft_delete_queryset(CountTaskHistory.objects.filter(task_id=instance.id))
            instance.soft_delete()

    def get_object(self):
        """
        زیرساخت سینک: اگر pk عددی نبود، به‌عنوان sync_id تفسیر می‌شود
        (کلاینت آفلاین ممکن است فقط sync_id پایدار را داشته باشد).
        """
        pk = self.kwargs.get(self.lookup_url_kwarg or self.lookup_field)
        if pk is not None and not str(pk).isdigit():
            from django.shortcuts import get_object_or_404
            queryset = self.filter_queryset(self.get_queryset())
            obj = get_object_or_404(queryset, sync_id=pk)
            self.check_object_permissions(self.request, obj)
            return obj
        return super().get_object()

    def update(self, request, *args, **kwargs):
        """
        تشخیص تداخل خوش‌بینانه: کلاینت می‌تواند base_updated_at (نسخه‌ای که
        رویش تغییر داده) بفرستد؛ اگر سرور جدیدتر باشد → 409 + رکورد سروری
        برای reconciliation سمت کلاینت. بدون این پارامتر، رفتار قبلی حفظ است.
        """
        base_raw = request.data.get('base_updated_at')
        if base_raw:
            from django.utils.dateparse import parse_datetime
            base = parse_datetime(str(base_raw))
            if base is None:
                return Response({'detail': 'base_updated_at نامعتبر است.'}, status=400)
            if timezone.is_naive(base):
                base = timezone.make_aware(base)
            instance = self.get_object()
            # تلورانس ۱ms برای جلوگیری از تداخل کاذب ناشی از گرد شدن میکروثانیه
            if (instance.updated_at - base).total_seconds() > 0.001:
                return Response({
                    'detail': 'conflict',
                    'server_record': self.get_serializer(instance).data,
                }, status=409)
        return super().update(request, *args, **kwargs)

    def perform_update(self, serializer):
        from rest_framework.exceptions import ValidationError
        instance = serializer.instance
        old_status = instance.status
        old_counted = str(instance.counted_balance)
        new_status = serializer.validated_data.get('status', old_status)

        # جلوگیری از جهش غیرمجاز وضعیت با متد مستقیم PATCH
        if new_status != old_status:
            allowed_patch_transitions = {
                'PENDING_COUNT': ['INITIAL_COUNT', 'PENDING_COUNT'],
                'INITIAL_COUNT': ['INITIAL_COUNT', 'PENDING_COUNT'],
                'SUPERVISOR_REJECTED': ['INITIAL_COUNT', 'SUPERVISOR_REJECTED'],
                'MANAGER_REJECTED': ['INITIAL_COUNT', 'MANAGER_REJECTED'],
            }
            allowed = allowed_patch_transitions.get(old_status, [])
            if new_status not in allowed:
                raise ValidationError({
                    'status': f'تغییر مستقیم وضعیت از {old_status} به {new_status} با متد PATCH مجاز نیست. لطفاً از اکشن‌های گردش‌کار استفاده کنید.'
                })

        updated_instance = serializer.save(modified_by=self.request.user)

        from accounts.audit_utils import log_audit_event
        from .models import CountTaskHistory

        note = ''
        if new_status in ['MANAGER_REJECTED', 'FINAL_APPROVED']:
            note = updated_instance.manager_note or ''
        elif new_status == 'SUPERVISOR_REJECTED':
            note = updated_instance.supervisor_note or ''
        elif new_status in ['COUNTED', 'INITIAL_COUNT']:
            note = updated_instance.counter_note or ''

        if old_status != new_status or old_counted != str(updated_instance.counted_balance):
            CountTaskHistory.objects.create(
                task=updated_instance,
                action_by=self.request.user if self.request.user.is_authenticated else None,
                action_type=new_status,
                counted_balance=updated_instance.counted_balance,
                note=note
            )

            try:
                item_label = updated_instance.item.fa_unic_code if updated_instance.item else f"#{updated_instance.id}"
                wh = updated_instance.item.warehouse if updated_instance.item else None
                log_audit_event(
                    module='counter' if new_status in ['INITIAL_COUNT', 'COUNTED', 'PENDING_COUNT'] else 'supervisor',
                    action='UPDATE' if old_status == new_status else 'STATUS_CHANGE',
                    target_model='CountTask',
                    target_object_id=updated_instance.id,
                    target_repr=f"شمارش کالا {item_label} ({new_status})",
                    warehouse=wh,
                    user=self.request.user if self.request.user.is_authenticated else None,
                    before_state={'status': old_status, 'counted_balance': old_counted},
                    after_state={'status': new_status, 'counted_balance': str(updated_instance.counted_balance), 'note': note}
                )
            except Exception:
                pass

    def get_queryset(self):
        user = self.request.user
        from django.db.models import Prefetch
        from .models import CountTaskHistory
        queryset = CountTask.objects.all().select_related(
            'item', 'counter', 'supervisor', 'assigned_manager', 'created_by', 'modified_by'
        ).prefetch_related(
            Prefetch('history', queryset=CountTaskHistory.objects.order_by('created_at').select_related('action_by')),
            # item_details کل ItemSerializer را صدا می‌زند؛ بدون این، هر تسک دو
            # کوئری عکس اضافه می‌زد.
            photo_prefetch('item__photos'),
        )
        
        as_role = self.request.query_params.get('as_role')
        warehouse_id = self.request.query_params.get('warehouse_id')
        status_filter = self.request.query_params.get('status')
        date_filter = self.request.query_params.get('date')
        q_filter = self.request.query_params.get('q')
        
        if warehouse_id and str(warehouse_id) not in ['ALL', '-1']:
            try:
                requested_wh = int(warehouse_id)
            except (TypeError, ValueError):
                return CountTask.objects.none()
            if not can_access_warehouse(user, requested_wh):
                return CountTask.objects.none()
            queryset = queryset.filter(item__warehouse_id=requested_wh)
        else:
            queryset = scope_queryset(queryset, user, field='item__warehouse_id')
        
        # 1. فیلتر بر اساس نقش (Role Filter)
        if as_role == 'counter':
            queryset = queryset.filter(counter=user)
        elif as_role == 'supervisor':
            queryset = queryset.filter(supervisor=user)
        elif as_role == 'manager':
            from django.db.models import Q
            queryset = queryset.filter(Q(assigned_manager=user) | Q(assigned_manager__isnull=True))
        elif as_role == 'tracking':
            # بازگرداندن همه موارد در حال شمارش برای پیگیری
            show_completed = self.request.query_params.get('show_completed', 'false').lower() == 'true'
            if not show_completed:
                queryset = queryset.exclude(status='FINAL_APPROVED')
        else:
            # Fallback to permission checking if as_role is not provided
            if user.is_superuser or user.has_perm('accounts.view_sys_manager_review') or user.has_perm('accounts.can_act_as_manager') or user.has_perm('inventory.can_act_as_manager'):
                from django.db.models import Q
                queryset = queryset.filter(Q(assigned_manager=user) | Q(assigned_manager__isnull=True))
            elif user.has_perm('accounts.view_sys_supervisor') or user.has_perm('accounts.can_act_as_supervisor') or user.has_perm('inventory.can_act_as_supervisor'):
                queryset = queryset.filter(supervisor=user)
            elif user.has_perm('accounts.view_sys_counter') or user.has_perm('accounts.can_act_as_counter') or user.has_perm('inventory.can_act_as_counter'):
                queryset = queryset.filter(counter=user)
            else:
                return CountTask.objects.none()

        # 2. فیلتر جستجوی متنی (Text Search Q)
        if q_filter:
            from django.db.models import Q
            q_clean = q_filter.strip()
            queryset = queryset.filter(
                Q(item__fa_unic_code__icontains=q_clean) |
                Q(item__description__icontains=q_clean) |
                Q(item__tag__icontains=q_clean) |
                Q(item__po__icontains=q_clean) |
                Q(item__pk_number__icontains=q_clean) |
                Q(item__new_location__icontains=q_clean)
            )

        # 3. فیلتر تاریخ (Date Filter)
        if date_filter and date_filter != 'all':
            from django.utils import timezone
            from datetime import timedelta
            now = timezone.localtime(timezone.now())
            today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
            if date_filter == 'today':
                queryset = queryset.filter(updated_at__gte=today_start)
            elif date_filter == 'yesterday':
                yesterday_start = today_start - timedelta(days=1)
                queryset = queryset.filter(updated_at__gte=yesterday_start, updated_at__lt=today_start)
            elif date_filter == 'week':
                week_start = today_start - timedelta(days=7)
                queryset = queryset.filter(updated_at__gte=week_start)

        # 4. فیلتر وضعیت (Status Filter)
        if status_filter and status_filter != 'all':
            from django.db.models import Q
            if as_role == 'counter':
                if status_filter == 'pending':
                    queryset = queryset.filter(status='PENDING_COUNT', counted_balance__isnull=True)
                elif status_filter == 'initial':
                    queryset = queryset.filter(Q(status='INITIAL_COUNT') | Q(status='PENDING_COUNT', counted_balance__isnull=False))
                elif status_filter == 'recount':
                    queryset = queryset.filter(status__in=['SUPERVISOR_REJECTED', 'MANAGER_REJECTED'])
                elif status_filter == 'completed':
                    queryset = queryset.filter(status__in=['COUNTED', 'SUPERVISOR_APPROVED', 'MANAGER_REVIEW', 'FINAL_APPROVED'])
            elif as_role in ['supervisor', 'doc_supervisor']:
                if status_filter == 'counted':
                    queryset = queryset.filter(status='COUNTED')
                elif status_filter == 'recount':
                    queryset = queryset.filter(status__in=['SUPERVISOR_REJECTED', 'MANAGER_REJECTED'])
                else:
                    queryset = queryset.filter(status=status_filter)
            else:
                queryset = queryset.filter(status=status_filter)
            
        return queryset

    @action(detail=False, methods=['get'])
    def pool_tasks(self, request):
        as_role = request.query_params.get('as_role')
        warehouse_id = request.query_params.get('warehouse_id')
        queryset = CountTask.objects.all().select_related('item', 'counter', 'supervisor', 'created_by', 'modified_by').prefetch_related(photo_prefetch('item__photos'))
        
        if warehouse_id:
            queryset = queryset.filter(item__warehouse_id=warehouse_id)
            
        if as_role == 'counter':
            queryset = queryset.filter(counter__isnull=True, status='PENDING_COUNT')
        elif as_role == 'supervisor':
            queryset = queryset.filter(supervisor__isnull=True, status='COUNTED')
        elif as_role == 'manager':
            queryset = queryset.filter(assigned_manager__isnull=True, status='MANAGER_REVIEW')
        else:
            return Response({'error': 'نقش نامعتبر است.'}, status=400)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['post'])
    def claim_tasks(self, request):
        task_ids = request.data.get('task_ids', [])
        as_role = request.data.get('as_role')

        if not task_ids or not as_role:
            return Response({'error': 'شناسه تسک‌ها یا نقش ارسال نشده است.'}, status=400)

        with transaction.atomic():
            if as_role == 'counter':
                tasks = CountTask.objects.select_for_update().filter(
                    id__in=task_ids,
                    counter__isnull=True,
                    status='PENDING_COUNT'
                )
                valid_task_ids = list(tasks.values_list('id', flat=True))
                if not valid_task_ids:
                    broadcast_count_task_update()
                    return Response({'success': False, 'claimed_count': 0, 'message': 'این کالا(ها) قبلاً توسط انبارگردان دیگری بر عهده گرفته شده است.'})

                from .models import Item
                item_ids = list(CountTask.objects.filter(id__in=valid_task_ids).values_list('item_id', flat=True))

                # Update CountTasks
                updated = CountTask.objects.filter(id__in=valid_task_ids).update(counter=request.user, updated_at=timezone.now())

                # Update Item field_assignee
                assignee_name = f"{request.user.first_name} {request.user.last_name}".strip() or request.user.username
                Item.objects.filter(id__in=item_ids).update(field_assignee=assignee_name, updated_at=timezone.now())

            elif as_role == 'supervisor':
                tasks = CountTask.objects.select_for_update().filter(
                    id__in=task_ids,
                    supervisor__isnull=True,
                    status='COUNTED'
                )
                updated = tasks.update(supervisor=request.user, updated_at=timezone.now())
            elif as_role == 'manager':
                tasks = CountTask.objects.select_for_update().filter(
                    id__in=task_ids,
                    assigned_manager__isnull=True,
                    status='MANAGER_REVIEW'
                )
                updated = tasks.update(assigned_manager=request.user, updated_at=timezone.now())
            else:
                return Response({'error': 'نقش نامعتبر است.'}, status=400)

        broadcast_count_task_update()
        return Response({'success': True, 'claimed_count': updated})

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user, modified_by=self.request.user)

    @action(detail=False, methods=['post'])
    def bulk_submit(self, request):
        user = request.user
        task_ids = request.data.get('task_ids', [])
        sync_ids = request.data.get('sync_ids', [])
        warehouse_id = request.data.get('warehouse_id') or request.query_params.get('warehouse_id')

        # کلاینت آفلاین ممکن است فقط sync_id پایدار را داشته باشد
        if sync_ids:
            task_ids = list(set(task_ids) | set(
                CountTask.objects.filter(sync_id__in=sync_ids).values_list('id', flat=True)
            ))

        with transaction.atomic():
            tasks_qs = CountTask.objects.select_for_update().filter(
                counter=user,
                status='INITIAL_COUNT',
                counted_balance__isnull=False
            )
            if task_ids:
                tasks_qs = tasks_qs.filter(id__in=task_ids)

            if warehouse_id and str(warehouse_id) not in ['ALL', '-1']:
                tasks_qs = tasks_qs.filter(item__warehouse_id=warehouse_id)

            tasks_list = list(tasks_qs.select_related('item'))

            if not tasks_list:
                # Idempotency: اگر تسک‌های درخواستی قبلاً ارسال شده‌اند (مثلاً retry صف
                # آفلاین پس از گم شدن پاسخ در تونل)، موفقیت no-op برگردان نه ابهام.
                if task_ids:
                    already = CountTask.objects.filter(
                        id__in=task_ids, counter=user,
                        status__in=['COUNTED', 'MANAGER_REVIEW', 'FINAL_APPROVED'],
                    ).count()
                    if already > 0:
                        return Response({
                            'message': f'{already} مورد قبلاً ارسال شده بود.',
                            'already_submitted': already,
                        })
                return Response({'message': 'هیچ موردی برای ارسال یافت نشد.'})
                
            from warehouses.services import get_setting
            wh_setting_cache = {}

            def get_warehouse_req_sup(wh_id):
                if wh_id not in wh_setting_cache:
                    wh_setting_cache[wh_id] = get_setting('require_supervisor_approval', wh_id)
                return wh_setting_cache[wh_id]
            
            from .models import CountTaskHistory
            histories = []
            counted_count = 0
            manager_count = 0
            for task in tasks_list:
                wh_id = task.item.warehouse_id if task.item else None
                req_sup_app = get_warehouse_req_sup(wh_id)
                task_req_sup = req_sup_app and not task.skip_supervisor
                target_status = 'COUNTED' if task_req_sup else 'MANAGER_REVIEW'
                if target_status == 'COUNTED':
                    counted_count += 1
                else:
                    manager_count += 1
                histories.append(CountTaskHistory(
                    task=task,
                    action_by=user,
                    action_type=target_status,
                    counted_balance=task.counted_balance,
                    note=task.counter_note
                ))
                task.status = target_status
                task.modified_by = user
                task.updated_at = timezone.now()  # bulk_update سیگنال auto_now را رد می‌کند

            count = len(tasks_list)
            if count > 0:
                CountTask.objects.bulk_update(tasks_list, ['status', 'modified_by', 'updated_at'])
            if histories:
                CountTaskHistory.objects.bulk_create(histories)
                
        if counted_count > 0 and manager_count > 0:
            msg = f'{count} مورد ارسال شد ({counted_count} مورد به سرپرست و {manager_count} مورد مستقیم به مدیر).'
        elif counted_count > 0:
            msg = f'{count} مورد به سرپرست ارسال شد.'
        else:
            msg = f'{count} مورد مستقیماً به مدیر ارسال شد.'
        broadcast_count_task_update()
        return Response({'message': msg})

    @action(detail=False, methods=['post'])
    def bulk_approve(self, request):
        user = request.user
        task_ids = request.data.get('task_ids', [])
        if not task_ids:
            return Response({'error': 'هیچ موردی انتخاب نشده است.'}, status=400)
            
        note = request.data.get('note', '')
        from .models import CountTaskHistory
        from django.db import transaction
        
        with transaction.atomic():
            tasks = CountTask.objects.select_for_update(of=('self',)).filter(
                id__in=task_ids, supervisor=user, status__in=['COUNTED', 'MANAGER_REJECTED']
            )
            histories = []
            tasks_list = list(tasks)
            for task in tasks_list:
                task.supervisor_note = note
                task.status = 'MANAGER_REVIEW'
                task.modified_by = user
                task.updated_at = timezone.now()
                histories.append(CountTaskHistory(
                    task=task,
                    action_by=user,
                    action_type='MANAGER_REVIEW',
                    counted_balance=task.counted_balance,
                    note=note
                ))
                
            if tasks_list:
                CountTask.objects.bulk_update(tasks_list, ['status', 'supervisor_note', 'modified_by', 'updated_at'])
            if histories:
                CountTaskHistory.objects.bulk_create(histories)
                
        broadcast_count_task_update()
        return Response({'message': f'{len(tasks_list)} مورد تایید و برای مدیر ارسال شد.'})

    @action(detail=False, methods=['post'])
    def bulk_manager_approve(self, request):
        user = request.user
        task_ids = request.data.get('task_ids', [])
        note = request.data.get('note', '')
        
        if not task_ids:
            return Response({'error': 'هیچ موردی انتخاب نشده است.'}, status=400)
            
        from .models import CountTaskHistory
        from django.db import transaction
        from decimal import Decimal, InvalidOperation
        from accounts.audit_utils import log_audit_event

        with transaction.atomic():
            tasks_qs = CountTask.objects.select_for_update(of=('self',)).filter(
                id__in=task_ids, status='MANAGER_REVIEW'
            ).select_related('item', 'item__warehouse')
            tasks = list(tasks_qs)
            if not tasks:
                return Response({'message': 'هیچ موردی در وضعیت بررسی مدیر یافت نشد.'})

            # بررسی دسترسی انبار برای کاربر (IDOR Check)
            if not user.is_superuser:
                from common.warehouse_scope import can_access_warehouse
                for task in tasks:
                    wh_id = task.item.warehouse_id if task.item else None
                    if not can_access_warehouse(user, wh_id):
                        return Response({'error': f'شما به انبار کالا #{task.id} دسترسی ندارید.'}, status=403)

            histories = []
            items_to_update = []
            wh_ids = set()
            
            for task in tasks:
                old_status = task.status
                old_counted = str(task.counted_balance)
                task.manager_note = note
                task.status = 'FINAL_APPROVED'
                task.modified_by = user
                task.updated_at = timezone.now()

                histories.append(CountTaskHistory(
                    task=task,
                    action_by=user,
                    action_type='FINAL_APPROVED',
                    counted_balance=task.counted_balance,
                    note=note
                ))
                
                # بروزرسانی کالای اصلی پس از تایید نهایی
                item = task.item
                if item:
                    if item.warehouse_id:
                        wh_ids.add(item.warehouse_id)
                    item.field_status = 'done'
                    if task.counted_balance is not None:
                        c_dec = _parse_decimal(task.counted_balance)
                        s_dec = _parse_decimal(item.bal4miv)
                        if c_dec is not None and s_dec is not None:
                            item.has_conflict = (c_dec != s_dec)
                        elif c_dec is not None:
                            item.has_conflict = (c_dec != Decimal('0'))
                        else:
                            item.has_conflict = False
                        item.inventory = task.counted_balance
                    item.modified_by = user
                    item.updated_at = timezone.now()
                    items_to_update.append(item)

                try:
                    item_label = item.fa_unic_code if item else f"#{task.id}"
                    wh = item.warehouse if item else None
                    log_audit_event(
                        module='manager',
                        action='APPROVE',
                        target_model='CountTask',
                        target_object_id=task.id,
                        target_repr=f"تایید نهایی شمارش کالا {item_label}",
                        warehouse=wh,
                        user=user if user.is_authenticated else None,
                        before_state={'status': old_status, 'counted_balance': old_counted},
                        after_state={'status': 'FINAL_APPROVED', 'counted_balance': str(task.counted_balance), 'note': note}
                    )
                except Exception:
                    pass
            
            CountTask.objects.bulk_update(tasks, ['status', 'manager_note', 'modified_by', 'updated_at'])
            count = len(tasks)
            
            if items_to_update:
                Item.objects.bulk_update(items_to_update, ['field_status', 'inventory', 'has_conflict', 'modified_by', 'updated_at'])
            
            if histories:
                CountTaskHistory.objects.bulk_create(histories)
            
        for wid in wh_ids:
            broadcast_count_task_update(warehouse_id=wid)
        if not wh_ids:
            broadcast_count_task_update()

        return Response({'message': f'{count} مورد به صورت گروهی تایید نهایی شد.'})

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        task = self.get_object()
        note = request.data.get('note') or request.data.get('reason', '')
        
        if task.status not in ['COUNTED', 'MANAGER_REJECTED']:
            return Response({'error': 'فقط موارد شمرده شده قابل رد هستند.'}, status=400)
            
        old_status = task.status
        task.status = 'SUPERVISOR_REJECTED'
        task.supervisor_note = note
        task.modified_by = request.user
        task.updated_at = timezone.now()
        task.save()
        
        from .models import CountTaskHistory
        from accounts.audit_utils import log_audit_event

        CountTaskHistory.objects.create(
            task=task,
            action_by=request.user,
            action_type='SUPERVISOR_REJECTED',
            counted_balance=task.counted_balance,
            note=note
        )

        try:
            item_label = task.item.fa_unic_code if task.item else f"#{task.id}"
            wh = task.item.warehouse if task.item else None
            log_audit_event(
                module='supervisor',
                action='REJECT',
                target_model='CountTask',
                target_object_id=task.id,
                target_repr=f"رد سرپرست برای بازشماری کالا {item_label}",
                warehouse=wh,
                user=request.user if request.user.is_authenticated else None,
                before_state={'status': old_status, 'counted_balance': str(task.counted_balance)},
                after_state={'status': 'SUPERVISOR_REJECTED', 'counted_balance': str(task.counted_balance), 'note': note}
            )
        except Exception:
            pass
        
        wh_id = task.item.warehouse_id if task.item else None
        broadcast_count_task_update(warehouse_id=wh_id)
        return Response({'message': 'مورد با موفقیت رد شد و به شمارشگر ارجاع داده شد.'})

    @action(detail=True, methods=['post'])
    def manager_reject(self, request, pk=None):
        """رد توسط مدیر — بسته به تنظیم سرپرست، به سرپرست یا انبارگردان برمی‌گردد"""
        task = self.get_object()
        note = request.data.get('note', '')
        
        if not note.strip():
            return Response({'error': 'لطفاً علت بازشماری را بنویسید.'}, status=400)
        
        if task.status != 'MANAGER_REVIEW':
            return Response({'error': 'فقط موارد در انتظار تایید مدیر قابل رد هستند.'}, status=400)
        
        from warehouses.services import get_setting
        from accounts.audit_utils import log_audit_event
        
        req_supervisor = get_setting('require_supervisor_approval', task.item.warehouse_id) if task.item else False
        old_status = task.status
        prev_counted = task.counted_balance

        if req_supervisor and task.supervisor:
            # ارسال به سرپرست (مقدار شمارش شده حفظ می‌شود تا سرپرست آن را بازبینی کند)
            target_status = 'MANAGER_REJECTED'
            target_msg = 'مورد برای بازشماری به سرپرست ارجاع شد.'
        else:
            # ارسال مستقیم به انبارگردان
            target_status = 'PENDING_COUNT'
            target_msg = 'مورد برای بازشماری مستقیماً به انبارگردان ارجاع شد.'
            task.counted_balance = None  # پاک کردن مقدار قبلی برای شمارش مجدد
        
        task.status = target_status
        task.manager_note = note
        task.modified_by = request.user
        task.updated_at = timezone.now()
        task.save()
        
        from .models import CountTaskHistory
        CountTaskHistory.objects.create(
            task=task,
            action_by=request.user,
            action_type='MANAGER_REJECTED',
            counted_balance=prev_counted,
            note=note
        )

        try:
            item_label = task.item.fa_unic_code if task.item else f"#{task.id}"
            wh = task.item.warehouse if task.item else None
            log_audit_event(
                module='manager',
                action='REJECT',
                target_model='CountTask',
                target_object_id=task.id,
                target_repr=f"رد مدیر برای بازشماری کالا {item_label}",
                warehouse=wh,
                user=request.user if request.user.is_authenticated else None,
                before_state={'status': old_status, 'counted_balance': str(prev_counted)},
                after_state={'status': target_status, 'counted_balance': str(task.counted_balance), 'note': note}
            )
        except Exception:
            pass
        
        wh_id = task.item.warehouse_id if task.item else None
        broadcast_count_task_update(warehouse_id=wh_id)
        return Response({'message': target_msg})

    @action(detail=False, methods=['post'])
    def bulk_reject(self, request):
        """رد گروهی توسط سرپرست و ارجاع به شمارشگر"""
        user = request.user
        task_ids = request.data.get('task_ids', [])
        note = request.data.get('note', '')
        
        if not task_ids:
            return Response({'error': 'هیچ موردی انتخاب نشده است.'}, status=400)
            
        from .models import CountTaskHistory
        from django.db import transaction
        from accounts.audit_utils import log_audit_event
        
        with transaction.atomic():
            tasks = CountTask.objects.select_for_update(of=('self',)).filter(
                id__in=task_ids, supervisor=user, status__in=['COUNTED', 'MANAGER_REJECTED']
            ).select_related('item', 'item__warehouse')
            histories = []
            tasks_list = list(tasks)
            wh_ids = set()

            for task in tasks_list:
                old_status = task.status
                if task.item and task.item.warehouse_id:
                    wh_ids.add(task.item.warehouse_id)
                task.supervisor_note = note
                task.status = 'SUPERVISOR_REJECTED'
                task.modified_by = user
                task.updated_at = timezone.now()
                histories.append(CountTaskHistory(
                    task=task,
                    action_by=user,
                    action_type='SUPERVISOR_REJECTED',
                    counted_balance=task.counted_balance,
                    note=note
                ))

                try:
                    item_label = task.item.fa_unic_code if task.item else f"#{task.id}"
                    wh = task.item.warehouse if task.item else None
                    log_audit_event(
                        module='supervisor',
                        action='REJECT',
                        target_model='CountTask',
                        target_object_id=task.id,
                        target_repr=f"رد سرپرست برای بازشماری کالا {item_label}",
                        warehouse=wh,
                        user=user if user.is_authenticated else None,
                        before_state={'status': old_status, 'counted_balance': str(task.counted_balance)},
                        after_state={'status': 'SUPERVISOR_REJECTED', 'counted_balance': str(task.counted_balance), 'note': note}
                    )
                except Exception:
                    pass
                
            if tasks_list:
                CountTask.objects.bulk_update(tasks_list, ['status', 'supervisor_note', 'modified_by', 'updated_at'])
            if histories:
                CountTaskHistory.objects.bulk_create(histories)
            
        for wid in wh_ids:
            broadcast_count_task_update(warehouse_id=wid)
        if not wh_ids:
            broadcast_count_task_update()

        return Response({'message': f'{len(tasks_list)} مورد با موفقیت رد شد و به شمارشگر ارجاع داده شد.'})

    @action(detail=False, methods=['post'])
    def bulk_manager_reject(self, request):
        """رد گروهی توسط مدیر — ارجاع به سرپرست یا شمارشگر بر اساس تنظیمات انبار"""
        user = request.user
        task_ids = request.data.get('task_ids', [])
        note = request.data.get('note', '')
        
        if not note.strip():
            return Response({'error': 'لطفاً علت بازشماری را بنویسید.'}, status=400)
            
        if not task_ids:
            return Response({'error': 'هیچ موردی انتخاب نشده است.'}, status=400)
            
        from warehouses.services import get_setting
        from .models import CountTaskHistory
        from django.db import transaction
        from accounts.audit_utils import log_audit_event

        with transaction.atomic():
            tasks = CountTask.objects.select_for_update(of=('self',)).filter(
                id__in=task_ids, status='MANAGER_REVIEW'
            ).select_related('item', 'item__warehouse')
            tasks_list = list(tasks)
            if not tasks_list:
                return Response({'message': 'هیچ موردی برای رد یافت نشد.'})

            # بررسی دسترسی انبار برای کاربر (IDOR Check)
            if not user.is_superuser:
                from common.warehouse_scope import can_access_warehouse
                for task in tasks_list:
                    wh_id = task.item.warehouse_id if task.item else None
                    if not can_access_warehouse(user, wh_id):
                        return Response({'error': f'شما به انبار کالا #{task.id} دسترسی ندارید.'}, status=403)

            histories = []
            wh_ids = set()

            for task in tasks_list:
                old_status = task.status
                prev_counted = task.counted_balance
                wh_id = task.item.warehouse_id if task.item else None
                if wh_id:
                    wh_ids.add(wh_id)
                req_supervisor = get_setting('require_supervisor_approval', wh_id) if wh_id else False
                if req_supervisor and task.supervisor:
                    target_status = 'MANAGER_REJECTED'
                else:
                    target_status = 'PENDING_COUNT'
                    task.counted_balance = None
                    
                task.status = target_status
                task.manager_note = note
                task.modified_by = user
                task.updated_at = timezone.now()
                
                histories.append(CountTaskHistory(
                    task=task,
                    action_by=user,
                    action_type='MANAGER_REJECTED',
                    counted_balance=prev_counted,
                    note=note
                ))

                try:
                    item_label = task.item.fa_unic_code if task.item else f"#{task.id}"
                    wh = task.item.warehouse if task.item else None
                    log_audit_event(
                        module='manager',
                        action='REJECT',
                        target_model='CountTask',
                        target_object_id=task.id,
                        target_repr=f"رد مدیر برای بازشماری کالا {item_label}",
                        warehouse=wh,
                        user=user if user.is_authenticated else None,
                        before_state={'status': old_status, 'counted_balance': str(prev_counted)},
                        after_state={'status': target_status, 'counted_balance': str(task.counted_balance), 'note': note}
                    )
                except Exception:
                    pass
                
            if tasks_list:
                CountTask.objects.bulk_update(tasks_list, ['status', 'manager_note', 'counted_balance', 'modified_by', 'updated_at'])
            if histories:
                CountTaskHistory.objects.bulk_create(histories)
            
        for wid in wh_ids:
            broadcast_count_task_update(warehouse_id=wid)
        if not wh_ids:
            broadcast_count_task_update()

        return Response({'message': f'{len(tasks_list)} مورد با موفقیت رد شد و جهت بازشماری ارجاع داده شد.'})

    @action(detail=False, methods=['post'])
    def bulk_cancel(self, request):
        """لغو تخصیص گروهی — رکوردهای PENDING_COUNT و INITIAL_COUNT مجاز هستند."""
        user = request.user
        task_ids = request.data.get('task_ids', [])
        if not task_ids:
            return Response({'error': 'هیچ رکوردی انتخاب نشده است.'}, status=400)

        from accounts.audit_utils import log_audit_event
        from .models import CountTaskHistory
        from django.db import transaction

        with transaction.atomic():
            all_tasks = CountTask.objects.select_for_update(of=('self',)).filter(id__in=task_ids).select_related('item', 'item__warehouse', 'counter')
            eligible_tasks = all_tasks.filter(status__in=['PENDING_COUNT', 'INITIAL_COUNT'])
            eligible_list = list(eligible_tasks)
            ineligible_count = all_tasks.count() - len(eligible_list)

            if not eligible_list:
                return Response(
                    {'error': 'هیچ‌یک از رکوردهای انتخاب شده قابل لغو تخصیص نیستند. فقط رکوردهای «در انتظار شمارش» و «شمارش اولیه» مجاز هستند.'},
                    status=400
                )

            item_ids = [t.item_id for t in eligible_list if t.item_id]
            task_ids_to_del = [t.id for t in eligible_list]
            wh_ids = set()

            for t in eligible_list:
                if t.item and t.item.warehouse_id:
                    wh_ids.add(t.item.warehouse_id)
                try:
                    item_label = t.item.fa_unic_code if t.item else f"#{t.id}"
                    wh = t.item.warehouse if t.item else None
                    log_audit_event(
                        module='dispatch',
                        action='CANCEL_ALLOCATION',
                        target_model='CountTask',
                        target_object_id=t.id,
                        target_repr=f"لغو تخصیص کالا {item_label}",
                        warehouse=wh,
                        user=user if user.is_authenticated else None,
                        before_state={'status': t.status, 'counter': t.counter.username if t.counter else None},
                        after_state={'status': 'CANCELLED_AND_DELETED'}
                    )
                except Exception:
                    pass

            # پاک کردن تاریخچه مربوطه و خود تسک‌ها به صورت نرم جهت حفظ سینک آفلاین
            soft_delete_queryset(CountTaskHistory.objects.filter(task_id__in=task_ids_to_del))
            deleted_count = len(eligible_list)
            soft_delete_queryset(CountTask.objects.filter(id__in=task_ids_to_del))

            # بازگرداندن وضعیت آیتم‌ها به استخر
            Item.objects.filter(id__in=item_ids).update(
                field_status='checking',
                field_assignee=None,
                updated_at=timezone.now()
            )

        msg = f'{deleted_count} وظیفه شمارش با موفقیت لغو تخصیص شد.'
        if ineligible_count > 0:
            msg += f' ({ineligible_count} رکورد به دلیل وضعیت نامعتبر نادیده گرفته شد.)'

        for wid in wh_ids:
            broadcast_count_task_update(warehouse_id=wid)
        if not wh_ids:
            broadcast_count_task_update()

        return Response({'message': msg})

    @action(detail=False, methods=['get'], url_path='get_export_columns')
    def get_export_columns(self, request):
        """لیست ستون‌های مجاز برای خروجی اکسل پیگیری شمارش"""
        columns = [
            {'key': 'warehouse_name',    'label': 'نام انبار'},
            {'key': 'fa_unic_code',      'label': 'کد یکتا'},
            {'key': 'description',       'label': 'شرح کالا'},
            {'key': 'counter_name',      'label': 'شمارنده'},
            {'key': 'supervisor_name',   'label': 'سرپرست'},
            {'key': 'manager_name',      'label': 'مدیر'},
            {'key': 'status',            'label': 'وضعیت'},
            {'key': 'counted_balance',   'label': 'موجودی شمارش شده'},
            {'key': 'inventory',         'label': 'موجودی سیستم'},
            {'key': 'difference',        'label': 'اختلاف'},
            {'key': 'created_at',        'label': 'تاریخ ایجاد'},
            {'key': 'updated_at',        'label': 'تاریخ بروزرسانی'},
            {'key': 'counter_note',      'label': 'یادداشت شمارنده'},
            {'key': 'supervisor_note',   'label': 'یادداشت سرپرست'},
            {'key': 'manager_note',      'label': 'یادداشت مدیر'},
        ]
        return Response(columns)

    @action(detail=False, methods=['post'], url_path='export_excel')
    def export_excel(self, request):
        """خروجی اکسل صفحه پیگیری شمارش با تاریخ شمسی و وضعیت فارسی"""
        import openpyxl
        from django.http import HttpResponse

        STATUS_FA = {
            'PENDING_COUNT':      'در انتظار شمارش',
            'INITIAL_COUNT':      'آماده ارسال (پیش‌نویس)',
            'COUNTED':            'شمارش شده (ارسال به سرپرست)',
            'SUPERVISOR_REVIEW':  'در بررسی سرپرست',
            'SUPERVISOR_APPROVED':'تایید سرپرست',
            'SUPERVISOR_REJECTED':'رد شده توسط سرپرست (بازشماری)',
            'MANAGER_REVIEW':     'در بررسی مدیر',
            'MANAGER_REJECTED':   'رد شده توسط مدیر (بازشماری)',
            'FINAL_APPROVED':     'تأیید نهایی',
        }

        def to_jalali(dt):
            if not dt:
                return ''
            try:
                import jdatetime
                from django.utils import timezone
                if timezone.is_aware(dt):
                    dt = timezone.localtime(dt)
                jdt = jdatetime.datetime.fromgregorian(datetime=dt)
                return jdt.strftime('%Y/%m/%d %H:%M:%S')
            except Exception as e:
                # اگر خطایی رخ داد، لاگ بگیریم تا اشکال‌زدایی راحت‌تر باشد
                import logging
                logging.getLogger(__name__).error(f"Date conversion error: {e}")
                return str(dt)

        data_scope    = request.data.get('data_scope', 'all')
        columns_scope = request.data.get('columns_scope', 'all_db')
        columns_list  = request.data.get('columns_list', [])

        # اعمال همان فیلترهای get_queryset
        from django.http import QueryDict
        original_query_params = request._request.GET
        try:
            q = QueryDict(mutable=True)
            for k, v in request.data.items():
                if isinstance(v, list):
                    q.setlist(k, [str(x) for x in v])
                elif v is not None:
                    q[k] = str(v)
            request._request.GET = q

            if data_scope == 'selected':
                selected_ids = request.data.get('selected_ids', [])
                queryset = self.get_queryset().filter(id__in=selected_ids)
            else:
                queryset = self.get_queryset()
        finally:
            request._request.GET = original_query_params

        # ستون‌ها
        ALL_COLUMNS = [
            {'key': 'warehouse_name',    'label': 'نام انبار'},
            {'key': 'fa_unic_code',      'label': 'کد یکتا'},
            {'key': 'description',       'label': 'شرح کالا'},
            {'key': 'counter_name',      'label': 'شمارنده'},
            {'key': 'supervisor_name',   'label': 'سرپرست'},
            {'key': 'manager_name',      'label': 'مدیر'},
            {'key': 'status',            'label': 'وضعیت'},
            {'key': 'counted_balance',   'label': 'موجودی شمارش شده'},
            {'key': 'inventory',         'label': 'موجودی سیستم'},
            {'key': 'difference',        'label': 'اختلاف'},
            {'key': 'created_at',        'label': 'تاریخ ایجاد'},
            {'key': 'updated_at',        'label': 'تاریخ بروزرسانی'},
            {'key': 'counter_note',      'label': 'یادداشت شمارنده'},
            {'key': 'supervisor_note',   'label': 'یادداشت سرپرست'},
            {'key': 'manager_note',      'label': 'یادداشت مدیر'},
        ]
        all_keys = [c['key'] for c in ALL_COLUMNS]
        key_to_label = {c['key']: c['label'] for c in ALL_COLUMNS}

        if columns_scope in ('visible', 'custom') and columns_list:
            selected_keys = [k for k in columns_list if k in all_keys]
            if not selected_keys:
                selected_keys = all_keys
        else:
            selected_keys = all_keys

        # بررسی وضعیت شمارش کور برای هر انبار جهت حفظ محرمانگی داده‌ها
        from warehouses.services import get_setting
        warehouse_blind_cache = {}

        def check_is_blind(task):
            wh_id = task.item.warehouse_id if task.item else None
            if wh_id not in warehouse_blind_cache:
                warehouse_blind_cache[wh_id] = (get_setting('blind_counting', wh_id) == 'blind')
            return warehouse_blind_cache[wh_id]

        def get_cell(task, key):
            if key == 'warehouse_name':
                if task.item and task.item.warehouse:
                    return getattr(task.item.warehouse, 'project_name', None) or task.item.warehouse.name or ''
                return ''
            elif key == 'fa_unic_code':
                return getattr(task.item, 'fa_unic_code', '') if task.item else ''
            elif key == 'description':
                return getattr(task.item, 'description', '') if task.item else ''
            elif key == 'counter_name':
                if task.counter:
                    return f"{task.counter.first_name} {task.counter.last_name}".strip() or task.counter.username
                return ''
            elif key == 'supervisor_name':
                if task.supervisor:
                    return f"{task.supervisor.first_name} {task.supervisor.last_name}".strip() or task.supervisor.username
                return ''
            elif key == 'manager_name':
                if task.assigned_manager:
                    return f"{task.assigned_manager.first_name} {task.assigned_manager.last_name}".strip() or task.assigned_manager.username
                return ''
            elif key == 'status':
                return STATUS_FA.get(task.status, task.status)
            elif key == 'counted_balance':
                return task.counted_balance if task.counted_balance is not None else ''
            elif key == 'inventory':
                if check_is_blind(task):
                    return ''
                base_bal = getattr(task.item, 'bal4miv', None) if task.item else None
                if base_bal is None and task.item:
                    base_bal = getattr(task.item, 'inventory', '')
                return base_bal if base_bal is not None else ''
            elif key == 'difference':
                if check_is_blind(task):
                    return ''
                base_bal = getattr(task.item, 'bal4miv', None) if task.item else None
                if base_bal is None and task.item:
                    base_bal = getattr(task.item, 'inventory', None)
                cnt = task.counted_balance
                if base_bal is not None and cnt is not None:
                    try:
                        return float(cnt) - float(base_bal)
                    except Exception:
                        return ''
                return ''
            elif key == 'created_at':
                return to_jalali(task.created_at)
            elif key == 'updated_at':
                return to_jalali(task.updated_at)
            elif key == 'counter_note':
                return task.counter_note or ''
            elif key == 'supervisor_note':
                return task.supervisor_note or ''
            elif key == 'manager_note':
                return task.manager_note or ''
            return ''

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'پیگیری شمارش'
        
        # راست‌چین کردن شیت
        ws.sheet_view.rightToLeft = True

        # استایل‌های هدر ۲ سطری
        header_fa_font = Font(name='Tahoma', bold=True, color='FFFFFF', size=11)
        header_fa_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid') # سرمه‌ای تیره
        header_en_font = Font(name='Consolas', size=9, color='64748B', italic=True) # فیلد دیتابیس
        header_en_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid') # طوسی روشن
        
        cell_font = Font(name='Tahoma', size=10)
        
        # رنگ‌های یکی‌درمیان ردیف‌ها
        fill_even = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
        fill_odd = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1')
        )

        # ساخت هدر ۲ سطری (سطر ۱: فارسی، سطر ۲: کلید فنی دیتابیس)
        headers_fa = [key_to_label[k] for k in selected_keys]
        headers_en = [k for k in selected_keys]
        ws.append(headers_fa)
        ws.append(headers_en)
        
        # استایل سطر اول
        for col_idx, cell in enumerate(ws[1], 1):
            cell.font = header_fa_font
            cell.fill = header_fa_fill
            cell.alignment = center_alignment
            cell.border = thin_border
            ws.row_dimensions[1].height = 28

        # استایل سطر دوم
        for col_idx, cell in enumerate(ws[2], 1):
            cell.font = header_en_font
            cell.fill = header_en_fill
            cell.alignment = center_alignment
            cell.border = thin_border
            ws.row_dimensions[2].height = 18

        # فریز کردن ۲ سطر اول (سطر ۳ به بعد اسکرول شود)
        ws.freeze_panes = 'A3'

        # محاسبه عرض اولیه بر اساس هدرها
        col_widths = {i: max(len(headers_fa[i]), len(headers_en[i])) + 6 for i in range(len(headers_fa))}

        # درج داده‌ها از سطر ۳
        row_idx = 3
        for task in queryset.select_related('item', 'item__warehouse', 'counter', 'supervisor', 'assigned_manager').iterator(chunk_size=500):
            row_data = [str(get_cell(task, k)) for k in selected_keys]
            ws.append(row_data)
            
            # تعیین رنگ پس‌زمینه ردیف
            current_fill = fill_even if row_idx % 2 == 0 else fill_odd
            
            # استایل‌دهی به سلول‌ها و محاسبه عرض
            for col_idx, val in enumerate(row_data):
                cell = ws.cell(row=row_idx, column=col_idx + 1)
                cell.font = cell_font
                cell.fill = current_fill
                cell.alignment = center_alignment
                cell.border = thin_border
                
                # بروزرسانی عرض ستون (محدود کردن به حداکثر 60)
                lines = val.split('\n')
                max_line_len = max([len(line) for line in lines]) if lines else 0
                if max_line_len > col_widths.get(col_idx, 10):
                    col_widths[col_idx] = min(max_line_len + 4, 60)
            
            ws.row_dimensions[row_idx].height = 22
            row_idx += 1

        # اعمال عرض ستون‌ها
        for col_idx, width in col_widths.items():
            ws.column_dimensions[get_column_letter(col_idx + 1)].width = max(width, 14)

        # افزودن قابلیت فیلتر روی هدرها
        if row_idx > 3:
            last_col_letter = get_column_letter(len(selected_keys))
            ws.auto_filter.ref = f"A1:{last_col_letter}{row_idx - 1}"

        from io import BytesIO
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="count_tracking_export.xlsx"'
        return response


class DocTaskViewSet(viewsets.ModelViewSet):
    serializer_class = DocTaskSerializer
    permission_classes = [IsAuthenticated]

    def get_permissions(self):
        from accounts.permissions import HasMenuAccess
        from rest_framework.permissions import IsAuthenticated
        
        if self.action in ['bulk_approve', 'reject']:
            return [HasMenuAccess('perm_doc_approve_action') | HasMenuAccess('view_wh_doc_approvals') | HasMenuAccess('can_act_as_doc_supervisor') | HasMenuAccess('can_act_as_manager') | HasMenuAccess('view_sys_manager_review')]
        elif self.action in ['bulk_manager_approve', 'manager_reject']:
            return [HasMenuAccess('perm_doc_approve_action') | HasMenuAccess('view_wh_doc_approvals') | HasMenuAccess('can_act_as_manager') | HasMenuAccess('view_sys_manager_review')]
        return [IsAuthenticated()]

    def get_queryset(self):
        user = self.request.user
        queryset = DocTask.objects.filter(item__is_deleted=False).select_related('item', 'doc_worker', 'doc_supervisor', 'created_by', 'modified_by').prefetch_related(photo_prefetch('item__photos'))
        
        as_role = self.request.query_params.get('as_role')
        warehouse_id = self.request.query_params.get('warehouse_id')
        
        if warehouse_id and str(warehouse_id) not in ['ALL', '-1']:
            queryset = queryset.filter(item__warehouse_id=warehouse_id)
        
        if as_role == 'doc_worker':
            queryset = queryset.filter(doc_worker=user)
        elif as_role == 'doc_supervisor':
            queryset = queryset.filter(doc_supervisor=user)
        elif as_role == 'manager':
            queryset = queryset.filter(assigned_manager=user, status='DOC_MANAGER_REVIEW')
        elif as_role == 'tracking':
            show_completed = self.request.query_params.get('show_completed', 'false').lower() == 'true'
            if not show_completed:
                queryset = queryset.exclude(status='DOC_FINAL_APPROVED')
        else:
            # Fallback
            from django.db.models import Q
            if user.is_superuser or user.groups.filter(name__in=['admin', 'manager']).exists() or user.has_perm('accounts.view_sys_manager_review'):
                queryset = queryset.filter(Q(assigned_manager=user) | Q(assigned_manager__isnull=True) | Q(doc_worker=user) | Q(doc_supervisor=user))
            else:
                queryset = queryset.filter(Q(doc_worker=user) | Q(doc_supervisor=user) | Q(assigned_manager=user))

        # اعمال فیلترهای تکمیلی جستجو، وضعیت و تاریخ
        q = self.request.query_params.get('q') or self.request.data.get('q')
        if q:
            q_clean = str(q).strip()
            from django.db.models import Q
            queryset = queryset.filter(
                Q(item__fa_unic_code__icontains=q_clean) |
                Q(item__description__icontains=q_clean) |
                Q(item__tag__icontains=q_clean) |
                Q(item__po__icontains=q_clean) |
                Q(item__pk_number__icontains=q_clean) |
                Q(item__new_location__icontains=q_clean) |
                Q(inv_rti_number__icontains=q_clean) |
                Q(added_rti_no__icontains=q_clean) |
                Q(doc_supplier__icontains=q_clean)
            )

        status_param = self.request.query_params.get('status') or self.request.data.get('status')
        if status_param and status_param != 'all':
            if status_param == 'untouched':
                queryset = queryset.filter(status='PENDING_DOC', inv_rti_number__isnull=True, added_rti_no__isnull=True, price_amount__isnull=True, doc_supplier__isnull=True)
            elif status_param == 'rejected':
                queryset = queryset.filter(status__in=['DOC_SUPERVISOR_REJECTED', 'DOC_MANAGER_REJECTED'])
            elif status_param == 'ready':
                queryset = queryset.filter(status='PENDING_DOC').exclude(inv_rti_number__isnull=True, added_rti_no__isnull=True, price_amount__isnull=True, doc_supplier__isnull=True)
            elif status_param == 'completed':
                queryset = queryset.filter(status='DOC_FINAL_APPROVED')
            else:
                queryset = queryset.filter(status=status_param)

        date_param = self.request.query_params.get('date') or self.request.data.get('date')
        if date_param and date_param != 'all':
            from django.utils import timezone
            import datetime
            now = timezone.now()
            today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
            if date_param == 'today':
                queryset = queryset.filter(updated_at__gte=today_start)
            elif date_param == 'yesterday':
                yesterday_start = today_start - datetime.timedelta(days=1)
                queryset = queryset.filter(updated_at__gte=yesterday_start, updated_at__lt=today_start)
            elif date_param == 'week':
                week_start = today_start - datetime.timedelta(days=7)
                queryset = queryset.filter(updated_at__gte=week_start)
            
        return queryset

    @action(detail=False, methods=['get'])
    def pool_tasks(self, request):
        as_role = request.query_params.get('as_role')
        warehouse_id = request.query_params.get('warehouse_id')
        queryset = DocTask.objects.filter(item__is_deleted=False).select_related('item', 'doc_worker', 'doc_supervisor', 'created_by', 'modified_by').prefetch_related(photo_prefetch('item__photos'))
        
        if warehouse_id and str(warehouse_id) not in ['ALL', '-1']:
            queryset = queryset.filter(item__warehouse_id=warehouse_id)
            
        if as_role == 'doc_worker':
            queryset = queryset.filter(doc_worker__isnull=True, status='PENDING_DOC')
        elif as_role == 'doc_supervisor':
            queryset = queryset.filter(doc_supervisor__isnull=True, status='DOC_PROCESSED')
        elif as_role == 'manager':
            queryset = queryset.filter(assigned_manager__isnull=True, status='DOC_MANAGER_REVIEW')
        else:
            return Response({'error': 'نقش نامعتبر است.'}, status=400)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['post'])
    def claim_tasks(self, request):
        task_ids = request.data.get('task_ids', [])
        sync_ids = request.data.get('sync_ids', [])
        as_role = request.data.get('as_role')

        if (not task_ids and not sync_ids) or not as_role:
            return Response({'error': 'لیست شناسه‌ها یا نقش ارسال نشده است.'}, status=400)

        from django.db.models import Q
        q_filter = Q()
        if task_ids:
            q_filter |= Q(id__in=task_ids)
        if sync_ids:
            q_filter |= Q(sync_id__in=sync_ids)

        with transaction.atomic():
            tasks = DocTask.objects.select_for_update().filter(q_filter).select_related('item')
            from .models import DocTaskHistory

            if as_role == 'doc_worker':
                tasks = tasks.filter(doc_worker__isnull=True, status='PENDING_DOC')
                valid_tasks = list(tasks)
                if not valid_tasks:
                    broadcast_doc_task_update()
                    return Response({'success': True, 'claimed_count': 0})

                valid_task_ids = [t.id for t in valid_tasks]
                item_ids = [t.item_id for t in valid_tasks if t.item_id]

                # Update DocTasks
                updated = DocTask.objects.filter(id__in=valid_task_ids).update(doc_worker=request.user, updated_at=timezone.now())

                # Update Item doc_assignee
                assignee_name = f"{request.user.first_name} {request.user.last_name}".strip() or request.user.username
                from .models import Item
                Item.objects.filter(id__in=item_ids).update(doc_assignee=assignee_name, updated_at=timezone.now())

                # Log History
                histories = [
                    DocTaskHistory(
                        task=t,
                        action_by=request.user,
                        action_type='CLAIMED',
                        note=f'بر عهده گرفته شد توسط کارشناس مالی ({assignee_name})',
                        data_snapshot=_create_doc_task_snapshot(t)
                    ) for t in valid_tasks
                ]
                DocTaskHistory.objects.bulk_create(histories)

            elif as_role == 'doc_supervisor':
                tasks = tasks.filter(doc_supervisor__isnull=True, status='DOC_PROCESSED')
                valid_tasks = list(tasks)
                if not valid_tasks:
                    broadcast_doc_task_update()
                    return Response({'success': True, 'claimed_count': 0})
                valid_task_ids = [t.id for t in valid_tasks]
                updated = DocTask.objects.filter(id__in=valid_task_ids).update(doc_supervisor=request.user, updated_at=timezone.now())
                supervisor_name = f"{request.user.first_name} {request.user.last_name}".strip() or request.user.username
                histories = [
                    DocTaskHistory(
                        task=t,
                        action_by=request.user,
                        action_type='CLAIMED',
                        note=f'بر عهده گرفته شد توسط سرپرست اسناد ({supervisor_name})',
                        data_snapshot=_create_doc_task_snapshot(t)
                    ) for t in valid_tasks
                ]
                DocTaskHistory.objects.bulk_create(histories)

            elif as_role == 'manager':
                tasks = tasks.filter(assigned_manager__isnull=True, status='DOC_MANAGER_REVIEW')
                valid_tasks = list(tasks)
                if not valid_tasks:
                    broadcast_doc_task_update()
                    return Response({'success': True, 'claimed_count': 0})
                valid_task_ids = [t.id for t in valid_tasks]
                updated = DocTask.objects.filter(id__in=valid_task_ids).update(assigned_manager=request.user, updated_at=timezone.now())
                manager_name = f"{request.user.first_name} {request.user.last_name}".strip() or request.user.username
                histories = [
                    DocTaskHistory(
                        task=t,
                        action_by=request.user,
                        action_type='CLAIMED',
                        note=f'بر عهده گرفته شد توسط مدیر ({manager_name})',
                        data_snapshot=_create_doc_task_snapshot(t)
                    ) for t in valid_tasks
                ]
                DocTaskHistory.objects.bulk_create(histories)
            else:
                return Response({'error': 'نقش نامعتبر است.'}, status=400)

        broadcast_doc_task_update()
        return Response({'success': True, 'claimed_count': updated})

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user, modified_by=self.request.user)

    def perform_update(self, serializer):
        instance = serializer.instance
        old_status = instance.status
        new_status = serializer.validated_data.get('status', old_status)
        
        updated_instance = serializer.save(modified_by=self.request.user)
        
        if old_status != new_status:
            from .models import DocTaskHistory
            
            note = ''
            if new_status in ['DOC_MANAGER_REJECTED', 'DOC_FINAL_APPROVED']:
                note = updated_instance.manager_note
            elif new_status == 'DOC_SUPERVISOR_REJECTED':
                note = updated_instance.supervisor_note
            elif new_status == 'DOC_PROCESSED':
                note = updated_instance.worker_note
                
            DocTaskHistory.objects.create(
                task=updated_instance,
                action_by=self.request.user,
                action_type=new_status,
                note=note,
                data_snapshot=_create_doc_task_snapshot(updated_instance)
            )

    @action(detail=False, methods=['post'])
    def bulk_submit(self, request):
        user = request.user
        task_ids = request.data.get('task_ids', [])
        sync_ids = request.data.get('sync_ids', [])
        warehouse_id = request.data.get('warehouse_id') or request.query_params.get('warehouse_id')
        
        from django.db.models import Q
        q_filter = Q()
        if task_ids:
            q_filter |= Q(id__in=task_ids)
        if sync_ids:
            q_filter |= Q(sync_id__in=sync_ids)
            
        with transaction.atomic():
            if task_ids or sync_ids:
                tasks = DocTask.objects.select_for_update().filter(q_filter, doc_worker=user, status__in=['PENDING_DOC', 'DOC_SUPERVISOR_REJECTED', 'DOC_MANAGER_REJECTED']).select_related('item')
            else:
                tasks = DocTask.objects.select_for_update().filter(doc_worker=user, status__in=['PENDING_DOC', 'DOC_SUPERVISOR_REJECTED', 'DOC_MANAGER_REJECTED']).select_related('item')
            
            if warehouse_id and str(warehouse_id) not in ['ALL', '-1']:
                tasks = tasks.filter(item__warehouse_id=warehouse_id)
            
            def has_doc_info(t):
                return bool(
                    t.status in ['DOC_SUPERVISOR_REJECTED', 'DOC_MANAGER_REJECTED'] or
                    t.inv_rti_number or t.added_rti_no or t.doc_supplier or
                    t.price_amount or t.total_value or t.similar_unit_price or
                    t.invoice_type or t.invoice_date or t.folder_address or
                    t.stamp or t.signature or (t.worker_note and str(t.worker_note).strip())
                )

            tasks_list = [t for t in tasks if has_doc_info(t)]
            if not tasks_list:
                if task_ids or sync_ids:
                    already = DocTask.objects.filter(
                        Q(id__in=task_ids) | Q(sync_id__in=sync_ids),
                        doc_worker=user,
                        status__in=['DOC_PROCESSED', 'DOC_MANAGER_REVIEW', 'DOC_FINAL_APPROVED']
                    ).count()
                    if already > 0:
                        return Response({
                            'message': f'{already} مورد قبلاً ارسال شده بود.',
                            'already_submitted': already
                        })
                return Response({'message': 'هیچ سند مالی تکمیل‌شده‌ای برای ارجاع یافت نشد.'})
                
            from warehouses.services import get_setting
            from .models import DocTaskHistory
            histories = []
            
            # بررسی تنظیم تایید سرپرست به تفکیک انبار هر تسک
            wh_settings_cache = {}
            target_status_counts = {'DOC_PROCESSED': 0, 'DOC_MANAGER_REVIEW': 0}
            
            now_time = timezone.now()
            for task in tasks_list:
                wh_id = task.item.warehouse_id if task.item else None
                if wh_id not in wh_settings_cache:
                    wh_settings_cache[wh_id] = get_setting('require_doc_supervisor_approval', wh_id)
                req_sup_app = wh_settings_cache[wh_id]
                
                task_req_sup = req_sup_app and not task.skip_supervisor
                target_status = 'DOC_PROCESSED' if task_req_sup else 'DOC_MANAGER_REVIEW'
                task.status = target_status
                task.modified_by = user
                task.updated_at = now_time
                target_status_counts[target_status] += 1
                
                histories.append(DocTaskHistory(
                    task=task,
                    action_by=user,
                    action_type=target_status,
                    note=task.worker_note,
                    data_snapshot=_create_doc_task_snapshot(task)
                ))
                
            count = len(tasks_list)
            if count > 0:
                DocTask.objects.bulk_update(tasks_list, ['status', 'modified_by', 'updated_at'])
                broadcast_doc_task_update()
            if histories:
                DocTaskHistory.objects.bulk_create(histories)
                
            if target_status_counts['DOC_PROCESSED'] > 0 and target_status_counts['DOC_MANAGER_REVIEW'] > 0:
                msg = f'{count} کالا ارسال شد ({target_status_counts["DOC_PROCESSED"]} مورد جهت بررسی سرپرست و {target_status_counts["DOC_MANAGER_REVIEW"]} مورد مستقیماً به مدیر).'
            elif target_status_counts['DOC_PROCESSED'] > 0:
                msg = f'{count} کالا جهت بررسی سرپرست ارسال شد.'
            else:
                msg = f'{count} کالا مستقیماً جهت بررسی مدیر ارسال شد.'
                
            return Response({'message': msg})

    @action(detail=False, methods=['post'])
    def bulk_approve(self, request):
        user = request.user
        task_ids = request.data.get('task_ids', [])
        if not task_ids:
            return Response({'error': 'هیچ کالایی انتخاب نشده است.'}, status=400)
            
        from django.db.models import Q
        from django.db import transaction
        from .models import DocTaskHistory
        note = request.data.get('note', '')
        
        with transaction.atomic():
            tasks = DocTask.objects.select_for_update().filter(id__in=task_ids, status__in=['DOC_PROCESSED', 'DOC_MANAGER_REJECTED'])
            if not (user.is_superuser or user.groups.filter(name__in=['admin', 'manager']).exists()):
                tasks = tasks.filter(Q(doc_supervisor=user) | Q(doc_supervisor__isnull=True))
            
            histories = []
            tasks_list = list(tasks)
            for task in tasks_list:
                task.supervisor_note = note
                task.status = 'DOC_MANAGER_REVIEW'
                task.modified_by = user
                task.updated_at = timezone.now()
                if task.doc_supervisor_id is None:
                    task.doc_supervisor = user
                histories.append(DocTaskHistory(
                    task=task,
                    action_by=user,
                    action_type='DOC_MANAGER_REVIEW',
                    note=note,
                    data_snapshot=_create_doc_task_snapshot(task)
                ))
                
            if tasks_list:
                DocTask.objects.bulk_update(tasks_list, ['status', 'modified_by', 'supervisor_note', 'doc_supervisor', 'updated_at'])
            if histories:
                DocTaskHistory.objects.bulk_create(histories)
                
        broadcast_doc_task_update()
        return Response({'message': f'{len(histories)} رکورد جهت تایید نهایی مدیر ارسال شد.'})

    @action(detail=False, methods=['post'])
    def reject(self, request):
        task_ids = request.data.get('task_ids', [])
        note = request.data.get('note', '')
        
        if not task_ids:
            return Response({'error': 'هیچ رکوردی انتخاب نشده است.'}, status=400)
            
        from django.db import transaction
        from .models import DocTaskHistory
        
        with transaction.atomic():
            tasks = DocTask.objects.select_for_update().filter(id__in=task_ids, status__in=['DOC_PROCESSED', 'DOC_MANAGER_REJECTED'])
            histories = []
            tasks_list = list(tasks)
            for task in tasks_list:
                task.supervisor_note = note
                task.status = 'DOC_SUPERVISOR_REJECTED'
                task.modified_by = request.user
                task.updated_at = timezone.now()
                histories.append(DocTaskHistory(
                    task=task,
                    action_by=request.user,
                    action_type='DOC_SUPERVISOR_REJECTED',
                    note=note,
                    data_snapshot=_create_doc_task_snapshot(task)
                ))
                
            if tasks_list:
                DocTask.objects.bulk_update(tasks_list, ['status', 'modified_by', 'supervisor_note', 'updated_at'])
            if histories:
                DocTaskHistory.objects.bulk_create(histories)
                
        broadcast_doc_task_update()
        return Response({'message': f'{len(histories)} رکورد به بررسی‌کننده اسناد برگشت داده شد.'})

    @action(detail=False, methods=['post'])
    def manager_reject(self, request):
        task_ids = request.data.get('task_ids', [])
        note = request.data.get('note', '')
        
        if not task_ids:
            return Response({'error': 'هیچ رکوردی انتخاب نشده است.'}, status=400)
            
        tasks = DocTask.objects.filter(id__in=task_ids, status='DOC_MANAGER_REVIEW')
        
        from .models import DocTaskHistory
        histories = []
        for task in tasks:
            task.manager_note = note
            task.status = 'DOC_MANAGER_REJECTED'
            task.modified_by = request.user
            histories.append(DocTaskHistory(
                task=task,
                action_by=request.user,
                action_type='DOC_MANAGER_REJECTED',
                note=note,
                data_snapshot=_create_doc_task_snapshot(task)
            ))
            
        if tasks:
            DocTask.objects.bulk_update(tasks, ['status', 'modified_by', 'manager_note'])
            DocTaskHistory.objects.bulk_create(histories)
            broadcast_doc_task_update()
            
        return Response({'message': f'{len(histories)} رکورد به سرپرست اسناد (یا بررسی‌کننده) برگشت داده شد.'})

    @action(detail=False, methods=['post'])
    def bulk_manager_approve(self, request):
        task_ids = request.data.get('task_ids', [])
        note = request.data.get('note', '')
        if not task_ids:
            return Response({'error': 'هیچ رکوردی انتخاب نشده است.'}, status=400)
            
        tasks_list = list(DocTask.objects.filter(id__in=task_ids, status='DOC_MANAGER_REVIEW').select_related('item'))
        if not tasks_list:
            return Response({'message': 'هیچ رکوردی در انتظار تایید مدیر یافت نشد.'})
        
        from .models import DocTaskHistory, Item
        from django.db import transaction
        
        histories = []
        items_to_update = []
        now = timezone.now()
        
        with transaction.atomic():
            for task in tasks_list:
                if note:
                    task.manager_note = note
                task.status = 'DOC_FINAL_APPROVED'
                task.modified_by = request.user
                histories.append(DocTaskHistory(
                    task=task,
                    action_by=request.user,
                    action_type='DOC_FINAL_APPROVED',
                    note=task.manager_note,
                    data_snapshot=_create_doc_task_snapshot(task)
                ))
                
                # انتقال و همگام‌سازی فیلدهای مالی به رکورد کالا (Item)
                item = task.item
                if item:
                    item.doc_status = 'done'
                    item.price_amount = task.price_amount
                    item.total_value = task.total_value
                    item.similar_unit_price = task.similar_unit_price
                    item.currency = task.currency
                    item.invoice_type = task.invoice_type
                    item.invoice_date = str(task.invoice_date) if task.invoice_date else None
                    item.inv_rti_number = task.inv_rti_number
                    item.added_rti_no = task.added_rti_no
                    item.page_row = str(task.page_row) if task.page_row is not None else None
                    item.invoice_page = str(task.invoice_page) if task.invoice_page is not None else None
                    item.doc_supplier = task.doc_supplier
                    item.folder_address = task.folder_address
                    if task.stamp is not None:
                        item.stamp = 'دارد' if task.stamp else 'ندارد'
                    if task.signature is not None:
                        item.signature = 'دارد' if task.signature else 'ندارد'
                    item.modified_by = request.user
                    item.updated_at = now
                    items_to_update.append(item)
                
            if tasks_list:
                DocTask.objects.bulk_update(tasks_list, ['status', 'modified_by', 'manager_note'])
                DocTaskHistory.objects.bulk_create(histories)
                if items_to_update:
                    Item.objects.bulk_update(items_to_update, [
                        'doc_status', 'price_amount', 'total_value', 'similar_unit_price',
                        'currency', 'invoice_type', 'invoice_date', 'inv_rti_number',
                        'added_rti_no', 'page_row', 'invoice_page', 'doc_supplier',
                        'folder_address', 'stamp', 'signature', 'modified_by', 'updated_at'
                    ])
                broadcast_doc_task_update()
            
        return Response({'message': f'{len(histories)} رکورد نهایی شد.'})
        
    @action(detail=False, methods=['post'])
    def bulk_cancel(self, request):
        task_ids = request.data.get('task_ids', [])
        if not task_ids:
            return Response({'error': 'هیچ رکوردی انتخاب نشده است.'}, status=400)
            
        tasks = DocTask.objects.filter(id__in=task_ids)
        item_ids = list(tasks.values_list('item_id', flat=True))
        
        from .models import Item
        deleted_count, _ = tasks.delete()
        Item.objects.filter(id__in=item_ids).update(doc_status='checking', doc_assignee=None, updated_at=timezone.now())
        
        broadcast_doc_task_update()
        return Response({'message': f'{deleted_count} وظیفه ارجاع اسناد با موفقیت لغو شد.'})

    @action(detail=False, methods=['get'], url_path='download_template')
    def download_template(self, request):
        """تولید فایل اکسل نمونه آزمایشی برای کارتابل مالی متناسب با فیلدهای قابل ویرایش"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
        from django.http import HttpResponse
        from warehouses.services import get_setting

        warehouse_id = request.query_params.get('warehouse_id')
        if warehouse_id and str(warehouse_id) in ['ALL', '-1', 'null', 'undefined']:
            warehouse_id = None
        num_wh_id = int(warehouse_id) if warehouse_id and str(warehouse_id).isdigit() else None

        # دریافت تنظیمات فیلدهای کارتابل مالی
        saved_perms = get_setting('field_permissions_doc', num_wh_id) or {}
        if isinstance(saved_perms, str):
            import json
            try:
                saved_perms = json.loads(saved_perms)
            except Exception:
                saved_perms = {}

        # لیست فیلدهای استاندارد مالی
        STANDARD_DOC_FIELDS = [
            {'key': 'price_amount',       'default_label': 'قیمت واحد',        'type': 'number',   'sample1': '2500000',               'sample2': '1850000'},
            {'key': 'similar_unit_price', 'default_label': 'قیمت کالای مشابه', 'type': 'number',   'sample1': '2400000',               'sample2': '1800000'},
            {'key': 'total_value',        'default_label': 'ارزش کل',          'type': 'number',   'sample1': '25000000',              'sample2': '18500000'},
            {'key': 'currency',           'default_label': 'ارز',              'type': 'currency', 'sample1': 'ریال',                  'sample2': 'دلار'},
            {'key': 'invoice_type',       'default_label': 'نوع فاکتور',        'type': 'inv_type', 'sample1': 'رسمی/مالیاتی',          'sample2': 'خریدهای داخلی'},
            {'key': 'invoice_date',       'default_label': 'تاریخ فاکتور',     'type': 'date',     'sample1': '1405/05/25 07:10',      'sample2': '1405/06/01 10:30'},
            {'key': 'inv_rti_number',     'default_label': 'شماره RTI فاکتور', 'type': 'text',     'sample1': 'RTI-1405-001',          'sample2': 'RTI-1405-002'},
            {'key': 'added_rti_no',       'default_label': 'شماره RTI افزوده‌شده', 'type': 'text', 'sample1': 'RTI-ADD-99',           'sample2': 'RTI-ADD-100'},
            {'key': 'invoice_page',       'default_label': 'صفحه فاکتور',      'type': 'number',   'sample1': '1',                     'sample2': '2'},
            {'key': 'page_row',           'default_label': 'ردیف فاکتور',      'type': 'number',   'sample1': '1',                     'sample2': '3'},
            {'key': 'doc_supplier',       'default_label': 'تأمین‌کننده',      'type': 'text',     'sample1': 'شرکت تأمین تجهیز پارس', 'sample2': 'شرکت مهندسی پویا'},
            {'key': 'folder_address',     'default_label': 'مسیر پوشه اسناد',  'type': 'text',     'sample1': 'Z:/Docs/Archive/1405',  'sample2': 'Z:/Docs/Archive/1405'},
            {'key': 'stamp',              'default_label': 'مهر',              'type': 'boolean',  'sample1': 'بله',                   'sample2': 'خیر'},
            {'key': 'signature',          'default_label': 'امضا',             'type': 'boolean',  'sample1': 'بله',                   'sample2': 'بله'},
            {'key': 'worker_note',        'default_label': 'یادداشت کارشناس',  'type': 'text',     'sample1': 'مدارک مالی مطابقت دارد', 'sample2': 'فاکتور ضمیمه شد'},
        ]

        # همیشه ستون شناساگر کد یکتا در ابتدا قرار می‌گیرد
        columns = [
            {'key': 'fa_unic_code', 'label': 'کد یکتا', 'type': 'text', 'sample1': 'FA-10001', 'sample2': 'FA-10002'}
        ]

        for f in STANDARD_DOC_FIELDS:
            cfg = saved_perms.get(f['key'], {}) if isinstance(saved_perms, dict) else {}
            is_editable = cfg.get('editable', True)
            is_visible = cfg.get('visible', True)
            if is_visible and is_editable:
                label = (cfg.get('custom_label') or '').strip() or f['default_label']
                columns.append({
                    'key': f['key'],
                    'label': label,
                    'type': f['type'],
                    'sample1': f['sample1'],
                    'sample2': f['sample2'],
                })

        # افزودن فیلدهای داینامیک فعال و قابل ویرایش انبار
        if num_wh_id:
            from .models import ItemFieldDefinition
            dyn_defs = ItemFieldDefinition.objects.filter(warehouse_id=num_wh_id, is_active=True)
            for d in dyn_defs:
                dyn_key = f"dyn_{d.name}"
                cfg = saved_perms.get(dyn_key, {}) if isinstance(saved_perms, dict) else {}
                is_editable = cfg.get('editable', True)
                is_visible = cfg.get('visible', True)
                if is_visible and is_editable:
                    label = (cfg.get('custom_label') or '').strip() or d.label or d.name
                    dtype = d.field_type
                    if dtype == 'number':
                        s1, s2 = '100', '250'
                    elif dtype == 'boolean':
                        s1, s2 = 'بله', 'خیر'
                    elif dtype == 'date':
                        s1, s2 = '1405/05/25', '1405/06/01'
                    else:
                        s1, s2 = 'مقدار نمونه ۱', 'مقدار نمونه ۲'
                    columns.append({
                        'key': dyn_key,
                        'label': label,
                        'type': 'boolean' if dtype == 'boolean' else ('number' if dtype == 'number' else ('date' if dtype == 'date' else 'text')),
                        'sample1': s1,
                        'sample2': s2,
                    })

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'قالب اسناد مالی'
        ws.sheet_view.rightToLeft = True

        header_font = Font(name='Tahoma', bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
        sample_font = Font(name='Tahoma', size=10)
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'), right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'), bottom=Side(style='thin', color='E2E8F0')
        )
        fill_row1 = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
        fill_row2 = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        headers = [c['label'] for c in columns]
        ws.append(headers)

        for col_idx, cell in enumerate(ws[1], 1):
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border

        row1_data = [c['sample1'] for c in columns]
        row2_data = [c['sample2'] for c in columns]
        ws.append(row1_data)
        ws.append(row2_data)

        for r_idx, (r_data, r_fill) in enumerate([(row1_data, fill_row1), (row2_data, fill_row2)], start=2):
            for c_idx, val in enumerate(r_data, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = sample_font
                cell.fill = r_fill
                cell.alignment = center_alignment
                cell.border = thin_border

        # اعمال کشوها با DataValidation
        bool_dv = DataValidation(type="list", formula1='"بله,خیر"', allow_blank=True)
        inv_dv = DataValidation(type="list", formula1='"رسمی/مالیاتی,خریدهای داخلی,خریدهای خارجی,امانی"', allow_blank=True)
        cur_dv = DataValidation(type="list", formula1='"ریال,دلار,یورو,سایر"', allow_blank=True)

        ws.add_data_validation(bool_dv)
        ws.add_data_validation(inv_dv)
        ws.add_data_validation(cur_dv)

        for col_idx, col in enumerate(columns, 1):
            col_letter = get_column_letter(col_idx)
            if col['type'] == 'boolean':
                bool_dv.add(f"{col_letter}2:{col_letter}500")
            elif col['type'] == 'inv_type':
                inv_dv.add(f"{col_letter}2:{col_letter}500")
            elif col['type'] == 'currency':
                cur_dv.add(f"{col_letter}2:{col_letter}500")

            max_len = max(len(str(col['label'])), len(str(col.get('sample1', ''))), len(str(col.get('sample2', ''))))
            ws.column_dimensions[col_letter].width = max(max_len + 6, 14)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Customs_Doc_Template.xlsx"'
        wb.save(response)
        return response

    @action(detail=False, methods=['get'], url_path='get_export_columns')
    def get_export_columns(self, request):
        """لیست ستون‌های مجاز برای خروجی اکسل کارتابل مالی و اسناد"""
        columns = [
            {'key': 'warehouse_name',        'label': 'نام انبار'},
            {'key': 'fa_unic_code',          'label': 'کد یکتا'},
            {'key': 'description',           'label': 'شرح کالا'},
            {'key': 'po',                    'label': 'شماره PO'},
            {'key': 'new_location',          'label': 'لوکیشن'},
            {'key': 'status',                'label': 'وضعیت'},
            {'key': 'doc_worker_name',       'label': 'کارشناس مالی'},
            {'key': 'doc_supervisor_name',   'label': 'سرپرست'},
            {'key': 'assigned_manager_name', 'label': 'مدیر'},
            {'key': 'inv_rti_number',        'label': 'شماره RTI فاکتور'},
            {'key': 'added_rti_no',          'label': 'شماره RTI افزوده‌شده'},
            {'key': 'invoice_type',          'label': 'نوع فاکتور'},
            {'key': 'invoice_date',          'label': 'تاریخ فاکتور'},
            {'key': 'invoice_page',          'label': 'صفحه فاکتور'},
            {'key': 'page_row',              'label': 'ردیف فاکتور'},
            {'key': 'doc_supplier',          'label': 'تأمین‌کننده'},
            {'key': 'price_amount',          'label': 'قیمت واحد'},
            {'key': 'similar_unit_price',    'label': 'قیمت کالای مشابه'},
            {'key': 'total_value',           'label': 'ارزش کل'},
            {'key': 'currency',              'label': 'ارز'},
            {'key': 'folder_address',        'label': 'مسیر پوشه اسناد'},
            {'key': 'stamp',                 'label': 'مهر'},
            {'key': 'signature',             'label': 'امضا'},
            {'key': 'worker_note',           'label': 'یادداشت کارشناس'},
            {'key': 'supervisor_note',       'label': 'یادداشت سرپرست'},
            {'key': 'manager_note',          'label': 'یادداشت مدیر'},
            {'key': 'created_at',            'label': 'تاریخ ایجاد'},
            {'key': 'updated_at',            'label': 'تاریخ بروزرسانی'},
        ]
        return Response(columns)

    @action(detail=False, methods=['post'], url_path='export_excel')
    def export_excel(self, request):
        """خروجی اکسل صفحه کارتابل مالی با تاریخ شمسی و وضعیت فارسی"""
        import openpyxl
        from django.http import HttpResponse

        STATUS_FA = {
            'PENDING_DOC':             'در انتظار بررسی',
            'DOC_PROCESSED':           'ارسال‌شده به سرپرست',
            'DOC_SUPERVISOR_REJECTED': 'رد سرپرست',
            'DOC_MANAGER_REVIEW':      'در بررسی مدیر',
            'DOC_MANAGER_REJECTED':    'رد مدیر',
            'DOC_FINAL_APPROVED':      'تأیید نهایی',
        }

        INVOICE_TYPE_FA = {
            'formal':      'رسمی/مالیاتی',
            'domestic':    'خریدهای داخلی',
            'foreign':     'خریدهای خارجی',
            'consignment': 'امانی',
        }

        CURRENCY_FA = {
            'IRR':   'ریال',
            'USD':   'دلار',
            'EUR':   'یورو',
            'OTHER': 'سایر',
        }

        def to_jalali(dt):
            if not dt:
                return ''
            try:
                import jdatetime
                from django.utils import timezone
                if timezone.is_aware(dt):
                    dt = timezone.localtime(dt)
                jdt = jdatetime.datetime.fromgregorian(datetime=dt)
                return jdt.strftime('%Y/%m/%d %H:%M')
            except Exception:
                return str(dt)

        data_scope    = request.data.get('data_scope', 'all')
        columns_scope = request.data.get('columns_scope', 'all_db')
        columns_list  = request.data.get('columns_list', [])

        from django.http import QueryDict
        original_query_params = request._request.GET
        try:
            q = QueryDict(mutable=True)
            for k, v in request.data.items():
                if isinstance(v, list):
                    q.setlist(k, [str(x) for x in v])
                elif v is not None:
                    q[k] = str(v)
            request._request.GET = q

            if data_scope == 'selected':
                selected_ids = request.data.get('selected_ids', [])
                queryset = self.get_queryset().filter(id__in=selected_ids)
            else:
                queryset = self.get_queryset()
        finally:
            request._request.GET = original_query_params

        ALL_COLUMNS = [
            {'key': 'warehouse_name',        'label': 'نام انبار'},
            {'key': 'fa_unic_code',          'label': 'کد یکتا'},
            {'key': 'description',           'label': 'شرح کالا'},
            {'key': 'po',                    'label': 'شماره PO'},
            {'key': 'new_location',          'label': 'لوکیشن'},
            {'key': 'status',                'label': 'وضعیت'},
            {'key': 'doc_worker_name',       'label': 'کارشناس مالی'},
            {'key': 'doc_supervisor_name',   'label': 'سرپرست'},
            {'key': 'assigned_manager_name', 'label': 'مدیر'},
            {'key': 'inv_rti_number',        'label': 'شماره RTI فاکتور'},
            {'key': 'added_rti_no',          'label': 'شماره RTI افزوده‌شده'},
            {'key': 'invoice_type',          'label': 'نوع فاکتور'},
            {'key': 'invoice_date',          'label': 'تاریخ فاکتور'},
            {'key': 'invoice_page',          'label': 'صفحه فاکتور'},
            {'key': 'page_row',              'label': 'ردیف فاکتور'},
            {'key': 'doc_supplier',          'label': 'تأمین‌کننده'},
            {'key': 'price_amount',          'label': 'قیمت واحد'},
            {'key': 'similar_unit_price',    'label': 'قیمت کالای مشابه'},
            {'key': 'total_value',           'label': 'ارزش کل'},
            {'key': 'currency',              'label': 'ارز'},
            {'key': 'folder_address',        'label': 'مسیر پوشه اسناد'},
            {'key': 'stamp',                 'label': 'مهر'},
            {'key': 'signature',             'label': 'امضا'},
            {'key': 'worker_note',           'label': 'یادداشت کارشناس'},
            {'key': 'supervisor_note',       'label': 'یادداشت سرپرست'},
            {'key': 'manager_note',          'label': 'یادداشت مدیر'},
            {'key': 'created_at',            'label': 'تاریخ ایجاد'},
            {'key': 'updated_at',            'label': 'تاریخ بروزرسانی'},
        ]
        all_keys = [c['key'] for c in ALL_COLUMNS]
        key_to_label = {c['key']: c['label'] for c in ALL_COLUMNS}

        if columns_scope in ('visible', 'custom') and columns_list:
            selected_keys = [k for k in columns_list if k in all_keys]
            if not selected_keys:
                selected_keys = all_keys
        else:
            selected_keys = all_keys

        def get_cell(task, key):
            if key == 'warehouse_name':
                if task.item and task.item.warehouse:
                    return getattr(task.item.warehouse, 'project_name', None) or task.item.warehouse.name or ''
                return ''
            elif key == 'fa_unic_code':
                return getattr(task.item, 'fa_unic_code', '') if task.item else ''
            elif key == 'description':
                return getattr(task.item, 'description', '') if task.item else ''
            elif key == 'po':
                return getattr(task.item, 'po', '') if task.item else ''
            elif key == 'new_location':
                return getattr(task.item, 'new_location', '') if task.item else ''
            elif key == 'status':
                return STATUS_FA.get(task.status, task.status)
            elif key == 'doc_worker_name':
                if task.doc_worker:
                    return f"{task.doc_worker.first_name} {task.doc_worker.last_name}".strip() or task.doc_worker.username
                return ''
            elif key == 'doc_supervisor_name':
                if task.doc_supervisor:
                    return f"{task.doc_supervisor.first_name} {task.doc_supervisor.last_name}".strip() or task.doc_supervisor.username
                return ''
            elif key == 'assigned_manager_name':
                if task.assigned_manager:
                    return f"{task.assigned_manager.first_name} {task.assigned_manager.last_name}".strip() or task.assigned_manager.username
                return ''
            elif key == 'inv_rti_number':
                return task.inv_rti_number or ''
            elif key == 'added_rti_no':
                return task.added_rti_no or ''
            elif key == 'invoice_type':
                return INVOICE_TYPE_FA.get(task.invoice_type, task.invoice_type or '')
            elif key == 'invoice_date':
                return task.invoice_date or ''
            elif key == 'invoice_page':
                return task.invoice_page if task.invoice_page is not None else ''
            elif key == 'page_row':
                return task.page_row if task.page_row is not None else ''
            elif key == 'doc_supplier':
                return task.doc_supplier or ''
            elif key == 'price_amount':
                return task.price_amount or ''
            elif key == 'similar_unit_price':
                return task.similar_unit_price or ''
            elif key == 'total_value':
                return task.total_value or ''
            elif key == 'currency':
                return CURRENCY_FA.get(task.currency, task.currency or '')
            elif key == 'folder_address':
                return task.folder_address or ''
            elif key == 'stamp':
                return 'بله' if task.stamp else 'خیر'
            elif key == 'signature':
                return 'بله' if task.signature else 'خیر'
            elif key == 'worker_note':
                return task.worker_note or ''
            elif key == 'supervisor_note':
                return task.supervisor_note or ''
            elif key == 'manager_note':
                return task.manager_note or ''
            elif key == 'created_at':
                return to_jalali(task.created_at)
            elif key == 'updated_at':
                return to_jalali(task.updated_at)
            return ''

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'کارتابل مالی'
        ws.sheet_view.rightToLeft = True

        header_font = Font(name='Tahoma', bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid') # سرمه‌ای دودی مدرن
        cell_font = Font(name='Tahoma', size=10)

        fill_even = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
        fill_odd = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'), right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'), bottom=Side(style='thin', color='E2E8F0')
        )

        headers = [key_to_label[k] for k in selected_keys]
        ws.append(headers)

        for col_idx, cell in enumerate(ws[1], 1):
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border

        ws.freeze_panes = 'A2'
        col_widths = {i: len(headers[i]) + 6 for i in range(len(headers))}

        row_idx = 2
        for task in queryset.select_related('item', 'item__warehouse', 'doc_worker', 'doc_supervisor', 'assigned_manager').iterator():
            row_data = [str(get_cell(task, k)) for k in selected_keys]
            ws.append(row_data)

            current_fill = fill_even if row_idx % 2 == 0 else fill_odd

            for col_idx, val in enumerate(row_data):
                cell = ws.cell(row=row_idx, column=col_idx + 1)
                cell.font = cell_font
                cell.fill = current_fill
                cell.alignment = center_alignment
                cell.border = thin_border

                lines = val.split('\n')
                max_line_len = max([len(line) for line in lines]) if lines else 0
                if max_line_len > col_widths.get(col_idx, 10):
                    col_widths[col_idx] = min(max_line_len + 4, 60)

            row_idx += 1

        for col_idx, width in col_widths.items():
            ws.column_dimensions[get_column_letter(col_idx + 1)].width = width

        if row_idx > 2:
            last_col_letter = get_column_letter(len(selected_keys))
            ws.auto_filter.ref = f"A1:{last_col_letter}{row_idx - 1}"

        from io import BytesIO
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="customs_cartable_export.xlsx"'
        return response

