import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from decimal import Decimal
from django.db import transaction
from django.db.models import Q
import jdatetime

from .models import (
    VehicleDriverProfile,
    VehicleTripLog,
    MonthlyWorkPeriod,
)
from common.date_utils import normalize_digits


def export_fleet_monthly_excel(warehouse_id=None, year_month=None):
    """
    تولید فایل اکسل استاندارد ۳۱ روزه تردد و کارکرد ماهانه ناوگان
    مطابق با معماری ۲ سطری هدر:
    سطر ۱: عناوین فارسی جهت نمایش به کاربر
    سطر ۲: کلیدهای دیتابیسی با فونت طوسی ریز
    فریز پنل در A3
    """
    if not year_month:
        today = jdatetime.date.today()
        year_month = f"{today.year:04d}/{today.month:02d}"

    # نرمال‌سازی سال و ماه
    ym_cleaned = normalize_digits(str(year_month)).strip().replace('-', '/')
    parts = ym_cleaned.split('/')
    if len(parts) >= 2:
        y, m = int(parts[0]), int(parts[1])
    else:
        y, m = 1405, 4
    year_month = f"{y:04d}/{m:02d}"

    if 1 <= m <= 6:
        days_in_month = 31
    elif 7 <= m <= 11:
        days_in_month = 30
    elif m == 12:
        days_in_month = 30 if jdatetime.date(y, 1, 1).isleap() else 29
    else:
        days_in_month = 31

    month_names = ["", "فروردین", "اردیبهشت", "خرداد", "تیر", "مرداد", "شهریور", "مهر", "آبان", "آذر", "دی", "بهمن", "اسفند"]
    month_name = month_names[m] if 1 <= m <= 12 else ""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"تردد ناوگان {m:02d}-{y}"
    ws.sheet_view.rightToLeft = True

    # استایل‌ها
    header_font = Font(name='Tahoma', size=10, bold=True, color='FFFFFF')
    key_font = Font(name='Courier New', size=8, bold=False, color='64748B')
    data_font = Font(name='Tahoma', size=9, bold=False)
    bold_data_font = Font(name='Tahoma', size=9, bold=True)
    
    header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid') # سرمه‌ای تیره
    friday_header_fill = PatternFill(start_color='991B1B', end_color='991B1B', fill_type='solid') # زرشکی برای جمعه
    key_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
    friday_col_fill = PatternFill(start_color='FFF1F2', end_color='FFF1F2', fill_type='solid') # قرمز بسیار روشن
    summary_fill = PatternFill(start_color='EEF2FF', end_color='EEF2FF', fill_type='solid') # نیلی ملایم

    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # ردیف‌های هدر
    row1_headers = ['ردیف', 'نام راننده', 'پلاک انتظامی', 'نوع خودرو', 'نوع مالکیت', 'نرخ پایه (تومان)']
    row2_keys = ['row_idx', 'driver_name', 'plate_number', 'vehicle_type', 'ownership_type', 'default_rate']

    friday_col_indexes = []
    weekday_short_names = ["ش", "ی", "د", "س", "چ", "پ", "ج"]
    
    col_idx = 7
    for d in range(1, days_in_month + 1):
        target_date = jdatetime.date(y, m, d)
        wday = target_date.weekday()
        is_friday = (wday == 6)
        if is_friday:
            friday_col_indexes.append(col_idx)

        w_name = weekday_short_names[wday]
        row1_headers.append(f"{d}\n{w_name}")
        row2_keys.append(f"day_{d:02d}")
        col_idx += 1

    row1_headers.extend(['جمع سرویس', 'جمع کل کرایه (تومان)', 'شماره شبا'])
    row2_keys.extend(['total_trips', 'total_amount', 'sheba_number'])

    # نوشتن هدر سطر ۱
    ws.append(row1_headers)
    ws.row_dimensions[1].height = 28
    for c_idx, val in enumerate(row1_headers, start=1):
        cell = ws.cell(row=1, column=c_idx)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        if c_idx in friday_col_indexes:
            cell.fill = friday_header_fill
        else:
            cell.fill = header_fill

    # نوشتن هدر سطر ۲ (کلیدهای سیستمی)
    ws.append(row2_keys)
    ws.row_dimensions[2].height = 16
    for c_idx, val in enumerate(row2_keys, start=1):
        cell = ws.cell(row=2, column=c_idx)
        cell.font = key_font
        cell.alignment = center_align
        cell.fill = key_fill
        cell.border = thin_border

    # واکشی اطلاعات خودروها و ترددها
    vehicles = VehicleDriverProfile.objects.filter(
        Q(assigned_warehouse_id=warehouse_id) | Q(assigned_warehouse__isnull=True),
        is_active=True
    ).order_by('driver_name')

    trips = VehicleTripLog.objects.filter(
        date_shamsi__startswith=year_month,
        is_deleted=False
    )
    if warehouse_id:
        trips = trips.filter(warehouse_id=warehouse_id)

    trip_map = {}
    for t in trips:
        try:
            d_num = int(t.date_shamsi.split('/')[2])
            trip_map[(t.vehicle_id, d_num)] = t
        except Exception:
            continue

    # نوشتن سطرهای داده
    current_row = 3
    for v_idx, v in enumerate(vehicles, start=1):
        row_data = [
            v_idx,
            v.driver_name,
            v.plate_number,
            v.get_vehicle_type_display(),
            v.get_ownership_type_display(),
            float(v.default_service_rate)
        ]

        total_trips = 0
        total_amount = 0.0

        for d in range(1, days_in_month + 1):
            t = trip_map.get((v.id, d))
            if t and t.trip_count > 0:
                row_data.append(t.trip_count)
                total_trips += t.trip_count
                if v.ownership_type in ['contract', 'personal']:
                    total_amount += float(t.total_amount)
            else:
                row_data.append(0)

        row_data.append(total_trips)
        row_data.append(total_amount if v.ownership_type != 'company' else 0)
        row_data.append(v.sheba_number or '')

        ws.append(row_data)
        ws.row_dimensions[current_row].height = 20

        # اعمال استایل سطرهای داده
        for c_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=current_row, column=c_idx)
            cell.font = data_font
            cell.border = thin_border
            
            if c_idx == 2:
                cell.alignment = right_align
            elif c_idx in [3, len(row_data)]:
                cell.alignment = left_align
            else:
                cell.alignment = center_align

            if c_idx in friday_col_indexes:
                cell.fill = friday_col_fill
            elif c_idx in [len(row_data) - 2, len(row_data) - 1]:
                cell.fill = summary_fill
                cell.font = bold_data_font

            if c_idx in [6, len(row_data) - 1]:
                cell.number_format = '#,##0'

        current_row += 1

    # تنظیم عرض ستون‌ها
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        if col[0].column in range(7, 7 + days_in_month):
            ws.column_dimensions[col_letter].width = 5
        elif col[0].column == 2:
            ws.column_dimensions[col_letter].width = 22
        elif col[0].column == 3:
            ws.column_dimensions[col_letter].width = 16
        elif col[0].column in [len(row1_headers) - 1, len(row1_headers)]:
            ws.column_dimensions[col_letter].width = 20
        else:
            ws.column_dimensions[col_letter].width = 14

    # فریز کردن پنل از سطر ۳
    ws.freeze_panes = 'A3'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def import_fleet_monthly_excel(file_obj, warehouse_id=None, year_month=None, user=None):
    """
    درون‌ریزی هوشمند فایل اکسل ۳۱ روزه تردد ناوگان و ذخیره در پایگاه داده
    """
    if not year_month:
        today = jdatetime.date.today()
        year_month = f"{today.year:04d}/{today.month:02d}"

    ym_cleaned = normalize_digits(str(year_month)).strip().replace('-', '/')
    parts = ym_cleaned.split('/')
    if len(parts) >= 2:
        y, m = int(parts[0]), int(parts[1])
    else:
        y, m = 1405, 4
    year_month = f"{y:04d}/{m:02d}"

    if 1 <= m <= 6:
        days_in_month = 31
    elif 7 <= m <= 11:
        days_in_month = 30
    elif m == 12:
        days_in_month = 30 if jdatetime.date(y, 1, 1).isleap() else 29
    else:
        days_in_month = 31

    # بررسی قفل دوره
    if warehouse_id:
        period = MonthlyWorkPeriod.objects.filter(warehouse_id=warehouse_id, year_month=year_month).first()
        if period and period.status in ['LOCKED', 'SUBMITTED', 'FINALIZED']:
            raise ValueError(f"دوره کارکرد {year_month} قفل شده است و امکان بارگذاری اکسل وجود ندارد.")

    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.active

    # خواندن کلیدهای سطر ۲
    key_col_map = {}
    for c_idx in range(1, ws.max_column + 1):
        k_val = ws.cell(row=2, column=c_idx).value
        if k_val:
            key_col_map[str(k_val).strip().lower()] = c_idx

    plate_col = key_col_map.get('plate_number', 3)
    driver_col = key_col_map.get('driver_name', 2)

    all_vehicles = {v.plate_number.replace(' ', ''): v for v in VehicleDriverProfile.objects.filter(is_active=True)}
    all_vehicles_by_name = {v.driver_name.strip(): v for v in VehicleDriverProfile.objects.filter(is_active=True)}

    updated_trips_count = 0
    with transaction.atomic():
        for r_idx in range(3, ws.max_row + 1):
            plate_val = str(ws.cell(row=r_idx, column=plate_col).value or '').strip().replace(' ', '')
            driver_val = str(ws.cell(row=r_idx, column=driver_col).value or '').strip()

            v_obj = all_vehicles.get(plate_val) or all_vehicles_by_name.get(driver_val)
            if not v_obj:
                continue

            target_wh = warehouse_id if warehouse_id else v_obj.assigned_warehouse_id

            # بررسی ستون‌های روزها (day_01 تا day_31)
            for d in range(1, days_in_month + 1):
                day_key = f"day_{d:02d}"
                c_idx = key_col_map.get(day_key, 6 + d)
                cell_val = ws.cell(row=r_idx, column=c_idx).value

                try:
                    trip_count = int(float(cell_val or 0))
                except (ValueError, TypeError):
                    trip_count = 0

                date_shamsi = f"{year_month}/{d:02d}"
                trip_obj = VehicleTripLog.objects.filter(
                    vehicle_id=v_obj.id,
                    warehouse_id=target_wh,
                    date_shamsi=date_shamsi,
                    is_deleted=False
                ).first()

                if trip_count > 0:
                    unit_rate = v_obj.default_service_rate
                    total_amount = unit_rate * trip_count
                    if trip_obj:
                        trip_obj.trip_count = trip_count
                        trip_obj.unit_rate = unit_rate
                        trip_obj.total_amount = total_amount
                        trip_obj.save()
                    else:
                        VehicleTripLog.objects.create(
                            vehicle_id=v_obj.id,
                            warehouse_id=target_wh,
                            date_shamsi=date_shamsi,
                            trip_count=trip_count,
                            unit_rate=unit_rate,
                            total_amount=total_amount,
                            created_by=user
                        )
                    updated_trips_count += 1
                else:
                    if trip_obj:
                        trip_obj.is_deleted = True
                        trip_obj.save()

    return {
        'message': f'اطلاعات تردد ماهانه با موفقیت بارگذاری شد ({updated_trips_count} سرویس ثبت/به‌روز شد).',
        'updated_count': updated_trips_count
    }
