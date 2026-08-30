# -*- coding: utf-8 -*-
"""
Fleet Settlement Engine and Group Bank Payment Diskette Generator.
Calculates monthly gross/net payouts for contracted/rented fleet vehicles,
and exports official Bank Meli group payment Excel files.
"""
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from decimal import Decimal
from django.db.models import Sum, F
from .models import VehicleDriverProfile, VehicleTripLog, MonthlyWorkPeriod


def calculate_monthly_fleet_settlement(warehouse_id=None, year_month=None):
    """
    محاسبه کارکرد و مبالغ تسویه ماهانه ناوگان خودرویی
    - خودروهای استیجاری و ملکی راننده: مشمول محاسبه کرایه و تسویه بانکی
    - خودروهای شرکتی: صرفاً ثبت تردد، بدون مبلغ تسویه (۰ تومان)
    """
    if not year_month:
        raise ValueError("پارامتر سال و ماه (year_month) الزامی است.")

    # بررسی وضعیت قفل دوره
    period_qs = MonthlyWorkPeriod.objects.filter(year_month=year_month)
    if warehouse_id:
        period_qs = period_qs.filter(warehouse_id=warehouse_id)
    period = period_qs.first()
    is_locked = bool(period and period.status in ('LOCKED', 'FINALIZED'))

    # لیست خودروها
    vehicle_qs = VehicleDriverProfile.objects.all()
    if warehouse_id:
        vehicle_qs = vehicle_qs.filter(assigned_warehouse_id=warehouse_id)

    # لاگ‌های تردد در بازه سال/ماه
    trip_qs = VehicleTripLog.objects.filter(date_shamsi__startswith=year_month)
    if warehouse_id:
        trip_qs = trip_qs.filter(warehouse_id=warehouse_id)

    # مپ ترددها بر اساس شناسه خودرو
    trips_by_vehicle = {}
    for t in trip_qs:
        v_id = t.vehicle_id
        if v_id not in trips_by_vehicle:
            trips_by_vehicle[v_id] = []
        trips_by_vehicle[v_id].append(t)

    records = []
    total_trips_sum = 0
    total_payable_sum = Decimal(0)
    contractor_count = 0

    for v in vehicle_qs:
        v_trips = trips_by_vehicle.get(v.id, [])
        trip_count = sum(t.trip_count for t in v_trips)
        gross_amount = sum((t.trip_count * t.unit_rate) for t in v_trips)
        
        # خودروهای شرکتی بدون مبلغ تسویه هستند
        is_payable = (v.ownership_type in ('contract', 'personal'))
        payable_amount = gross_amount if is_payable else Decimal(0)

        total_trips_sum += trip_count
        if is_payable:
            total_payable_sum += payable_amount
            if trip_count > 0:
                contractor_count += 1

        records.append({
            'vehicle_id': v.id,
            'driver_name': v.driver_name,
            'driver_national_code': v.driver_national_code or '',
            'driver_phone': v.driver_phone or '',
            'plate_number': v.plate_number,
            'vehicle_type': v.vehicle_type,
            'vehicle_type_display': v.get_vehicle_type_display(),
            'ownership_type': v.ownership_type,
            'ownership_type_display': v.get_ownership_type_display(),
            'default_service_rate': float(v.default_service_rate or 0),
            'bank_name': v.bank_name or 'بانک ملی',
            'account_number': v.account_number or '',
            'sheba_number': v.sheba_number or '',
            'trip_count': trip_count,
            'gross_amount': float(gross_amount),
            'payable_amount': float(payable_amount),
            'is_payable': is_payable,
            'is_active': v.is_active
        })

    # مرتب‌سازی بر اساس نام راننده
    records.sort(key=lambda r: r['driver_name'])

    summary = {
        'year_month': year_month,
        'warehouse_id': warehouse_id,
        'is_locked': is_locked,
        'period_status': period.status if period else 'OPEN',
        'total_vehicles_count': len(records),
        'active_contractor_count': contractor_count,
        'total_fleet_trips': total_trips_sum,
        'total_payable_settlement': float(total_payable_sum)
    }

    return {
        'summary': summary,
        'records': records
    }


def generate_fleet_bank_meli_excel(settlement_data: dict, year_month_title: str) -> io.BytesIO:
    """
    تولید فایل اکسل پرداخت گروهی بانک ملی برای رانندگان ناوگان
    مطابق ساختار رسمی پورتال بانک ملی (۷ ستونه)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fleet_Bank_Payment"
    ws.views.sheetView[0].rightToLeft = True

    headers = [
        "ردیف",
        "مبلغ (ریال)",
        "شماره حساب ملی / شبا مقصد",
        "نام گیرنده (راننده/پیمانکار)",
        "توضیحات",
        "شناسه واریز (اختیاری)",
        "کد ملی راننده"
    ]

    header_font = Font(name="Tahoma", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    border_thin = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    ws.append(headers)
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = border_thin

    records = settlement_data.get('records', [])
    row_idx = 1
    for rec in records:
        # فقط رانندگان دارای مبلغ تسویه مثبت
        if not rec.get('is_payable', True):
            continue
        payable_toman = rec.get('payable_amount', 0)
        if payable_toman <= 0:
            continue

        # تبدیل تومان به ریال در خروجی دیسکت بانکی (هر تومان = ۱۰ ریال)
        payable_rial = int(round(payable_toman * 10))

        acc_or_sheba = (rec.get('sheba_number') or rec.get('account_number') or '').strip()
        driver_name = rec.get('driver_name', '')
        desc = f"تسویه کرایه حمل ناوگان {year_month_title} - {rec.get('plate_number', '')}"
        national_code = rec.get('driver_national_code', '')

        ws.append([
            row_idx,
            payable_rial,
            acc_or_sheba,
            driver_name,
            desc,
            "",
            national_code
        ])

        current_row = row_idx + 1
        for col_num in range(1, 8):
            c = ws.cell(row=current_row, column=col_num)
            c.border = border_thin
            c.font = Font(name="Tahoma", size=9)
            if col_num in (1, 2, 3, 6, 7):
                c.alignment = align_center
                if col_num == 2:
                    c.number_format = '#,##0'
            else:
                c.alignment = align_right

        row_idx += 1

    # تنظیم عرض ستون‌ها
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 28
    ws.column_dimensions['E'].width = 38
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 18

    # فریز ردیف اول
    ws.freeze_panes = 'A2'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
