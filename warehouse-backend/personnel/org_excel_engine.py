"""
موتور استاندارد ورود و خروج اکسل ساختار سازمانی و طرف‌حساب‌های مالی
(Financial Projects, Project Sections & Counterparties Excel Engine)
مطابق با استاندارد Rule 6:
- ساختار ۲ سطری هدر (سطر ۱: عناوین فارسی، سطر ۲: کلیدهای دیتابیسی با فونت طوسی)
- فریز پنل در A3 (freeze_panes = 'A3')
- راست‌به‌چپ (rightToLeft = True)
- تبدیل زمان‌ها به تاریخ و زمان شمسی
"""

import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.db import transaction
import jdatetime

from .models import FinancialProject, ProjectSection, Counterparty
from common.date_utils import normalize_digits


def _format_datetime_shamsi(dt):
    if not dt:
        return '-'
    try:
        jdt = jdatetime.datetime.fromgregorian(datetime=dt)
        return jdt.strftime('%Y/%m/%d %H:%M:%S')
    except Exception:
        return str(dt)


def _apply_2row_header(ws, columns, header_bg='059669'):
    """
    اعمال هدر دو سطری استاندارد شرکت:
    سطر ۱: عناوین فارسی
    سطر ۲: کلیدهای سیستمی دیتابیسی
    """
    f_title = Font(name='B Nazanin', size=11, bold=True, color='FFFFFF')
    f_key = Font(name='Segoe UI', size=9, bold=True, color='CBD5E1')
    fill_title = PatternFill(start_color=header_bg, end_color=header_bg, fill_type='solid')
    fill_key = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')

    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = 'A3'
    ws.sheet_view.rightToLeft = True

    for col_idx, col in enumerate(columns, 1):
        # سطر ۱
        c1 = ws.cell(row=1, column=col_idx, value=col['title'])
        c1.font = f_title
        c1.fill = fill_title
        c1.alignment = align_center
        c1.border = thin_border

        # سطر ۲
        c2 = ws.cell(row=2, column=col_idx, value=col['key'])
        c2.font = f_key
        c2.fill = fill_key
        c2.alignment = align_center
        c2.border = thin_border

        ws.column_dimensions[get_column_letter(col_idx)].width = col.get('width', 18)


# -------------------------------------------------------------
# ۱. پروژه‌های مالی (Financial Projects)
# -------------------------------------------------------------
PROJECT_COLUMNS = [
    {'title': 'کد پروژه', 'key': 'code', 'width': 15},
    {'title': 'نام پروژه مالی / عملیاتی', 'key': 'name', 'width': 28},
    {'title': 'توضیحات و دامنه فعالیت', 'key': 'description', 'width': 35},
    {'title': 'وضعیت فعال (بله/خیر)', 'key': 'is_active', 'width': 16},
    {'title': 'تعداد بخش‌ها', 'key': 'sections_count', 'width': 14},
    {'title': 'تاریخ ایجاد', 'key': 'created_at', 'width': 22},
]

def export_projects_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'پروژه‌های مالی'
    _apply_2row_header(ws, PROJECT_COLUMNS, header_bg='4338CA')  # رنگ ایندیگو

    font_data = Font(name='B Nazanin', size=11)
    align_data = Alignment(horizontal='center', vertical='center')
    align_text = Alignment(horizontal='right', vertical='center')

    for row_idx, p in enumerate(FinancialProject.objects.all().order_by('code'), 3):
        ws.cell(row=row_idx, column=1, value=p.code).alignment = align_data
        ws.cell(row=row_idx, column=2, value=p.name).alignment = align_text
        ws.cell(row=row_idx, column=3, value=p.description or '-').alignment = align_text
        ws.cell(row=row_idx, column=4, value='بله' if p.is_active else 'خیر').alignment = align_data
        ws.cell(row=row_idx, column=5, value=p.sections.count()).alignment = align_data
        ws.cell(row=row_idx, column=6, value=_format_datetime_shamsi(p.created_at)).alignment = align_data

        for c in range(1, 7):
            ws.cell(row=row_idx, column=c).font = font_data

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="financial_projects.xlsx"'
    return response


# -------------------------------------------------------------
# ۲. بخش‌ها و دپارتمان‌ها (Project Sections)
# -------------------------------------------------------------
SECTION_COLUMNS = [
    {'title': 'کد پروژه مادر', 'key': 'project_code', 'width': 16},
    {'title': 'نام پروژه مادر', 'key': 'project_name', 'width': 24},
    {'title': 'کد بخش', 'key': 'code', 'width': 15},
    {'title': 'نام بخش / دپارتمان', 'key': 'name', 'width': 28},
    {'title': 'وضعیت فعال (بله/خیر)', 'key': 'is_active', 'width': 16},
    {'title': 'تاریخ ایجاد', 'key': 'created_at', 'width': 22},
]

def export_sections_excel(project_id=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'بخش‌های پروژه'
    _apply_2row_header(ws, SECTION_COLUMNS, header_bg='0D9488')  # سبز کله‌غازی

    qs = ProjectSection.objects.select_related('project').all()
    if project_id:
        qs = qs.filter(project_id=project_id)
    qs = qs.order_by('project__code', 'code')

    font_data = Font(name='B Nazanin', size=11)
    align_data = Alignment(horizontal='center', vertical='center')
    align_text = Alignment(horizontal='right', vertical='center')

    for row_idx, s in enumerate(qs, 3):
        ws.cell(row=row_idx, column=1, value=s.project.code if s.project else '-').alignment = align_data
        ws.cell(row=row_idx, column=2, value=s.project.name if s.project else '-').alignment = align_text
        ws.cell(row=row_idx, column=3, value=s.code).alignment = align_data
        ws.cell(row=row_idx, column=4, value=s.name).alignment = align_text
        ws.cell(row=row_idx, column=5, value='بله' if s.is_active else 'خیر').alignment = align_data
        ws.cell(row=row_idx, column=6, value=_format_datetime_shamsi(s.created_at)).alignment = align_data

        for c in range(1, 7):
            ws.cell(row=row_idx, column=c).font = font_data

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="project_sections.xlsx"'
    return response


# -------------------------------------------------------------
# ۳. طرف‌حساب‌های مالی (Counterparties)
# -------------------------------------------------------------
COUNTERPARTY_COLUMNS = [
    {'title': 'نام طرف‌حساب', 'key': 'name', 'width': 26},
    {'title': 'نوع طرف‌حساب', 'key': 'counterparty_type', 'width': 18},
    {'title': 'کد ملی / شناسه اقتصادی', 'key': 'national_id', 'width': 20},
    {'title': 'تلفن تماس', 'key': 'phone', 'width': 18},
    {'title': 'نام بانک عامل', 'key': 'bank_name', 'width': 18},
    {'title': 'شماره حساب', 'key': 'account_number', 'width': 20},
    {'title': 'شماره شبا (۲۴ رقمی)', 'key': 'sheba_number', 'width': 28},
    {'title': 'بخش منتسب', 'key': 'section_name', 'width': 22},
    {'title': 'پروژه مادر', 'key': 'project_name', 'width': 22},
    {'title': 'وضعیت فعال', 'key': 'is_active', 'width': 14},
    {'title': 'تاریخ ثبت', 'key': 'created_at', 'width': 22},
]

def export_counterparties_excel(section_id=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'طرف‌حساب‌های مالی'
    _apply_2row_header(ws, COUNTERPARTY_COLUMNS, header_bg='059669')  # سبز زمردی

    qs = Counterparty.objects.select_related('section', 'section__project').all()
    if section_id:
        qs = qs.filter(section_id=section_id)
    qs = qs.order_by('name')

    font_data = Font(name='B Nazanin', size=11)
    align_data = Alignment(horizontal='center', vertical='center')
    align_text = Alignment(horizontal='right', vertical='center')

    type_map = dict(Counterparty.TYPE_CHOICES)

    for row_idx, cp in enumerate(qs, 3):
        sheba_display = f"IR{cp.sheba_number}" if cp.sheba_number else '-'
        ws.cell(row=row_idx, column=1, value=cp.name).alignment = align_text
        ws.cell(row=row_idx, column=2, value=type_map.get(cp.counterparty_type, cp.counterparty_type)).alignment = align_data
        ws.cell(row=row_idx, column=3, value=cp.national_id or '-').alignment = align_data
        ws.cell(row=row_idx, column=4, value=cp.phone or '-').alignment = align_data
        ws.cell(row=row_idx, column=5, value=cp.bank_name or '-').alignment = align_text
        ws.cell(row=row_idx, column=6, value=cp.account_number or '-').alignment = align_data
        ws.cell(row=row_idx, column=7, value=sheba_display).alignment = align_data
        ws.cell(row=row_idx, column=8, value=cp.section.name if cp.section else 'عمومی').alignment = align_text
        ws.cell(row=row_idx, column=9, value=cp.section.project.name if (cp.section and cp.section.project) else 'سراسری').alignment = align_text
        ws.cell(row=row_idx, column=10, value='بله' if cp.is_active else 'خیر').alignment = align_data
        ws.cell(row=row_idx, column=11, value=_format_datetime_shamsi(cp.created_at)).alignment = align_data

        for c in range(1, 12):
            ws.cell(row=row_idx, column=c).font = font_data

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="counterparties.xlsx"'
    return response


# -------------------------------------------------------------
# دانلود قالب‌های آماده برای اکسل (Templates)
# -------------------------------------------------------------
def download_org_template(entity_type='counterparties'):
    wb = openpyxl.Workbook()
    ws = wb.active

    if entity_type == 'projects':
        ws.title = 'قالب پروژه‌ها'
        _apply_2row_header(ws, PROJECT_COLUMNS[:4], header_bg='4338CA')
        # ردیف نمونه
        ws.cell(row=3, column=1, value='PRJ-01')
        ws.cell(row=3, column=2, value='پروژه عملیات مرکزی')
        ws.cell(row=3, column=3, value='مدیریت کل عملیات و انبارها')
        ws.cell(row=3, column=4, value='بله')
        filename = 'projects_template.xlsx'

    elif entity_type == 'sections':
        ws.title = 'قالب بخش‌ها'
        _apply_2row_header(ws, SECTION_COLUMNS[:4], header_bg='0D9488')
        ws.cell(row=3, column=1, value='PRJ-01')
        ws.cell(row=3, column=2, value='پروژه عملیات مرکزی')
        ws.cell(row=3, column=3, value='SEC-101')
        ws.cell(row=3, column=4, value='دپارتمان لجستیک و ترابری')
        filename = 'sections_template.xlsx'

    else:
        ws.title = 'قالب طرف‌حساب‌ها'
        _apply_2row_header(ws, COUNTERPARTY_COLUMNS[:7], header_bg='059669')
        ws.cell(row=3, column=1, value='شرکت حمل‌ونقل پیشتاز')
        ws.cell(row=3, column=2, value='driver')
        ws.cell(row=3, column=3, value='10101234567')
        ws.cell(row=3, column=4, value='09121111111')
        ws.cell(row=3, column=5, value='بانک ملت')
        ws.cell(row=3, column=6, value='0101111111001')
        ws.cell(row=3, column=7, value='IR060120000000000101111111')
        filename = 'counterparties_template.xlsx'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# -------------------------------------------------------------
# پارس و ورود دسته‌جمعی داده‌ها از فایل اکسل (Bulk Import)
# -------------------------------------------------------------
def import_counterparties_from_excel(file_obj):
    """
    خواندن و ثبت دسته‌جمعی طرف‌های حساب از اکسل
    """
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.active

    created = 0
    updated = 0
    errors = []

    # خواندن از ردیف ۳ به بعد (ردیف ۱ و ۲ هدر هستند)
    for row_idx in range(3, ws.max_row + 1):
        name = ws.cell(row=row_idx, column=1).value
        if not name or str(name).strip() == '':
            continue
        name = str(name).strip()

        raw_type = ws.cell(row=row_idx, column=2).value or 'other'
        cp_type = str(raw_type).strip().lower()
        if cp_type not in dict(Counterparty.TYPE_CHOICES):
            # تبدیل عناوین فارسی به کلید سیستمی
            f_map = {
                'راننده': 'driver', 'مالک خودرو': 'driver',
                'تعمیرگاه': 'repair_shop', 'تعمیرگاه و قطعات': 'repair_shop',
                'جایگاه سوخت': 'fuel_station', 'بنزین': 'fuel_station',
                'پیمانکار': 'contractor', 'خدماتی': 'contractor',
                'سایر': 'other'
            }
            cp_type = f_map.get(cp_type, 'other')

        national_id = normalize_digits(str(ws.cell(row=row_idx, column=3).value or '')).strip()
        phone = normalize_digits(str(ws.cell(row=row_idx, column=4).value or '')).strip()
        bank_name = str(ws.cell(row=row_idx, column=5).value or '').strip()
        account_number = normalize_digits(str(ws.cell(row=row_idx, column=6).value or '')).strip()
        raw_sheba = normalize_digits(str(ws.cell(row=row_idx, column=7).value or '')).strip()
        sheba_number = raw_sheba.upper().replace('IR', '').strip()

        try:
            with transaction.atomic():
                cp, is_new = Counterparty.objects.update_or_create(
                    name=name,
                    defaults={
                        'counterparty_type': cp_type,
                        'national_id': national_id or None,
                        'phone': phone or None,
                        'bank_name': bank_name or None,
                        'account_number': account_number or None,
                        'sheba_number': sheba_number or None,
                        'is_active': True
                    }
                )
                if is_new:
                    created += 1
                else:
                    updated += 1
        except Exception as e:
            errors.append(f"ردیف {row_idx} ({name}): {str(e)}")

    return {
        'success': created > 0 or updated > 0 or len(errors) == 0,
        'summary': {
            'total_rows': created + updated + len(errors),
            'created': created,
            'updated': updated,
            'skipped': len(errors)
        },
        'errors': [{'row': idx, 'message': err} for idx, err in enumerate(errors, 1)]
    }
