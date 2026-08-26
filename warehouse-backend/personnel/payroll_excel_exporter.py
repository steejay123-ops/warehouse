# -*- coding: utf-8 -*-
"""
2-Row Header Excel Exporter for Monthly Payroll (conforming to Project Excel Standards and sheet `تیر`).
Row 1: Persian display titles
Row 2: Database field keys in light slate font
Freeze panes at A3
"""
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def generate_monthly_payroll_excel(payroll_records, period_title="تیر 1405") -> bytes:
    """
    Generates 2-Row Header Excel workbook with all 58 columns matching sheet `تیر`.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    safe_title = str(period_title).replace('/', '-').replace(':', '-').replace('\\', '-')
    ws.title = f"حقوق {safe_title}"[:31]
    ws.views.sheetView[0].rightToLeft = True

    # 58 Columns definition: (Persian title, DB key)
    columns = [
        ("گروه", "status_category"),
        ("ردیف", "row_number"),
        ("maliat", "include_in_tax"),
        ("بیمه", "include_in_insurance"),
        ("bank", "include_in_bank"),
        ("کد ملی", "national_code"),
        ("نام و نام خانوادگی", "full_name"),
        ("وضعیت تاهل", "marital_status"),
        ("تعداد فرزند", "children_count"),
        ("کارکرد قرارداد شده", "contract_hours"),
        ("حقوق قرارداد شده", "contract_salary"),
        ("گروه شغلی", "job_grade"),
        ("مزد روزانه", "daily_wage"),
        ("پایه سنواتی", "daily_seniority"),
        ("تعداد سال کارکرد", "years_of_service"),
        ("مزد مبنا", "base_daily_rate"),
        ("ساعت کارکرد", "worked_hours"),
        ("روز بیمه", "insurance_days"),
        ("ساعت اضافه کار", "overtime_hours"),
        ("روز جمعه کاری", "friday_work_days"),
        ("ماموریت", "mission_days"),
        ("مالیات", "income_tax"),
        ("کسر مساعده", "advance_payment_deduction"),
        ("سایر مزایا", "other_allowances"),
        ("جمعه کاری", "friday_work_amount"),
        ("اضافه کار", "overtime_amount"),
        ("هزینه سفر", "travel_cost_amount"),
        ("حق ماموریت", "mission_amount"),
        ("عیدی و پاداش", "bonus_amount"),
        ("مرخصی", "leave_amount"),
        ("بن خانوار", "food_allowance"),
        ("حق مسکن", "housing_allowance"),
        ("حق تاهل", "marital_allowance"),
        ("فوق العاده ایاب و ذهاب", "transport_allowance"),
        ("فوق العاده جذب بازار", "market_attraction_allowance"),
        ("فوق العاده بدی آب و هوا", "bad_weather_allowance"),
        ("فوق العاده دوری از خانواده", "remote_hardship_allowance"),
        ("فوق العاده مناطق پارسیان و عسلویه", "south_pars_allowance"),
        ("حق اولاد", "child_allowance"),
        ("حق سنوات", "seniority_allowance"),
        ("مزایا مشمول", "total_taxable_allowances"),
        ("جمع مزایا مشمول غیر مستمر", "total_non_continuous_taxable_allowances"),
        ("مزایا مستمر مشمول مالیات", "continuous_taxable_allowances"),
        ("جمع مزایا غیر مشمول", "total_non_taxable_allowances"),
        ("جمع پایه سنواتی", "total_seniority_accumulated"),
        ("بیمه سهم کارگر", "worker_insurance"),
        ("بیمه سهم کارفرما", "employer_insurance"),
        ("بیمه بیکاری", "unemployment_insurance"),
        ("جمع بیمه", "total_insurance"),
        ("حقوق پایه", "base_salary"),
        ("حقوق و مزایای مستمر مالیات", "continuous_taxable_salary_allowances"),
        ("جمع حقوق و مزایا مشمول بیمه", "total_insurable_salary_allowances"),
        ("حقوق ناخالص", "gross_salary"),
        ("کسورات", "total_deductions"),
        ("حقوق خالص", "net_salary"),
        ("شماره حساب", "bank_account_number"),
        ("قابل پرداخت", "payable_amount"),
        ("چک مالیات", "tax_check_discrepancy"),
    ]

    # Row 1: Persian display titles
    row1_titles = [c[0] for c in columns]
    # Row 2: Database field keys in light slate
    row2_keys = [c[1] for c in columns]

    ws.append(row1_titles)
    ws.append(row2_keys)

    # Styles
    f_row1 = Font(name="B Nazanin", size=11, bold=True, color="FFFFFF")
    fill_row1 = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Dark Navy
    
    f_row2 = Font(name="Consolas", size=9, bold=False, color="64748B") # Light Slate
    fill_row2 = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center")
    border_thin = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # Apply Header Styles
    for col_idx in range(1, len(columns) + 1):
        c1 = ws.cell(row=1, column=col_idx)
        c1.font = f_row1
        c1.fill = fill_row1
        c1.alignment = align_center
        c1.border = border_thin

        c2 = ws.cell(row=2, column=col_idx)
        c2.font = f_row2
        c2.fill = fill_row2
        c2.alignment = align_center
        c2.border = border_thin

    # Append Data Rows
    for r_idx, rec in enumerate(payroll_records, start=3):
        row_vals = []
        for col_name, col_key in columns:
            val = getattr(rec, col_key, '')
            if isinstance(val, bool):
                val = 1 if val else 0
            row_vals.append(val)
        ws.append(row_vals)

        # Style data cells
        for col_idx in range(1, len(columns) + 1):
            cell = ws.cell(row=r_idx, column=col_idx)
            cell.border = border_thin
            cell.font = Font(name="B Nazanin", size=10)
            col_key = columns[col_idx - 1][1]
            if col_key in ('row_number', 'national_code', 'job_grade', 'bank_account_number', 'marital_status'):
                cell.alignment = align_center
            elif isinstance(cell.value, (int, float)) or (hasattr(cell.value, 'as_integer_ratio')):
                cell.alignment = align_center
                if col_key not in ('row_number', 'children_count', 'years_of_service', 'insurance_days', 'include_in_tax', 'include_in_insurance', 'include_in_bank'):
                    cell.number_format = '#,##0'
            else:
                cell.alignment = align_right

    # Freeze Panes at A3
    ws.freeze_panes = 'A3'

    # Auto-adjust column widths
    for col_idx in range(1, len(columns) + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(14, len(columns[col_idx-1][0]) * 2 + 2)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
