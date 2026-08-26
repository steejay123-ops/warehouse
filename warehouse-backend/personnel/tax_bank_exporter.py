# -*- coding: utf-8 -*-
"""
Tax and Bank Exporter Module for Personnel Payroll.
Generates Tax WH (Monthly Salary) & WP (Personnel Info) text files, and Bank Meli group payment text/Excel files.
"""
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from .models import MonthlyPayrollRecord, TaxRuleSettings, BankExportSettings


def _clean_date_digits(date_str: str) -> str:
    """Helper to convert '1365/03/08' or '1365-03-08' to '13650308'."""
    if not date_str:
        return ''
    return str(date_str).replace('/', '').replace('-', '').strip()


def generate_wp_tax_content(payroll_records, tax_settings=None) -> str:
    """
    Generates WP (Personnel Info) monthly tax text file content (UTF-8 with BOM).
    Conforms to TaxFileHelp.pdf version 1.7.0.4 (23 comma-delimited fields).
    """
    lines = []
    taxable_records = [r for r in payroll_records if getattr(r, 'include_in_tax', True)]

    for rec in taxable_records:
        p = rec.personnel
        fields = [
            str(getattr(p, 'nationality_code', '1') or '1').strip(),                      # 1. ملیت (1=ایرانی / 2=خارجی)
            str(p.national_code or '').strip().zfill(10),                                 # 2. کد ملی (10 رقم)
            str(p.first_name or '').strip(),                                              # 3. نام
            str(p.last_name or '').strip(),                                               # 4. نام خانوادگی
            str(p.father_name or '').strip(),                                             # 5. نام پدر
            _clean_date_digits(p.birth_date or ''),                                       # 6. تاریخ تولد (YYYYMMDD)
            str(p.id_number or p.national_code or '').strip(),                            # 7. شماره شناسنامه
            str(p.birth_place or p.issue_place or '').strip(),                            # 8. محل تولد
            str(getattr(p, 'education_level', '5') or '5').strip(),                       # 9. مدرک تحصیلی (1 تا 7)
            str(getattr(p, 'insurance_type', '2') or '2').strip(),                        # 10. نوع بیمه (1 تا 5)
            str(p.insurance_number or '').strip(),                                        # 11. شماره بیمه (10 رقمی)
            str(getattr(p, 'insurance_name', 'تامین اجتماعی') or 'تامین اجتماعی').strip(), # 12. نام بیمه
            str(getattr(p, 'exemption_type', '1') or '1').strip(),                        # 13. نوع معافیت (1 تا 11 / 10+11)
            str(getattr(p, 'citizenship_country_code', '103') or '103').strip(),          # 14. کشور محل تابعیت (103=ایران)
            str(getattr(p, 'residence_country_code', '103') or '103').strip(),            # 15. کشور محل زندگی (103=ایران)
            '1' if p.postal_code else '1',                                                # 16. کد پستی
            '1' if p.address else '1',                                                    # 17. آدرس
            str(getattr(p, 'job_category', '5') or '5').strip(),                          # 18. رسته (5=انبارداری / 9=ترابری / 15=فنی / 2=اداری / 17=کارگری)
            str(p.job_title or '').strip(),                                               # 19. سمت شغلی
            str(getattr(p, 'employment_type', '2') or '2').strip(),                       # 20. نوع استخدام (1 تا 12)
            _clean_date_digits(p.start_date or '14040901'),                               # 21. تاریخ شروع به کار (YYYYMMDD)
            _clean_date_digits(p.end_date or ''),                                         # 22. تاریخ پایان کار (در صورت تسویه/خروج)
            _clean_date_digits(p.retirement_date or ''),                                  # 23. تاریخ بازنشستگی
        ]
        lines.append(",".join(fields))

    return "\ufeff" + "\r\n".join(lines)


def generate_wh_tax_content(payroll_records, tax_settings=None) -> str:
    """
    Generates WH (Monthly Salary Summary) tax text file content (UTF-8 with BOM).
    Conforms to TaxFileHelp.pdf version 1.7.0.4 (39 comma-delimited fields).
    Uses individual personnel tax settings with workshop defaults fallback.
    """
    lines = []
    taxable_records = [r for r in payroll_records if getattr(r, 'include_in_tax', True)]

    # Workshop Global Defaults
    d_pay_type = getattr(tax_settings, 'payment_type', '1') or '1'
    d_srv_loc = getattr(tax_settings, 'service_location', '1') or '1'
    d_exc = getattr(tax_settings, 'exceptions', '1') or '1'
    d_curr_type = getattr(tax_settings, 'currency_type', '84') or '84'
    d_curr_rate = int(getattr(tax_settings, 'currency_exchange_rate', 1) or 1)
    d_house_type = getattr(tax_settings, 'housing_benefit_type', '1') or '1'
    d_veh_type = getattr(tax_settings, 'vehicle_benefit_type', '1') or '1'

    for rec in taxable_records:
        p = rec.personnel
        
        # Individual Settings with Workshop Fallback (never empty or undefined)
        pay_type = getattr(p, 'tax_payment_type', None) or d_pay_type
        srv_loc = getattr(p, 'tax_service_location', None) or d_srv_loc
        exc = getattr(p, 'tax_exceptions', None) or d_exc
        curr_type = getattr(p, 'tax_currency_type', None) or d_curr_type
        curr_rate = int(getattr(p, 'tax_currency_exchange_rate', None) or d_curr_rate)
        house_type = getattr(p, 'tax_housing_benefit_type', None) or d_house_type
        veh_type = getattr(p, 'tax_vehicle_benefit_type', None) or d_veh_type

        fields = [
            str(p.national_code or '').strip().zfill(10), # 1. کد ملی (10 رقم)
            str(pay_type).strip(),                        # 2. نوع پرداختی (جدول 8)
            str(srv_loc).strip(),                         # 3. وضعیت محل خدمت (جدول 9)
            str(exc).strip(),                             # 4. استثنائات بودجه (جدول 10)
            str(curr_type).strip(),                       # 5. نوع ارز (جدول 13: 84=ریال)
            str(curr_rate).strip(),                       # 6. نرخ تسعیر ارز (1)
            str(int(round(rec.continuous_taxable_salary_allowances or 0))), # 7. جمع ناخالص مستمر نقدی ماه جاری
            '0',                                          # 8. مستمر نقدی معوق
            str(house_type).strip(),                      # 9. مسکن (جدول 11: 1=عدم استفاده / 2=با اثاثیه / 3=بدون اثاثیه)
            '0',                                          # 10. کسر مسکن
            str(veh_type).strip(),                        # 11. اتومبیل اختصاصی (جدول 12: 1=عدم استفاده / 2=با راننده / 3=بدون راننده)
            '0',                                          # 12. کسر اتومبیل
            '0',                                          # 13. قیمت تمام شده مستمر غیرنقدی
            '0',                                          # 14. مستمر غیرنقدی معوق
            '0',                                          # 15. حق‌الزحمه / مشاوره / داوری
            '0',                                          # 16. قراردادهای پژوهشی
            str(int(round(rec.overtime_amount or 0))),    # 17. اضافه کاری ریالی
            str(int(round(rec.travel_cost_amount or 0))), # 18. هزینه سفر ریالی
            str(int(round(rec.mission_amount or 0))),     # 19. فوق العاده مسافرت (ماموریت)
            '0',                                          # 20. کارانه
            '0',                                          # 21. پاداش (غیر از سال و پایان خدمت)
            '0',                                          # 22. پاداش آخر سال
            str(int(round(rec.bonus_amount or 0))),       # 23. عیدی سالانه ریالی
            '0',                                          # 24. پاداش پایان خدمت
            '0',                                          # 25. خسارت اخراج
            '0',                                          # 26. بازخرید خدمت
            str(int(round(rec.seniority_allowance or 0))),# 27. حق سنوات ریالی
            str(int(round(rec.leave_amount or 0))),       # 28. حقوق ایام مرخصی استفاده نشده
            str(int(round(rec.other_allowances or 0))),   # 29. سایر حقوق و مزایای غیرمستمر نقدی
            '0',                                          # 30. غیرمستمر نقدی معوق
            '0',                                          # 31. غیرمستمر غیرنقدی
            '0',                                          # 32. غیرمستمر غیرنقدی معوق
            str(int(round(rec.worker_insurance or 0))),   # 33. بیمه سهم کارگر
            '0',                                          # 34. بیمه های عمر و زندگی موضوع ماده 137
            '0',                                          # 35. خالص پرداختی به حقوق‌بگیر
            '0',                                          # 36. حق التدریس / حق پژوهش
            '0',                                          # 37. حق کشیک
            '0',                                          # 38. رفاهی و انگیزشی و بهره‌وری
            '0',                                          # 39. حق‌السعی (به استثنای مزد و پاداش)
        ]
        lines.append(",".join(fields))

    return "\ufeff" + "\r\n".join(lines)


def generate_bank_meli_excel(payroll_records, bank_settings, year_month_title="تیر 1405") -> bytes:
    """
    Generates Bank Meli Group Payment Excel File (.xlsx) matching BankTable and BankMeli macro structure.
    Headers: ردیف | مبلغ (ریال) | شماره حساب ملی / شبا مقصد | نام گیرنده | توضیحات | شناسه واریز | کد ملی
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bank_Payment"
    ws.views.sheetView[0].rightToLeft = True

    headers = [
        "ردیف",
        "مبلغ (ریال)",
        "شماره حساب ملی / شبا مقصد",
        "نام گیرنده",
        "توضیحات",
        "شناسه واریز (اختیاری - حداکثر 35 کاراکتر)",
        "کد ملی (اختیاری)"
    ]

    header_font = Font(name="B Nazanin", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    border_thin = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    ws.append(headers)
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = border_thin

    deposit_desc_tpl = bank_settings.deposit_description_template if bank_settings else "انبارداری حقوق {month_name}"
    deposit_id_default = bank_settings.default_deposit_id if bank_settings else ""

    bank_records = [r for r in payroll_records if getattr(r, 'include_in_bank', True)]
    row_idx = 1
    for rec in bank_records:
        p = rec.personnel
        payable = int(round(rec.payable_amount or 0))
        if payable <= 0:
            continue
        
        acc_no = (p.account_number or p.sheba_number or '').strip()
        desc = deposit_desc_tpl.format(month_name=year_month_title, fiscal_year="")
        
        ws.append([
            row_idx,
            payable,
            acc_no,
            p.full_name,
            desc,
            deposit_id_default,
            p.national_code
        ])
        
        # Style row
        current_row = row_idx + 1
        for col_num in range(1, 8):
            c = ws.cell(row=current_row, column=col_num)
            c.border = border_thin
            c.font = Font(name="B Nazanin", size=10)
            if col_num in (1, 2, 3, 6, 7):
                c.alignment = align_center
                if col_num == 2:
                    c.number_format = '#,##0'
            else:
                c.alignment = align_right

        row_idx += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 24
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
