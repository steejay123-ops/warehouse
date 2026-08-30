from django.test import TestCase
from django.contrib.auth import get_user_model
from rest_framework.test import APIClient
from rest_framework import status
from personnel.models import (
    PersonnelProfile,
    DailyAttendance,
    MonthlyWorkPeriod,
    VehicleDriverProfile,
    VehicleTripLog,
)
from warehouses.models import Warehouse

User = get_user_model()

class AttendanceAndFleetAuditTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.user = User.objects.create_superuser(
            username='admin_audit',
            password='Password123!',
            email='audit@example.com'
        )
        self.client.force_authenticate(user=self.user)

        self.warehouse1 = Warehouse.objects.create(name='انبار مرکزی', is_active=True)
        self.warehouse2 = Warehouse.objects.create(name='انبار غرب', is_active=True)

        self.personnel1 = PersonnelProfile.objects.create(
            first_name='علی',
            last_name='رضایی',
            national_code='0012345678',
            assigned_warehouse=self.warehouse1,
            daily_base_wage=5000000,
            is_active=True
        )

        self.personnel_floating = PersonnelProfile.objects.create(
            first_name='محمد',
            last_name='شناور',
            national_code='0098765432',
            assigned_warehouse=None,
            daily_base_wage=5000000,
            is_active=True
        )

        self.vehicle1 = VehicleDriverProfile.objects.create(
            driver_name='حسن راننده',
            plate_number='12ع345-67',
            assigned_warehouse=self.warehouse1,
            default_service_rate=1500000,
            is_active=True
        )

    def test_01_bulk_save_preserves_assigned_warehouse_when_floating_or_null(self):
        """
        تست اصلاح انتساب خودکار انبار در ثبت سراسری بدون نال شدن warehouse_id
        """
        payload = {
            'warehouse_id': None,
            'date_shamsi': '1405/06/08',
            'is_override': True,
            'items': [
                {
                    'personnel_id': self.personnel1.id,
                    'warehouse_id': None,
                    'status': 'PRESENT_10H',
                    'effective_hours': 10,
                    'overtime_hours': 2,
                    'is_friday_work': False,
                    'is_mission': False,
                    'advance_payment': 0,
                    'notes': 'تست انتساب انبار'
                }
            ]
        }
        res = self.client.post('/api/personnel/attendance/bulk-save/', payload, format='json')
        self.assertEqual(res.status_code, status.HTTP_200_OK)

        att = DailyAttendance.objects.get(personnel=self.personnel1, date_shamsi='1405/06/08')
        self.assertEqual(att.warehouse_id, self.warehouse1.id)

    def test_02_absent_and_leave_hours_are_sanitized_to_zero(self):
        """
        تست ریست و پاکسازی خودکار ساعات موثر و اضافه‌کار در وضعیت غیبت و مرخصی
        """
        payload = {
            'warehouse_id': self.warehouse1.id,
            'date_shamsi': '1405/06/09',
            'is_override': True,
            'items': [
                {
                    'personnel_id': self.personnel1.id,
                    'warehouse_id': self.warehouse1.id,
                    'status': 'ABSENT',
                    'effective_hours': 10,  # ارسال اشتباه ساعت برای غایب
                    'overtime_hours': 4,    # ارسال اشتباه اضافه‌کار برای غایب
                    'is_friday_work': True,
                    'is_mission': True,
                    'advance_payment': 0,
                    'notes': 'غایب غیرموجه'
                }
            ]
        }
        res = self.client.post('/api/personnel/attendance/bulk-save/', payload, format='json')
        self.assertEqual(res.status_code, status.HTTP_200_OK)

        att = DailyAttendance.objects.get(personnel=self.personnel1, date_shamsi='1405/06/09')
        self.assertEqual(att.status, 'ABSENT')
        self.assertEqual(float(att.effective_hours), 0.0)
        self.assertEqual(float(att.overtime_hours), 0.0)
        self.assertFalse(att.is_friday_work)
        self.assertFalse(att.is_mission)

    def test_03_get_matrix_does_not_propagate_friday_work_to_non_fridays(self):
        """
        تست قطع سرایت پیش‌فرض جمعه‌کاری روز قبل به روزهای عادی بعد در متد get_matrix
        """
        DailyAttendance.objects.create(
            personnel=self.personnel1,
            warehouse=self.warehouse1,
            date_shamsi='1405/06/06',
            status='FRIDAY_WORK',
            effective_hours=10,
            overtime_hours=0,
            is_friday_work=True,
            created_by=self.user
        )

        # فراخوانی ماتریس برای روز بعد 1405/06/07 (شنبه - غیر جمعه)
        res = self.client.get('/api/personnel/attendance/matrix/?date_shamsi=1405/06/07')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        rows = res.data['rows']
        p1_row = next(r for r in rows if r['personnel_id'] == self.personnel1.id)

        self.assertNotEqual(p1_row['status'], 'FRIDAY_WORK')
        self.assertFalse(p1_row['is_friday_work'])

    def test_04_vehicle_bulk_save_supports_nullable_warehouse_id(self):
        """
        تست پشتیبانی از ذخیره سراسری و شناور تردد ناوگان با warehouse_id اختیاری
        """
        payload = {
            'warehouse_id': None,
            'date_shamsi': '1405/06/08',
            'items': [
                {
                    'vehicle_id': self.vehicle1.id,
                    'trip_count': 3,
                    'unit_rate': 1500000,
                    'dispatch_reference': 'DSP-9988',
                    'origin_destination': 'انبار به کارخانه',
                    'notes': 'تست تردد بدون انبار سراسری'
                }
            ]
        }
        res = self.client.post('/api/personnel/trips/bulk-save/', payload, format='json')
        self.assertEqual(res.status_code, status.HTTP_200_OK)

        trip = VehicleTripLog.objects.get(vehicle=self.vehicle1, date_shamsi='1405/06/08')
        self.assertEqual(trip.trip_count, 3)
        self.assertEqual(trip.warehouse_id, self.warehouse1.id)
        self.assertEqual(trip.total_amount, 4500000)

    def test_05_monthly_grid_bulk_save_sanitizes_hours_and_preserves_warehouse(self):
        """
        تست ذخیره ماتریس ۳۱ روزه و پاکسازی ساعات در وضعیت غیبت/مرخصی
        """
        payload = {
            'warehouse_id': None,
            'year_month': '1405/06',
            'is_override': True,
            'items': [
                {
                    'personnel_id': self.personnel1.id,
                    'day': 10,
                    'status': 'LEAVE',
                    'effective_hours': 10,
                    'overtime_hours': 3,
                    'is_friday_work': False,
                    'is_mission': False,
                    'advance_payment': 0,
                    'notes': 'مرخصی استحقاقی'
                }
            ]
        }
        res = self.client.post('/api/personnel/attendance/bulk-save-monthly-grid/', payload, format='json')
        self.assertEqual(res.status_code, status.HTTP_200_OK)

        att = DailyAttendance.objects.get(personnel=self.personnel1, date_shamsi='1405/06/10')
        self.assertEqual(att.status, 'LEAVE')
        self.assertEqual(float(att.effective_hours), 0.0)
        self.assertEqual(float(att.overtime_hours), 0.0)
        self.assertEqual(att.warehouse_id, self.warehouse1.id)

    def test_06_vehicle_monthly_grid_retrieval_and_structure(self):
        """
        تست واکشی ماتریس ۳۱ روزه ناوگان و ساختار ستون‌های روزانه
        """
        VehicleTripLog.objects.create(
            vehicle=self.vehicle1,
            warehouse=self.warehouse1,
            date_shamsi='1405/06/05',
            trip_count=2,
            unit_rate=1500000,
            total_amount=3000000,
            dispatch_reference='BL-101',
            origin_destination='شیراز به اصفهان',
            created_by=self.user
        )

        res = self.client.get(f'/api/personnel/trips/monthly-grid/?warehouse_id={self.warehouse1.id}&year_month=1405/06')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res.data['days_in_month'], 31)
        self.assertEqual(len(res.data['days_meta']), 31)

        rows = res.data['rows']
        self.assertTrue(len(rows) >= 1)
        v1_row = next(r for r in rows if r['vehicle_id'] == self.vehicle1.id)
        self.assertEqual(v1_row['total_trips'], 2)
        self.assertEqual(v1_row['total_amount'], 3000000)

        # بررسی سلول روز ۵
        day5 = next(d for d in v1_row['days'] if d['day'] == 5)
        self.assertEqual(day5['trip_count'], 2)
        self.assertEqual(day5['dispatch_reference'], 'BL-101')

    def test_07_vehicle_update_day_trip(self):
        """
        تست اکشن ویرایش سریع روزانه از روی تقویم ماهانه ناوگان
        """
        payload = {
            'vehicle_id': self.vehicle1.id,
            'warehouse_id': self.warehouse1.id,
            'date_shamsi': '1405/06/15',
            'trip_count': 4,
            'unit_rate': 2000000,
            'dispatch_reference': 'BL-202',
            'origin_destination': 'مرکز به شعبه ۲',
            'notes': 'سرویس ویژه'
        }
        res = self.client.post('/api/personnel/trips/update-day-trip/', payload, format='json')
        self.assertEqual(res.status_code, status.HTTP_200_OK)

        trip = VehicleTripLog.objects.get(vehicle=self.vehicle1, date_shamsi='1405/06/15')
        self.assertEqual(trip.trip_count, 4)
        self.assertEqual(trip.unit_rate, 2000000)
        self.assertEqual(trip.total_amount, 8000000)
        self.assertEqual(trip.dispatch_reference, 'BL-202')

    def test_08_vehicle_bulk_save_monthly_grid(self):
        """
        تست ثبت دسته‌جمعی ماتریس ۳۱ روزه ناوگان
        """
        payload = {
            'warehouse_id': self.warehouse1.id,
            'year_month': '1405/06',
            'items': [
                {
                    'vehicle_id': self.vehicle1.id,
                    'day': 20,
                    'trip_count': 3,
                    'unit_rate': 1500000,
                    'dispatch_reference': 'BL-303',
                    'origin_destination': 'مسیر تست',
                    'notes': 'تست بالک'
                }
            ]
        }
        res = self.client.post('/api/personnel/trips/bulk-save-monthly-grid/', payload, format='json')
        self.assertEqual(res.status_code, status.HTTP_200_OK)

        trip = VehicleTripLog.objects.get(vehicle=self.vehicle1, date_shamsi='1405/06/20')
        self.assertEqual(trip.trip_count, 3)
        self.assertEqual(trip.total_amount, 4500000)

    def test_09_rbac_permissions_distinct_for_four_pillars(self):
        """
        تست تفکیک کامل ۴ مجوز مستقل RBAC برای پورتال‌های ۴ گانه
        """
        from django.contrib.auth.models import Permission
        perm_codenames = [
            'view_sys_personnel_attendance',
            'view_sys_fleet_attendance',
            'view_sys_payroll',
            'view_sys_fleet_settlement'
        ]
        for code in perm_codenames:
            perm = Permission.objects.filter(codename=code).first()
            self.assertIsNotNone(perm, f"Permission {code} should exist in auth_permission table")

    def test_10_export_fleet_monthly_excel_structure_and_freeze(self):
        """
        تست تولید اکسل ۳۱ روزه ناوگان، فریز A3، و هدر ۲ سطری
        """
        from personnel.fleet_excel_engine import export_fleet_monthly_excel
        import openpyxl

        VehicleTripLog.objects.create(
            vehicle=self.vehicle1,
            warehouse=self.warehouse1,
            date_shamsi='1405/06/07',
            trip_count=5,
            unit_rate=1500000,
            total_amount=7500000,
            created_by=self.user
        )

        excel_buf = export_fleet_monthly_excel(warehouse_id=self.warehouse1.id, year_month='1405/06')
        self.assertIsNotNone(excel_buf)

        wb = openpyxl.load_workbook(excel_buf, data_only=True)
        ws = wb.active

        # بررسی ساختار ۲ سطری
        self.assertEqual(ws.cell(row=1, column=2).value, 'نام راننده')
        self.assertEqual(ws.cell(row=2, column=2).value, 'driver_name')
        self.assertEqual(ws.freeze_panes, 'A3')

        # بررسی سطر داده
        self.assertEqual(ws.cell(row=3, column=2).value, self.vehicle1.driver_name)
        # روز 7 (ستون 6 + 7 = 13)
        self.assertEqual(ws.cell(row=3, column=13).value, 5)

    def test_11_import_fleet_monthly_excel_roundtrip(self):
        """
        تست ایمپورت فایل اکسل تولید شده و ذخیره ترددها در دیتابیس
        """
        from personnel.fleet_excel_engine import export_fleet_monthly_excel, import_fleet_monthly_excel
        import io

        # ۱. ابتدا اکسل را خروجی می‌گیریم
        excel_buf = export_fleet_monthly_excel(warehouse_id=self.warehouse1.id, year_month='1405/06')

        # ۲. مقدار روز ۱۰ را در اکسل دستکاری می‌کنیم
        import openpyxl
        wb = openpyxl.load_workbook(excel_buf)
        ws = wb.active
        # ستون روز ۱۰ = ستون ۱۶ (6 + 10)
        ws.cell(row=3, column=16).value = 6
        modified_buf = io.BytesIO()
        wb.save(modified_buf)
        modified_buf.seek(0)

        # ۳. فایل دستکاری شده را درون‌ریزی می‌کنیم
        res = import_fleet_monthly_excel(
            file_obj=modified_buf,
            warehouse_id=self.warehouse1.id,
            year_month='1405/06',
            user=self.user
        )
        self.assertTrue(res['updated_count'] >= 1)

        # ۴. بررسی ثبت در دیتابیس
        trip = VehicleTripLog.objects.get(vehicle=self.vehicle1, date_shamsi='1405/06/10')
        self.assertEqual(trip.trip_count, 6)
        self.assertEqual(trip.total_amount, 6 * self.vehicle1.default_service_rate)

    def test_12_import_fleet_monthly_excel_rejects_locked_period(self):
        """
        تست عدم اجازه ورود اطلاعات اکسل در دوره‌های قفل‌شده
        """
        from personnel.fleet_excel_engine import import_fleet_monthly_excel
        import io

        MonthlyWorkPeriod.objects.create(
            warehouse=self.warehouse1,
            year_month='1405/06',
            status='LOCKED',
            locked_by=self.user
        )

        dummy_buf = io.BytesIO()
        with self.assertRaises(ValueError) as ctx:
            import_fleet_monthly_excel(
                file_obj=dummy_buf,
                warehouse_id=self.warehouse1.id,
                year_month='1405/06',
                user=self.user
            )
        self.assertIn('قفل شده است', str(ctx.exception))

    def test_13_vehicle_trip_audit_log_created_on_update(self):
        """
        تست ثبت خودکار رکورد در VehicleTripAuditLog هنگام ویرایش تردد ناوگان
        """
        from personnel.models import VehicleTripAuditLog

        # ثبت اولیه
        payload_create = {
            'vehicle_id': self.vehicle1.id,
            'warehouse_id': self.warehouse1.id,
            'date_shamsi': '1405/06/18',
            'trip_count': 2,
            'unit_rate': 1500000,
            'notes': 'ثبت اول'
        }
        res1 = self.client.post('/api/personnel/trips/update-day-trip/', payload_create, format='json')
        self.assertEqual(res1.status_code, status.HTTP_200_OK)

        log1 = VehicleTripAuditLog.objects.filter(trip__vehicle=self.vehicle1, date_shamsi='1405/06/18', field_name='create').first()
        self.assertIsNotNone(log1)
        self.assertEqual(log1.new_value, '2')

        # ویرایش تعداد سرویس
        payload_update = {
            'vehicle_id': self.vehicle1.id,
            'warehouse_id': self.warehouse1.id,
            'date_shamsi': '1405/06/18',
            'trip_count': 5,
            'unit_rate': 1500000,
            'notes': 'افزایش سرویس'
        }
        res2 = self.client.post('/api/personnel/trips/update-day-trip/', payload_update, format='json')
        self.assertEqual(res2.status_code, status.HTTP_200_OK)

        log2 = VehicleTripAuditLog.objects.filter(trip__vehicle=self.vehicle1, date_shamsi='1405/06/18', field_name='trip_count').first()
        self.assertIsNotNone(log2)
        self.assertEqual(log2.old_value, '2')
        self.assertEqual(log2.new_value, '5')

    def test_14_vehicle_driver_profile_crud_and_audit_api(self):
        """
        تست اندپوینت‌های CRUD پرونده خودرو و واکشی لاگ‌های ممیزی
        """
        # ایجاد خودرو جدید
        payload = {
            'driver_name': 'رضا راننده آزمایشی',
            'driver_national_code': '1234567890',
            'driver_phone': '09123456789',
            'plate_number': '99ط111-22',
            'vehicle_type': 'khavar',
            'ownership_type': 'contract',
            'default_service_rate': 2500000,
            'sheba_number': 'IR050170000000123456789012',
            'is_active': True
        }
        res = self.client.post('/api/personnel/vehicles/', payload, format='json')
        self.assertEqual(res.status_code, status.HTTP_201_CREATED)
        new_v_id = res.data['id']

        # دریافت لیست audit-logs
        res_audit = self.client.get(f'/api/personnel/trips/audit-logs/?vehicle_id={self.vehicle1.id}')
        self.assertEqual(res_audit.status_code, status.HTTP_200_OK)
        self.assertIsInstance(res_audit.data, list)

    def test_15_calculate_monthly_fleet_settlement_engine(self):
        """
        تست موتور محاسبه تسویه ناوگان (فاز ۴)
        تفکیک خودروهای استیجاری/ملکی (دارای تسویه) از خودروهای شرکتی (تسویه ۰)
        """
        from personnel.fleet_settlement_engine import calculate_monthly_fleet_settlement
        from personnel.models import VehicleTripLog

        # ثبت تردد برای خودرو ۱ (استیجاری) و خودرو ۲ (شرکتی)
        vehicle_company = VehicleDriverProfile.objects.create(
            driver_name='راننده شرکتی',
            plate_number='88ب999-11',
            vehicle_type='nissan',
            ownership_type='company',
            assigned_warehouse=self.warehouse1,
            is_active=True
        )

        VehicleTripLog.objects.create(
            vehicle=self.vehicle1,
            warehouse=self.warehouse1,
            date_shamsi='1405/06/10',
            trip_count=3,
            unit_rate=1500000,
            created_by=self.user
        )
        VehicleTripLog.objects.create(
            vehicle=vehicle_company,
            warehouse=self.warehouse1,
            date_shamsi='1405/06/10',
            trip_count=4,
            unit_rate=0,
            created_by=self.user
        )

        res = calculate_monthly_fleet_settlement(warehouse_id=self.warehouse1.id, year_month='1405/06')
        summary = res['summary']
        records = res['records']

        self.assertEqual(summary['total_fleet_trips'], 7)
        # خودرو ۱ استیجاری: 3 * 1,500,000 = 4,500,000
        self.assertEqual(summary['total_payable_settlement'], 4500000.0)

        rec1 = next(r for r in records if r['vehicle_id'] == self.vehicle1.id)
        rec2 = next(r for r in records if r['vehicle_id'] == vehicle_company.id)

        self.assertEqual(rec1['payable_amount'], 4500000.0)
        self.assertTrue(rec1['is_payable'])

        # خودرو ۲ شرکتی است -> بدون مبلغ تسویه
        self.assertEqual(rec2['payable_amount'], 0.0)
        self.assertFalse(rec2['is_payable'])

    def test_16_generate_fleet_bank_meli_excel(self):
        """
        تست تولید فایل اکسل پرداخت گروهی بانک ملی برای ناوگان
        """
        from personnel.fleet_settlement_engine import calculate_monthly_fleet_settlement, generate_fleet_bank_meli_excel
        import openpyxl

        data = calculate_monthly_fleet_settlement(warehouse_id=self.warehouse1.id, year_month='1405/06')
        excel_buf = generate_fleet_bank_meli_excel(data, '1405/06')
        self.assertIsNotNone(excel_buf)

        wb = openpyxl.load_workbook(excel_buf)
        ws = wb.active
        self.assertEqual(ws.title, "Fleet_Bank_Payment")
        self.assertEqual(ws.cell(row=1, column=1).value, "ردیف")
        self.assertEqual(ws.cell(row=1, column=2).value, "مبلغ (ریال)")
        self.assertEqual(ws.cell(row=1, column=3).value, "شماره حساب ملی / شبا مقصد")

    def test_17_fleet_settlement_api_and_bank_excel_download(self):
        """
        تست اندپوینت‌های API تسویه ناوگان و دانلود اکسل بانک ملی
        """
        # محاسبه تسویه از طریق API
        res_calc = self.client.get(f'/api/personnel/fleet-settlement/calculate/?year_month=1405/06&warehouse_id={self.warehouse1.id}')
        self.assertEqual(res_calc.status_code, status.HTTP_200_OK)
        self.assertIn('summary', res_calc.data)
        self.assertIn('records', res_calc.data)

        # دانلود اکسل بانک ملی
        res_excel = self.client.get(f'/api/personnel/fleet-settlement/export-bank-excel/?year_month=1405/06&warehouse_id={self.warehouse1.id}')
        self.assertEqual(res_excel.status_code, status.HTTP_200_OK)
        self.assertEqual(res_excel['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')



