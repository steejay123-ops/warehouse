import io
import openpyxl
from django.test import TestCase
from django.contrib.auth import get_user_model
from rest_framework.test import APIClient
from rest_framework import status
from personnel.models import (
    MonthlyWorkPeriod,
    VehicleDriverProfile,
    VehicleTripLog,
    VehicleTripAuditLog,
)
from warehouses.models import Warehouse
from personnel.fleet_settlement_engine import (
    calculate_monthly_fleet_settlement,
    generate_fleet_bank_meli_excel,
)
from personnel.fleet_excel_engine import (
    export_fleet_monthly_excel,
    import_fleet_monthly_excel,
)

User = get_user_model()


class FleetComprehensiveTenVehiclesTests(TestCase):
    """
    مجموعه تست‌های جامع و پیشرفته کارکرد ناوگان و ماشین‌آلات
    همراه با دیتای ۱۰ خودرو در سناریوهای متنوع (استیجاری، ملکی، شرکتی، تسویه و اکسل)
    """

    def setUp(self):
        self.client = APIClient()
        self.user = User.objects.create_superuser(
            username='fleet_admin',
            password='Password123!',
            email='fleet_admin@example.com'
        )
        self.client.force_authenticate(user=self.user)

        # ایجاد ۲ انبار
        self.wh1 = Warehouse.objects.create(name='انبار مرکزی تهران', is_active=True)
        self.wh2 = Warehouse.objects.create(name='انبار ترانزیت بندر', is_active=True)

        # ایجاد ۱۰ خودرو/راننده با مشخصات واقعی و سناریوهای مختلف
        self.vehicles_data = [
            # ۱. استیجاری - نیسان
            {
                'driver_name': 'رضا صادقی',
                'plate_number': '11ع111-11',
                'driver_phone': '09121111111',
                'driver_national_code': '0011111111',
                'bank_name': 'ملی',
                'account_number': '0101111111001',
                'sheba_number': 'IR110170000000111111111001',
                'vehicle_type': 'nissan',
                'ownership_type': 'contract',
                'assigned_warehouse': self.wh1,
                'default_service_rate': 1200000,
                'is_active': True
            },
            # ۲. استیجاری - خاور
            {
                'driver_name': 'محمود کریمی',
                'plate_number': '22ع222-22',
                'driver_phone': '09122222222',
                'driver_national_code': '0022222222',
                'bank_name': 'ملی',
                'account_number': '0102222222002',
                'sheba_number': 'IR220170000000222222222002',
                'vehicle_type': 'khavar',
                'ownership_type': 'contract',
                'assigned_warehouse': self.wh1,
                'default_service_rate': 2500000,
                'is_active': True
            },
            # ۳. استیجاری - تریلی
            {
                'driver_name': 'بهرام رادمنش',
                'plate_number': '33ع333-33',
                'driver_phone': '09123333333',
                'driver_national_code': '0033333333',
                'bank_name': 'ملت',
                'account_number': '0103333333003',
                'sheba_number': 'IR330120000000333333333003',
                'vehicle_type': 'trailer',
                'ownership_type': 'contract',
                'assigned_warehouse': self.wh2,
                'default_service_rate': 7000000,
                'is_active': True
            },
            # ۴. ملکی/شخصی - وانت پیکان
            {
                'driver_name': 'اصغر مرادی',
                'plate_number': '44ع444-44',
                'driver_phone': '09124444444',
                'driver_national_code': '0044444444',
                'bank_name': 'ملی',
                'account_number': '0104444444004',
                'sheba_number': 'IR440170000000444444444004',
                'vehicle_type': 'pickup',
                'ownership_type': 'personal',
                'assigned_warehouse': self.wh1,
                'default_service_rate': 900000,
                'is_active': True
            },
            # ۵. شرکتی - خاور ۶ تن (حقوق‌بگیر، بدون تسویه کرایه‌ای)
            {
                'driver_name': 'قاسم سلیمانیان',
                'plate_number': '55ع555-55',
                'driver_phone': '09125555555',
                'driver_national_code': '0055555555',
                'bank_name': 'صادرات',
                'account_number': '0105555555005',
                'sheba_number': 'IR550190000000555555555005',
                'vehicle_type': 'khavar',
                'ownership_type': 'company',
                'assigned_warehouse': self.wh1,
                'default_service_rate': 0,
                'is_active': True
            },
            # ۶. شرکتی - لیفتراک انبار (بدون تسویه)
            {
                'driver_name': 'علی اکبر رضوانی (اپراتور لیفتراک)',
                'plate_number': 'LF-901',
                'driver_phone': '09126666666',
                'driver_national_code': '0066666666',
                'bank_name': 'ملی',
                'account_number': '0106666666006',
                'sheba_number': 'IR660170000000666666666006',
                'vehicle_type': 'other',
                'ownership_type': 'company',
                'assigned_warehouse': self.wh1,
                'default_service_rate': 0,
                'is_active': True
            },
            # ۷. استیجاری - ایسوزو (شناور / بدون انتساب اولیه انبار)
            {
                'driver_name': 'داوود حیدری',
                'plate_number': '77ع777-77',
                'driver_phone': '09127777777',
                'driver_national_code': '0077777777',
                'bank_name': 'ملی',
                'account_number': '0107777777007',
                'sheba_number': 'IR770170000000777777777007',
                'vehicle_type': 'truck',
                'ownership_type': 'contract',
                'assigned_warehouse': None,
                'default_service_rate': 3200000,
                'is_active': True
            },
            # ۸. استیجاری - تریلی کفی
            {
                'driver_name': 'جواد میرزایی',
                'plate_number': '88ع888-88',
                'driver_phone': '09128888888',
                'driver_national_code': '0088888888',
                'bank_name': 'تجارت',
                'account_number': '0108888888008',
                'sheba_number': 'IR880180000000888888888008',
                'vehicle_type': 'trailer',
                'ownership_type': 'contract',
                'assigned_warehouse': self.wh2,
                'default_service_rate': 6500000,
                'is_active': True
            },
            # ۹. استیجاری - نیسان یخچالدار
            {
                'driver_name': 'سعید تقوی',
                'plate_number': '99ع999-99',
                'driver_phone': '09129999999',
                'driver_national_code': '0099999999',
                'bank_name': 'ملی',
                'account_number': '0109999999009',
                'sheba_number': 'IR990170000000999999999009',
                'vehicle_type': 'nissan',
                'ownership_type': 'contract',
                'assigned_warehouse': self.wh1,
                'default_service_rate': 1800000,
                'is_active': True
            },
            # ۱۰. ملکی - وانت مزدا
            {
                'driver_name': 'حمید گودرزی',
                'plate_number': '10ع100-10',
                'driver_phone': '09120000010',
                'driver_national_code': '0010000010',
                'bank_name': 'ملی',
                'account_number': '0101000010010',
                'sheba_number': 'IR100170000000100000100010',
                'vehicle_type': 'pickup',
                'ownership_type': 'personal',
                'assigned_warehouse': self.wh2,
                'default_service_rate': 1100000,
                'is_active': True
            },
        ]

        self.vehicles = [VehicleDriverProfile.objects.create(**vd) for vd in self.vehicles_data]
        self.year_month = '1405/06'

    def test_01_verify_ten_vehicles_created(self):
        """۱. بررسی ایجاد موفقیت‌آمیز حداقل ۱۰ خودرو و راننده در دیتابیس"""
        self.assertEqual(VehicleDriverProfile.objects.count(), 10)
        contract_count = VehicleDriverProfile.objects.filter(ownership_type='contract').count()
        company_count = VehicleDriverProfile.objects.filter(ownership_type='company').count()
        personal_count = VehicleDriverProfile.objects.filter(ownership_type='personal').count()
        self.assertEqual(contract_count, 6)
        self.assertEqual(company_count, 2)
        self.assertEqual(personal_count, 2)

    def test_02_daily_trips_recording_and_audit_logging(self):
        """۲. ثبت تردد روزانه برای خودروها و بررسی لاگ ممیزی خودکار"""
        v1 = self.vehicles[0]  # رضا صادقی - نیسان

        payload = {
            'vehicle_id': v1.id,
            'warehouse_id': self.wh1.id,
            'date_shamsi': '1405/06/05',
            'trip_count': 3,
            'unit_rate': 1200000,
            'dispatch_reference': 'DISP-101',
            'origin_destination': 'تهران به کرج',
            'notes': '۳ سرویس مواد اولیه'
        }
        res = self.client.post('/api/personnel/trips/update-day-trip/', payload, format='json')
        self.assertEqual(res.status_code, status.HTTP_200_OK)

        trip = VehicleTripLog.objects.get(vehicle=v1, date_shamsi='1405/06/05')
        self.assertEqual(trip.trip_count, 3)
        self.assertEqual(trip.total_amount, 3 * 1200000)

        # بررسی لاگ ممیزی برای ایجاد رکورد جدید
        audit_logs = VehicleTripAuditLog.objects.filter(trip=trip)
        self.assertTrue(audit_logs.exists())

        # سناریوی ویرایش نرخ و تعداد سرویس
        update_payload = {
            'vehicle_id': v1.id,
            'warehouse_id': self.wh1.id,
            'date_shamsi': '1405/06/05',
            'trip_count': 4,
            'unit_rate': 1300000,
            'notes': 'افزایش نرخ به دلیل سختی مسیر'
        }
        res_update = self.client.post('/api/personnel/trips/update-day-trip/', update_payload, format='json')
        self.assertEqual(res_update.status_code, status.HTTP_200_OK)

        trip.refresh_from_db()
        self.assertEqual(trip.trip_count, 4)
        self.assertEqual(trip.unit_rate, 1300000)
        self.assertEqual(trip.total_amount, 4 * 1300000)

        # بررسی لاگ‌های ویرایش
        rate_log = VehicleTripAuditLog.objects.filter(trip=trip, field_name='unit_rate').latest('changed_at')
        self.assertEqual(rate_log.old_value, '1200000')
        self.assertEqual(rate_log.new_value, '1300000')

    def test_03_monthly_grid_calculation_across_multiple_vehicles(self):
        """۳. تست تجمیع ماتریس ۳۱ روزه ماهانه برای چندین خودرو"""
        # ثبت تردد روزهای مختلف برای خودروهای ۱ تا ۵
        VehicleTripLog.objects.create(
            vehicle=self.vehicles[0], warehouse=self.wh1, date_shamsi='1405/06/01',
            trip_count=2, unit_rate=1200000, total_amount=2400000
        )
        VehicleTripLog.objects.create(
            vehicle=self.vehicles[0], warehouse=self.wh1, date_shamsi='1405/06/02',
            trip_count=3, unit_rate=1200000, total_amount=3600000
        )
        VehicleTripLog.objects.create(
            vehicle=self.vehicles[1], warehouse=self.wh1, date_shamsi='1405/06/01',
            trip_count=1, unit_rate=2500000, total_amount=2500000
        )

        res = self.client.get(f'/api/personnel/trips/monthly-grid/?year_month=1405/06&warehouse_id={self.wh1.id}')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        data = res.json()

        self.assertIn('days_meta', data)
        self.assertIn('rows', data)
        self.assertEqual(len(data['days_meta']), 31)

        # جستجوی سطر راننده اول
        row1 = next(r for r in data['rows'] if r['vehicle_id'] == self.vehicles[0].id)
        self.assertEqual(row1['total_trips'], 5)
        self.assertEqual(row1['total_amount'], 6000000)

    def test_04_period_lock_guards_for_fleet_trips(self):
        """۴. تست قفل دوره و گارد جلوگیری از تغییرات بدون اجازه سوپریوزر"""
        period = MonthlyWorkPeriod.objects.create(
            year_month='1405/06',
            warehouse=self.wh1,
            status='LOCKED',
            locked_by=self.user
        )

        # ایجاد کاربر عادی بدون حق آنلاک
        normal_user = User.objects.create_user(username='normal_user', password='Password123!')
        self.client.force_authenticate(user=normal_user)

        payload = {
            'vehicle_id': self.vehicles[0].id,
            'warehouse_id': self.wh1.id,
            'date_shamsi': '1405/06/10',
            'trip_count': 2,
            'unit_rate': 1200000,
        }
        res = self.client.post('/api/personnel/trips/update-day-trip/', payload, format='json')
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('قفل', res.json()['error'])

        # تست با سوپریوزر و پارامتر is_override
        self.client.force_authenticate(user=self.user)
        payload['is_override'] = True
        res_override = self.client.post('/api/personnel/trips/update-day-trip/', payload, format='json')
        self.assertEqual(res_override.status_code, status.HTTP_200_OK)

    def test_05_fleet_settlement_calculation_engine_all_ten_vehicles(self):
        """۵. تست جامع موتور تسویه مالی ناوگان و تفکیک مالکیت برای تمام ۱۰ خودرو"""
        # ثبت کارکرد برای خودروهای مختلف در ماه ۱۴۰۵/۰۶
        # ۱. رضا صادقی (استیجاری): ۱۰ سرویس * ۱،۲۰۰،۰۰۰ = ۱۲،۰۰۰،۰۰۰
        VehicleTripLog.objects.create(
            vehicle=self.vehicles[0], warehouse=self.wh1, date_shamsi='1405/06/01',
            trip_count=10, unit_rate=1200000, total_amount=12000000
        )
        # ۲. محمود کریمی (استیجاری): ۴ سرویس * ۲،۵۰۰،۰۰۰ = ۱۰،۰۰۰،۰۰۰
        VehicleTripLog.objects.create(
            vehicle=self.vehicles[1], warehouse=self.wh1, date_shamsi='1405/06/02',
            trip_count=4, unit_rate=2500000, total_amount=10000000
        )
        # ۳. بهرام رادمنش (استیجاری): ۲ سرویس * ۷،۰۰۰،۰۰۰ = ۱۴،۰۰۰،۰۰۰
        VehicleTripLog.objects.create(
            vehicle=self.vehicles[2], warehouse=self.wh2, date_shamsi='1405/06/03',
            trip_count=2, unit_rate=7000000, total_amount=14000000
        )
        # ۴. اصغر مرادی (شخصی/ملکی): ۵ سرویس * ۹۰۰،۰۰۰ = ۴،۵۰۰،۰۰۰
        VehicleTripLog.objects.create(
            vehicle=self.vehicles[3], warehouse=self.wh1, date_shamsi='1405/06/04',
            trip_count=5, unit_rate=900000, total_amount=4500000
        )
        # ۵. قاسم سلیمانیان (شرکتی): ۲۰ سرویس * ۰ = ۰ (مشمول تسویه نیست)
        VehicleTripLog.objects.create(
            vehicle=self.vehicles[4], warehouse=self.wh1, date_shamsi='1405/06/05',
            trip_count=20, unit_rate=0, total_amount=0
        )
        # ۶. علی اکبر رضوانی (لیفتراک شرکتی): ۱۵ سرویس * ۰ = ۰
        VehicleTripLog.objects.create(
            vehicle=self.vehicles[5], warehouse=self.wh1, date_shamsi='1405/06/06',
            trip_count=15, unit_rate=0, total_amount=0
        )

        res = calculate_monthly_fleet_settlement(warehouse_id=None, year_month='1405/06')
        summary = res['summary']
        records = res['records']

        # بررسی جمع کل سرویس‌ها: ۱۰ + ۴ + ۲ + ۵ + ۲۰ + ۱۵ = ۵۶ سرویس
        self.assertEqual(summary['total_fleet_trips'], 56)

        # بررسی سرجمع مبالغ قابل تسویه: ۱۲م + ۱۰م + ۱۴م + ۴.۵م = ۴۰،۵۰۰،۰۰۰ تومان
        self.assertEqual(summary['total_payable_settlement'], 40500000)

        # بررسی وضعیت تسویه راننده شرکتی (باید مبلغ صفر و مشمول تسویه False باشد)
        company_rec = next(r for r in records if r['vehicle_id'] == self.vehicles[4].id)
        self.assertEqual(company_rec['payable_amount'], 0)
        self.assertFalse(company_rec['is_payable'])
        self.assertEqual(company_rec['trip_count'], 20)

        # بررسی وضعیت تسویه راننده استیجاری
        contract_rec = next(r for r in records if r['vehicle_id'] == self.vehicles[0].id)
        self.assertEqual(contract_rec['payable_amount'], 12000000)
        self.assertTrue(contract_rec['is_payable'])

    def test_06_bank_meli_group_payment_excel_generation(self):
        """۶. تست تولید فایل اکسل پرداخت گروهی بانک ملی برای رانندگان"""
        # ثبت تردد برای راننده استیجاری
        VehicleTripLog.objects.create(
            vehicle=self.vehicles[0], warehouse=self.wh1, date_shamsi='1405/06/15',
            trip_count=5, unit_rate=1200000, total_amount=6000000
        )

        settlement_res = calculate_monthly_fleet_settlement(warehouse_id=None, year_month='1405/06')
        excel_stream = generate_fleet_bank_meli_excel(settlement_res, 'شهریور ۱۴۰۵')
        wb = openpyxl.load_workbook(excel_stream)
        ws = wb.active

        # بررسی ساختار هدر بانک ملی
        headers = [ws.cell(1, col).value for col in range(1, 8)]
        self.assertEqual(headers[0], 'ردیف')
        self.assertEqual(headers[1], 'مبلغ (ریال)')
        self.assertEqual(headers[2], 'شماره حساب ملی / شبا مقصد')
        self.assertEqual(headers[3], 'نام گیرنده (راننده/پیمانکار)')

        # بررسی ردیف داده‌ها (مبلغ باید به ریال باشد یعنی ۶،۰۰۰،۰۰۰ * ۱۰ = ۶۰،۰۰۰،۰۰۰ ریال)
        self.assertEqual(ws.cell(2, 1).value, 1)
        self.assertEqual(ws.cell(2, 2).value, 60000000)
        self.assertEqual(ws.cell(2, 3).value, self.vehicles[0].sheba_number)
        self.assertEqual(ws.cell(2, 4).value, self.vehicles[0].driver_name)

    def test_07_two_way_31_day_excel_export_and_import(self):
        """۷. تست چرخه رفت و برگشت اکسل ماتریسی ۳۱ روزه ناوگان"""
        # ثبت اولیه در دیتابیس
        VehicleTripLog.objects.create(
            vehicle=self.vehicles[0], warehouse=self.wh1, date_shamsi='1405/06/01',
            trip_count=2, unit_rate=1200000, total_amount=2400000
        )

        # اکسپورت به فایل اکسل
        stream_out = export_fleet_monthly_excel(warehouse_id=self.wh1.id, year_month='1405/06')
        wb = openpyxl.load_workbook(stream_out)
        ws = wb.active

        # پیدا کردن سطر مربوط به خودروی اول بر اساس پلاک در اکسل
        target_row = None
        for r in range(3, ws.max_row + 1):
            if ws.cell(r, 3).value == self.vehicles[0].plate_number:
                target_row = r
                break
        self.assertIsNotNone(target_row)
        ws.cell(target_row, 8).value = 4  # ستون ۸ مربوط به روز دوم (day_02)

        # ذخیره در بایت‌استریم
        stream_in = io.BytesIO()
        wb.save(stream_in)
        stream_in.seek(0)

        # ایمپورت مجدد اکسل به دیتابیس
        res_import = import_fleet_monthly_excel(
            file_obj=stream_in,
            warehouse_id=self.wh1.id,
            year_month='1405/06',
            user=self.user
        )
        self.assertGreater(res_import['updated_count'], 0)

        # بررسی ثبت شدن رکورد روز دوم در دیتابیس
        day2_trip = VehicleTripLog.objects.get(vehicle=self.vehicles[0], date_shamsi='1405/06/02')
        self.assertEqual(day2_trip.trip_count, 4)
        self.assertEqual(day2_trip.total_amount, 4 * 1200000)
