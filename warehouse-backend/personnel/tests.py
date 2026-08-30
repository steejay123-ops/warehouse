import os
from django.conf import settings
from django.test import TestCase
from django.core.management import call_command
from django.contrib.auth import get_user_model
from rest_framework.test import APIClient
from warehouses.models import Warehouse
from .models import (
    PersonnelProfile,
    VehicleDriverProfile,
    MonthlyWorkPeriod,
    DailyAttendance,
    AttendanceAuditLog,
    VehicleTripLog,
    PayrollYearlySettings,
    MonthlyPayrollRecord
)

User = get_user_model()


class PersonnelAppTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.user = User.objects.create_superuser(
            username='admin_test',
            password='testpassword123',
            national_code='1234567890'
        )
        self.client.force_authenticate(user=self.user)
        call_command('seed_payroll_settings')
        
        self.warehouse = Warehouse.objects.create(
            name='انبار مرکزی پارس',
            code='WH-PARS-01'
        )
        
        self.personnel1 = PersonnelProfile.objects.create(
            first_name='حسین',
            last_name='اکبری',
            national_code='2452304301',
            job_title='مدیر پروژه',
            daily_base_wage=6572696,
            assigned_warehouse=self.warehouse
        )
        
        self.personnel2 = PersonnelProfile.objects.create(
            first_name='علیرضا',
            last_name='عالیشوندی',
            national_code='4260006134',
            job_title='کمک انباردار',
            daily_base_wage=6044057,
            assigned_warehouse=self.warehouse
        )
        
        self.vehicle1 = VehicleDriverProfile.objects.create(
            plate_number='12-الف-345-ایران-72',
            vehicle_type='nissan',
            driver_name='محمد رضایی',
            default_service_rate=2500000,
            assigned_warehouse=self.warehouse
        )

    def test_personnel_hourly_rate(self):
        """تست نرخ ساعتی معادل ۱۰ ساعت در روز"""
        self.assertEqual(self.personnel1.hourly_rate, 657269.6)

    def test_matrix_get_and_bulk_save(self):
        """تست واکشی ماتریس و ذخیره سریع گروهی کارکرد"""
        import jdatetime
        today = jdatetime.date.today()
        date_str = f"{today.year:04d}/{today.month:02d}/{today.day:02d}"
        res = self.client.get(f'/api/personnel/attendance/matrix/?warehouse_id={self.warehouse.id}&date_shamsi={date_str}')
        self.assertEqual(res.status_code, 200)
        self.assertEqual(len(res.data['rows']), 2)
        self.assertEqual(res.data['rows'][0]['status'], '')
        self.assertEqual(res.data['rows'][0]['effective_hours'], 0.0)
        self.assertFalse(res.data['rows'][0]['is_existing'])

        # ذخیره گروهی
        save_payload = {
            'warehouse_id': self.warehouse.id,
            'date_shamsi': date_str,
            'items': [
                {
                    'personnel_id': self.personnel1.id,
                    'status': 'PRESENT_10H',
                    'effective_hours': 10.0,
                    'overtime_hours': 2.5,
                    'is_friday_work': False,
                    'is_mission': False,
                    'advance_payment': 0,
                    'notes': 'تست اضافه‌کار'
                },
                {
                    'personnel_id': self.personnel2.id,
                    'status': 'HALF_5H',
                    'effective_hours': 5.0,
                    'overtime_hours': 0.0,
                    'is_friday_work': False,
                    'is_mission': False,
                    'advance_payment': 500000,
                    'notes': 'نیمه‌وقت'
                }
            ]
        }
        res_save = self.client.post('/api/personnel/attendance/bulk-save/', save_payload, format='json')
        self.assertEqual(res_save.status_code, 200)
        self.assertEqual(res_save.data['saved_count'], 2)

        # ویرایش مجدد و بررسی ثبت لاگ ممیزی
        save_payload['items'][0]['overtime_hours'] = 4.0
        res_update = self.client.post('/api/personnel/attendance/bulk-save/', save_payload, format='json')
        self.assertEqual(res_update.status_code, 200)
        self.assertEqual(res_update.data['updated_count'], 2)
        
        # بررسی ثبت لاگ ممیزی
        logs = AttendanceAuditLog.objects.filter(field_name='overtime_hours')
        self.assertTrue(logs.exists())
        self.assertIn(logs.first().new_value, ['4.00', '4.0'])

    def test_vehicle_trips_bulk_save_and_summary(self):
        """تست ثبت سریع سرویس‌های خودرو و دریافت خلاصه ماهانه"""
        date_str = '1404/05/10'
        trip_payload = {
            'warehouse_id': self.warehouse.id,
            'date_shamsi': date_str,
            'items': [
                {
                    'vehicle_id': self.vehicle1.id,
                    'trip_count': 3,
                    'unit_rate': 2500000,
                    'dispatch_reference': 'DISP-1001',
                    'origin_destination': 'انبار به سایت عسلویه'
                }
            ]
        }
        res = self.client.post('/api/personnel/trips/bulk-save/', trip_payload, format='json')
        self.assertEqual(res.status_code, 200)
        self.assertEqual(res.data['saved_count'], 1)

        # بررسی خلاصه ماهانه
        res_summary = self.client.get(f'/api/personnel/trips/monthly-summary/?warehouse_id={self.warehouse.id}&year_month=1404/05')
        self.assertEqual(res_summary.status_code, 200)
        self.assertEqual(res_summary.data['summary'][0]['total_trips'], 3)
        self.assertEqual(res_summary.data['summary'][0]['total_payable'], 7500000)

    def test_period_lock(self):
        """تست قفل شدن دوره و ممانعت از ویرایش"""
        date_str = '1404/05/10'
        period = MonthlyWorkPeriod.objects.create(
            warehouse=self.warehouse,
            year_month='1404/05',
            status='OPEN'
        )
        res_lock = self.client.post(f'/api/personnel/periods/{period.id}/lock/')
        self.assertEqual(res_lock.status_code, 200)
        period.refresh_from_db()
        self.assertEqual(period.status, 'LOCKED')

        # تلاش برای ثبت در دوره قفل شده
        payload = {
            'warehouse_id': self.warehouse.id,
            'date_shamsi': date_str,
            'items': []
        }
        res_err = self.client.post('/api/personnel/attendance/bulk-save/', payload, format='json')
        self.assertEqual(res_err.status_code, 400)

    def test_national_code_padding_and_wage_linkage(self):
        """تست پدینگ خودکار کد ملی و محاسبه خودکار مزد مبنا و نرخ ساعتی"""
        p = PersonnelProfile.objects.create(
            first_name='سعید',
            last_name='مرادی',
            national_code='245230430',  # 9 رقمی با صفر ابتدایی افتاده
            daily_base_wage=5000000,
            daily_seniority_bonus=200000
        )
        self.assertEqual(p.national_code, '0245230430')
        self.assertEqual(p.base_daily_rate, 5200000)
        self.assertEqual(p.hourly_rate, 520000.0)

    def test_excel_import_endpoint(self):
        """تست اندپوینت درون‌ریزی فایل اکسل شرکت"""
        excel_path = 'E:/warehouse project/Documents/SalaryCalculation/حقوق تیر ماه انبارداری.xlsm'
        if not os.path.exists(excel_path):
            excel_path = 'E:/warehouse project/حقوق تیر ماه انبارداری.xlsm'
        with open(excel_path, 'rb') as fp:
            res = self.client.post(
                f'/api/personnel/profiles/import-excel/?warehouse_id={self.warehouse.id}',
                {'file': fp},
                format='multipart'
            )
        self.assertEqual(res.status_code, 200)
        self.assertGreater(res.data['created_count'] + res.data['updated_count'], 40)
        
        # بررسی ثبت شدن پرسنل شیت Emp_info
        hossein = PersonnelProfile.objects.get(national_code='2452304301')
        self.assertEqual(hossein.first_name, 'حسین')
        self.assertEqual(hossein.last_name, 'اکبری')
        self.assertEqual(hossein.job_code, '027212')
        self.assertEqual(hossein.status_category, 'نفرات شرکتی')
        self.assertEqual(hossein.contract_hours, 230)

    def test_payroll_yearly_settings_and_job_grades(self):
        """تست تنظیمات سالانه حقوق و دریافت نرخ گروه شغلی"""
        res_settings = self.client.get('/api/personnel/settings/active-or-year/?year=1405')
        self.assertEqual(res_settings.status_code, 200)
        self.assertEqual(res_settings.data['fiscal_year'], '1405')

        # تست واکشی گروه شغلی ۱۹
        res_grade = self.client.get('/api/personnel/settings/job-grade-rate/?grade=19&year=1405')
        self.assertEqual(res_grade.status_code, 200)
        self.assertEqual(res_grade.data['daily_base_wage'], 6572696.0)
        self.assertEqual(res_grade.data['daily_seniority_bonus'], 171867.0)
        self.assertEqual(res_grade.data['hourly_rate'], 657269.6)

    def test_monthly_payroll_calculation_and_exports_flow(self):
        """تست جامع محاسبه حقوق ماهانه، تولید دیسکت‌های بیمه، فایل‌های مالیات و بانک"""
        # 1. محاسبه حقوق دوره تیر ماه ۱۴۰۵
        calc_payload = {
            'warehouse_id': self.warehouse.id,
            'year_month': '1405/04'
        }
        res_calc = self.client.post('/api/personnel/monthly-payroll/calculate-period/', calc_payload, format='json')
        self.assertEqual(res_calc.status_code, 200)
        self.assertGreater(len(res_calc.data['records']), 0)
        period_id = res_calc.data['period_id']

        # 2. تست صدور دیسکت‌های بیمه (ZIP حاوی DSKWOR00.DBF و DSKKAR00.DBF)
        res_dsk = self.client.get(f'/api/personnel/monthly-payroll/export-dsk-zip/?period_id={period_id}')
        self.assertEqual(res_dsk.status_code, 200)
        self.assertEqual(res_dsk['Content-Type'], 'application/zip')
        self.assertGreater(len(res_dsk.content), 1000)

        # 3. تست صدور فایل‌های مالیات حقوق (WH و WP)
        res_wh = self.client.get(f'/api/personnel/monthly-payroll/export-tax-wh/?period_id={period_id}')
        self.assertEqual(res_wh.status_code, 200)
        self.assertTrue(res_wh.content.startswith(b'\xef\xbb\xbf')) # UTF-8 BOM

        res_wp = self.client.get(f'/api/personnel/monthly-payroll/export-tax-wp/?period_id={period_id}')
        self.assertEqual(res_wp.status_code, 200)
        self.assertTrue(res_wp.content.startswith(b'\xef\xbb\xbf')) # UTF-8 BOM

        # 4. تست صدور فایل اکسل پرداخت بانک ملی
        res_bank = self.client.get(f'/api/personnel/monthly-payroll/export-bank-excel/?period_id={period_id}')
        self.assertEqual(res_bank.status_code, 200)
        self.assertEqual(res_bank['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # 5. تست صدور اکسل ۲ سطری استاندارد ۵۸ ستون حقوق
        res_excel = self.client.get(f'/api/personnel/monthly-payroll/export-monthly-excel/?period_id={period_id}')
        self.assertEqual(res_excel.status_code, 200)
        self.assertEqual(res_excel['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    def test_surplus_allocation_and_zero_tax_check_discrepancy(self):
        """تست اختصاصی تسهیم مازاد و صفر بودن چک مالیات (ستون ۵۸)"""
        period = MonthlyWorkPeriod.objects.create(
            year_month='1405/04',
            warehouse=self.warehouse,
            status='OPEN'
        )
        self.personnel1.contract_hours = 230
        self.personnel1.contract_base_salary = 270000000 # 27 Million Tomans
        self.personnel1.job_grade = '19'
        self.personnel1.save()

        # ثبت کارکرد کامل ۲۳۰ ساعت
        DailyAttendance.objects.create(
            personnel=self.personnel1,
            warehouse=self.warehouse,
            date_shamsi='1405/04/01',
            effective_hours=230,
            status='PRESENT_10H'
        )

        from personnel.payroll_engine import PayrollCalculationEngine
        engine = PayrollCalculationEngine(period=period)
        records = engine.calculate_period(period)
        self.assertEqual(len(records), 2)

        record1 = next(r for r in records if r.personnel_id == self.personnel1.id)
        # ستون ۵۸ (چک مالیات) باید دقیقاً ۰ باشد
        self.assertEqual(record1.tax_check_discrepancy, 0)
        # حقوق ناخالص باید ۲۷۰ میلیون ریال باشد
        self.assertEqual(record1.gross_salary, 270000000)
        # اضافه‌کار و سفر/ماموریت باید مقادیر مثبت تسهیم‌شده داشته باشند
        self.assertGreater(record1.overtime_amount, 0)
        self.assertGreater(record1.travel_cost_amount + record1.mission_amount, 0)

    def test_import_tax_excel_flow(self):
        """تست اختصاصی درون‌ریزی فایل اکسل مالیات دارایی و تطبیق با کدملی"""
        self.personnel1.national_code = '2450718214'
        self.personnel1.save()

        period = MonthlyWorkPeriod.objects.create(
            year_month='1405/04',
            warehouse=self.warehouse,
            status='OPEN'
        )
        from personnel.payroll_engine import PayrollCalculationEngine
        PayrollCalculationEngine(period=period).calculate_period(period)

        possible_paths = [
            r'E:\warehouse project\Documents\SalaryCalculation\TaxCalculatedByTheTaxOffice.xlsx',
            r'E:\warehouse project\Documents\SalaryCalculation\نمونه مالیات محاسبه شده.xlsx',
            os.path.join(settings.BASE_DIR, '..', 'Documents', 'SalaryCalculation', 'TaxCalculatedByTheTaxOffice.xlsx'),
            os.path.join(settings.BASE_DIR, '..', 'Documents', 'SalaryCalculation', 'نمونه مالیات محاسبه شده.xlsx'),
            r'E:\warehouse project\نمونه مالیات محاسبه شده.xlsx',
        ]
        tax_file_path = next((p for p in possible_paths if os.path.exists(p)), possible_paths[0])
        with open(tax_file_path, 'rb') as fp:
            res = self.client.post(
                '/api/personnel/monthly-payroll/import-tax-excel/',
                {'file': fp, 'period_id': period.id},
                format='multipart'
            )
        self.assertEqual(res.status_code, 200)
        self.assertGreater(res.data['matched_count'], 0)

        # بررسی به‌روزرسانی مالیات پرسنل با کدملی 2452304301
        rec = MonthlyPayrollRecord.objects.get(period=period, personnel=self.personnel1)
        self.assertTrue(rec.is_tax_imported)
        self.assertEqual(rec.tax_source_type, 'IMPORTED_EXCEL')

    def test_wh_and_wp_tax_file_exports(self):
        """تست اختصاصی ساختار فایل‌های WP (۲۳ ستون) و WH (۳۹ ستون) با تنظیمات انفرادی پرسنل"""
        from personnel.tax_bank_exporter import generate_wp_tax_content, generate_wh_tax_content

        # تنظیم مشخصات پرسنل با کدهای اختصاصی دارایی 1.7.0.4
        self.personnel1.education_level = '7' # فوق دکتری
        self.personnel1.nationality_code = '1'
        self.personnel1.exemption_type = '11' # معافیت جوانی جمعیت
        self.personnel1.job_category = '5' # انبارداری
        self.personnel1.employment_type = '2' # شرکتی
        self.personnel1.tax_payment_type = '2' # ارزی
        self.personnel1.tax_service_location = '2' # مناطق محروم
        self.personnel1.tax_exceptions = '4' # کارکنان عملیاتی نفت
        self.personnel1.tax_currency_type = '40' # دلار
        self.personnel1.tax_currency_exchange_rate = 1.00
        self.personnel1.tax_housing_benefit_type = '2' # مسکن با اثاثیه
        self.personnel1.tax_vehicle_benefit_type = '2' # با راننده
        self.personnel1.save()

        period = MonthlyWorkPeriod.objects.create(
            year_month='1405/04',
            warehouse=self.warehouse,
            status='OPEN'
        )
        rec = MonthlyPayrollRecord.objects.create(
            period=period,
            personnel=self.personnel1,
            continuous_taxable_salary_allowances=250000000,
            overtime_amount=40000000,
            travel_cost_amount=20000000,
            mission_amount=15000000,
            bonus_amount=30000000,
            seniority_allowance=15000000,
            worker_insurance=45000000,
            include_in_tax=True
        )

        wp_content = generate_wp_tax_content([rec])
        wh_content = generate_wh_tax_content([rec])

        # بررسی ۲۳ ستون فایل WP
        wp_lines = [l for l in wp_content.replace('\ufeff', '').split('\r\n') if l.strip()]
        self.assertEqual(len(wp_lines), 1)
        wp_cols = wp_lines[0].split(',')
        self.assertEqual(len(wp_cols), 23)
        self.assertEqual(wp_cols[0], '1') # ملیت
        self.assertEqual(wp_cols[1], '2452304301') # کدملی
        self.assertEqual(wp_cols[8], '7') # مدرک تحصیلی فوق دکتری
        self.assertEqual(wp_cols[12], '11') # معافیت ماده ۱۸
        self.assertEqual(wp_cols[17], '5') # رسته انبارداری

        # بررسی ۳۹ ستون فایل WH
        wh_lines = [l for l in wh_content.replace('\ufeff', '').split('\r\n') if l.strip()]
        self.assertEqual(len(wh_lines), 1)
        wh_cols = wh_lines[0].split(',')
        self.assertEqual(len(wh_cols), 39)
        self.assertEqual(wh_cols[0], '2452304301') # کدملی
        self.assertEqual(wh_cols[1], '2') # نوع پرداختی ارزی
        self.assertEqual(wh_cols[2], '2') # محل خدمت مناطق محروم
        self.assertEqual(wh_cols[3], '4') # استثنائات نفت
        self.assertEqual(wh_cols[4], '40') # دلار
        self.assertEqual(wh_cols[6], '250000000') # ناخالص مستمر
        self.assertEqual(wh_cols[8], '2') # مسکن با اثاثیه
        self.assertEqual(wh_cols[10], '2') # اتومبیل با راننده
        self.assertEqual(wh_cols[16], '40000000') # اضافه کار
        self.assertEqual(wh_cols[17], '20000000') # هزینه سفر
        self.assertEqual(wh_cols[18], '15000000') # ماموریت
        self.assertEqual(wh_cols[22], '30000000') # عیدی
        self.assertEqual(wh_cols[26], '15000000') # سنوات
        self.assertEqual(wh_cols[32], '45000000') # بیمه کارگر


class PersonnelAttendanceWindowTestCase(TestCase):
    def setUp(self):
        import jdatetime
        self.client = APIClient()
        self.user = User.objects.create_superuser(
            username='admin_attendance_test',
            password='testpassword123',
            national_code='9876543210'
        )
        self.client.force_authenticate(user=self.user)
        
        self.warehouse = Warehouse.objects.create(
            name='انبار عسلویه ۱',
            code='WH-ASALUYEH-01'
        )
        
        self.personnel = PersonnelProfile.objects.create(
            first_name='سعید',
            last_name='مرادی',
            national_code='1271234567',
            job_title='انباردار',
            daily_base_wage=5000000,
            assigned_warehouse=self.warehouse
        )
        
        today = jdatetime.date.today()
        self.today_shamsi = today.strftime('%Y/%m/%d')
        self.year_month = self.today_shamsi[:7]
        self.current_year = str(today.year)
        
        self.settings = PayrollYearlySettings.objects.create(
            fiscal_year=self.current_year,
            title=f'تنظیمات سال {self.current_year}',
            is_active=True,
            attendance_edit_past_days=3,
            attendance_edit_future_days=0
        )

    def test_get_monthly_grid_success(self):
        """تست استخراج شیت ماهانه کارکرد پرسنل با احتساب جمعه‌ها و تنظیمات بازه"""
        url = f'/api/personnel/attendance/monthly-grid/?warehouse_id={self.warehouse.id}&year_month={self.year_month}'
        res = self.client.get(url)
        self.assertEqual(res.status_code, 200)
        data = res.json()
        self.assertIn('days_meta', data)
        self.assertIn('rows', data)
        self.assertGreaterEqual(len(data['days_meta']), 29)
        self.assertEqual(len(data['rows']), 1)
        self.assertEqual(data['rows'][0]['personnel_id'], self.personnel.id)

    def test_past_date_window_rejection(self):
        """تست مسدودسازی ثبت کارکرد برای بیش از ۳ روز گذشته"""
        import jdatetime
        past_date = (jdatetime.date.today() - jdatetime.timedelta(days=10)).strftime('%Y/%m/%d')
        payload = {
            'warehouse_id': self.warehouse.id,
            'date_shamsi': past_date,
            'items': [{
                'personnel_id': self.personnel.id,
                'status': 'PRESENT_10H',
                'effective_hours': 10,
                'overtime_hours': 0,
                'is_friday_work': False,
                'is_mission': False
            }]
        }
        res = self.client.post('/api/personnel/attendance/bulk-save/', payload, format='json')
        self.assertEqual(res.status_code, 400)
        self.assertIn('مجاز نیست', str(res.json()))

    def test_future_date_window_rejection(self):
        """تست مسدودسازی ثبت کارکرد برای تاریخ آینده وقتی سقف آینده ۰ است"""
        import jdatetime
        future_date = (jdatetime.date.today() + jdatetime.timedelta(days=2)).strftime('%Y/%m/%d')
        payload = {
            'warehouse_id': self.warehouse.id,
            'date_shamsi': future_date,
            'items': [{
                'personnel_id': self.personnel.id,
                'status': 'PRESENT_10H',
                'effective_hours': 10,
                'overtime_hours': 0,
                'is_friday_work': False,
                'is_mission': False
            }]
        }
        res = self.client.post('/api/personnel/attendance/bulk-save/', payload, format='json')
        self.assertEqual(res.status_code, 400)
        self.assertIn('آینده', str(res.json()))

    def test_today_date_allowed(self):
        """تست ثبت موفق کارکرد در تاریخ مجاز (امروز)"""
        payload = {
            'warehouse_id': self.warehouse.id,
            'date_shamsi': self.today_shamsi,
            'items': [{
                'personnel_id': self.personnel.id,
                'status': 'PRESENT_10H',
                'effective_hours': 10,
                'overtime_hours': 2,
                'is_friday_work': False,
                'is_mission': False,
                'advance_payment': 500000,
                'notes': 'تست حضور روزانه'
            }]
        }
        res = self.client.post('/api/personnel/attendance/bulk-save/', payload, format='json')
        self.assertEqual(res.status_code, 200)
        
        att = DailyAttendance.objects.filter(
            warehouse=self.warehouse,
            personnel=self.personnel,
            date_shamsi=self.today_shamsi
        ).first()
        self.assertIsNotNone(att)
        self.assertEqual(float(att.effective_hours), 10.0)
        self.assertEqual(float(att.overtime_hours), 2.0)

    def test_bulk_save_monthly_grid(self):
        """تست ذخیره گروهی شیت ماهانه در تاریخ مجاز"""
        import jdatetime
        today = jdatetime.date.today()
        today_day = today.day
        payload = {
            'warehouse_id': self.warehouse.id,
            'year_month': self.year_month,
            'items': [{
                'personnel_id': self.personnel.id,
                'day': today_day,
                'status': 'PRESENT_10H',
                'effective_hours': 10,
                'overtime_hours': 3,
                'is_friday_work': False,
                'is_mission': False,
                'advance_payment': 0,
                'notes': 'ثبت از طریق شیت ماهانه'
            }]
        }
        res = self.client.post('/api/personnel/attendance/bulk-save-monthly-grid/', payload, format='json')
        self.assertEqual(res.status_code, 200)
        
        att = DailyAttendance.objects.filter(
            warehouse=self.warehouse,
            personnel=self.personnel,
            date_shamsi=f"{self.year_month}/{today_day:02d}"
        ).first()
        self.assertIsNotNone(att)
        self.assertEqual(float(att.overtime_hours), 3.0)


class PersonnelAttendanceFiveRequirementsTestCase(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.user = User.objects.create_superuser(
            username='admin_attend_tests',
            password='testpassword123',
            national_code='9876543210'
        )
        self.client.force_authenticate(user=self.user)
        call_command('seed_payroll_settings')

        self.warehouse = Warehouse.objects.create(
            name='انبار جنوب',
            code='WH-SOUTH-01'
        )
        self.personnel = PersonnelProfile.objects.create(
            first_name='مهرداد',
            last_name='مرادی',
            national_code='1112223334',
            job_title='کارشناس فنی',
            assigned_warehouse=self.warehouse
        )

    def test_8_digit_date_format_accepted(self):
        """تست پذیرش تاریخ ۸ رقمی بدون اسلش مثل 14050607 بدون بروز خطای فرمت نامعتبر"""
        import jdatetime
        today = jdatetime.date.today()
        # تاریخ امروز به صورت ۸ رقم پیوسته
        today_8_digits = f"{today.year:04d}{today.month:02d}{today.day:02d}"
        
        res = self.client.get(f'/api/personnel/attendance/matrix/?warehouse_id={self.warehouse.id}&date_shamsi={today_8_digits}')
        self.assertEqual(res.status_code, 200)
        self.assertEqual(res.data['date_shamsi'], f"{today.year:04d}/{today.month:02d}/{today.day:02d}")

    def test_unrecorded_days_not_marked_present_in_monthly_grid(self):
        """تست اینکه روزهای ثبت‌نشده در شیت ماهانه حاضر زده نمی‌شوند و جمع ساعات صفر است"""
        import jdatetime
        today = jdatetime.date.today()
        year_month = f"{today.year:04d}/{today.month:02d}"

        res = self.client.get(f'/api/personnel/attendance/monthly-grid/?warehouse_id={self.warehouse.id}&year_month={year_month}')
        self.assertEqual(res.status_code, 200)
        p_row = next(r for r in res.data['rows'] if r['personnel_id'] == self.personnel.id)
        # جمع ساعات روزهای ثبت‌نشده باید صفر باشد
        self.assertEqual(p_row['total_hours'], 0.0)
        self.assertEqual(p_row['present_days'], 0)
        # وضعیت روزهای ثبت‌نشده نباید PRESENT_10H باشد
        first_day = p_row['days'][0]
        self.assertEqual(first_day['status'], '')
        self.assertEqual(first_day['effective_hours'], 0.0)
        self.assertFalse(first_day['is_existing'])

    def test_unrecorded_days_clean_slate_in_matrix(self):
        """تست اینکه روزهای ثبت‌نشده یا پاکسازی‌شده با وضعیت خالی و ساعات صفر لود می‌شوند و مقادیر کهنه بازنمی‌گردند"""
        import jdatetime
        today = jdatetime.date.today()
        yesterday = today - jdatetime.timedelta(days=1)
        yesterday_str = f"{yesterday.year:04d}/{yesterday.month:02d}/{yesterday.day:02d}"
        today_str = f"{today.year:04d}/{today.month:02d}/{today.day:02d}"

        DailyAttendance.objects.create(
            personnel=self.personnel,
            warehouse=self.warehouse,
            date_shamsi=yesterday_str,
            status='LEAVE',
            effective_hours=0.0,
            overtime_hours=0.0
        )

        # واکشی ماتریس برای امروز (که هنوز ثبت نشده یا پاکسازی شده)
        res = self.client.get(f'/api/personnel/attendance/matrix/?warehouse_id={self.warehouse.id}&date_shamsi={today_str}')
        self.assertEqual(res.status_code, 200)
        p_row = next(r for r in res.data['rows'] if r['personnel_id'] == self.personnel.id)
        self.assertFalse(p_row['is_existing'])
        # باید وضعیت تمیز و خالی باشد تا دکمه پاکسازی یا لود روزهای ثبت‌نشده بدون دستکاری عمل کند
        self.assertEqual(p_row['status'], '')
        self.assertEqual(p_row['effective_hours'], 0.0)

    def test_warehouse_independent_attendance(self):
        """تست واکشی و ثبت کارکرد مستقل از انبار (بدون warehouse_id)"""
        import jdatetime
        today = jdatetime.date.today()
        today_str = f"{today.year:04d}/{today.month:02d}/{today.day:02d}"

        # فراخوانی بدون warehouse_id
        res = self.client.get(f'/api/personnel/attendance/matrix/?date_shamsi={today_str}')
        self.assertEqual(res.status_code, 200)
        self.assertTrue(any(r['personnel_id'] == self.personnel.id for r in res.data['rows']))

        # ذخیره کارکرد بدون ارسال انبار
        payload = {
            'date_shamsi': today_str,
            'items': [{
                'personnel_id': self.personnel.id,
                'status': 'MISSION',
                'effective_hours': 10.0,
                'overtime_hours': 1.0,
                'is_friday_work': False,
                'is_mission': True,
                'advance_payment': 0,
                'notes': 'ماموریت خارج شهر'
            }]
        }
        res_save = self.client.post('/api/personnel/attendance/bulk-save/', payload, format='json')
        self.assertEqual(res_save.status_code, 200)

        # بررسی در شیت ماهانه مستقل از انبار
        year_month = f"{today.year:04d}/{today.month:02d}"
        res_monthly = self.client.get(f'/api/personnel/attendance/monthly-grid/?year_month={year_month}')
        self.assertEqual(res_monthly.status_code, 200)
        p_row = next(r for r in res_monthly.data['rows'] if r['personnel_id'] == self.personnel.id)
        self.assertEqual(p_row['total_hours'], 10.0)

    def test_admin_override_date_window(self):
        """تست دور زدن بازه تاریخی توسط مدیر ارشد با فلگ is_override"""
        import jdatetime
        past_date = (jdatetime.date.today() - jdatetime.timedelta(days=20)).strftime('%Y/%m/%d')
        payload = {
            'warehouse_id': self.warehouse.id,
            'date_shamsi': past_date,
            'is_override': True,
            'items': [{
                'personnel_id': self.personnel.id,
                'status': 'PRESENT_10H',
                'effective_hours': 10,
                'overtime_hours': 0,
                'is_friday_work': False,
                'is_mission': False
            }]
        }
        res = self.client.post('/api/personnel/attendance/bulk-save/', payload, format='json')
        self.assertEqual(res.status_code, 200)

    def test_monthly_grid_save_with_existing_past_records(self):
        """تست اینکه رکوردهای ثبت‌شده گذشته در شیت ماهانه مانع ذخیره روزهای جاری نمی‌شوند"""
        import jdatetime
        today = jdatetime.date.today()
        past_day = 1
        past_date = f"{today.year:04d}/{today.month:02d}/{past_day:02d}"
        
        # ثبت رکورد قدیمی از قبل
        DailyAttendance.objects.create(
            personnel=self.personnel,
            warehouse=self.warehouse,
            date_shamsi=past_date,
            status='PRESENT_10H',
            effective_hours=10,
            overtime_hours=0
        )

        year_month = f"{today.year:04d}/{today.month:02d}"
        payload = {
            'warehouse_id': self.warehouse.id,
            'year_month': year_month,
            'items': [
                # رکورد گذشته دست‌نخورده (Unchanged)
                {
                    'personnel_id': self.personnel.id,
                    'day': past_day,
                    'status': 'PRESENT_10H',
                    'effective_hours': 10,
                    'overtime_hours': 0,
                    'is_friday_work': False,
                    'is_mission': False,
                    'advance_payment': 0,
                    'notes': ''
                },
                # رکورد امروز (جدید و مجاز)
                {
                    'personnel_id': self.personnel.id,
                    'day': today.day,
                    'status': 'PRESENT_10H',
                    'effective_hours': 10,
                    'overtime_hours': 2,
                    'is_friday_work': False,
                    'is_mission': False,
                    'advance_payment': 0,
                    'notes': 'ثبت کارکرد امروز'
                }
            ]
        }
        res = self.client.post('/api/personnel/attendance/bulk-save-monthly-grid/', payload, format='json')
        self.assertEqual(res.status_code, 200)

    def test_lock_enforcement_in_all_warehouses_mode(self):
        """تست قفل بودن انبار پرسنل حتی در حالت فیلتر تمام انبارها (warehouse_id=None)"""
        import jdatetime
        today = jdatetime.date.today()
        year_month = f"{today.year:04d}/{today.month:02d}"
        today_str = f"{today.year:04d}/{today.month:02d}/{today.day:02d}"

        # قفل کردن دوره برای انبار پرسنل
        period, _ = MonthlyWorkPeriod.objects.get_or_create(
            warehouse=self.warehouse,
            year_month=year_month,
            defaults={'status': 'LOCKED'}
        )
        period.status = 'LOCKED'
        period.save()

        # تلاش برای ثبت بدون warehouse_id
        payload = {
            'date_shamsi': today_str,
            'items': [{
                'personnel_id': self.personnel.id,
                'status': 'PRESENT_10H',
                'effective_hours': 10,
                'overtime_hours': 0,
                'is_friday_work': False,
                'is_mission': False
            }]
        }
        res = self.client.post('/api/personnel/attendance/bulk-save/', payload, format='json')
        self.assertEqual(res.status_code, 400)
        self.assertIn('قفل', str(res.json()))

    def test_two_step_approval_workflow(self):
        """تست گردش کار دو مرحله‌ای تایید کارکرد (ارسال، تایید، رد و بازگشایی)"""
        import jdatetime
        today = jdatetime.date.today()
        year_month = f"{today.year:04d}/{today.month:02d}"

        # ۱. سرپرست ارسال می‌کند (submit)
        res_submit = self.client.post('/api/personnel/periods/workflow-action/', {
            'warehouse_id': self.warehouse.id,
            'year_month': year_month,
            'action': 'submit',
            'notes': 'کارکرد تیرماه آماده بررسی مالی است'
        }, format='json')
        self.assertEqual(res_submit.status_code, 200)
        self.assertEqual(res_submit.data['period']['status'], 'SUBMITTED')

        # ۲. مدیر مالی رد می‌کند با ذکر علت (reject)
        res_reject = self.client.post('/api/personnel/periods/workflow-action/', {
            'warehouse_id': self.warehouse.id,
            'year_month': year_month,
            'action': 'reject',
            'notes': 'اضافه‌کار روز ۱۵ نامعتبر است'
        }, format='json')
        self.assertEqual(res_reject.status_code, 200)
        self.assertEqual(res_reject.data['period']['status'], 'REJECTED')
        self.assertEqual(res_reject.data['period']['rejection_reason'], 'اضافه‌کار روز ۱۵ نامعتبر است')

        # ۳. سرپرست اصلاح کرده و دوباره ارسال می‌کند
        res_submit2 = self.client.post('/api/personnel/periods/workflow-action/', {
            'warehouse_id': self.warehouse.id,
            'year_month': year_month,
            'action': 'submit',
            'notes': 'اصلاح انجام شد'
        }, format='json')
        self.assertEqual(res_submit2.status_code, 200)
        self.assertEqual(res_submit2.data['period']['status'], 'SUBMITTED')

        # ۴. مدیر مالی تایید نهایی و قفل می‌کند (approve)
        res_approve = self.client.post('/api/personnel/periods/workflow-action/', {
            'warehouse_id': self.warehouse.id,
            'year_month': year_month,
            'action': 'approve'
        }, format='json')
        self.assertEqual(res_approve.status_code, 200)
        self.assertEqual(res_approve.data['period']['status'], 'LOCKED')

    def test_export_monthly_grid_excel(self):
        """تست خروجی فایل اکسل استاندارد دو سطری شیت ماهانه"""
        import jdatetime
        today = jdatetime.date.today()
        year_month = f"{today.year:04d}/{today.month:02d}"

        res = self.client.get(f'/api/personnel/attendance/export-monthly-excel/?warehouse_id={self.warehouse.id}&year_month={year_month}')
        self.assertEqual(res.status_code, 200)
        self.assertEqual(res['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.assertIn('Timesheet_', res['Content-Disposition'])

    def test_monthly_grid_anomaly_detection(self):
        """تست شناسایی خودکار ناهنجاری‌ها مانند کارکرد بیش از ۱۶ ساعت در شیت ماهانه"""
        import jdatetime
        today = jdatetime.date.today()
        year_month = f"{today.year:04d}/{today.month:02d}"
        date_str = f"{year_month}/10"

        # ثبت کارکرد ۱۸ ساعته (بیش از سقف ۱۶ ساعت)
        DailyAttendance.objects.create(
            personnel=self.personnel,
            warehouse=self.warehouse,
            date_shamsi=date_str,
            status='PRESENT_10H',
            effective_hours=10,
            overtime_hours=8  # ۱۰ + ۸ = ۱۸ ساعت
        )

        res = self.client.get(f'/api/personnel/attendance/monthly-grid/?warehouse_id={self.warehouse.id}&year_month={year_month}')
        self.assertEqual(res.status_code, 200)
        anomalies = res.data.get('anomalies', [])
        self.assertTrue(any(a['type'] == 'EXCESSIVE_HOURS' for a in anomalies))

    def test_import_monthly_excel_leading_zeros_and_audit(self):
        """تست درون‌ریزی اکسل با کد ملی دارای صفر ابتدایی، انتساب پریود و ثبت لاگ ممیزی"""
        import io
        import openpyxl
        import jdatetime

        today = jdatetime.date.today()
        year_month = f"{today.year:04d}/{today.month:02d}"

        # به پرسنل تست یک کد ملی با صفر اول می‌دهیم
        self.personnel.national_code = '0012345678'
        self.personnel.save()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value='ردیف')
        ws.cell(row=1, column=2, value='کد ملی')
        ws.cell(row=1, column=3, value='نام پرسنل')
        ws.cell(row=1, column=4, value='سمت')
        ws.cell(row=1, column=5, value='انبار')
        ws.cell(row=1, column=6, value='روز ۰۱')

        ws.cell(row=2, column=1, value='row_num')
        ws.cell(row=2, column=2, value='national_code')
        ws.cell(row=2, column=3, value='full_name')
        ws.cell(row=2, column=4, value='job_title')
        ws.cell(row=2, column=5, value='warehouse_name')
        ws.cell(row=2, column=6, value='day_01')

        # ردیف ۳: اکسل صفر اول را می‌اندازد و ۱۲۳۴۵۶۷۸ ذخیره می‌کند
        ws.cell(row=3, column=1, value=1)
        ws.cell(row=3, column=2, value=12345678)  # عددی بدون صفر اول
        ws.cell(row=3, column=3, value=self.personnel.full_name)
        ws.cell(row=3, column=4, value='کارگر')
        ws.cell(row=3, column=5, value=self.warehouse.name)
        ws.cell(row=3, column=6, value=10)

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)

        from django.core.files.uploadedfile import SimpleUploadedFile
        excel_file = SimpleUploadedFile(
            'timesheet.xlsx',
            stream.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        res = self.client.post('/api/personnel/attendance/import-monthly-excel/', {
            'file': excel_file,
            'year_month': year_month,
            'warehouse_id': self.warehouse.id,
            'is_override': 'true'
        }, format='multipart')

        self.assertEqual(res.status_code, 200)
        self.assertEqual(res.data['matched_personnel_count'], 1)
        self.assertEqual(res.data['updated_cells'], 1)

        # بررسی ایجاد رکورد، انتساب period و ثبت لاگ ممیزی
        att = DailyAttendance.objects.filter(personnel=self.personnel, date_shamsi=f"{year_month}/01").first()
        self.assertIsNotNone(att)
        self.assertIsNotNone(att.period)
        self.assertEqual(float(att.effective_hours), 10.0)

        # بررسی ثبت لاگ
        logs = AttendanceAuditLog.objects.filter(attendance=att)
        self.assertTrue(logs.exists())

    def test_rejected_period_date_window_bypass(self):
        """تست لغو سقف روزهای گذشته در صورت REJECTED بودن دوره کارکرد"""
        from personnel.views import validate_attendance_date_window
        import jdatetime
        today = jdatetime.date.today()
        # تاریخی در ۲۰ روز قبل
        past_date = today - jdatetime.timedelta(days=20)
        year_month = f"{past_date.year:04d}/{past_date.month:02d}"
        date_str = f"{year_month}/{past_date.day:02d}"

        # دوره‌ای با وضعیت REJECTED ایجاد می‌کنیم
        MonthlyWorkPeriod.objects.create(
            warehouse=self.warehouse,
            year_month=year_month,
            status='REJECTED',
            rejection_reason='لطفا اصلاح فرمایید'
        )

        # نباید خطای سقف روز گذشته بدهد
        is_valid = validate_attendance_date_window(date_str, user=self.user, warehouse_id=self.warehouse.id)
        self.assertTrue(is_valid)











