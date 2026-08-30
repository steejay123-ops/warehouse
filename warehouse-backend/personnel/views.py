from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.http import HttpResponse
from django.db import transaction
from django.db.models import Sum, Count, Q
from django.utils import timezone
import io
import zipfile
import openpyxl
from decimal import Decimal
from common.date_utils import format_to_shamsi_str, normalize_digits
import jdatetime

from .models import (
    PersonnelProfile,
    VehicleDriverProfile,
    MonthlyWorkPeriod,
    DailyAttendance,
    AttendanceAuditLog,
    VehicleTripLog,
    PayrollYearlySettings,
    JobGradeTier,
    WorkshopInsuranceSettings,
    TaxRuleSettings,
    BankExportSettings,
    MonthlyPayrollRecord
)
from .serializers import (
    PersonnelProfileSerializer,
    VehicleDriverProfileSerializer,
    DailyAttendanceSerializer,
    BulkAttendanceMatrixSerializer,
    BulkMonthlyAttendanceGridSerializer,
    MonthlyGridItemInputSerializer,
    AttendanceAuditLogSerializer,
    VehicleTripLogSerializer,
    BulkVehicleTripMatrixSerializer,
    MonthlyWorkPeriodSerializer,
    PayrollYearlySettingsSerializer,
    JobGradeTierSerializer,
    WorkshopInsuranceSettingsSerializer,
    TaxRuleSettingsSerializer,
    BankExportSettingsSerializer,
    MonthlyPayrollRecordSerializer
)
from .payroll_engine import calculate_monthly_payroll_for_period
from .dbf_generator import generate_dskkar_bytes, generate_dskwor_bytes
from .tax_bank_exporter import generate_wh_tax_content, generate_wp_tax_content, generate_bank_meli_excel
from .payroll_excel_exporter import generate_monthly_payroll_excel


class PersonnelProfileViewSet(viewsets.ModelViewSet):
    queryset = PersonnelProfile.objects.all().select_related('assigned_warehouse', 'user')
    serializer_class = PersonnelProfileSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = None

    def get_queryset(self):
        qs = super().get_queryset()
        warehouse_id = self.request.query_params.get('warehouse_id')
        if warehouse_id:
            qs = qs.filter(Q(assigned_warehouse_id=warehouse_id) | Q(assigned_warehouse__isnull=True))
        
        is_active = self.request.query_params.get('is_active')
        if is_active is not None:
            qs = qs.filter(is_active=is_active.lower() in ['true', '1'])
            
        search = self.request.query_params.get('search')
        if search:
            qs = qs.filter(
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(national_code__icontains=search) |
                Q(job_title__icontains=search)
            )
        return qs

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    def destroy(self, request, *args, **kwargs):
        """
        حذف پرسنل:
        اگر پرسنل دارای کارکرد ثبت‌شده یا فیش حقوقی باشد، جهت حفظ یکپارچگی مالی
        امکان حذف مستقیم وجود ندارد (باید غیرفعال شود).
        اما اگر هنوز هیچ کارکردی برای این فرد ثبت نشده است، اجازه حذف کامل داده می‌شود.
        """
        instance = self.get_object()
        has_attendance = DailyAttendance.objects.filter(personnel=instance, is_deleted=False).exists()
        has_payroll = MonthlyPayrollRecord.objects.filter(personnel=instance).exists()

        if has_attendance or has_payroll:
            return Response(
                {'error': f'پرسنل «{instance.full_name}» دارای سابقه کارکرد یا حقوق ثبت‌شده است و نمی‌توان پرونده او را حذف کرد. می‌توانید وضعیت او را غیرفعال نمایید.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        full_name = instance.full_name
        instance.delete()
        return Response({'message': f'پرسنل «{full_name}» با موفقیت از سامانه حذف گردید.'}, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post', 'POST', 'get', 'GET'], url_path='import-excel')
    def import_excel(self, request):
        """
        درون‌ریزی مستقیم شیت Emp_info از فایل اکسل شرکت (Upsert بر مبنای کد ملی ۱۰ رقمی)
        """
        import openpyxl
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'error': 'فایل اکسل الزامی است.'}, status=status.HTTP_400_BAD_REQUEST)
        
        warehouse_id = request.data.get('warehouse_id') or request.query_params.get('warehouse_id')
        
        try:
            wb = openpyxl.load_workbook(file_obj, data_only=True)
        except Exception as e:
            return Response({'error': f'خطا در باز کردن فایل اکسل: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)
        
        target_sheet = None
        for name in wb.sheetnames:
            if name.strip().lower() in ['emp_info', 'emp info', 'empinfo', 'پرسنل']:
                target_sheet = wb[name]
                break
        if not target_sheet:
            target_sheet = wb.active

        ws = target_sheet
        
        # پیدا کردن ردیف هدر
        header_map = {}
        header_row_idx = 1
        for r_idx in range(1, min(ws.max_row + 1, 5)):
            row_vals = [ws.cell(row=r_idx, column=c).value for c in range(1, ws.max_column + 1)]
            row_str = ' '.join([str(v) for v in row_vals if v is not None])
            if 'کد ملی' in row_str or 'نام خانوادگی' in row_str:
                header_row_idx = r_idx
                break

        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=header_row_idx, column=col_idx).value
            if val:
                cleaned = str(val).strip().replace('  ', ' ')
                header_map[cleaned] = col_idx

        if 'کد ملی' not in header_map and 'نام خانوادگی' not in header_map:
            return Response({'error': 'شیت پرسنلی یا ستون‌های الزامی (کد ملی، نام خانوادگی) در فایل یافت نشد.'}, status=status.HTTP_400_BAD_REQUEST)

        created_count = 0
        updated_count = 0
        errors = []

        with transaction.atomic():
            for r in range(header_row_idx + 1, ws.max_row + 1):
                nat_col = header_map.get('کد ملی')
                if not nat_col:
                    continue
                raw_nc = ws.cell(row=r, column=nat_col).value
                if raw_nc is None or str(raw_nc).strip() == '':
                    continue
                
                nc_clean = str(raw_nc).split('.')[0].strip().zfill(10)
                if not nc_clean.isdigit() or len(nc_clean) != 10:
                    errors.append(f'ردیف {r}: کد ملی نامعتبر است ({raw_nc})')
                    continue

                def get_v(key, default=''):
                    c_idx = header_map.get(key)
                    if c_idx:
                        val = ws.cell(row=r, column=c_idx).value
                        if val is not None:
                            return str(val).strip()
                    return default

                def get_num(key, default=0):
                    c_idx = header_map.get(key)
                    if c_idx:
                        val = ws.cell(row=r, column=c_idx).value
                        if val is not None:
                            try:
                                return float(val)
                            except (ValueError, TypeError):
                                pass
                    return default

                def get_bool(key, default=True):
                    c_idx = header_map.get(key)
                    if c_idx:
                        val = ws.cell(row=r, column=c_idx).value
                        if val is not None:
                            return str(val).strip() in ['1', 'true', 'True', 'بله']
                    return default

                first_name = get_v('نام')
                last_name = get_v('نام خانوادگی')
                if not last_name:
                    continue

                profile_defaults = {
                    'first_name': first_name,
                    'last_name': last_name,
                    'father_name': get_v('نام پدر') or get_v('نام  پدر'),
                    'birth_date': get_v('تاریخ تولد'),
                    'id_number': get_v('شماره شناسنامه'),
                    'birth_place': get_v('محل تولد'),
                    'education_level': get_v('مدرک تحصیلی', '5'),
                    'insurance_type': get_v('نوع بیمه', '2'),
                    'insurance_number': get_v('شماره بیمه'),
                    'insurance_name': get_v('نام بیمه', 'تامین اجتماعی'),
                    'exemption_type': get_v('نوع معافیت', '1'),
                    'job_category': get_v('رسته', '15'),
                    'job_title': get_v('سمت شغل') or get_v('عنوان شغل') or 'پرسنل انبار',
                    'employment_type': get_v('نوع استخدام', '2'),
                    'start_date': get_v('تاریخ شروع به کار'),
                    'end_date': get_v('تاریخ پایان کار'),
                    'retirement_date': get_v('تاریخ بازنشستگی'),
                    'job_code': get_v('کد شغل').zfill(6) if get_v('کد شغل') else '',
                    'id_series': get_v('مسلسل شناسنامه'),
                    'id_serial': get_v('سریال'),
                    'issue_place': get_v('محل صدور'),
                    'issue_date': get_v('تاریخ صدور'),
                    'gender': get_v('جنسیت', 'مرد'),
                    'phone_number': get_v('موبایل') or get_v('شماره همراه') or get_v('شماره تماس'),
                    'account_number': get_v('شماره حساب'),
                    'sheba_number': get_v('شماره شبا') or get_v('شبا'),
                    'bank_name': get_v('نام بانک') or get_v('بانک'),
                    'address': get_v('آدرس') or get_v('ادرس') or get_v('آدرس محل سکونت'),
                    'postal_code': get_v('کد پستی') or get_v('کدپستی'),
                    'marital_status': 'married' if 'متاهل' in get_v('وضعیت تاهل') or 'متأهل' in get_v('وضعیت تاهل') else 'single',
                    'children_count': int(get_num('تعداد فرزند', 0)),
                    'contract_hours': get_num('ساعت کار قرارداد شده', 230),
                    'contract_base_salary': get_num('حقوق قرارداد شده', 0),
                    'job_grade': get_v('گروه شغلی', '19'),
                    'daily_base_wage': get_num('مزد روزانه', 0),
                    'daily_seniority_bonus': get_num('پایه سنواتی', 0),
                    'base_daily_rate': get_num('مزد مبنا', 0),
                    'base_years_experience': int(get_num('تعداد سال کارکرد', 0)),
                    'status_category': get_v('وضعیت', 'نفرات شرکتی'),
                    'group_status': get_v('گروه', 'شاغل'),
                    'include_in_tax': get_bool('maliat', True),
                    'include_in_insurance': get_bool('بیمه', True),
                    'include_in_bank': get_bool('bank', True),
                    'created_by': request.user if request.user and request.user.is_authenticated else None
                }
                if warehouse_id and str(warehouse_id).isdigit():
                    profile_defaults['assigned_warehouse_id'] = int(warehouse_id)

                obj, created = PersonnelProfile.objects.update_or_create(
                    national_code=nc_clean,
                    defaults=profile_defaults
                )
                if created:
                    created_count += 1
                else:
                    updated_count += 1

        return Response({
            'message': f'درون‌ریزی پرسنل با موفقیت انجام شد ({created_count} پرسنل جدید، {updated_count} به‌روزرسانی).',
            'created_count': created_count,
            'updated_count': updated_count,
            'errors': errors
        })


class VehicleDriverProfileViewSet(viewsets.ModelViewSet):
    queryset = VehicleDriverProfile.objects.all().select_related('assigned_warehouse', 'user')
    serializer_class = VehicleDriverProfileSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = None

    def get_queryset(self):
        qs = super().get_queryset()
        warehouse_id = self.request.query_params.get('warehouse_id')
        if warehouse_id:
            qs = qs.filter(Q(assigned_warehouse_id=warehouse_id) | Q(assigned_warehouse__isnull=True))
            
        is_active = self.request.query_params.get('is_active')
        if is_active is not None:
            qs = qs.filter(is_active=is_active.lower() in ['true', '1'])
            
        search = self.request.query_params.get('search')
        if search:
            qs = qs.filter(
                Q(driver_name__icontains=search) |
                Q(plate_number__icontains=search) |
                Q(driver_national_code__icontains=search)
            )
        return qs

def normalize_attendance_date(date_str: str) -> str:
    """
    نرمال‌سازی تاریخ شمسی و پشتیبانی کامل از ورودی‌های پیوسته ۸ رقمی (14050607)
    و تبدیل خودکار به استاندارد YYYY/MM/DD بدون بروز خطا
    """
    if not date_str:
        return ""
    cleaned = normalize_digits(str(date_str)).strip()
    digits_only = ''.join(c for c in cleaned if c.isdigit())
    if len(digits_only) == 8 and ('/' not in cleaned and '-' not in cleaned):
        return f"{digits_only[:4]}/{digits_only[4:6]}/{digits_only[6:8]}"
    shamsi_str = format_to_shamsi_str(cleaned)
    if shamsi_str and len(shamsi_str.split('/')) == 3:
        return shamsi_str
    parts = cleaned.replace('-', '/').split('/')
    if len(parts) == 3:
        try:
            return f"{int(parts[0]):04d}/{int(parts[1]):02d}/{int(parts[2]):02d}"
        except Exception:
            pass
    return date_str


def validate_attendance_date_window(date_shamsi: str, user=None, allow_override: bool = False, warehouse_id=None):
    """
    اعتبارسنجی بازه مجاز ثبت و ویرایش کارکرد بر مبنای تنظیمات سالانه حقوق
    -1 به معنای نامحدود، 0 به معنای فقط امروز، اعداد مثبت به معنای سقف روز مجاز
    پشتیبانی کامل از رشته‌های ۸ رقمی پیوسته (مانند 14050607)
    مدیر ارشد سیستم (Superuser) با فلگ allow_override می‌تواند محدودیت زمانی را دور بزند.
    در صورتی که دوره کارکرد در وضعیت REJECTED باشد، جهت امکان اصلاح مغایرت‌ها توسط سرپرست، سقف روزهای گذشته لغو می‌شود.
    """
    if allow_override and user and (getattr(user, 'is_superuser', False) or getattr(user, 'is_staff', False) or user.has_perm('personnel.can_override_attendance_lock')):
        return True

    import jdatetime
    from rest_framework.exceptions import ValidationError as DRFValidationError
    
    if not date_shamsi:
        raise DRFValidationError("تاریخ شمسی معتبر نیست.")
    
    norm_date = normalize_attendance_date(date_shamsi)
    try:
        parts = [int(p) for p in norm_date.split('/')]
        if len(parts) != 3:
            raise ValueError
        target_date = jdatetime.date(parts[0], parts[1], parts[2])
    except Exception:
        raise DRFValidationError(f"فرمت تاریخ شمسی نامعتبر است: {date_shamsi}")

    today = jdatetime.date.today()
    delta_days = (target_date - today).days

    fiscal_year = str(parts[0])
    settings = PayrollYearlySettings.objects.filter(fiscal_year=fiscal_year, is_active=True).first()
    if not settings:
        settings = PayrollYearlySettings.objects.filter(is_active=True).first()

    past_limit = settings.attendance_edit_past_days if settings else 3
    future_limit = settings.attendance_edit_future_days if settings else 0

    # بررسی روزهای گذشته
    if delta_days < 0:
        year_month = f"{parts[0]:04d}/{parts[1]:02d}"
        rej_period_qs = MonthlyWorkPeriod.objects.filter(year_month=year_month, status='REJECTED')
        if warehouse_id:
            rej_period_qs = rej_period_qs.filter(warehouse_id=warehouse_id)
        if rej_period_qs.exists():
            return True

        past_days = abs(delta_days)
        if past_limit is not None and past_limit != -1:
            if past_days > past_limit:
                if past_limit == 0:
                    raise DRFValidationError(f"طبق تنظیمات سیستم، فقط امکان ثبت کارکرد برای تاریخ امروز ({today.strftime('%Y/%m/%d')}) وجود دارد و ویرایش سوابق گذشته مجاز نیست.")
                raise DRFValidationError(f"ویرایش کارکرد برای {past_days} روز قبل مجاز نیست. حداکثر سقف مجاز طبق تنظیمات مدیر {past_limit} روز است.")

    # بررسی روزهای آینده
    elif delta_days > 0:
        future_days = delta_days
        if future_limit is not None and future_limit != -1:
            if future_days > future_limit:
                if future_limit == 0:
                    raise DRFValidationError(f"ثبت کارکرد برای تاریخ‌های آینده ({norm_date}) مجاز نیست.")
                raise DRFValidationError(f"ثبت کارکرد برای {future_days} روز آینده مجاز نیست. حداکثر سقف مجاز طبق تنظیمات مدیر {future_limit} روز است.")

    return True


def is_date_shamsi_friday(date_shamsi_str):
    """
    بررسی جمعه بودن یک تاریخ شمسی
    """
    try:
        import jdatetime
        cleaned = normalize_digits(str(date_shamsi_str)).strip().replace('-', '/')
        parts = [int(p) for p in cleaned.split('/') if p.strip()]
        if len(parts) >= 3:
            return jdatetime.date(parts[0], parts[1], parts[2]).weekday() == 6
    except Exception:
        pass
    return False


def broadcast_attendance_updated(warehouse_id=None, date_shamsi=None, year_month=None, message=None, sender_id=None, client_tab_id=None, extra_data=None):
    """
    ارسال بلادرنگ رویداد ثبت/ویرایش کارکرد به کانال وب‌سوکت سراسری
    """
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        channel_layer = get_channel_layer()
        if channel_layer is not None:
            payload = {
                'type': 'send_notification',
                'type_str': 'attendance_updated',
                'message': message or 'کارکرد پرسنل به‌روزرسانی شد.',
                'date_shamsi': date_shamsi,
                'year_month': year_month,
                'sender_id': sender_id,
                'client_tab_id': client_tab_id,
            }
            if warehouse_id is not None:
                payload['warehouse_id'] = warehouse_id
            if extra_data and isinstance(extra_data, dict):
                payload.update(extra_data)

            async_to_sync(channel_layer.group_send)(
                'global_notifications',
                payload
            )
    except Exception as e:
        import logging
        logging.getLogger(__name__).warning(f"[WebSocket] Error broadcasting attendance_updated: {e}")


class DailyAttendanceViewSet(viewsets.ModelViewSet):
    queryset = DailyAttendance.objects.filter(is_deleted=False).select_related('personnel', 'warehouse', 'period')
    serializer_class = DailyAttendanceSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = None

    def get_queryset(self):
        qs = super().get_queryset()
        warehouse_id = self.request.query_params.get('warehouse_id')
        if warehouse_id:
            qs = qs.filter(warehouse_id=warehouse_id)
        date_shamsi = self.request.query_params.get('date_shamsi')
        if date_shamsi:
            qs = qs.filter(date_shamsi=date_shamsi)
        personnel_id = self.request.query_params.get('personnel_id')
        if personnel_id:
            qs = qs.filter(personnel_id=personnel_id)
        return qs

    @action(detail=False, methods=['get'], url_path='matrix')
    def get_matrix(self, request):
        """
        دریافت داده‌های ماتریسی حضور و غیاب (پشتیبانی از مستقل از انبار یا فیلتر انبار خاص)
        همراه با استخراج هوشمند آخرین وضعیت ثبت‌شده فرد در روز قبل به عنوان پیش‌فرض
        """
        raw_wh = request.query_params.get('warehouse_id')
        warehouse_id = int(raw_wh) if (raw_wh and str(raw_wh).upper() not in ['ALL', '0', 'NONE', '']) else None
        raw_date = request.query_params.get('date_shamsi')
        if not raw_date:
            return Response({'error': 'پارامتر date_shamsi الزامی است.'}, status=status.HTTP_400_BAD_REQUEST)

        date_shamsi = normalize_attendance_date(raw_date)
        year_month = date_shamsi[:7]

        # بررسی وضعیت قفل دوره
        if warehouse_id:
            period = MonthlyWorkPeriod.objects.filter(warehouse_id=warehouse_id, year_month=year_month).first()
            is_locked = period.status == 'LOCKED' if period else False
            period_status = period.status if period else 'OPEN'
        else:
            is_locked = False
            period_status = 'OPEN'

        # لیست پرسنل فعال یا پرسنلی که در این تاریخ کارکرد دارند (مستقل از انبار یا فیلتر انبار خاص)
        if warehouse_id:
            personnel_list = PersonnelProfile.objects.filter(
                Q(assigned_warehouse_id=warehouse_id) | Q(assigned_warehouse__isnull=True),
                Q(is_active=True) | Q(daily_attendances__date_shamsi=date_shamsi, daily_attendances__is_deleted=False)
            ).distinct().select_related('assigned_warehouse').order_by('last_name', 'first_name')
            attendances_qs = DailyAttendance.objects.filter(
                Q(warehouse_id=warehouse_id) | Q(personnel__in=personnel_list),
                date_shamsi=date_shamsi,
                is_deleted=False
            ).select_related('personnel', 'warehouse').order_by('id')
            # اولویت با رکوردهایی است که اختصاصاً برای همین انبار ثبت شده‌اند
            existing_attendances = {}
            for att in attendances_qs:
                if att.personnel_id not in existing_attendances or att.warehouse_id == warehouse_id:
                    existing_attendances[att.personnel_id] = att
        else:
            personnel_list = PersonnelProfile.objects.filter(
                Q(is_active=True) | Q(daily_attendances__date_shamsi=date_shamsi, daily_attendances__is_deleted=False)
            ).distinct().select_related('assigned_warehouse').order_by('last_name', 'first_name')
            attendances_qs = DailyAttendance.objects.filter(
                date_shamsi=date_shamsi,
                is_deleted=False
            ).select_related('personnel', 'warehouse').order_by('id')
            existing_attendances = {att.personnel_id: att for att in attendances_qs}

        is_today_friday = is_date_shamsi_friday(date_shamsi)

        matrix_rows = []
        for p in personnel_list:
            att = existing_attendances.get(p.id)
            if att:
                matrix_rows.append({
                    'personnel_id': p.id,
                    'full_name': p.full_name,
                    'national_code': p.national_code,
                    'job_title': p.job_title,
                    'is_active': p.is_active,
                    'warehouse_id': att.warehouse_id,
                    'warehouse_name': att.warehouse.name if att.warehouse else (p.assigned_warehouse.name if p.assigned_warehouse else 'شناور'),
                    'attendance_id': att.id,
                    'status': att.status,
                    'effective_hours': float(att.effective_hours),
                    'overtime_hours': float(att.overtime_hours),
                    'is_friday_work': att.is_friday_work,
                    'is_mission': att.is_mission,
                    'advance_payment': float(att.advance_payment),
                    'notes': att.notes or '',
                    'is_existing': True
                })
            else:
                # روزهای ثبت‌نشده یا پاکسازی‌شده با وضعیت خالی و ساعت ۰ بازگردانده می‌شوند
                matrix_rows.append({
                    'personnel_id': p.id,
                    'full_name': p.full_name,
                    'national_code': p.national_code,
                    'job_title': p.job_title,
                    'is_active': p.is_active,
                    'warehouse_id': p.assigned_warehouse_id,
                    'warehouse_name': p.assigned_warehouse.name if p.assigned_warehouse else 'شناور',
                    'attendance_id': None,
                    'status': '',
                    'effective_hours': 0.0,
                    'overtime_hours': 0.0,
                    'is_friday_work': is_today_friday,
                    'is_mission': False,
                    'advance_payment': 0.0,
                    'notes': '',
                    'is_existing': False
                })

        return Response({
            'warehouse_id': warehouse_id,
            'date_shamsi': date_shamsi,
            'year_month': year_month,
            'is_locked': is_locked,
            'period_status': period_status,
            'rows': matrix_rows
        })

    @action(detail=False, methods=['post'], url_path='bulk-save')
    def bulk_save(self, request):
        """
        ذخیره گروهی و سریع ماتریس حضور و غیاب پرسنل در پایان روز
        (پشتیبانی کامل از کارکرد مستقل از انبار و لاگ ممیزی)
        """
        serializer = BulkAttendanceMatrixSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        raw_wh = serializer.validated_data.get('warehouse_id')
        warehouse_id = int(raw_wh) if (raw_wh and str(raw_wh).upper() not in ['ALL', '0', 'NONE', '']) else None
        raw_date = serializer.validated_data['date_shamsi']
        date_shamsi = normalize_attendance_date(raw_date)
        items = serializer.validated_data['items']

        allow_override = bool(request.data.get('is_override', False) or request.headers.get('X-Admin-Override') == 'true')
        is_admin_override = allow_override and (getattr(request.user, 'is_superuser', False) or request.user.has_perm('personnel.can_override_attendance_lock'))

        # اعتبارسنجی بازه زمانی مجاز بر مبنای تنظیمات مدیر
        try:
            validate_attendance_date_window(date_shamsi, user=request.user, allow_override=is_admin_override, warehouse_id=warehouse_id)
        except Exception as e:
            err_msg = getattr(e, 'detail', str(e))
            if isinstance(err_msg, list):
                err_msg = err_msg[0]
            return Response({'error': str(err_msg)}, status=status.HTTP_400_BAD_REQUEST)

        year_month = date_shamsi[:7]
        period = None
        if warehouse_id:
            period, _ = MonthlyWorkPeriod.objects.get_or_create(
                warehouse_id=warehouse_id,
                year_month=year_month,
                defaults={'status': 'OPEN'}
            )
            if period.status == 'LOCKED' and not is_admin_override:
                return Response({'error': f'دوره کارکرد {year_month} قفل شده است و امکان ثبت یا ویرایش وجود ندارد.'}, status=status.HTTP_400_BAD_REQUEST)

        saved_count = 0
        updated_count = 0

        with transaction.atomic():
            for item in items:
                personnel_id = item['personnel_id']
                p_obj = PersonnelProfile.objects.filter(id=personnel_id).first()
                if not p_obj:
                    continue

                item_wh = item.get('warehouse_id')
                check_wh = item_wh if item_wh is not None else (warehouse_id if warehouse_id else p_obj.assigned_warehouse_id)
                target_period = period
                if not target_period and check_wh:
                    target_period = MonthlyWorkPeriod.objects.filter(
                        warehouse_id=check_wh,
                        year_month=year_month
                    ).first()
                if target_period and target_period.status == 'LOCKED' and not is_admin_override:
                    return Response({'error': f'دوره کارکرد {year_month} برای انبار پرسنل ({p_obj.full_name}) قفل شده است.'}, status=status.HTTP_400_BAD_REQUEST)

                target_wh = item_wh if item_wh is not None else (warehouse_id if warehouse_id else p_obj.assigned_warehouse_id)

                status_val = item['status']
                eff_h = float(item.get('effective_hours', 0))
                ot_h = float(item.get('overtime_hours', 0))
                is_fri = bool(item.get('is_friday_work', False))
                is_mis = bool(item.get('is_mission', False))
                adv_pay = float(item.get('advance_payment', 0))
                notes_val = item.get('notes', '') or ''

                # اگر وضعیت خالی و ساعت صفر باشد (پاکسازی‌شده)، رکورد دیتابیس باید حذف گردد
                if not status_val and eff_h == 0:
                    att_obj = DailyAttendance.objects.select_for_update().filter(
                        personnel_id=personnel_id,
                        date_shamsi=date_shamsi,
                        is_deleted=False
                    ).first()
                    if att_obj:
                        att_obj.is_deleted = True
                        att_obj.modified_by = request.user
                        att_obj.save()
                        AttendanceAuditLog.objects.create(
                            attendance=att_obj,
                            personnel_name=p_obj.full_name,
                            date_shamsi=date_shamsi,
                            changed_by=request.user,
                            field_name='is_deleted',
                            old_value='False',
                            new_value='True',
                            reason='پاکسازی کارکرد روزانه'
                        )
                        updated_count += 1
                    continue

                # اعتبارسنجی و ریست منطقی ساعات در صورت غیبت یا مرخصی
                if status_val in ['ABSENT', 'LEAVE']:
                    eff_h = 0.0
                    ot_h = 0.0
                    is_fri = False
                    is_mis = False

                # پیدا کردن رکورد فعال پرسنل در این تاریخ (مستقل از انبار قبلی)
                att_obj = DailyAttendance.objects.select_for_update().filter(
                    personnel_id=personnel_id,
                    date_shamsi=date_shamsi,
                    is_deleted=False
                ).first()

                if att_obj:
                    # بررسی تغییرات جهت ثبت در Audit Log
                    changes = []
                    item_sanitized = {
                        'status': status_val,
                        'effective_hours': eff_h,
                        'overtime_hours': ot_h,
                        'is_friday_work': is_fri,
                        'is_mission': is_mis,
                        'advance_payment': adv_pay,
                        'notes': notes_val
                    }
                    fields_to_check = ['status', 'effective_hours', 'overtime_hours', 'is_friday_work', 'is_mission', 'advance_payment', 'notes']
                    for f in fields_to_check:
                        new_val = item_sanitized.get(f)
                        old_val = getattr(att_obj, f)
                        is_diff = False
                        if f in ['effective_hours', 'overtime_hours', 'advance_payment']:
                            if float(new_val or 0) != float(old_val or 0):
                                is_diff = True
                        elif f in ['is_friday_work', 'is_mission']:
                            if bool(new_val) != bool(old_val):
                                is_diff = True
                        else:
                            if str(new_val or '') != str(old_val or ''):
                                is_diff = True
                        if is_diff:
                            changes.append((f, str(old_val), str(new_val)))

                    # اعمال تغییرات
                    att_obj.status = status_val
                    att_obj.effective_hours = eff_h
                    att_obj.overtime_hours = ot_h
                    att_obj.is_friday_work = is_fri
                    att_obj.is_mission = is_mis
                    att_obj.advance_payment = adv_pay
                    att_obj.notes = notes_val
                    att_obj.warehouse_id = target_wh
                    att_obj.modified_by = request.user
                    if target_period:
                        att_obj.period = target_period
                    att_obj.save()

                    # پاکسازی رکوردهای تکراری قدیمی احتمالی برای این روز
                    DailyAttendance.objects.filter(
                        personnel_id=personnel_id,
                        date_shamsi=date_shamsi,
                        is_deleted=False
                    ).exclude(id=att_obj.id).delete()

                    # ثبت لاگ‌های ممیزی
                    for f_name, o_val, n_val in changes:
                        AttendanceAuditLog.objects.create(
                            attendance=att_obj,
                            personnel_name=p_obj.full_name,
                            date_shamsi=date_shamsi,
                            changed_by=request.user,
                            field_name=f_name,
                            old_value=o_val,
                            new_value=n_val,
                            reason='ویرایش ماتریسی کارکرد روزانه'
                        )
                    updated_count += 1
                else:
                    # ایجاد رکورد جدید
                    att_obj = DailyAttendance.objects.create(
                        personnel_id=personnel_id,
                        warehouse_id=target_wh,
                        date_shamsi=date_shamsi,
                        status=status_val,
                        effective_hours=eff_h,
                        overtime_hours=ot_h,
                        is_friday_work=is_fri,
                        is_mission=is_mis,
                        advance_payment=adv_pay,
                        notes=notes_val,
                        period=target_period or period,
                        created_by=request.user,
                        modified_by=request.user
                    )
                    saved_count += 1

        client_tab_id = request.data.get('client_tab_id') or request.headers.get('X-Client-Tab-ID')
        broadcast_attendance_updated(
            warehouse_id=warehouse_id,
            date_shamsi=date_shamsi,
            year_month=year_month,
            sender_id=request.user.id if request.user and request.user.is_authenticated else None,
            client_tab_id=client_tab_id,
            message=f'کارکرد پرسنل برای تاریخ {date_shamsi} به‌روزرسانی شد.'
        )

        return Response({
            'message': f'اطلاعات کارکرد با موفقیت ذخیره شد ({saved_count} جدید، {updated_count} بروزرسانی).',
            'saved_count': saved_count,
            'updated_count': updated_count
        })

    @action(detail=False, methods=['post'], url_path='clear-day')
    def clear_day(self, request):
        """
        پاکسازی کامل و اصولی کارکرد روزانه پرسنل برای یک تاریخ مشخص
        """
        date_shamsi = request.data.get('date_shamsi')
        if not date_shamsi:
            return Response({'error': 'پارامتر date_shamsi الزامی است.'}, status=status.HTTP_400_BAD_REQUEST)

        raw_wh = request.data.get('warehouse_id')
        warehouse_id = int(raw_wh) if (raw_wh and str(raw_wh).upper() not in ['ALL', '0', 'NONE', '']) else None
        personnel_ids = request.data.get('personnel_ids')

        # اعتبارسنجی بازه مجاز ویرایش تاریخ
        is_admin_override = request.user.is_superuser or (request.user.role and request.user.role.name in ['Admin', 'Manager']) or request.user.has_perm('personnel.can_override_attendance_lock')
        if not is_admin_override:
            validate_attendance_date_window(date_shamsi, user=request.user, allow_override=is_admin_override, warehouse_id=warehouse_id)

        qs = DailyAttendance.objects.filter(date_shamsi=date_shamsi, is_deleted=False)
        if warehouse_id:
            qs = qs.filter(warehouse_id=warehouse_id)
        if personnel_ids and isinstance(personnel_ids, list) and len(personnel_ids) > 0:
            qs = qs.filter(personnel_id__in=personnel_ids)

        deleted_count = 0
        with transaction.atomic():
            for att in qs.select_for_update():
                att.is_deleted = True
                att.modified_by = request.user
                att.save()
                AttendanceAuditLog.objects.create(
                    attendance=att,
                    personnel_name=att.personnel.full_name if att.personnel else '',
                    date_shamsi=date_shamsi,
                    changed_by=request.user,
                    field_name='is_deleted',
                    old_value='False',
                    new_value='True',
                    reason='پاکسازی کارکرد روزانه'
                )
                deleted_count += 1

        ym = date_shamsi[:7] if len(date_shamsi) >= 7 else None
        client_tab_id = request.data.get('client_tab_id') or request.headers.get('X-Client-Tab-ID')
        broadcast_attendance_updated(
            warehouse_id=warehouse_id,
            date_shamsi=date_shamsi,
            year_month=ym,
            sender_id=request.user.id if request.user and request.user.is_authenticated else None,
            client_tab_id=client_tab_id,
            message=f'کارکرد روز {date_shamsi} برای {deleted_count} نفر پاکسازی شد.'
        )

        return Response({
            'message': f'کارکرد روز {date_shamsi} برای {deleted_count} نفر با موفقیت پاکسازی شد.',
            'cleared_count': deleted_count
        })

    @action(detail=False, methods=['get'], url_path='monthly-grid')
    def get_monthly_grid(self, request):
        """
        دریافت ماتریس کارکرد ۳۱ روزه کلیه پرسنل در یک ماه مشخص (نمای شیت ماهانه)
        پشتیبانی کامل از کارکرد مستقل از انبار و اصلاح وضعیت روزهای ثبت‌نشده (نباید به عنوان حاضر بزند)
        """
        import jdatetime
        raw_wh = request.query_params.get('warehouse_id')
        warehouse_id = int(raw_wh) if (raw_wh and str(raw_wh).upper() not in ['ALL', '0', 'NONE', '']) else None
        raw_year_month = request.query_params.get('year_month')
        if not raw_year_month:
            return Response({'error': 'پارامتر year_month الزامی است.'}, status=status.HTTP_400_BAD_REQUEST)

        # نرمال‌سازی سال و ماه (پشتیبانی از 140504 یا 1405/04)
        ym_cleaned = normalize_digits(str(raw_year_month)).strip().replace('-', '/')
        parts = ym_cleaned.split('/')
        if len(parts) >= 2:
            y, m = int(parts[0]), int(parts[1])
            year_month = f"{y:04d}/{m:02d}"
        else:
            digits_only = ''.join(c for c in ym_cleaned if c.isdigit())
            if len(digits_only) >= 6:
                y, m = int(digits_only[:4]), int(digits_only[4:6])
                year_month = f"{y:04d}/{m:02d}"
            else:
                return Response({'error': 'فرمت year_month نامعتبر است (مثال: 1405/04).'}, status=status.HTTP_400_BAD_REQUEST)

        # محاسبه تعداد روزهای ماه خورشیدی
        if 1 <= m <= 6:
            days_in_month = 31
        elif 7 <= m <= 11:
            days_in_month = 30
        elif m == 12:
            days_in_month = 30 if jdatetime.date(y, 1, 1).isleap() else 29
        else:
            return Response({'error': 'شماره ماه باید بین ۱ تا ۱۲ باشد.'}, status=status.HTTP_400_BAD_REQUEST)

        month_names = ["", "فروردین", "اردیبهشت", "خرداد", "تیر", "مرداد", "شهریور", "مهر", "آبان", "آذر", "دی", "بهمن", "اسفند"]
        month_name = month_names[m] if 1 <= m <= 12 else ""

        is_locked = False
        period_status = 'OPEN'
        period_obj = None
        if warehouse_id:
            period_obj = MonthlyWorkPeriod.objects.filter(warehouse_id=warehouse_id, year_month=year_month).first()
            if period_obj:
                period_status = period_obj.status
                is_locked = period_obj.status in ['LOCKED', 'SUBMITTED', 'FINALIZED']

        # واکشی تنظیمات بازه ویرایش
        settings = PayrollYearlySettings.objects.filter(fiscal_year=str(y), is_active=True).first()
        if not settings:
            settings = PayrollYearlySettings.objects.filter(is_active=True).first()
        past_limit = settings.attendance_edit_past_days if settings else 3
        future_limit = settings.attendance_edit_future_days if settings else 0

        today = jdatetime.date.today()

        days_meta = []
        for d in range(1, days_in_month + 1):
            target_date = jdatetime.date(y, m, d)
            weekday = target_date.weekday()  # 0=شنبه ... 6=جمعه
            is_friday = (weekday == 6)
            weekday_names = ["ش", "ی", "د", "س", "چ", "پ", "ج"]
            full_weekday_names = ["شنبه", "یکشنبه", "دوشنبه", "سه‌شنبه", "چهارشنبه", "پنج‌شنبه", "جمعه"]
            
            delta_days = (target_date - today).days
            is_editable = not is_locked
            if is_editable:
                if delta_days < 0 and past_limit is not None and past_limit != -1 and abs(delta_days) > past_limit:
                    is_editable = False
                elif delta_days > 0 and future_limit is not None and future_limit != -1 and delta_days > future_limit:
                    is_editable = False

            date_str = f"{year_month}/{d:02d}"
            days_meta.append({
                'day': d,
                'date_shamsi': date_str,
                'weekday_short': weekday_names[weekday],
                'weekday_name': full_weekday_names[weekday],
                'is_friday': is_friday,
                'is_editable': is_editable,
                'is_today': (target_date == today)
            })

        # پرسنل فعال یا پرسنل غیرفعالی که در این ماه کارکرد دارند (مستقل از انبار یا فیلتر انبار خاص)
        if warehouse_id:
            personnel_list = PersonnelProfile.objects.filter(
                Q(assigned_warehouse_id=warehouse_id) | Q(assigned_warehouse__isnull=True),
                Q(is_active=True) | Q(daily_attendances__date_shamsi__startswith=year_month, daily_attendances__is_deleted=False)
            ).distinct().order_by('last_name', 'first_name')
            attendances = DailyAttendance.objects.filter(
                Q(warehouse_id=warehouse_id) | Q(personnel__in=personnel_list),
                date_shamsi__startswith=year_month,
                is_deleted=False
            ).order_by('id')
        else:
            personnel_list = PersonnelProfile.objects.filter(
                Q(is_active=True) | Q(daily_attendances__date_shamsi__startswith=year_month, daily_attendances__is_deleted=False)
            ).distinct().order_by('last_name', 'first_name')
            attendances = DailyAttendance.objects.filter(
                date_shamsi__startswith=year_month,
                is_deleted=False
            ).order_by('id')

        att_map = {}
        for a in attendances:
            try:
                d_num = int(a.date_shamsi.split('/')[2])
                att_map[(a.personnel_id, d_num)] = a
            except Exception:
                continue

        status_display_map = {
            'PRESENT_10H': 'حاضر ۱۰س',
            'HALF_5H': 'نیمه‌وقت ۵س',
            'ABSENT': 'غایب',
            'LEAVE': 'مرخصی',
            'MISSION': 'ماموریت',
            'FRIDAY_WORK': 'جمعه‌کاری',
            'CUSTOM': 'سفارشی'
        }

        grid_rows = []
        anomalies = []

        for p in personnel_list:
            p_days = []
            total_hours = 0.0
            total_overtime = 0.0
            present_days_count = 0
            recorded_days = []

            for dm in days_meta:
                d = dm['day']
                att = att_map.get((p.id, d))
                if att:
                    eff_h = float(att.effective_hours)
                    ovt_h = float(att.overtime_hours)
                    adv = float(att.advance_payment or 0)
                    st_disp = status_display_map.get(att.status, att.status)

                    p_days.append({
                        'day': d,
                        'date_shamsi': dm['date_shamsi'],
                        'status': att.status,
                        'status_display': st_disp,
                        'effective_hours': eff_h,
                        'overtime_hours': ovt_h,
                        'is_friday_work': att.is_friday_work,
                        'is_mission': att.is_mission,
                        'advance_payment': adv,
                        'notes': att.notes or '',
                        'is_existing': True,
                        'attendance_id': att.id
                    })
                    total_hours += eff_h
                    total_overtime += ovt_h
                    if att.status and att.status != 'ABSENT' and eff_h > 0:
                        present_days_count += 1
                    recorded_days.append(d)

                    # ۱. ناهنجاری کارکرد بیش از ۱۶ ساعت
                    if (eff_h + ovt_h) > 16:
                        anomalies.append({
                            'type': 'EXCESSIVE_HOURS',
                            'severity': 'high',
                            'personnel_id': p.id,
                            'full_name': p.full_name,
                            'day': d,
                            'message': f"کارکرد بیش از سقف مجاز ({eff_h + ovt_h:.1f} ساعت) برای {p.full_name} در روز {d} ثبت شده است."
                        })

                    # ۲. تضاد اضافه‌کار با غیبت یا مرخصی
                    if att.status in ['ABSENT', 'LEAVE'] and ovt_h > 0:
                        anomalies.append({
                            'type': 'CONFLICT_OVERTIME',
                            'severity': 'high',
                            'personnel_id': p.id,
                            'full_name': p.full_name,
                            'day': d,
                            'message': f"ثبت اضافه‌کار ({ovt_h:.1f} ساعت) در روز {d} برای {p.full_name} با وضعیت «{st_disp}» مغایرت دارد."
                        })

                    # ۳. مساعده نقدی در روز غیبت
                    if att.status == 'ABSENT' and adv > 0:
                        anomalies.append({
                            'type': 'ADVANCE_ABSENT',
                            'severity': 'medium',
                            'personnel_id': p.id,
                            'full_name': p.full_name,
                            'day': d,
                            'message': f"ثبت مساعده ({adv:,.0f} ریال) برای {p.full_name} در روز {d} با وضعیت غایب مغایرت دارد."
                        })
                else:
                    p_days.append({
                        'day': d,
                        'date_shamsi': dm['date_shamsi'],
                        'status': '',
                        'status_display': '—',
                        'effective_hours': 0.0,
                        'overtime_hours': 0.0,
                        'is_friday_work': False,
                        'is_mission': False,
                        'advance_payment': 0.0,
                        'notes': '',
                        'is_existing': False,
                        'attendance_id': None
                    })

            # ۴. شناسایی روزهای خالی و ثبت‌نشده در میان ماه
            if len(recorded_days) >= 2:
                min_d, max_d = min(recorded_days), max(recorded_days)
                for d_i in range(min_d + 1, max_d):
                    if d_i not in recorded_days:
                        dm_meta = next((dm for dm in days_meta if dm['day'] == d_i), None)
                        if dm_meta and not dm_meta['is_friday']:
                            anomalies.append({
                                'type': 'GAP_DAY',
                                'severity': 'medium',
                                'personnel_id': p.id,
                                'full_name': p.full_name,
                                'day': d_i,
                                'message': f"روز کاری {d_i} برای {p.full_name} تعیین وضعیت نشده و خالی مانده است."
                            })

            grid_rows.append({
                'personnel_id': p.id,
                'full_name': p.full_name,
                'national_code': p.national_code,
                'job_title': p.job_title,
                'is_active': p.is_active,
                'total_hours': round(total_hours, 2),
                'total_overtime': round(total_overtime, 2),
                'present_days': present_days_count,
                'days': p_days
            })

        period_info = {
            'id': period_obj.id if period_obj else None,
            'status': period_status,
            'status_display': period_obj.get_status_display() if period_obj else 'باز (امکان ثبت و ویرایش)',
            'is_locked': is_locked,
            'submitted_at': period_obj.submitted_at if period_obj else None,
            'submitted_by_name': f"{period_obj.submitted_by.first_name} {period_obj.submitted_by.last_name}".strip() if period_obj and period_obj.submitted_by else None,
            'submission_notes': period_obj.submission_notes if period_obj else '',
            'rejection_reason': period_obj.rejection_reason if period_obj else '',
            'locked_at': period_obj.locked_at if period_obj else None,
            'locked_by_name': f"{period_obj.locked_by.first_name} {period_obj.locked_by.last_name}".strip() if period_obj and period_obj.locked_by else None,
        }

        return Response({
            'warehouse_id': warehouse_id,
            'year_month': year_month,
            'month_name': month_name,
            'days_in_month': days_in_month,
            'is_locked': is_locked,
            'period_status': period_status,
            'period_info': period_info,
            'settings_window': {
                'past_days': past_limit,
                'future_days': future_limit
            },
            'days_meta': days_meta,
            'rows': grid_rows,
            'anomalies': anomalies
        })

    @action(detail=False, methods=['post'], url_path='bulk-save-monthly-grid')
    def bulk_save_monthly_grid(self, request):
        """
        ذخیره گروهی سلول‌های تغییریافته شیت ماهانه
        (پشتیبانی از مستقل از انبار، با اعتبارسنجی بازه تاریخی و ثبت لاگ ممیزی)
        """
        serializer = BulkMonthlyAttendanceGridSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        raw_wh = serializer.validated_data.get('warehouse_id')
        warehouse_id = int(raw_wh) if (raw_wh and str(raw_wh).upper() not in ['ALL', '0', 'NONE', '']) else None
        raw_ym = serializer.validated_data['year_month']
        ym_cleaned = normalize_digits(str(raw_ym)).strip().replace('-', '/')
        parts = ym_cleaned.split('/')
        if len(parts) >= 2:
            year_month = f"{int(parts[0]):04d}/{int(parts[1]):02d}"
        else:
            year_month = raw_ym

        items = serializer.validated_data['items']

        allow_override = bool(request.data.get('is_override', False) or request.headers.get('X-Admin-Override') == 'true')
        is_admin_override = allow_override and (getattr(request.user, 'is_superuser', False) or request.user.has_perm('personnel.can_override_attendance_lock'))
        period = None
        if warehouse_id:
            period, _ = MonthlyWorkPeriod.objects.get_or_create(
                warehouse_id=warehouse_id,
                year_month=year_month,
                defaults={'status': 'OPEN'}
            )
            if period.status == 'LOCKED' and not is_admin_override:
                return Response({'error': f'دوره کارکرد {year_month} قفل شده است و امکان ویرایش وجود ندارد.'}, status=status.HTTP_400_BAD_REQUEST)

        # واکشی یکجای رکوردهای موجود ماه جاری جهت تشخیص تغییرات (فقط رکوردهای تغییریافته نیاز به ولیدیشن تاریخ دارند)
        existing_qs = DailyAttendance.objects.filter(
            date_shamsi__startswith=year_month,
            is_deleted=False
        )
        if warehouse_id:
            existing_qs = existing_qs.filter(warehouse_id=warehouse_id)
        existing_map = {(att.personnel_id, att.date_shamsi): att for att in existing_qs}

        # اعتبارسنجی فقط برای روزهای جدید یا تغییریافته
        for item in items:
            if not item.get('status') and float(item.get('effective_hours', 0)) == 0:
                continue
            day_str = f"{item['day']:02d}"
            date_shamsi = f"{year_month}/{day_str}"
            personnel_id = item['personnel_id']

            existing_att = existing_map.get((personnel_id, date_shamsi))
            is_changed = True
            if existing_att:
                status_val = item.get('status') or ('PRESENT_10H' if float(item.get('effective_hours', 0)) >= 10 else ('HALF_5H' if float(item.get('effective_hours', 0)) >= 5 else 'CUSTOM'))
                if (
                    str(existing_att.status) == str(status_val) and
                    float(existing_att.effective_hours) == float(item.get('effective_hours', 0)) and
                    float(existing_att.overtime_hours) == float(item.get('overtime_hours', 0)) and
                    bool(existing_att.is_friday_work) == bool(item.get('is_friday_work', False)) and
                    bool(existing_att.is_mission) == bool(item.get('is_mission', False)) and
                    float(existing_att.advance_payment) == float(item.get('advance_payment', 0)) and
                    (existing_att.notes or '') == (item.get('notes', '') or '')
                ):
                    is_changed = False

            if is_changed and not is_admin_override:
                try:
                    validate_attendance_date_window(date_shamsi, user=request.user, allow_override=is_admin_override, warehouse_id=warehouse_id)
                except Exception as e:
                    err_msg = getattr(e, 'detail', str(e))
                    if isinstance(err_msg, list):
                        err_msg = err_msg[0]
                    return Response({'error': f"خطا در روز {item['day']}: {err_msg}"}, status=status.HTTP_400_BAD_REQUEST)

        saved_count = 0
        updated_count = 0

        with transaction.atomic():
            for item in items:
                status_val = item.get('status') or ''
                eff_h = float(item.get('effective_hours', 0))
                day_str = f"{item['day']:02d}"
                date_shamsi = f"{year_month}/{day_str}"
                personnel_id = item['personnel_id']

                p_obj = PersonnelProfile.objects.filter(id=personnel_id).first()
                if not p_obj:
                    continue

                check_wh = warehouse_id if warehouse_id else p_obj.assigned_warehouse_id
                target_period = period
                if not target_period and check_wh:
                    target_period = MonthlyWorkPeriod.objects.filter(
                        warehouse_id=check_wh,
                        year_month=year_month
                    ).first()
                if target_period and target_period.status == 'LOCKED' and not is_admin_override:
                    return Response({'error': f'دوره کارکرد {year_month} برای انبار پرسنل ({p_obj.full_name}) قفل شده است.'}, status=status.HTTP_400_BAD_REQUEST)

                target_wh = warehouse_id if warehouse_id else p_obj.assigned_warehouse_id

                # پیدا کردن رکورد فعال پرسنل در این تاریخ (مستقل از انبار قبلی)
                att_obj = DailyAttendance.objects.select_for_update().filter(
                    personnel_id=personnel_id,
                    date_shamsi=date_shamsi,
                    is_deleted=False
                ).first()

                if not att_obj and not status_val and eff_h == 0:
                    continue

                if not status_val:
                    status_val = 'PRESENT_10H' if eff_h >= 10 else ('HALF_5H' if eff_h >= 5 else 'CUSTOM')

                ot_h = float(item.get('overtime_hours', 0))
                is_fri = bool(item.get('is_friday_work', False))
                is_mis = bool(item.get('is_mission', False))
                adv_pay = float(item.get('advance_payment', 0))
                notes_val = item.get('notes', '') or ''

                # اعتبارسنجی و ریست منطقی ساعات در صورت غیبت یا مرخصی
                if status_val in ['ABSENT', 'LEAVE']:
                    eff_h = 0.0
                    ot_h = 0.0
                    is_fri = False
                    is_mis = False

                if att_obj:
                    changes = []
                    item_sanitized = {
                        'status': status_val,
                        'effective_hours': eff_h,
                        'overtime_hours': ot_h,
                        'is_friday_work': is_fri,
                        'is_mission': is_mis,
                        'advance_payment': adv_pay,
                        'notes': notes_val
                    }
                    fields_to_check = ['status', 'effective_hours', 'overtime_hours', 'is_friday_work', 'is_mission', 'advance_payment', 'notes']
                    for f in fields_to_check:
                        new_val = item_sanitized.get(f)
                        old_val = getattr(att_obj, f)
                        is_diff = False
                        if f in ['effective_hours', 'overtime_hours', 'advance_payment']:
                            if float(new_val or 0) != float(old_val or 0):
                                is_diff = True
                        elif f in ['is_friday_work', 'is_mission']:
                            if bool(new_val) != bool(old_val):
                                is_diff = True
                        else:
                            if str(new_val or '') != str(old_val or ''):
                                is_diff = True
                        if is_diff:
                            changes.append((f, str(old_val), str(new_val)))

                    # اگر تغییری نکرده بود از ذخیره و لاگ تکراری صرف‌نظر می‌کنیم
                    if not changes:
                        continue

                    att_obj.status = status_val
                    att_obj.effective_hours = eff_h
                    att_obj.overtime_hours = ot_h
                    att_obj.is_friday_work = is_fri
                    att_obj.is_mission = is_mis
                    att_obj.advance_payment = adv_pay
                    att_obj.notes = notes_val
                    if target_wh:
                        att_obj.warehouse_id = target_wh
                    att_obj.modified_by = request.user
                    if target_period:
                        att_obj.period = target_period
                    att_obj.save()

                    # پاکسازی رکوردهای تکراری قدیمی احتمالی برای این روز
                    DailyAttendance.objects.filter(
                        personnel_id=personnel_id,
                        date_shamsi=date_shamsi,
                        is_deleted=False
                    ).exclude(id=att_obj.id).delete()

                    for f_name, o_val, n_val in changes:
                        AttendanceAuditLog.objects.create(
                            attendance=att_obj,
                            personnel_name=p_obj.full_name,
                            date_shamsi=date_shamsi,
                            changed_by=request.user,
                            field_name=f_name,
                            old_value=o_val,
                            new_value=n_val,
                            reason='ویرایش شیت ماهانه کارکرد'
                        )
                    updated_count += 1
                else:
                    att_obj = DailyAttendance.objects.create(
                        personnel_id=personnel_id,
                        warehouse_id=target_wh,
                        date_shamsi=date_shamsi,
                        status=status_val,
                        effective_hours=eff_h,
                        overtime_hours=ot_h,
                        is_friday_work=is_fri,
                        is_mission=is_mis,
                        advance_payment=adv_pay,
                        notes=notes_val,
                        period=target_period or period,
                        created_by=request.user,
                        modified_by=request.user
                    )
                    saved_count += 1

        client_tab_id = request.data.get('client_tab_id') or request.headers.get('X-Client-Tab-ID')
        broadcast_attendance_updated(
            warehouse_id=warehouse_id,
            date_shamsi=f"{year_month}/01",
            year_month=year_month,
            sender_id=request.user.id if request.user and request.user.is_authenticated else None,
            client_tab_id=client_tab_id,
            message=f'شیت کارکرد ماهانه {year_month} به‌روزرسانی شد.'
        )

        return Response({
            'message': f'اطلاعات شیت ماهانه با موفقیت ذخیره شد ({saved_count} جدید، {updated_count} بروزرسانی).',
            'saved_count': saved_count,
            'updated_count': updated_count
        })

    @action(detail=False, methods=['get'], url_path='export-monthly-excel')
    def export_monthly_excel(self, request):
        """
        خروجی استاندارد دو سطری اکسل از شیت ماهانه کارکرد پرسنل
        (شامل فرمول‌های جمع، استایل زیبا و فریز ستون‌ها طبق استاندارد)
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        raw_wh = request.query_params.get('warehouse_id')
        warehouse_id = int(raw_wh) if (raw_wh and str(raw_wh).upper() not in ['ALL', '0', 'NONE', '']) else None
        raw_ym = request.query_params.get('year_month')
        if not raw_ym:
            return Response({'error': 'پارامتر year_month الزامی است.'}, status=status.HTTP_400_BAD_REQUEST)

        ym_cleaned = normalize_digits(str(raw_ym)).strip().replace('-', '/')
        parts = ym_cleaned.split('/')
        if len(parts) >= 2:
            y, m = int(parts[0]), int(parts[1])
            year_month = f"{y:04d}/{m:02d}"
        else:
            return Response({'error': 'فرمت year_month نامعتبر است.'}, status=status.HTTP_400_BAD_REQUEST)

        if 1 <= m <= 6:
            days_in_month = 31
        elif 7 <= m <= 11:
            days_in_month = 30
        elif m == 12:
            days_in_month = 30 if jdatetime.date(y, 1, 1).isleap() else 29
        else:
            return Response({'error': 'شماره ماه باید بین ۱ تا ۱۲ باشد.'}, status=status.HTTP_400_BAD_REQUEST)

        month_names = ["", "فروردین", "اردیبهشت", "خرداد", "تیر", "مرداد", "شهریور", "مهر", "آبان", "آذر", "دی", "بهمن", "اسفند"]
        month_name = month_names[m] if 1 <= m <= 12 else ""

        fridays = set()
        weekday_names = ["ش", "ی", "د", "س", "چ", "پ", "ج"]
        day_headers = []
        for d in range(1, days_in_month + 1):
            t_date = jdatetime.date(y, m, d)
            w = t_date.weekday()
            if w == 6:
                fridays.add(d)
            day_headers.append(f"{d} ({weekday_names[w]})")

        if warehouse_id:
            personnel_list = PersonnelProfile.objects.filter(
                Q(assigned_warehouse_id=warehouse_id) | Q(assigned_warehouse__isnull=True),
                is_active=True
            ).select_related('assigned_warehouse').order_by('last_name', 'first_name')
            attendances = DailyAttendance.objects.filter(
                warehouse_id=warehouse_id,
                date_shamsi__startswith=year_month,
                is_deleted=False
            )
        else:
            personnel_list = PersonnelProfile.objects.filter(is_active=True).select_related('assigned_warehouse').order_by('last_name', 'first_name')
            attendances = DailyAttendance.objects.filter(
                date_shamsi__startswith=year_month,
                is_deleted=False
            )

        att_map = {(a.personnel_id, int(a.date_shamsi.split('/')[2])): a for a in attendances}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"کارکرد {month_name} {y}"
        ws.sheet_view.rightToLeft = True

        font_header1 = Font(name='B Nazanin', size=11, bold=True, color='FFFFFF')
        font_header2 = Font(name='Calibri', size=9, bold=False, color='94A3B8')
        font_body = Font(name='B Nazanin', size=10, bold=False, color='0F172A')
        font_bold = Font(name='B Nazanin', size=10, bold=True, color='0F172A')
        font_mono = Font(name='Consolas', size=10, bold=True, color='0F172A')

        fill_header1 = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
        fill_header2 = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
        fill_friday = PatternFill(start_color='F3E8FF', end_color='F3E8FF', fill_type='solid')
        fill_total = PatternFill(start_color='EEF2FF', end_color='EEF2FF', fill_type='solid')
        fill_stripe = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')

        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_right = Alignment(horizontal='right', vertical='center')

        row1_titles = ['ردیف', 'کد ملی', 'نام و نام خانوادگی', 'عنوان شغلی', 'انبار اختصاصی'] + day_headers + ['جمع ساعت کارکرد', 'جمع اضافه کار', 'مجموع کل ساعات (کارکرد + اضافه‌کار)', 'تعداد روز حضور']
        row2_keys = ['row_num', 'national_code', 'full_name', 'job_title', 'warehouse_name'] + [f'day_{d:02d}' for d in range(1, days_in_month + 1)] + ['total_hours', 'total_overtime', 'total_hours_sum', 'present_days']

        for col_idx, (t1, t2) in enumerate(zip(row1_titles, row2_keys), start=1):
            c1 = ws.cell(row=1, column=col_idx, value=t1)
            c1.font = font_header1
            c1.fill = fill_header1
            c1.alignment = align_center
            c1.border = thin_border

            c2 = ws.cell(row=2, column=col_idx, value=t2)
            c2.font = font_header2
            c2.fill = fill_header2
            c2.alignment = align_center
            c2.border = thin_border

        for p_idx, p in enumerate(personnel_list, start=1):
            row_num = p_idx + 2
            is_even = (p_idx % 2 == 0)

            ws.cell(row=row_num, column=1, value=p_idx).alignment = align_center
            ws.cell(row=row_num, column=2, value=p.national_code or '').alignment = align_center
            ws.cell(row=row_num, column=3, value=p.full_name).alignment = align_right
            ws.cell(row=row_num, column=4, value=p.job_title or '').alignment = align_right
            ws.cell(row=row_num, column=5, value=p.assigned_warehouse.name if p.assigned_warehouse else 'تمام انبارها').alignment = align_center

            tot_hours = 0.0
            tot_ot = 0.0
            p_days = 0

            for d in range(1, days_in_month + 1):
                col_idx = 5 + d
                att = att_map.get((p.id, d))
                cell = ws.cell(row=row_num, column=col_idx)
                cell.alignment = align_center

                if att:
                    h = float(att.effective_hours)
                    ot = float(att.overtime_hours)
                    tot_hours += h
                    tot_ot += ot
                    if att.status and att.status != 'ABSENT' and h > 0:
                        p_days += 1

                    if att.status == 'LEAVE':
                        cell.value = 'مرخصی'
                    elif att.status == 'ABSENT':
                        cell.value = 'غایب'
                    elif att.status == 'MISSION':
                        cell.value = f"مامور ({h:g})"
                    elif att.status == 'FRIDAY_WORK':
                        cell.value = f"جمعه ({h:g})"
                    elif h > 0:
                        cell.value = h
                    else:
                        cell.value = ''
                else:
                    cell.value = ''

                if d in fridays:
                    cell.fill = fill_friday
                elif is_even:
                    cell.fill = fill_stripe

                cell.border = thin_border
                cell.font = font_body

            first_col_letter = get_column_letter(6)
            last_col_letter = get_column_letter(5 + days_in_month)
            col_tot_letter = get_column_letter(6 + days_in_month)
            col_ot_letter = get_column_letter(7 + days_in_month)

            c_tot = ws.cell(row=row_num, column=6 + days_in_month, value=f"=SUM({first_col_letter}{row_num}:{last_col_letter}{row_num})")
            c_tot.font = font_bold
            c_tot.fill = fill_total
            c_tot.alignment = align_center
            c_tot.border = thin_border

            c_ot = ws.cell(row=row_num, column=7 + days_in_month, value=tot_ot)
            c_ot.font = font_mono
            c_ot.alignment = align_center
            c_ot.border = thin_border

            c_tot_all = ws.cell(row=row_num, column=8 + days_in_month, value=f"={col_tot_letter}{row_num}+{col_ot_letter}{row_num}")
            c_tot_all.font = font_bold
            c_tot_all.fill = fill_total
            c_tot_all.alignment = align_center
            c_tot_all.border = thin_border

            c_pdays = ws.cell(row=row_num, column=9 + days_in_month, value=f'=COUNTIF({first_col_letter}{row_num}:{last_col_letter}{row_num}, ">0")')
            c_pdays.font = font_bold
            c_pdays.alignment = align_center
            c_pdays.border = thin_border

            for c_i in range(1, 6):
                cell_i = ws.cell(row=row_num, column=c_i)
                cell_i.font = font_bold if c_i == 3 else font_body
                cell_i.border = thin_border
                if is_even and cell_i.fill.fill_type is None:
                    cell_i.fill = fill_stripe

        ws.freeze_panes = 'F3'
        ws.row_dimensions[1].height = 28
        ws.row_dimensions[2].height = 18

        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 22
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 16
        for d in range(1, days_in_month + 1):
            col_letter = get_column_letter(5 + d)
            ws.column_dimensions[col_letter].width = 8
        ws.column_dimensions[get_column_letter(6 + days_in_month)].width = 14
        ws.column_dimensions[get_column_letter(7 + days_in_month)].width = 13
        ws.column_dimensions[get_column_letter(8 + days_in_month)].width = 20
        ws.column_dimensions[get_column_letter(9 + days_in_month)].width = 13

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"Timesheet_{year_month.replace('/', '_')}.xlsx"
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @action(detail=False, methods=['post'], url_path='import-monthly-excel')
    def import_monthly_excel(self, request):
        """
        بارگذاری مستقیم و هوشمند شیت ماهانه از فایل اکسل
        (تطبیق بر اساس ردیف کلیدها و کد ملی پرسنل همراه با ذخیره اتمیک)
        """
        import openpyxl
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'error': 'فایل اکسل الزامی است.'}, status=status.HTTP_400_BAD_REQUEST)

        raw_wh = request.data.get('warehouse_id')
        warehouse_id = int(raw_wh) if (raw_wh and str(raw_wh).upper() not in ['ALL', '0', 'NONE', '']) else None
        raw_ym = request.data.get('year_month')
        if not raw_ym:
            return Response({'error': 'پارامتر year_month الزامی است.'}, status=status.HTTP_400_BAD_REQUEST)

        ym_cleaned = normalize_digits(str(raw_ym)).strip().replace('-', '/')
        parts = ym_cleaned.split('/')
        if len(parts) >= 2:
            year_month = f"{int(parts[0]):04d}/{int(parts[1]):02d}"
        else:
            return Response({'error': 'فرمت year_month نامعتبر است.'}, status=status.HTTP_400_BAD_REQUEST)

        allow_override = bool(request.data.get('is_override', False) or request.headers.get('X-Admin-Override') == 'true')
        is_admin_override = allow_override and (getattr(request.user, 'is_superuser', False) or request.user.has_perm('personnel.can_override_attendance_lock'))

        if warehouse_id:
            period = MonthlyWorkPeriod.objects.filter(warehouse_id=warehouse_id, year_month=year_month).first()
            if period and period.status in ['LOCKED', 'SUBMITTED', 'FINALIZED'] and not is_admin_override:
                return Response({'error': f'دوره کارکرد {year_month} در وضعیت «{period.get_status_display()}» است و امکان بارگذاری وجود ندارد.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(file_obj, data_only=True)
            ws = wb.active
        except Exception as e:
            return Response({'error': f'خطا در باز کردن فایل اکسل: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

        nat_col_idx = None
        day_cols = {}

        for col_idx in range(1, ws.max_column + 1):
            k2 = str(ws.cell(row=2, column=col_idx).value or '').strip()
            k1 = str(ws.cell(row=1, column=col_idx).value or '').strip()

            if k2 == 'national_code' or 'کد ملی' in k1:
                nat_col_idx = col_idx
            elif k2.startswith('day_'):
                try:
                    d_num = int(k2.replace('day_', ''))
                    day_cols[d_num] = col_idx
                except ValueError:
                    pass
            elif k1.startswith('روز '):
                try:
                    d_num = int(k1.replace('روز ', '').split()[0].split('(')[0])
                    day_cols[d_num] = col_idx
                except ValueError:
                    pass

        if not nat_col_idx:
            return Response({'error': 'ستون کد ملی پرسنل در فایل اکسل یافت نشد.'}, status=status.HTTP_400_BAD_REQUEST)
        if not day_cols:
            return Response({'error': 'هیچ ستونی برای روزهای ماه در فایل اکسل شناسایی نشد.'}, status=status.HTTP_400_BAD_REQUEST)

        personnel_map = {
            normalize_digits(str(p.national_code)).split('.')[0].strip().zfill(10): p
            for p in PersonnelProfile.objects.filter(is_active=True)
            if p.national_code
        }

        updated_cells = 0
        matched_personnel_count = 0

        with transaction.atomic():
            for row_idx in range(3, ws.max_row + 1):
                raw_nat = ws.cell(row=row_idx, column=nat_col_idx).value
                if raw_nat is None or str(raw_nat).strip() == '':
                    continue
                raw_nat_str = str(raw_nat).split('.')[0].strip()
                nat_code = normalize_digits(raw_nat_str).strip().zfill(10)
                p_obj = personnel_map.get(nat_code)
                if not p_obj:
                    continue

                matched_personnel_count += 1
                target_wh = warehouse_id if warehouse_id else p_obj.assigned_warehouse_id

                target_period = None
                if target_wh:
                    target_period, _ = MonthlyWorkPeriod.objects.get_or_create(
                        warehouse_id=target_wh,
                        year_month=year_month,
                        defaults={'status': 'OPEN'}
                    )
                    if target_period and target_period.status in ['LOCKED', 'SUBMITTED', 'FINALIZED'] and not is_admin_override:
                        return Response({
                            'error': f'امکان بارگذاری اکسل وجود ندارد؛ دوره کارکرد {year_month} برای انبار پرسنل ({p_obj.full_name}) در وضعیت «{target_period.get_status_display()}» است.'
                        }, status=status.HTTP_400_BAD_REQUEST)

                for d_num, col_i in day_cols.items():
                    val = ws.cell(row=row_idx, column=col_i).value
                    if val is None or str(val).strip() == '' or str(val).strip() == '—':
                        continue

                    val_str = str(val).strip()
                    val_str_clean = normalize_digits(val_str)
                    date_shamsi = f"{year_month}/{d_num:02d}"

                    status_val = 'PRESENT_10H'
                    eff_hours = 10.0
                    ot_hours = 0.0
                    is_friday = False
                    is_mission = False

                    try:
                        num = float(val_str_clean)
                        if num == 10:
                            status_val = 'PRESENT_10H'
                            eff_hours = 10.0
                        elif num == 5:
                            status_val = 'HALF_5H'
                            eff_hours = 5.0
                        elif num == 0:
                            status_val = 'ABSENT'
                            eff_hours = 0.0
                        else:
                            status_val = 'CUSTOM'
                            eff_hours = num
                    except ValueError:
                        if 'مرخصی' in val_str:
                            status_val = 'LEAVE'
                            eff_hours = 0.0
                        elif 'غایب' in val_str:
                            status_val = 'ABSENT'
                            eff_hours = 0.0
                        elif 'مامور' in val_str:
                            status_val = 'MISSION'
                            is_mission = True
                            eff_hours = 10.0
                        elif 'جمعه' in val_str:
                            status_val = 'FRIDAY_WORK'
                            is_friday = True
                            eff_hours = 10.0
                        elif 'نیمه' in val_str:
                            status_val = 'HALF_5H'
                            eff_hours = 5.0
                        elif 'حاضر' in val_str:
                            status_val = 'PRESENT_10H'
                            eff_hours = 10.0

                    att_filter = {'personnel_id': p_obj.id, 'date_shamsi': date_shamsi, 'is_deleted': False}
                    if target_wh:
                        att_filter['warehouse_id'] = target_wh

                    att_obj = DailyAttendance.objects.select_for_update().filter(**att_filter).first()
                    if not att_obj and not target_wh:
                        att_obj = DailyAttendance.objects.select_for_update().filter(personnel_id=p_obj.id, date_shamsi=date_shamsi, is_deleted=False).first()

                    # اعتبارسنجی بازه تاریخی در صورت تغییر یا ایجاد
                    is_changed = False
                    if not att_obj:
                        is_changed = True
                    else:
                        if (att_obj.status != status_val or float(att_obj.effective_hours) != eff_hours or
                            att_obj.is_friday_work != is_friday or att_obj.is_mission != is_mission):
                            is_changed = True

                    if is_changed and not is_admin_override:
                        try:
                            validate_attendance_date_window(
                                date_shamsi=date_shamsi,
                                user=request.user,
                                allow_override=is_admin_override,
                                warehouse_id=target_wh
                            )
                        except Exception as e:
                            err_msg = getattr(e, 'detail', str(e))
                            if isinstance(err_msg, list):
                                err_msg = err_msg[0]
                            return Response({
                                'error': f"خطا در ردیف پرسنل {p_obj.full_name} ({date_shamsi}): {err_msg}"
                            }, status=status.HTTP_400_BAD_REQUEST)

                    if att_obj:
                        changes = []
                        if att_obj.status != status_val:
                            changes.append(('status', str(att_obj.status), str(status_val)))
                        if float(att_obj.effective_hours) != eff_hours:
                            changes.append(('effective_hours', str(att_obj.effective_hours), str(eff_hours)))
                        if att_obj.is_friday_work != is_friday:
                            changes.append(('is_friday_work', str(att_obj.is_friday_work), str(is_friday)))
                        if att_obj.is_mission != is_mission:
                            changes.append(('is_mission', str(att_obj.is_mission), str(is_mission)))

                        if changes:
                            att_obj.status = status_val
                            att_obj.effective_hours = eff_hours
                            att_obj.is_friday_work = is_friday
                            att_obj.is_mission = is_mission
                            if target_wh:
                                att_obj.warehouse_id = target_wh
                            if target_period:
                                att_obj.period = target_period
                            att_obj.modified_by = request.user
                            att_obj.save()

                            for f_name, o_val, n_val in changes:
                                AttendanceAuditLog.objects.create(
                                    attendance=att_obj,
                                    personnel_name=p_obj.full_name,
                                    date_shamsi=date_shamsi,
                                    changed_by=request.user,
                                    field_name=f_name,
                                    old_value=o_val,
                                    new_value=n_val,
                                    reason='ویرایش از طریق بارگذاری اکسل شیت ماهانه'
                                )
                            updated_cells += 1
                    else:
                        att_obj = DailyAttendance.objects.create(
                            personnel=p_obj,
                            warehouse_id=target_wh,
                            period=target_period,
                            date_shamsi=date_shamsi,
                            status=status_val,
                            effective_hours=eff_hours,
                            overtime_hours=ot_hours,
                            is_friday_work=is_friday,
                            is_mission=is_mission,
                            advance_payment=0,
                            created_by=request.user,
                            modified_by=request.user
                        )
                        AttendanceAuditLog.objects.create(
                            attendance=att_obj,
                            personnel_name=p_obj.full_name,
                            date_shamsi=date_shamsi,
                            changed_by=request.user,
                            field_name='all',
                            old_value=None,
                            new_value=f"{status_val} ({eff_hours}h)",
                            reason='ایجاد از طریق بارگذاری اکسل شیت ماهانه'
                        )
                        updated_cells += 1

        broadcast_attendance_updated(
            warehouse_id=warehouse_id,
            date_shamsi=f"{year_month}/01",
            year_month=year_month,
            sender_id=request.user.id if request.user and request.user.is_authenticated else None,
            message=f'اطلاعات شیت کارکرد ماهانه {year_month} از فایل اکسل بارگذاری و به‌روزرسانی شد.'
        )

        return Response({
            'message': f'اطلاعات شیت کارکرد با موفقیت از اکسل بارگذاری شد ({matched_personnel_count} پرسنل، {updated_cells} سلول کارکرد ثبت گردید).',
            'matched_personnel_count': matched_personnel_count,
            'updated_cells': updated_cells
        })

    @action(detail=False, methods=['get'], url_path='monthly-summary')
    def get_monthly_summary(self, request):
        """
        گزارش تجمیعی ماهانه کارکرد پرسنل
        """
        warehouse_id = request.query_params.get('warehouse_id')
        year_month = request.query_params.get('year_month')
        if not warehouse_id or not year_month:
            return Response({'error': 'warehouse_id و year_month الزامی هستند.'}, status=status.HTTP_400_BAD_REQUEST)

        personnel_list = PersonnelProfile.objects.filter(
            Q(assigned_warehouse_id=warehouse_id) | Q(assigned_warehouse__isnull=True),
            is_active=True
        ).order_by('last_name', 'first_name')

        attendances = DailyAttendance.objects.filter(
            warehouse_id=warehouse_id,
            date_shamsi__startswith=year_month,
            is_deleted=False
        )

        period = MonthlyWorkPeriod.objects.filter(warehouse_id=warehouse_id, year_month=year_month).first()

        summary_rows = []
        for p in personnel_list:
            p_atts = [a for a in attendances if a.personnel_id == p.id]
            
            total_hours = sum(float(a.effective_hours) for a in p_atts)
            total_overtime = sum(float(a.overtime_hours) for a in p_atts)
            friday_days = sum(1 for a in p_atts if a.is_friday_work)
            mission_days = sum(1 for a in p_atts if a.is_mission)
            absent_days = sum(1 for a in p_atts if a.status == 'ABSENT')
            leave_days = sum(1 for a in p_atts if a.status == 'LEAVE')
            present_days = sum(1 for a in p_atts if a.status in ['PRESENT_10H', 'HALF_5H', 'FRIDAY_WORK', 'MISSION', 'CUSTOM'] and a.effective_hours > 0)
            total_advances = sum(float(a.advance_payment) for a in p_atts)
            
            # تبدیل به معادل روز (تقسیم بر ۱۰)
            equivalent_days = round(total_hours / 10.0, 2)
            
            # برآورد دستمزد ناخالص کارکرد
            daily_wage = float(p.daily_base_wage)
            hourly_wage = float(p.hourly_rate)
            gross_base_pay = round(equivalent_days * daily_wage, 0)
            overtime_pay = round(total_overtime * (hourly_wage * 1.4), 0)

            summary_rows.append({
                'personnel_id': p.id,
                'full_name': p.full_name,
                'national_code': p.national_code,
                'job_title': p.job_title,
                'contract_type': p.get_contract_type_display(),
                'daily_base_wage': daily_wage,
                'total_hours': round(total_hours, 1),
                'equivalent_days': equivalent_days,
                'present_days': present_days,
                'total_overtime_hours': round(total_overtime, 1),
                'friday_days': friday_days,
                'mission_days': mission_days,
                'absent_days': absent_days,
                'leave_days': leave_days,
                'total_advances': total_advances,
                'gross_base_pay': gross_base_pay,
                'overtime_pay': overtime_pay,
                'estimated_total_pay': gross_base_pay + overtime_pay - total_advances
            })

        return Response({
            'warehouse_id': int(warehouse_id),
            'year_month': year_month,
            'period_status': period.status if period else 'OPEN',
            'is_locked': period.status == 'LOCKED' if period else False,
            'summary': summary_rows
        })


class VehicleTripViewSet(viewsets.ModelViewSet):
    queryset = VehicleTripLog.objects.filter(is_deleted=False).select_related('vehicle', 'warehouse', 'period')
    serializer_class = VehicleTripLogSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = None

    def get_queryset(self):
        qs = super().get_queryset()
        warehouse_id = self.request.query_params.get('warehouse_id')
        if warehouse_id:
            qs = qs.filter(warehouse_id=warehouse_id)
        date_shamsi = self.request.query_params.get('date_shamsi')
        if date_shamsi:
            qs = qs.filter(date_shamsi=date_shamsi)
        vehicle_id = self.request.query_params.get('vehicle_id')
        if vehicle_id:
            qs = qs.filter(vehicle_id=vehicle_id)
        return qs

    @action(detail=False, methods=['get'], url_path='matrix')
    def get_matrix(self, request):
        """
        دریافت لیست سریع خودروها برای ثبت سرویس روزانه
        """
        warehouse_id = request.query_params.get('warehouse_id')
        date_shamsi = request.query_params.get('date_shamsi')
        if not warehouse_id or not date_shamsi:
            return Response({'error': 'warehouse_id و date_shamsi الزامی هستند.'}, status=status.HTTP_400_BAD_REQUEST)

        vehicles = VehicleDriverProfile.objects.filter(
            Q(assigned_warehouse_id=warehouse_id) | Q(assigned_warehouse__isnull=True),
            is_active=True
        ).order_by('driver_name')

        existing_trips = {
            t.vehicle_id: t
            for t in VehicleTripLog.objects.filter(
                warehouse_id=warehouse_id,
                date_shamsi=date_shamsi,
                is_deleted=False
            )
        }

        matrix_rows = []
        for v in vehicles:
            t = existing_trips.get(v.id)
            if t:
                matrix_rows.append({
                    'vehicle_id': v.id,
                    'driver_name': v.driver_name,
                    'plate_number': v.plate_number,
                    'vehicle_type_display': v.get_vehicle_type_display(),
                    'default_rate': float(v.default_service_rate),
                    'trip_id': t.id,
                    'trip_count': t.trip_count,
                    'unit_rate': float(t.unit_rate),
                    'total_amount': float(t.total_amount),
                    'dispatch_reference': t.dispatch_reference or '',
                    'origin_destination': t.origin_destination or '',
                    'notes': t.notes or '',
                    'is_existing': True
                })
            else:
                matrix_rows.append({
                    'vehicle_id': v.id,
                    'driver_name': v.driver_name,
                    'plate_number': v.plate_number,
                    'vehicle_type_display': v.get_vehicle_type_display(),
                    'default_rate': float(v.default_service_rate),
                    'trip_id': None,
                    'trip_count': 0,
                    'unit_rate': float(v.default_service_rate),
                    'total_amount': 0.0,
                    'dispatch_reference': '',
                    'origin_destination': '',
                    'notes': '',
                    'is_existing': False
                })

        return Response({
            'warehouse_id': int(warehouse_id),
            'date_shamsi': date_shamsi,
            'rows': matrix_rows
        })

    @action(detail=False, methods=['post'], url_path='bulk-save')
    def bulk_save(self, request):
        """
        ثبت سریع سرویس‌های ناوگان در پایان روز
        (پشتیبانی از ذخیره سراسری ناوگان و ارسال برودکست وب‌سوکت)
        """
        serializer = BulkVehicleTripMatrixSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        raw_wh = serializer.validated_data.get('warehouse_id')
        warehouse_id = int(raw_wh) if (raw_wh and str(raw_wh).upper() not in ['ALL', '0', 'NONE', '']) else None
        date_shamsi = serializer.validated_data['date_shamsi']
        items = serializer.validated_data['items']

        year_month = date_shamsi[:7]
        period = None
        if warehouse_id:
            period, _ = MonthlyWorkPeriod.objects.get_or_create(
                warehouse_id=warehouse_id,
                year_month=year_month,
                defaults={'status': 'OPEN'}
            )
            if period.status == 'LOCKED':
                return Response({'error': f'دوره {year_month} قفل شده است.'}, status=status.HTTP_400_BAD_REQUEST)

        saved_count = 0
        with transaction.atomic():
            for item in items:
                vehicle_id = item['vehicle_id']
                trip_count = item.get('trip_count', 0)
                
                v_obj = VehicleDriverProfile.objects.filter(id=vehicle_id).first()
                if not v_obj:
                    continue

                target_wh = warehouse_id if warehouse_id else v_obj.assigned_warehouse_id
                target_period = period
                if not target_period and target_wh:
                    target_period, _ = MonthlyWorkPeriod.objects.get_or_create(
                        warehouse_id=target_wh,
                        year_month=year_month,
                        defaults={'status': 'OPEN'}
                    )
                if target_period and target_period.status == 'LOCKED':
                    return Response({'error': f'دوره {year_month} برای انبار خودرو ({v_obj.driver_name}) قفل شده است.'}, status=status.HTTP_400_BAD_REQUEST)

                trip_obj = VehicleTripLog.objects.filter(
                    vehicle_id=vehicle_id,
                    warehouse_id=target_wh,
                    date_shamsi=date_shamsi,
                    is_deleted=False
                ).first()

                unit_rate = item.get('unit_rate', v_obj.default_service_rate)
                total_amount = unit_rate * trip_count

                if trip_count > 0:
                    if trip_obj:
                        trip_obj.trip_count = trip_count
                        trip_obj.unit_rate = unit_rate
                        trip_obj.total_amount = total_amount
                        trip_obj.dispatch_reference = item.get('dispatch_reference', '')
                        trip_obj.origin_destination = item.get('origin_destination', '')
                        trip_obj.notes = item.get('notes', '')
                        trip_obj.period = target_period
                        trip_obj.save()
                    else:
                        VehicleTripLog.objects.create(
                            vehicle_id=vehicle_id,
                            warehouse_id=target_wh,
                            date_shamsi=date_shamsi,
                            trip_count=trip_count,
                            unit_rate=unit_rate,
                            total_amount=total_amount,
                            dispatch_reference=item.get('dispatch_reference', ''),
                            origin_destination=item.get('origin_destination', ''),
                            notes=item.get('notes', ''),
                            period=target_period,
                            created_by=request.user
                        )
                    saved_count += 1
                else:
                    # اگر صفر بود و قبلاً ثبت شده بود، حذف منطقی
                    if trip_obj:
                        trip_obj.is_deleted = True
                        trip_obj.save()

        # ارسال برودکست وب‌سوکت برای ناوگان
        try:
            from channels.layers import get_channel_layer
            from asgiref.sync import async_to_sync
            channel_layer = get_channel_layer()
            client_tab_id = request.data.get('client_tab_id') or request.headers.get('X-Client-Tab-ID')
            if channel_layer:
                async_to_sync(channel_layer.group_send)(
                    'global_notifications',
                    {
                        'type': 'send_notification',
                        'type_str': 'fleet_trips_updated',
                        'message': f'تردد ناوگان برای تاریخ {date_shamsi} ثبت/به‌روزرسانی شد.',
                        'warehouse_id': warehouse_id,
                        'date_shamsi': date_shamsi,
                        'year_month': year_month,
                        'sender_id': request.user.id if request.user and request.user.is_authenticated else None,
                        'client_tab_id': client_tab_id
                    }
                )
        except Exception as e:
            import logging
            logging.getLogger(__name__).warning(f"[WebSocket] Error broadcasting fleet_trips_updated: {e}")

        return Response({
            'message': 'سرویس‌های خودروها با موفقیت ذخیره شدند.',
            'saved_count': saved_count
        })

    @action(detail=False, methods=['get'], url_path='monthly-summary')
    def get_monthly_summary(self, request):
        """
        گزارش کاردکس و عملکرد ماهانه خودروها
        """
        warehouse_id = request.query_params.get('warehouse_id')
        year_month = request.query_params.get('year_month')
        if not warehouse_id or not year_month:
            return Response({'error': 'warehouse_id و year_month الزامی هستند.'}, status=status.HTTP_400_BAD_REQUEST)

        vehicles = VehicleDriverProfile.objects.filter(
            Q(assigned_warehouse_id=warehouse_id) | Q(assigned_warehouse__isnull=True),
            is_active=True
        ).order_by('driver_name')

        trips = VehicleTripLog.objects.filter(
            warehouse_id=warehouse_id,
            date_shamsi__startswith=year_month,
            is_deleted=False
        )

        summary_rows = []
        for v in vehicles:
            v_trips = [t for t in trips if t.vehicle_id == v.id]
            total_trips = sum(t.trip_count for t in v_trips)
            total_payable = sum(float(t.total_amount) for t in v_trips)
            active_days = len(set(t.date_shamsi for t in v_trips if t.trip_count > 0))

            summary_rows.append({
                'vehicle_id': v.id,
                'driver_name': v.driver_name,
                'plate_number': v.plate_number,
                'vehicle_type': v.get_vehicle_type_display(),
                'default_rate': float(v.default_service_rate),
                'sheba_number': v.sheba_number or '',
                'active_days': active_days,
                'total_trips': total_trips,
                'total_payable': total_payable
            })

        return Response({
            'warehouse_id': int(warehouse_id),
            'year_month': year_month,
            'summary': summary_rows
        })


class MonthlyWorkPeriodViewSet(viewsets.ModelViewSet):
    queryset = MonthlyWorkPeriod.objects.all().select_related('warehouse', 'locked_by')
    serializer_class = MonthlyWorkPeriodSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = None

    def get_queryset(self):
        qs = super().get_queryset()
        warehouse_id = self.request.query_params.get('warehouse_id')
        if warehouse_id:
            qs = qs.filter(warehouse_id=warehouse_id)
        return qs

    @action(detail=True, methods=['post'], url_path='lock')
    def lock_period(self, request, pk=None):
        period = self.get_object()
        period.status = 'LOCKED'
        period.locked_at = timezone.now()
        period.locked_by = request.user
        period.save()
        return Response({'message': f'دوره {period.year_month} با موفقیت قفل شد.'})

    @action(detail=True, methods=['post'], url_path='unlock')
    def unlock_period(self, request, pk=None):
        if not request.user.is_superuser:
            return Response({'error': 'فقط مدیر ارشد سیستم اجازه بازگشایی دوره را دارد.'}, status=status.HTTP_403_FORBIDDEN)
        period = self.get_object()
        period.status = 'OPEN'
        period.locked_at = None
        period.locked_by = None
        period.save()
        return Response({'message': f'دوره {period.year_month} با موفقیت بازگشایی شد.'})

    @action(detail=True, methods=['post'], url_path='submit')
    def submit_period(self, request, pk=None):
        period = self.get_object()
        if period.status == 'LOCKED':
            return Response({'error': 'دوره از قبل قفل شده است.'}, status=status.HTTP_400_BAD_REQUEST)
        period.status = 'SUBMITTED'
        period.submitted_at = timezone.now()
        period.submitted_by = request.user
        period.submission_notes = request.data.get('notes', '')
        period.rejection_reason = None
        period.save()
        return Response({
            'message': f'دوره {period.year_month} جهت بررسی مالی ارسال گردید.',
            'status': period.status,
            'status_display': period.get_status_display()
        })

    @action(detail=True, methods=['post'], url_path='approve')
    def approve_period(self, request, pk=None):
        is_financial = request.user.is_superuser or request.user.groups.filter(name__in=['Accountant', 'Finance', 'مدیر مالی']).exists() or request.user.has_perm('personnel.can_override_attendance_lock')
        if not is_financial:
            return Response({'error': 'فقط حسابدار یا مدیر مالی مجاز به تایید نهایی دوره است.'}, status=status.HTTP_403_FORBIDDEN)
        period = self.get_object()
        period.status = 'LOCKED'
        period.locked_at = timezone.now()
        period.locked_by = request.user
        period.save()
        return Response({
            'message': f'دوره {period.year_month} تایید نهایی و قفل شد.',
            'status': period.status,
            'status_display': period.get_status_display()
        })

    @action(detail=True, methods=['post'], url_path='reject')
    def reject_period(self, request, pk=None):
        is_financial = request.user.is_superuser or request.user.groups.filter(name__in=['Accountant', 'Finance', 'مدیر مالی']).exists() or request.user.has_perm('personnel.can_override_attendance_lock')
        if not is_financial:
            return Response({'error': 'فقط حسابدار یا مدیر مالی مجاز به رد کارکرد است.'}, status=status.HTTP_403_FORBIDDEN)
        reason = request.data.get('reason', '').strip()
        if not reason:
            return Response({'error': 'درج علت رد دوره برای اطلاع سرپرست الزامی است.'}, status=status.HTTP_400_BAD_REQUEST)
        period = self.get_object()
        period.status = 'REJECTED'
        period.rejection_reason = reason
        period.locked_at = None
        period.locked_by = None
        period.save()
        return Response({
            'message': f'دوره {period.year_month} رد شد و جهت بازبینی به سرپرست بازگشت.',
            'status': period.status,
            'status_display': period.get_status_display()
        })

    @action(detail=False, methods=['post'], url_path='workflow-action')
    def workflow_action(self, request):
        """
        اقدام گردش کار کارکرد بر مبنای warehouse_id و year_month (پشتیبانی از فراخوانی سریع بدون دانستن pk)
        """
        warehouse_id = request.data.get('warehouse_id')
        year_month = request.data.get('year_month')
        action_type = request.data.get('action')
        notes = request.data.get('notes', '')

        if not warehouse_id or not year_month or not action_type:
            return Response({'error': 'پارامترهای warehouse_id، year_month و action الزامی هستند.'}, status=status.HTTP_400_BAD_REQUEST)

        period, _ = MonthlyWorkPeriod.objects.get_or_create(
            warehouse_id=warehouse_id,
            year_month=year_month,
            defaults={'status': 'OPEN'}
        )

        is_financial = request.user.is_superuser or request.user.groups.filter(name__in=['Accountant', 'Finance', 'مدیر مالی']).exists() or request.user.has_perm('personnel.can_override_attendance_lock')

        if action_type == 'submit':
            if period.status == 'LOCKED':
                return Response({'error': 'دوره از قبل قفل شده است.'}, status=status.HTTP_400_BAD_REQUEST)
            period.status = 'SUBMITTED'
            period.submitted_at = timezone.now()
            period.submitted_by = request.user
            period.submission_notes = notes
            period.rejection_reason = None
            period.save()
            return Response({
                'message': f'کارکرد دوره {year_month} جهت بررسی مالی ارسال گردید.',
                'period': MonthlyWorkPeriodSerializer(period).data
            })

        elif action_type == 'approve':
            if not is_financial:
                return Response({'error': 'فقط حسابدار یا مدیر مالی مجاز به تایید نهایی و قفل دوره است.'}, status=status.HTTP_403_FORBIDDEN)
            period.status = 'LOCKED'
            period.locked_at = timezone.now()
            period.locked_by = request.user
            period.save()
            return Response({
                'message': f'کارکرد دوره {year_month} با موفقیت تایید نهایی و قفل شد.',
                'period': MonthlyWorkPeriodSerializer(period).data
            })

        elif action_type == 'reject':
            if not is_financial:
                return Response({'error': 'فقط حسابدار یا مدیر مالی مجاز به رد کارکرد است.'}, status=status.HTTP_403_FORBIDDEN)
            if not notes.strip():
                return Response({'error': 'درج علت رد دوره الزامی است.'}, status=status.HTTP_400_BAD_REQUEST)
            period.status = 'REJECTED'
            period.rejection_reason = notes.strip()
            period.locked_at = None
            period.locked_by = None
            period.save()
            return Response({
                'message': f'دوره {year_month} رد شد و جهت بازبینی به سرپرست بازگردانده شد.',
                'period': MonthlyWorkPeriodSerializer(period).data
            })

        elif action_type == 'unlock':
            is_unlock_authorized = (
                request.user.is_superuser or
                request.user.groups.filter(name__in=['Accountant', 'Finance', 'مدیر مالی']).exists() or
                request.user.has_perm('personnel.can_override_attendance_lock')
            )
            if not is_unlock_authorized:
                return Response({'error': 'فقط مدیر ارشد سیستم یا مدیر مالی مجاز به بازگشایی دوره است.'}, status=status.HTTP_403_FORBIDDEN)
            period.status = 'OPEN'
            period.locked_at = None
            period.locked_by = None
            period.save()
            return Response({
                'message': f'دوره {year_month} با موفقیت بازگشایی شد.',
                'period': MonthlyWorkPeriodSerializer(period).data
            })

        return Response({'error': 'نوع اقدام ناشناخته است.'}, status=status.HTTP_400_BAD_REQUEST)


# ══════════════════════════════════════════════════════════════════════════════
# ── ۷. ویوست تنظیمات جامع سالانه و ۲۰ گروه شغلی ───────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════

class PayrollYearlySettingsViewSet(viewsets.ModelViewSet):
    queryset = PayrollYearlySettings.objects.all().prefetch_related(
        'job_grades', 'workshop_insurance', 'tax_settings', 'bank_export_settings'
    )
    serializer_class = PayrollYearlySettingsSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = None

    @action(detail=False, methods=['get'], url_path='active-or-year')
    def get_active_or_year(self, request):
        year = request.query_params.get('year', '1405')
        settings_obj = PayrollYearlySettings.objects.filter(fiscal_year=year).first()
        if not settings_obj:
            settings_obj = PayrollYearlySettings.objects.filter(is_active=True).first()
        if not settings_obj:
            return Response({'error': 'تنظیماتی یافت نشد.'}, status=status.HTTP_404_NOT_FOUND)
        return Response(PayrollYearlySettingsSerializer(settings_obj).data)

    @action(detail=False, methods=['get'], url_path='job-grade-rate')
    def get_job_grade_rate(self, request):
        """
        دریافت خودکار مزد روزانه و پایه سنواتی بر اساس گروه شغلی (۱ تا ۲۰)
        """
        grade = request.query_params.get('grade', '19')
        year = request.query_params.get('year', '1405')
        
        tier = JobGradeTier.objects.filter(
            yearly_settings__fiscal_year=year,
            grade_number=int(grade)
        ).first()

        if not tier:
            tier = JobGradeTier.objects.filter(grade_number=int(grade)).first()

        if tier:
            return Response({
                'grade_number': tier.grade_number,
                'daily_base_wage': float(tier.daily_base_wage),
                'daily_seniority_bonus': float(tier.daily_seniority_bonus),
                'hourly_rate': round(float(tier.daily_base_wage) / 10.0, 2)
            })
        return Response({'error': 'گروه شغلی در جدول تنظیمات یافت نشد.'}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=True, methods=['post'], url_path='update-all')
    def update_all_tabs(self, request, pk=None):
        """
        به‌روزرسانی یکپارچه هر ۵ تب تنظیمات توسط حسابدار
        """
        settings_obj = self.get_object()
        data = request.data

        with transaction.atomic():
            # 1. Update main fields
            for field in [
                'monthly_food_allowance', 'monthly_housing_allowance', 'monthly_spouse_allowance',
                'monthly_child_allowance', 'shift_percent', 'transport_help_percent',
                'transport_fixed_amount', 'specialist_attraction_percent', 'bad_weather_percent',
                'remote_hardship_percent', 'south_pars_percent', 'travel_cost_per_day',
                'worker_insurance_rate', 'employer_insurance_rate', 'unemployment_insurance_rate',
                'surplus_overtime_percent', 'attendance_edit_past_days', 'attendance_edit_future_days'
            ]:
                if field in data:
                    setattr(settings_obj, field, data[field])
            settings_obj.save()

            # 2. Update Workshop Insurance
            ws_data = data.get('workshop_insurance')
            if ws_data and hasattr(settings_obj, 'workshop_insurance'):
                ws = settings_obj.workshop_insurance
                for f in ['workshop_code', 'workshop_name', 'employer_name', 'workshop_address', 'list_type', 'list_number', 'default_dsk_rate', 'default_mon_pym']:
                    if f in ws_data:
                        setattr(ws, f, ws_data[f])
                ws.save()

            # 3. Update Tax Settings
            tax_data = data.get('tax_settings')
            if tax_data and hasattr(settings_obj, 'tax_settings'):
                ts = settings_obj.tax_settings
                for f in ['payment_type', 'service_location', 'exceptions', 'currency_type', 'currency_exchange_rate', 'housing_benefit_type', 'vehicle_benefit_type']:
                    if f in tax_data:
                        setattr(ts, f, tax_data[f])
                ts.save()

            # 4. Update Bank Settings
            bank_data = data.get('bank_export_settings')
            if bank_data and hasattr(settings_obj, 'bank_export_settings'):
                bs = settings_obj.bank_export_settings
                for f in ['bank_name', 'source_account_number', 'default_deposit_id', 'deposit_description_template']:
                    if f in bank_data:
                        setattr(bs, f, bank_data[f])
                bs.save()

            # 5. Update Job Grades
            jg_list = data.get('job_grades', [])
            for jg_item in jg_list:
                grade_num = jg_item.get('grade_number')
                if grade_num:
                    JobGradeTier.objects.update_or_create(
                        yearly_settings=settings_obj,
                        grade_number=int(grade_num),
                        defaults={
                            'daily_base_wage': jg_item.get('daily_base_wage', 0),
                            'daily_seniority_bonus': jg_item.get('daily_seniority_bonus', 0),
                        }
                    )

        return Response({
            'message': 'تنظیمات پایه حقوق و دستمزد با موفقیت به‌روزرسانی شد.',
            'settings': PayrollYearlySettingsSerializer(settings_obj).data
        })


# ══════════════════════════════════════════════════════════════════════════════
# ── ۸. ویوست محاسبه ماهانه حقوق و صدور دیسکت‌ها و گزارشات ۵۸ ستونی ────────────
# ══════════════════════════════════════════════════════════════════════════════

class MonthlyPayrollViewSet(viewsets.ModelViewSet):
    queryset = MonthlyPayrollRecord.objects.all().select_related('period', 'personnel', 'period__warehouse')
    serializer_class = MonthlyPayrollRecordSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = None

    def get_queryset(self):
        qs = super().get_queryset()
        period_id = self.request.query_params.get('period_id')
        if period_id:
            qs = qs.filter(period_id=period_id)
        
        warehouse_id = self.request.query_params.get('warehouse_id')
        if warehouse_id:
            qs = qs.filter(period__warehouse_id=warehouse_id)
            
        year_month = self.request.query_params.get('year_month')
        if year_month:
            qs = qs.filter(period__year_month=year_month)

        return qs.order_by('row_number')

    @action(detail=False, methods=['post'], url_path='calculate-period')
    def calculate_period(self, request):
        """
        تجمیع کارکرد و محاسبه خودکار ۵۸ ستون حقوق برای دوره انتخابی
        """
        warehouse_id = request.data.get('warehouse_id')
        year_month = request.data.get('year_month')
        
        if not warehouse_id or not year_month:
            return Response({'error': 'شناسه انبار و سال/ماه الزامی است.'}, status=status.HTTP_400_BAD_REQUEST)

        # Get or create work period
        period, _ = MonthlyWorkPeriod.objects.get_or_create(
            warehouse_id=int(warehouse_id),
            year_month=str(year_month).strip(),
            defaults={'status': 'OPEN'}
        )

        if period.status == 'LOCKED':
            return Response({'error': f'دوره {period.year_month} قفل شده است و امکان محاسبه مجدد ندارد.'}, status=status.HTTP_400_BAD_REQUEST)

        records = calculate_monthly_payroll_for_period(period, user=request.user)
        serialized = MonthlyPayrollRecordSerializer(records, many=True).data

        # Summary Metrics
        total_gross = sum(float(r.gross_salary) for r in records)
        total_payable = sum(float(r.payable_amount) for r in records)
        total_insurance = sum(float(r.total_insurance) for r in records)
        total_tax = sum(float(r.income_tax) for r in records)

        return Response({
            'message': f'محاسبه حقوق دوره {period.year_month} برای {len(records)} نفر پرسنل با موفقیت انجام شد.',
            'period_id': period.id,
            'period_status': period.status,
            'summary': {
                'total_personnel': len(records),
                'total_gross': total_gross,
                'total_payable': total_payable,
                'total_insurance': total_insurance,
                'total_tax': total_tax,
            },
            'records': serialized
        })

    @action(detail=False, methods=['get'], url_path='export-dsk-zip')
    def export_dsk_zip(self, request):
        """
        صدور پکیج فشرده ZIP شامل دو دیسکت استاندارد DSKWOR00.DBF و DSKKAR00.DBF
        """
        period_id = request.query_params.get('period_id')
        if not period_id:
            return Response({'error': 'شناسه دوره (period_id) الزامی است.'}, status=status.HTTP_400_BAD_REQUEST)

        period = MonthlyWorkPeriod.objects.filter(id=period_id).first()
        if not period:
            return Response({'error': 'دوره یافت نشد.'}, status=status.HTTP_404_NOT_FOUND)

        records = list(MonthlyPayrollRecord.objects.filter(period=period).select_related('personnel'))
        if not records:
            return Response({'error': 'رکوردی برای این دوره ثبت یا محاسبه نشده است.'}, status=status.HTTP_400_BAD_REQUEST)

        year_str = period.year_month.split('/')[0] if '/' in period.year_month else '1405'
        month_str = period.year_month.split('/')[1] if '/' in period.year_month else '04'
        fiscal_year_short = int(year_str[-2:]) # 05
        month_num = int(month_str)

        settings_obj = PayrollYearlySettings.objects.filter(fiscal_year=year_str).first()
        ws_settings = getattr(settings_obj, 'workshop_insurance', None) if settings_obj else None
        if not ws_settings:
            ws_settings = WorkshopInsuranceSettings.objects.first()
        if not ws_settings:
            ws_settings = WorkshopInsuranceSettings(
                workshop_code='4894290013',
                workshop_name='شرکت پیمانکاری پارس',
                employer_name='مدیریت انبارداری',
                workshop_address='عسلویه، منطقه ویژه اقتصادی'
            )

        insured_recs = [r for r in records if r.include_in_insurance]
        
        # Calculate sums for DSKKAR
        sum_days = sum(r.insurance_days for r in insured_recs)
        sum_daily = sum(int(r.daily_wage) for r in insured_recs)
        sum_base = sum(int(r.base_salary) for r in insured_recs)
        sum_maz = sum(int(r.total_taxable_allowances) for r in insured_recs)
        sum_mash = sum(int(r.total_insurable_salary_allowances) for r in insured_recs)
        sum_totl = sum(int(r.gross_salary) for r in insured_recs)
        sum_bime = sum(int(r.worker_insurance) for r in insured_recs)
        sum_koso = sum(int(r.employer_insurance) for r in insured_recs)
        sum_bic = sum(int(r.unemployment_insurance) for r in insured_recs)
        sum_inc = sum(int(r.total_seniority_accumulated) for r in insured_recs)
        sum_spouse = sum(int(r.marital_allowance) for r in insured_recs)

        month_names = {1: 'فروردین', 2: 'اردیبهشت', 3: 'خرداد', 4: 'تیر', 5: 'مرداد', 6: 'شهریور', 7: 'مهر', 8: 'آبان', 9: 'آذر', 10: 'دی', 11: 'بهمن', 12: 'اسفند'}
        m_name = month_names.get(month_num, 'تیر')
        list_desc = f"انبارداری بیمه {m_name} {year_str}"

        dskkar_bytes = generate_dskkar_bytes(
            workshop_settings=ws_settings,
            fiscal_year_short=fiscal_year_short,
            month_num=month_num,
            list_number='01',
            list_description=list_desc,
            records_count=len(insured_recs),
            sum_insurance_days=sum_days,
            sum_daily_wage=sum_daily,
            sum_base_salary=sum_base,
            sum_insurable_allowances=sum_maz,
            sum_insurable_total=sum_mash,
            sum_gross_salary=sum_totl,
            sum_employee_insurance=sum_bime,
            sum_employer_insurance=sum_koso,
            sum_unemployment_insurance=sum_bic,
            sum_seniority=sum_inc,
            sum_spouse_allowance=sum_spouse
        )

        dskwor_bytes = generate_dskwor_bytes(
            workshop_code=ws_settings.workshop_code,
            fiscal_year_short=fiscal_year_short,
            month_num=month_num,
            list_number='01',
            payroll_records=records
        )

        # Build ZIP in memory
        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as z:
            z.writestr("DSKKAR00.DBF", dskkar_bytes)
            z.writestr("DSKWOR00.DBF", dskwor_bytes)

        zip_val = zip_buf.getvalue()
        resp = HttpResponse(zip_val, content_type='application/zip')
        resp['Content-Disposition'] = f'attachment; filename="Bimeh_Diskettes_{year_str}_{month_str}.zip"'
        return resp

    @action(detail=False, methods=['get'], url_path='export-tax-wh')
    def export_tax_wh(self, request):
        """
        صدور فایل متنی WH مالیات حقوق
        """
        period_id = request.query_params.get('period_id')
        records = list(MonthlyPayrollRecord.objects.filter(period_id=period_id).select_related('personnel'))
        if not records:
            return Response({'error': 'رکوردی یافت نشد.'}, status=status.HTTP_404_NOT_FOUND)

        settings_obj = PayrollYearlySettings.objects.filter(is_active=True).first()
        tax_settings = getattr(settings_obj, 'tax_settings', None)

        content = generate_wh_tax_content(records, tax_settings)
        period = records[0].period
        month_code = period.year_month.replace('/', '')[2:] if period else '140504'
        filename = f"WH{month_code}.txt"

        resp = HttpResponse(content.encode('utf-8'), content_type='text/plain; charset=utf-8')
        resp['Content-Disposition'] = f'attachment; filename="{filename}"'
        return resp

    @action(detail=False, methods=['get'], url_path='export-tax-wp')
    def export_tax_wp(self, request):
        """
        صدور فایل متنی WP پرسنل مالیاتی
        """
        period_id = request.query_params.get('period_id')
        records = list(MonthlyPayrollRecord.objects.filter(period_id=period_id).select_related('personnel'))
        if not records:
            return Response({'error': 'رکوردی یافت نشد.'}, status=status.HTTP_404_NOT_FOUND)

        settings_obj = PayrollYearlySettings.objects.filter(is_active=True).first()
        tax_settings = getattr(settings_obj, 'tax_settings', None)

        content = generate_wp_tax_content(records, tax_settings)
        period = records[0].period
        month_code = period.year_month.replace('/', '')[2:] if period else '140504'
        filename = f"WP{month_code}.txt"

        resp = HttpResponse(content.encode('utf-8'), content_type='text/plain; charset=utf-8')
        resp['Content-Disposition'] = f'attachment; filename="{filename}"'
        return resp

    @action(detail=False, methods=['get'], url_path='export-bank-excel')
    def export_bank_excel(self, request):
        """
        صدور فایل اکسل پرداخت گروهی بانک ملی
        """
        period_id = request.query_params.get('period_id')
        records = list(MonthlyPayrollRecord.objects.filter(period_id=period_id).select_related('personnel'))
        if not records:
            return Response({'error': 'رکوردی یافت نشد.'}, status=status.HTTP_404_NOT_FOUND)

        settings_obj = PayrollYearlySettings.objects.filter(is_active=True).first()
        bank_settings = getattr(settings_obj, 'bank_export_settings', None)
        period = records[0].period

        excel_bytes = generate_bank_meli_excel(records, bank_settings, year_month_title=period.year_month)
        filename = f"Bank_Payment_{period.year_month.replace('/', '_')}.xlsx"

        resp = HttpResponse(excel_bytes, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename="{filename}"'
        return resp

    @action(detail=False, methods=['get'], url_path='export-monthly-excel')
    def export_monthly_excel(self, request):
        """
        صدور فایل اکسل ۲ سطری استاندارد ۵۸ ستون شیت ماهانه
        """
        period_id = request.query_params.get('period_id')
        records = list(MonthlyPayrollRecord.objects.filter(period_id=period_id).select_related('personnel'))
        if not records:
            return Response({'error': 'رکوردی یافت نشد.'}, status=status.HTTP_404_NOT_FOUND)

        period = records[0].period
        excel_bytes = generate_monthly_payroll_excel(records, period_title=period.year_month)
        filename = f"Payroll_Report_{period.year_month.replace('/', '_')}.xlsx"

        resp = HttpResponse(excel_bytes, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename="{filename}"'
        return resp

    @action(detail=False, methods=['post'], url_path='import-tax-excel')
    def import_tax_excel(self, request):
        """
        درون‌ریزی فایل اکسل مالیات محاسبه‌شده توسط دارایی (نمونه مالیات محاسبه شده.xlsx)
        و تطبیق خودکار با ستون مالیات و متادیتای رکوردهای دوره بر اساس کد ملی
        """
        excel_file = request.FILES.get('file')
        period_id = request.data.get('period_id') or request.query_params.get('period_id')
        year_month = request.data.get('year_month') or request.query_params.get('year_month')

        if not excel_file:
            return Response({'error': 'لطفاً فایل اکسل مالیات دارایی را انتخاب کنید.'}, status=status.HTTP_400_BAD_REQUEST)

        # Locate target period
        if period_id:
            try:
                p_id = int(str(period_id).strip())
                period = MonthlyWorkPeriod.objects.filter(id=p_id).first()
            except Exception:
                period = MonthlyWorkPeriod.objects.filter(id=period_id).first()
        elif year_month:
            period = MonthlyWorkPeriod.objects.filter(year_month=year_month).first()
        else:
            period = MonthlyWorkPeriod.objects.order_by('-created_at').first()

        if not period:
            return Response({'error': 'دوره ماهانه مورد نظر یافت نشد.'}, status=status.HTTP_404_NOT_FOUND)

        try:
            file_bytes = excel_file.read()
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            ws = wb.active

            # Find header indices
            col_map = {}
            for c in range(1, ws.max_column + 1):
                header_val = str(ws.cell(1, c).value or '').strip()
                clean_h = header_val.replace(' ', '').replace('_', '')
                if 'کدملی' in clean_h or 'کد_ملی' in clean_h:
                    col_map['national_code'] = c
                elif 'مالیاتمحاسبهشده' in clean_h or 'مالیات' in clean_h:
                    col_map['tax_amount'] = c
                elif 'تعدادماهاعمالشدهدرمعافیت' in clean_h or 'تعدادماه' in clean_h:
                    col_map['exemption_months'] = c
                elif 'بیشازیککارفرما' in clean_h or 'چندکارفرما' in clean_h:
                    col_map['multiple_employers'] = c

            if 'national_code' not in col_map or 'tax_amount' not in col_map:
                col_map.setdefault('national_code', 1)
                col_map.setdefault('tax_amount', 4)
                col_map.setdefault('exemption_months', 5)
                col_map.setdefault('multiple_employers', 6)

            matched_count = 0
            unmatched_rows = []
            total_imported_tax = Decimal(0)

            records = MonthlyPayrollRecord.objects.filter(period=period).select_related('personnel')
            record_map = {str(r.national_code).strip().zfill(10): r for r in records}

            with transaction.atomic():
                for r_idx in range(2, ws.max_row + 1):
                    raw_code = ws.cell(r_idx, col_map['national_code']).value
                    if raw_code is None:
                        continue
                    
                    code_str = str(raw_code).split('.')[0].strip().zfill(10)
                    if not code_str or not code_str.isdigit():
                        continue

                    raw_tax = ws.cell(r_idx, col_map['tax_amount']).value or 0
                    try:
                        tax_val = Decimal(str(raw_tax).replace(',', '').strip())
                    except Exception:
                        tax_val = Decimal(0)

                    raw_months = ws.cell(r_idx, col_map.get('exemption_months', 5)).value if 'exemption_months' in col_map else 1
                    try:
                        ex_months = int(raw_months or 1)
                    except Exception:
                        ex_months = 1

                    raw_mult = ws.cell(r_idx, col_map.get('multiple_employers', 6)).value if 'multiple_employers' in col_map else 'خیر'
                    has_mult = str(raw_mult).strip() in ['بله', 'true', 'True', '1', 'YES', 'yes']

                    if code_str in record_map:
                        rec = record_map[code_str]
                        rec.income_tax = tax_val
                        rec.tax_source_type = 'IMPORTED_EXCEL'
                        rec.tax_exemption_months = ex_months
                        rec.has_multiple_employers = has_mult
                        rec.is_tax_imported = True

                        rec.net_salary = rec.gross_salary - rec.worker_insurance - rec.income_tax
                        rec.payable_amount = rec.net_salary - rec.advance_payment_deduction
                        rec.save()

                        matched_count += 1
                        total_imported_tax += tax_val
                    else:
                        unmatched_rows.append({
                            'row': r_idx,
                            'national_code': code_str,
                            'tax_amount': float(tax_val)
                        })

            return Response({
                'message': f'اطلاعات مالیاتی با موفقیت بارگذاری شد. {matched_count} پرسنل تطبیق داده شدند.',
                'matched_count': matched_count,
                'unmatched_count': len(unmatched_rows),
                'unmatched_rows': unmatched_rows,
                'total_imported_tax': float(total_imported_tax),
                'period_year_month': period.year_month
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                'error': f'خطا در پردازش فایل اکسل مالیات: {str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)


