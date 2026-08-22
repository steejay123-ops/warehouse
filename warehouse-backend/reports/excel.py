"""
خروجی Excel گزارش‌ساز — ترکیبی، بدون سقف ردیف و با استایل

- نتایج کوچک (تا SYNC_ROW_LIMIT): تولید همزمان و دانلود فوری (با ردیف‌های زبرا).
- نتایج بزرگ: ReportExportJob — اگر worker جدا (run_export_worker) زنده باشد job
  فقط در صف می‌ماند و worker برمی‌دارد؛ وگرنه fallback به threading.Thread فعلی.
  فایل با Workbook(write_only=True) و iterator(chunk_size) ساخته می‌شود
  (مصرف RAM ثابت)، فرانت درصد پیشرفت را poll می‌کند و در پایان دانلود streaming
  انجام می‌شود. به این ترتیب هیچ درخواست HTTP طولانی‌ای نداریم و تایم‌اوت
  Cloudflare Tunnel بی‌اثر است.

استایل در حالت write_only فقط از طریق WriteOnlyCell ممکن است و
column_dimensions باید قبل از append اولین ردیف ست شود. freeze_panes و
auto_filter بین نسخه‌های openpyxl رفتار متفاوتی در این حالت دارند → try/except.
"""
import datetime as _dt
import threading
import uuid
from decimal import Decimal

import openpyxl
from openpyxl.cell import WriteOnlyCell
from openpyxl.utils import get_column_letter
from django.db import connection
from django.db.models import Q
from django.utils import timezone

from django.http import HttpResponse

from .models import ExportWorkerStatus, ReportExportJob

SYNC_ROW_LIMIT = 10_000
CHUNK_SIZE = 2000
WORKER_ALIVE_WINDOW = 15  # ثانیه — نبض تازه‌تر از این یعنی worker زنده است

from common.excel_utils import (
    HEADER_FILL, HEADER_FONT, KEY_FONT, TITLE_FONT, ZEBRA_FILL, CENTER, 
    NUMBER_FORMATS, get_cell_value, styled_cell, set_column_widths, freeze_header_panes,
    jalali_now_str
)


def _write_workbook(ws_target, qs, columns, progress_cb=None, total=0,
                    report_name='گزارش', zebra=False):
    """
    نوشتن شیت کامل: ردیف ۱ عنوان گزارش، ردیف ۲ برچسب فارسی، ردیف ۳ کلید سیستمی + داده.
    zebra فقط برای نتایج کوچک (هزینه WriteOnlyCell به‌ازای هر سلول).
    """
    set_column_widths(ws_target, columns, is_write_only=True)
    freeze_header_panes(ws_target, row=4, is_write_only=True)

    # ردیف ۱: عنوان گزارش + زمان
    title_text = f"{report_name} — {jalali_now_str()}"
    title_cell = styled_cell(ws_target, title_text, is_write_only=True)
    title_cell.font = TITLE_FONT
    ws_target.append([title_cell])

    # ردیف ۲: برچسب فارسی
    header_row = []
    for c in columns:
        cell = styled_cell(ws_target, c['label'], is_write_only=True)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        header_row.append(cell)
    ws_target.append(header_row)

    # ردیف ۳: کلید سیستمی (برای import/دیباگ)
    key_row = []
    for c in columns:
        cell = styled_cell(ws_target, c['key'], is_write_only=True)
        cell.font = KEY_FONT
        cell.alignment = CENTER
        key_row.append(cell)
    ws_target.append(key_row)

    keys = [c.get('source') or c['key'] for c in columns]
    formats = [NUMBER_FORMATS.get(c.get('type')) for c in columns]
    needs_cell = [f is not None for f in formats]

    written = 0
    for row in qs.iterator(chunk_size=CHUNK_SIZE):
        fill = ZEBRA_FILL if (zebra and written % 2 == 1) else None
        out = []
        for j, k in enumerate(keys):
            v = get_cell_value(row.get(k))
            if fill is not None or needs_cell[j]:
                out.append(styled_cell(ws_target, v, number_format=formats[j], fill=fill, is_write_only=True))
            else:
                out.append(v)
        ws_target.append(out)

        written += 1
        if progress_cb and written % CHUNK_SIZE == 0:
            progress_cb(written, total)
    if progress_cb:
        progress_cb(written, total)


    # فیلتر روی محدوده داده (سربرگ فیلتر = ردیف کلیدها) — بعد از داده مجاز است
    last_col = get_column_letter(len(columns)) if columns else 'A'
    try:
        ws_target.auto_filter.ref = f'A3:{last_col}{3 + written}'
    except Exception:
        pass
    return written


def _make_workbook():
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet(title='Report')
    try:
        ws.sheet_view.rightToLeft = True
    except Exception:
        pass  # در حالت write_only بعضی نسخه‌ها sheet_view ندارند — حیاتی نیست
    return wb, ws


def sync_excel_response(qs, columns, filename='report.xlsx', report_name='گزارش'):
    """تولید همزمان برای نتایج کوچک — دانلود فوری."""
    wb, ws = _make_workbook()
    _write_workbook(ws, qs, columns, report_name=report_name, zebra=True)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def worker_alive():
    """آیا worker جدا (run_export_worker) نبض تازه دارد؟"""
    st = ExportWorkerStatus.objects.filter(pk=1).first()
    if st is None:
        return False
    return (timezone.now() - st.alive_at).total_seconds() < WORKER_ALIVE_WINDOW


def start_export_job(user, spec, report_name, total_rows):
    """
    ساخت job. اگر worker زنده باشد job در صف می‌ماند تا worker بردارد
    (پایدار در برابر ری‌استارت)؛ وگرنه fallback به thread پس‌زمینه فعلی.
    """
    job = ReportExportJob.objects.create(
        owner=user, spec=spec, report_name=report_name or 'report',
        status='pending', total_rows=total_rows,
    )
    if not worker_alive():
        t = threading.Thread(target=_run_export_job, args=(job.pk, user.pk), daemon=True)
        t.start()
    return job


def _run_export_job(job_pk, user_pk):
    """بدنه پردازش (thread یا worker) — کوئری را دوباره از spec می‌سازد و فایل را می‌نویسد."""
    from accounts.models import CustomUser
    from .engine import ReportEngine, ReportError

    try:
        job = ReportExportJob.objects.get(pk=job_pk)
        user = CustomUser.objects.get(pk=user_pk)
        eng = ReportEngine(user, job.spec)
        qs, columns, total = eng.export_queryset()

        job.total_rows = total
        job.status = 'running'
        job.progress = 0
        job.heartbeat_at = timezone.now()
        job.save(update_fields=['total_rows', 'status', 'progress', 'heartbeat_at'])

        last_hb = [timezone.now()]

        def on_progress(done, tot):
            pct = int(done / tot * 100) if tot else 0
            now = timezone.now()
            # ثبت نبض در DB هر ۵ ثانیه
            if (now - last_hb[0]).total_seconds() >= 5:
                last_hb[0] = now
                ReportExportJob.objects.filter(pk=job_pk).update(
                    progress=pct, heartbeat_at=now,
                )

        wb, ws = _make_workbook()
        _write_workbook(
            ws, qs, columns, report_name=job.report_name,
            total=total, progress_cb=on_progress, zebra=True,
        )

        exports_dir = ReportExportJob.exports_dir()
        fname = f'report_{job_pk}_{uuid.uuid4().hex[:8]}.xlsx'
        fpath = exports_dir / fname
        wb.save(str(fpath))

        ReportExportJob.objects.filter(pk=job_pk).update(
            status='done',
            progress=100,
            file_path=f'report_exports/{fname}',
            heartbeat_at=timezone.now(),
            finished_at=timezone.now(),
        )
    except ReportError as e:
        ReportExportJob.objects.filter(pk=job_pk).update(
            status='failed', error_message=e.message, finished_at=timezone.now(),
        )
    except Exception as e:  # noqa: BLE001 — هر خطای غیرمنتظره باید status را ببندد
        ReportExportJob.objects.filter(pk=job_pk).update(
            status='failed', error_message=str(e)[:1000], finished_at=timezone.now(),
        )
    finally:
        # اتصال DB مخصوص این thread/پروسه — داخل تراکنش (تست‌ها) بسته نشود
        if not connection.in_atomic_block:
            connection.close()


def cleanup_old_jobs(max_age_hours=24):
    """پاک‌سازی فرصت‌طلبانه فایل‌ها و jobهای قدیمی + jobهای واقعاً مرده."""
    cutoff = timezone.now() - timezone.timedelta(hours=max_age_hours)
    old = ReportExportJob.objects.filter(created_at__lt=cutoff)
    for job in old:
        p = job.absolute_file_path
        if p is not None and p.exists():
            try:
                p.unlink()
            except OSError:
                pass
    old.delete()

    # اسکن فایل‌های یتیم روی دیسک که رکوردی ندارند یا بیش از ۲۴ ساعت قدمت دارند
    try:
        exports_dir = ReportExportJob.exports_dir()
        if exports_dir.exists():
            for f in exports_dir.glob('report_*.xlsx'):
                try:
                    mtime = timezone.datetime.fromtimestamp(f.stat().st_mtime, tz=timezone.utc)
                    if mtime < cutoff:
                        f.unlink()
                except OSError:
                    pass
    except Exception:
        pass

    # running بدون نبض تازه (بیش از ۱ ساعت) → مرده؛ نبض ملاک است نه زمان شروع
    stale_cutoff = timezone.now() - timezone.timedelta(hours=1)
    ReportExportJob.objects.filter(status='running').filter(
        Q(heartbeat_at__lt=stale_cutoff)
        | Q(heartbeat_at__isnull=True, created_at__lt=stale_cutoff),
    ).update(status='failed', error_message='پردازش ناتمام ماند؛ دوباره اجرا کنید.')

    # pending قدیمی فقط وقتی worker زنده نیست (با worker زنده، صف سالم است)
    if not worker_alive():
        ReportExportJob.objects.filter(
            status='pending', created_at__lt=stale_cutoff,
        ).update(status='failed', error_message='پروسه سرور ری‌استارت شد؛ دوباره اجرا کنید.')
