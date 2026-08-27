"""
تست‌های گزارش‌ساز — تمرکز روی مرزهای امنیتی و صحت موتور کوئری.
اجرا: python manage.py test reports
"""
import io
from decimal import Decimal
from unittest.mock import patch

import openpyxl
from django.contrib.auth.models import Permission
from django.test import TestCase
from django.utils import timezone
from rest_framework.test import APIClient

from accounts.models import CustomUser
from inventory.models import CountTask, CountTaskHistory, Item, ItemFieldDefinition
from warehouses.models import SystemSetting, Warehouse

from .engine import ReportEngine, ReportError
from .excel import _run_export_job, cleanup_old_jobs, start_export_job
from .models import ExportWorkerStatus, ReportExportJob, ReportTemplate
from .registry import get_registry


def _perm(codename):
    return Permission.objects.get(codename=codename, content_type__app_label='accounts')


def make_user(username, perms=(), superuser=False, warehouses=(), first_name='', last_name=''):
    u = CustomUser.objects.create_user(
        username=username, password='x', is_superuser=superuser,
        first_name=first_name, last_name=last_name,
    )
    for c in perms:
        u.user_permissions.add(_perm(c))
    for w in warehouses:
        u.assigned_warehouses.add(w)
    # کش مجوزهای جنگو را دور بزن
    return CustomUser.objects.get(pk=u.pk)


class BaseReportTest(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.wh1 = Warehouse.objects.create(name='انبار ۱', project_name='پروژه الف')
        cls.wh2 = Warehouse.objects.create(name='انبار ۲')

        # شمارش کور را برای این تست‌ها خاموش کن (پیش‌فرض سیستم blind است)
        SystemSetting.objects.create(key='blind_counting', value='normal', warehouse=None)

        cls.counter_user = make_user('counter1', first_name='علی', last_name='رضایی')
        cls.supervisor_user = make_user('super1', first_name='محمد', last_name='محمدی')

        cls.i1 = Item.objects.create(
            warehouse=cls.wh1, fa_unic_code='FA-001', description='پیچ فولادی',
            vendor='زیمنس', inventory=Decimal('10'), dynamic_data={'coating': '5.5'},
        )
        cls.i2 = Item.objects.create(
            warehouse=cls.wh1, fa_unic_code='FA-002', description='مهره',
            vendor='ABB', inventory=Decimal('4'), dynamic_data={'coating': '2'},
        )
        cls.i3 = Item.objects.create(
            warehouse=cls.wh2, fa_unic_code='FA-003', description='واشر',
            vendor='زیمنس', inventory=Decimal('7'), dynamic_data={},
        )
        ItemFieldDefinition.objects.create(
            warehouse=cls.wh1, name='coating', label='پوشش', field_type='number',
        )

        cls.ct1 = CountTask.objects.create(
            item=cls.i1, counter=cls.counter_user, supervisor=cls.supervisor_user,
            status='COUNTED', counted_balance=Decimal('10'),
        )
        cls.ct2 = CountTask.objects.create(
            item=cls.i2, counter=cls.counter_user, supervisor=cls.supervisor_user,
            status='PENDING_COUNT', counted_balance=None,
        )

        cls.admin = make_user('admin', superuser=True)
        cls.viewer = make_user(
            'viewer', perms=('view_sys_reports', 'view_wh_docs', 'view_sys_supervisor'), warehouses=(cls.wh1,),
        )
        cls.no_access = make_user('noaccess', perms=('view_sys_reports',))


class EngineSecurityTests(BaseReportTest):
    def test_unknown_entity_rejected(self):
        with self.assertRaises(ReportError):
            ReportEngine(self.admin, {'entity': 'nope'})

    def test_entity_permission_403(self):
        with self.assertRaises(ReportError) as ctx:
            ReportEngine(self.no_access, {'entity': 'items'})
        self.assertEqual(ctx.exception.status, 403)

    def test_unknown_field_rejected(self):
        eng = ReportEngine(self.admin, {'entity': 'items', 'fields': ['warehouse__manager__password']})
        with self.assertRaises(ReportError):
            eng.build()

    def test_raw_orm_path_in_filter_rejected(self):
        eng = ReportEngine(self.admin, {
            'entity': 'items', 'fields': ['fa_unic_code'],
            'filters': {'field': 'created_by__password', 'operator': 'icontains', 'value': 'x'},
        })
        with self.assertRaises(ReportError):
            eng.build()

    def test_warehouse_scope_enforced(self):
        # کاربر فقط انبار ۱ را دارد — ردیف انبار ۲ هرگز نمی‌آید
        res = ReportEngine(self.viewer, {'entity': 'items', 'fields': ['fa_unic_code']}).run()
        codes = {r['fa_unic_code'] for r in res['rows']}
        self.assertEqual(codes, {'FA-001', 'FA-002'})

    def test_foreign_warehouse_id_403(self):
        with self.assertRaises(ReportError) as ctx:
            ReportEngine(self.viewer, {
                'entity': 'items', 'warehouse_id': self.wh2.id, 'fields': ['fa_unic_code'],
            })
        self.assertEqual(ctx.exception.status, 403)

    def test_bad_warehouse_id_400(self):
        with self.assertRaises(ReportError) as ctx:
            ReportEngine(self.admin, {'entity': 'items', 'warehouse_id': 'abc'})
        self.assertEqual(ctx.exception.status, 400)

    def test_sensitive_field_hidden_without_bypass(self):
        SystemSetting.objects.create(
            key='SENSITIVE_EXCEL_FIELDS', value=['inventory'], warehouse=None,
        )
        eng = ReportEngine(self.viewer, {'entity': 'items', 'fields': ['inventory']})
        with self.assertRaises(ReportError):
            eng.build()
        # superuser می‌بیند
        res = ReportEngine(self.admin, {'entity': 'items', 'fields': ['inventory']}).run()
        self.assertEqual(res['count'], 3)

    def test_blind_counting_hides_balance(self):
        SystemSetting.objects.filter(key='blind_counting').update(value='blind')
        eng = ReportEngine(self.viewer, {'entity': 'items', 'fields': ['inventory']})
        with self.assertRaises(ReportError):
            eng.build()


class EngineQueryTests(BaseReportTest):
    def test_and_or_filter_tree(self):
        spec = {
            'entity': 'items', 'fields': ['fa_unic_code'],
            'filters': {'op': 'AND', 'children': [
                {'field': 'inventory', 'operator': 'gte', 'value': 5},
                {'op': 'OR', 'children': [
                    {'field': 'vendor', 'operator': 'icontains', 'value': 'زیمنس'},
                    {'field': 'description', 'operator': 'icontains', 'value': 'مهره'},
                ]},
            ]},
        }
        res = ReportEngine(self.admin, spec).run()
        codes = {r['fa_unic_code'] for r in res['rows']}
        self.assertEqual(codes, {'FA-001', 'FA-003'})

    def test_invalid_coercion_400_not_500(self):
        spec = {
            'entity': 'items', 'fields': ['fa_unic_code'],
            'filters': {'field': 'inventory', 'operator': 'gte', 'value': 'متن'},
        }
        with self.assertRaises(ReportError):
            ReportEngine(self.admin, spec).build()

    def test_invalid_operator_for_type(self):
        spec = {
            'entity': 'items', 'fields': ['fa_unic_code'],
            'filters': {'field': 'inventory', 'operator': 'icontains', 'value': '1'},
        }
        with self.assertRaises(ReportError):
            ReportEngine(self.admin, spec).build()

    def test_group_by_with_sum(self):
        spec = {
            'entity': 'items',
            'group_by': ['warehouse__name'],
            'aggregations': [{'field': 'inventory', 'fn': 'sum', 'alias': 'total_balance'}],
            'sort': [{'field': 'total_balance', 'dir': 'desc'}],
        }
        res = ReportEngine(self.admin, spec).run()
        self.assertEqual(res['count'], 2)
        first = res['rows'][0]
        self.assertEqual(first['warehouse__name'], 'انبار ۱')
        self.assertEqual(Decimal(first['total_balance']), Decimal('14'))

    def test_dynamic_jsonb_filter_and_sum(self):
        spec = {
            'entity': 'items', 'warehouse_id': self.wh1.id,
            'group_by': ['warehouse__name'],
            'aggregations': [{'field': 'dyn__coating', 'fn': 'sum', 'alias': 'sum_coating'}],
            'filters': {'field': 'dyn__coating', 'operator': 'gte', 'value': 1},
        }
        res = ReportEngine(self.admin, spec).run()
        self.assertEqual(res['count'], 1)
        self.assertEqual(Decimal(res['rows'][0]['sum_coating']), Decimal('7.5'))

    def test_dynamic_field_requires_warehouse(self):
        eng = ReportEngine(self.admin, {'entity': 'items', 'fields': ['dyn__coating']})
        with self.assertRaises(ReportError):
            eng.build()

    def test_alias_validation(self):
        spec = {
            'entity': 'items', 'group_by': ['warehouse__name'],
            'aggregations': [{'field': 'inventory', 'fn': 'sum', 'alias': 'DROP TABLE'}],
        }
        with self.assertRaises(ReportError):
            ReportEngine(self.admin, spec).build()

    def test_sort_only_on_selected(self):
        spec = {
            'entity': 'items', 'fields': ['fa_unic_code'],
            'sort': [{'field': 'inventory', 'dir': 'desc'}],
        }
        with self.assertRaises(ReportError):
            ReportEngine(self.admin, spec).build()

    def test_not_on_leaf(self):
        spec = {
            'entity': 'items', 'fields': ['fa_unic_code'],
            'filters': {'field': 'vendor', 'operator': 'icontains',
                        'value': 'زیمنس', 'not': True},
        }
        res = ReportEngine(self.admin, spec).run()
        codes = {r['fa_unic_code'] for r in res['rows']}
        self.assertEqual(codes, {'FA-002'})

    def test_not_on_group(self):
        # NOT(vendor~زیمنس OR balance>=7) → فقط FA-002 (ABB، balance=4)
        spec = {
            'entity': 'items', 'fields': ['fa_unic_code'],
            'filters': {'op': 'OR', 'not': True, 'children': [
                {'field': 'vendor', 'operator': 'icontains', 'value': 'زیمنس'},
                {'field': 'inventory', 'operator': 'gte', 'value': 7},
            ]},
        }
        res = ReportEngine(self.admin, spec).run()
        codes = {r['fa_unic_code'] for r in res['rows']}
        self.assertEqual(codes, {'FA-002'})

    def test_having_basic(self):
        spec = {
            'entity': 'items',
            'group_by': ['warehouse__name'],
            'aggregations': [{'field': 'inventory', 'fn': 'sum', 'alias': 'total_balance'}],
            'having': [{'alias': 'total_balance', 'operator': 'gte', 'value': 10}],
        }
        res = ReportEngine(self.admin, spec).run()
        self.assertEqual(res['count'], 1)
        row = res['rows'][0]
        self.assertEqual(row['warehouse__name'], 'انبار ۱')
        self.assertEqual(Decimal(row['total_balance']), Decimal('14'))

    def test_having_between(self):
        spec = {
            'entity': 'items',
            'group_by': ['warehouse__name'],
            'aggregations': [{'field': 'inventory', 'fn': 'sum', 'alias': 'total_balance'}],
            'having': [{'alias': 'total_balance', 'operator': 'between', 'value': [5, 10]}],
        }
        res = ReportEngine(self.admin, spec).run()
        self.assertEqual(res['count'], 1)
        self.assertEqual(res['rows'][0]['warehouse__name'], 'انبار ۲')

    def test_having_unknown_alias_400(self):
        spec = {
            'entity': 'items',
            'group_by': ['warehouse__name'],
            'aggregations': [{'field': 'inventory', 'fn': 'sum', 'alias': 'total_balance'}],
            'having': [{'alias': 'inventory', 'operator': 'gte', 'value': 1}],
        }
        with self.assertRaises(ReportError) as ctx:
            ReportEngine(self.admin, spec).build()
        self.assertEqual(ctx.exception.status, 400)

    def test_having_without_group_by_400(self):
        spec = {
            'entity': 'items', 'fields': ['fa_unic_code'],
            'having': [{'alias': 'x', 'operator': 'gte', 'value': 1}],
        }
        with self.assertRaises(ReportError) as ctx:
            ReportEngine(self.admin, spec).build()
        self.assertEqual(ctx.exception.status, 400)

    def test_having_non_numeric_value_400(self):
        spec = {
            'entity': 'items',
            'group_by': ['warehouse__name'],
            'aggregations': [{'field': 'inventory', 'fn': 'sum', 'alias': 'total_balance'}],
            'having': [{'alias': 'total_balance', 'operator': 'gte', 'value': 'متن'}],
        }
        with self.assertRaises(ReportError) as ctx:
            ReportEngine(self.admin, spec).build()
        self.assertEqual(ctx.exception.status, 400)
        self.assertIn('HAVING', ctx.exception.message)

    def test_users_m2m_count_no_cartesian(self):
        res = ReportEngine(self.admin, {
            'entity': 'users', 'fields': ['username', 'assigned_warehouses_count'],
        }).run()
        by_name = {r['username']: r for r in res['rows']}
        self.assertEqual(by_name['viewer']['assigned_warehouses_count'], 1)
        # هر کاربر دقیقاً یک ردیف — تکثیر دکارتی رخ نداده
        self.assertEqual(res['count'], CustomUser.objects.count())


class ApiTests(BaseReportTest):
    def setUp(self):
        self.client = APIClient()

    def test_entities_filtered_by_permission(self):
        self.client.force_authenticate(self.viewer)
        r = self.client.get('/api/reports/entities/')
        self.assertEqual(r.status_code, 200)
        keys = {e['key'] for e in r.json()}
        self.assertIn('items', keys)
        self.assertNotIn('users', keys)

    def test_reports_menu_gate(self):
        u = make_user('nomenu', perms=('view_wh_docs',))
        self.client.force_authenticate(u)
        r = self.client.get('/api/reports/entities/')
        self.assertEqual(r.status_code, 403)

    def test_run_endpoint(self):
        self.client.force_authenticate(self.viewer)
        r = self.client.post('/api/reports/run/', {
            'entity': 'items', 'fields': ['fa_unic_code', 'description'],
        }, format='json')
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r.json()['count'], 2)

    def test_run_bad_field_400(self):
        self.client.force_authenticate(self.viewer)
        r = self.client.post('/api/reports/run/', {
            'entity': 'items', 'fields': ['warehouse__manager__password'],
        }, format='json')
        self.assertEqual(r.status_code, 400)

    def test_export_requires_reports_perm(self):
        user_without_reports = make_user('noreports', perms=('view_wh_docs',))
        self.client.force_authenticate(user_without_reports)
        r = self.client.post('/api/reports/export/', {
            'entity': 'items', 'fields': ['fa_unic_code'],
        }, format='json')
        self.assertEqual(r.status_code, 403)

        # کاربری که دسترسی گزارش‌ساز دارد، می‌تواند خروجی بگیرد
        self.client.force_authenticate(self.viewer)
        r = self.client.post('/api/reports/export/', {
            'entity': 'items', 'fields': ['fa_unic_code'],
        }, format='json')
        self.assertEqual(r.status_code, 200)

    def test_export_sync_small(self):
        self.client.force_authenticate(self.admin)
        r = self.client.post('/api/reports/export/', {
            'entity': 'items', 'fields': ['fa_unic_code'],
        }, format='json')
        self.assertEqual(r.status_code, 200)
        self.assertIn('spreadsheetml', r['Content-Type'])

    def test_template_visibility(self):
        other = make_user('other', perms=('view_sys_reports',))
        ReportTemplate.objects.create(
            name='خصوصی', entity='items', owner=self.viewer, spec={'entity': 'items'},
        )
        pub = ReportTemplate.objects.create(
            name='عمومی', entity='items', owner=self.viewer, spec={'entity': 'items'},
            is_public=True,
        )
        self.client.force_authenticate(other)
        r = self.client.get('/api/reports/templates/')
        names = {t['name'] for t in r.json()}
        self.assertEqual(names, {'عمومی'})
        # غیرمالک نمی‌تواند قالب عمومی را ویرایش/حذف کند
        r = self.client.patch(f'/api/reports/templates/{pub.id}/', {'name': 'x'}, format='json')
        self.assertEqual(r.status_code, 403)
        r = self.client.delete(f'/api/reports/templates/{pub.id}/')
        self.assertEqual(r.status_code, 403)

    def test_template_create_sets_owner(self):
        self.client.force_authenticate(self.viewer)
        r = self.client.post('/api/reports/templates/', {
            'name': 'گزارشم', 'entity': 'items', 'spec': {'entity': 'items'},
        }, format='json')
        self.assertEqual(r.status_code, 201)
        self.assertEqual(ReportTemplate.objects.get(pk=r.json()['id']).owner, self.viewer)

    def test_run_with_not_and_having(self):
        self.client.force_authenticate(self.admin)
        r = self.client.post('/api/reports/run/', {
            'entity': 'items',
            'group_by': ['warehouse__name'],
            'aggregations': [{'field': 'inventory', 'fn': 'sum', 'alias': 'total_balance'}],
            'filters': {'field': 'vendor', 'operator': 'icontains',
                        'value': 'ABB', 'not': True},
            'having': [{'alias': 'total_balance', 'operator': 'gte', 'value': 8}],
        }, format='json')
        self.assertEqual(r.status_code, 200)
        body = r.json()
        # بدون ABB: انبار ۱ جمع ۱۰، انبار ۲ جمع ۷ → having>=8 فقط انبار ۱
        self.assertEqual(body['count'], 1)
        self.assertEqual(body['rows'][0]['warehouse__name'], 'انبار ۱')

    def test_template_bad_entity_rejected(self):
        self.client.force_authenticate(self.viewer)
        r = self.client.post('/api/reports/templates/', {
            'name': 'بد', 'entity': 'nope', 'spec': {},
        }, format='json')
        self.assertEqual(r.status_code, 400)


class ExcelStyleTests(BaseReportTest):
    def setUp(self):
        self.client = APIClient()
        self.client.force_authenticate(self.admin)

    def test_sync_excel_styled_and_openable(self):
        r = self.client.post('/api/reports/export/', {
            'entity': 'items', 'fields': ['fa_unic_code', 'inventory'],
            'report_name': 'گزارش تست',
        }, format='json')
        self.assertEqual(r.status_code, 200)

        ws = openpyxl.load_workbook(io.BytesIO(r.content))['Report']
        # ردیف ۱ عنوان، ردیف ۲ برچسب (بولد روی fill سرمه‌ای)، ردیف ۳ کلید، داده از ۴
        self.assertTrue(str(ws['A1'].value).startswith('گزارش تست'))
        self.assertTrue(ws['A2'].font.bold)
        self.assertIn('4F46E5', str(ws['A2'].fill.fgColor.rgb))
        self.assertEqual(ws['A3'].value, 'fa_unic_code')
        self.assertEqual(ws.max_row, 3 + 3)  # ۳ ردیف سربرگ + ۳ آیتم
        # فرمت هزارگان روی ستون عددی balance
        self.assertEqual(ws['B4'].number_format, '#,##0.###')
        # زبرا روی ردیف دوم داده
        self.assertIn('F1F5F9', str(ws['A5'].fill.fgColor.rgb))
        # فریز و فیلتر (در write_only فقط از راه Pane/auto_filter.ref ممکن است)
        self.assertEqual(ws.freeze_panes, 'A4')
        self.assertEqual(ws.auto_filter.ref, 'A3:B6')
        self.assertTrue(ws.sheet_view.rightToLeft)

    def test_excel_with_joined_fields(self):
        r = self.client.post('/api/reports/export/', {
            'entity': 'items',
            'joins': [{'to': 'count_tasks', 'type': 'inner', 'as': 'count_tasks'}],
            'fields': ['fa_unic_code', 'count_tasks.status'],
            'warehouse_id': self.wh1.id,
            'report_name': 'گزارش پیوندی',
        }, format='json')
        self.assertEqual(r.status_code, 200)
        ws = openpyxl.load_workbook(io.BytesIO(r.content))['Report']
        self.assertEqual(ws['B3'].value, 'count_tasks.status')
        # بررسی وجود مقدار در ردیف داده (ردیف ۴)
        self.assertIn(ws['B4'].value, ['COUNTED', 'PENDING_COUNT'])


class PdfExportTests(BaseReportTest):
    def setUp(self):
        self.client = APIClient()
        self.client.force_authenticate(self.admin)

    def _spec(self, **extra):
        return {'entity': 'items', 'fields': ['fa_unic_code', 'vendor'],
                'format': 'pdf', 'report_name': 'گزارش اقلام', **extra}

    def test_pdf_returned(self):
        r = self.client.post('/api/reports/export/', self._spec(), format='json')
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r['Content-Type'], 'application/pdf')
        self.assertTrue(r.content.startswith(b'%PDF'))

    def test_pdf_row_limit_persian_error(self):
        with patch('reports.pdf.PDF_ROW_LIMIT', 2):
            r = self.client.post('/api/reports/export/', self._spec(), format='json')
        self.assertEqual(r.status_code, 400)
        self.assertIn('Excel', r.json()['error'])

    def test_pdf_column_limit(self):
        with patch('reports.pdf.PDF_MAX_COLUMNS', 1):
            r = self.client.post('/api/reports/export/', self._spec(), format='json')
        self.assertEqual(r.status_code, 400)
        self.assertIn('ستون', r.json()['error'])

    def test_pdf_grouped_report(self):
        spec = {
            'entity': 'items', 'format': 'pdf', 'report_name': 'گروهی',
            'group_by': ['warehouse__name'],
            'aggregations': [{'field': 'inventory', 'fn': 'sum', 'alias': 'total'}],
        }
        r = self.client.post('/api/reports/export/', spec, format='json')
        self.assertEqual(r.status_code, 200)
        self.assertTrue(r.content.startswith(b'%PDF'))

    def test_pdf_zero_rows(self):
        spec = self._spec(filters={'field': 'vendor', 'operator': 'eq', 'value': 'NON_EXISTENT'})
        r = self.client.post('/api/reports/export/', spec, format='json')
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r['Content-Type'], 'application/pdf')
        self.assertTrue(r.content.startswith(b'%PDF'))

    def test_pdf_with_joined_fields(self):
        spec = {
            'entity': 'items', 'format': 'pdf', 'report_name': 'اقلام و تسک‌ها',
            'joins': [{'to': 'count_tasks', 'type': 'inner', 'as': 'count_tasks'}],
            'fields': ['fa_unic_code', 'count_tasks.status'],
            'warehouse_id': self.wh1.id,
        }
        r = self.client.post('/api/reports/export/', spec, format='json')
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r['Content-Type'], 'application/pdf')
        self.assertTrue(r.content.startswith(b'%PDF'))


class ExportWorkerTests(BaseReportTest):
    SPEC = {'entity': 'items', 'fields': ['fa_unic_code']}

    def test_run_export_job_done_with_heartbeat(self):
        job = ReportExportJob.objects.create(
            owner=self.admin, spec=self.SPEC, report_name='r',
            status='pending', total_rows=3,
        )
        _run_export_job(job.pk, self.admin.pk)
        job.refresh_from_db()
        self.assertEqual(job.status, 'done')
        self.assertEqual(job.progress, 100)
        self.assertIsNotNone(job.heartbeat_at)
        p = job.absolute_file_path
        self.assertTrue(p.exists())
        p.unlink()

    def test_job_queued_when_worker_alive(self):
        ExportWorkerStatus.objects.create(pk=1, alive_at=timezone.now())
        job = start_export_job(self.admin, self.SPEC, 'r', 20_000)
        job.refresh_from_db()
        # worker زنده → thread ساخته نمی‌شود؛ job در صف می‌ماند تا worker بردارد
        self.assertEqual(job.status, 'pending')

    def test_orphan_recovery_by_heartbeat_not_age(self):
        from .management.commands.run_export_worker import Command
        old = timezone.now() - timezone.timedelta(minutes=10)
        dead = ReportExportJob.objects.create(
            owner=self.admin, spec=self.SPEC, status='running', heartbeat_at=old,
        )
        alive = ReportExportJob.objects.create(
            owner=self.admin, spec=self.SPEC, status='running',
            heartbeat_at=timezone.now(),
        )
        # job بزرگ زنده (نبض تازه) نباید یتیم فرض شود حتی اگر قدیمی شروع شده باشد
        ReportExportJob.objects.filter(pk=alive.pk).update(
            created_at=timezone.now() - timezone.timedelta(hours=2),
        )
        import io
        Command(stdout=io.StringIO())._recover_orphans()
        dead.refresh_from_db()
        alive.refresh_from_db()
        self.assertEqual(dead.status, 'pending')
        self.assertEqual(alive.status, 'running')

    def test_cleanup_respects_live_queue(self):
        ExportWorkerStatus.objects.create(pk=1, alive_at=timezone.now())
        stale_running = ReportExportJob.objects.create(
            owner=self.admin, spec=self.SPEC, status='running',
            heartbeat_at=timezone.now() - timezone.timedelta(hours=2),
        )
        queued = ReportExportJob.objects.create(
            owner=self.admin, spec=self.SPEC, status='pending',
        )
        ReportExportJob.objects.filter(pk=queued.pk).update(
            created_at=timezone.now() - timezone.timedelta(hours=2),
        )
        cleanup_old_jobs()
        stale_running.refresh_from_db()
        queued.refresh_from_db()
        self.assertEqual(stale_running.status, 'failed')
        # با worker زنده، صف pending سالم است و failed نمی‌شود
        self.assertEqual(queued.status, 'pending')

    def test_orphan_poison_pill_fails_after_max_attempts(self):
        from .management.commands.run_export_worker import Command
        old = timezone.now() - timezone.timedelta(minutes=10)
        poison = ReportExportJob.objects.create(
            owner=self.admin, spec=self.SPEC, status='running',
            heartbeat_at=old, attempts=2,
        )
        import io
        Command(stdout=io.StringIO())._recover_orphans()
        poison.refresh_from_db()
        self.assertEqual(poison.status, 'failed')
        self.assertEqual(poison.attempts, 3)
        self.assertIn('مسموم', poison.error_message)


class JoinQueryTests(BaseReportTest):
    def test_join_output_key_mapping(self):
        """کلیدهای ستون پیوندی در دیکشنری خروجی باید با فرمت کلاینت (count_tasks.status) باشند."""
        spec = {
            'entity': 'items',
            'joins': [{'to': 'count_tasks', 'type': 'inner', 'as': 'count_tasks'}],
            'fields': ['fa_unic_code', 'count_tasks.status'],
            'warehouse_id': self.wh1.id,
        }
        res = ReportEngine(self.admin, spec).run()
        self.assertEqual(res['count'], 2)
        row = res['rows'][0]
        self.assertIn('fa_unic_code', row)
        self.assertIn('count_tasks.status', row)
        self.assertNotIn('count_tasks_fr__status', row)

    def test_join_calculated_annotation(self):
        """فیلد محاسباتی counter_name روی جدول پیوندی بدون خطای FieldError کار کند و نام شخص را بسازد."""
        spec = {
            'entity': 'items',
            'joins': [{'to': 'count_tasks', 'type': 'inner', 'as': 'count_tasks'}],
            'fields': ['fa_unic_code', 'count_tasks.counter_name', 'count_tasks.supervisor_name'],
            'warehouse_id': self.wh1.id,
        }
        res = ReportEngine(self.admin, spec).run()
        self.assertEqual(res['count'], 2)
        row = res['rows'][0]
        self.assertIn('count_tasks.counter_name', row)
        self.assertEqual(row['count_tasks.counter_name'], 'علی رضایی')
        self.assertEqual(row['count_tasks.supervisor_name'], 'محمد محمدی')

    def test_join_filter_and_sort(self):
        """فیلتر و مرتب‌سازی روی ستون‌های پیوندی به درستی کار کند."""
        spec = {
            'entity': 'items',
            'joins': [{'to': 'count_tasks', 'type': 'left', 'as': 'count_tasks'}],
            'fields': ['fa_unic_code', 'count_tasks.status'],
            'filters': {'field': 'count_tasks.status', 'operator': 'eq', 'value': 'COUNTED'},
            'sort': [{'field': 'fa_unic_code', 'dir': 'asc'}],
        }
        res = ReportEngine(self.admin, spec).run()
        self.assertEqual(res['count'], 1)
        self.assertEqual(res['rows'][0]['fa_unic_code'], 'FA-001')
        self.assertEqual(res['rows'][0]['count_tasks.status'], 'COUNTED')

    def test_join_filter_only_exists_with_calculated_field(self):
        """فیلتر روی فیلد محاسباتی جدول پیوندی وقتی جدول در خروجی نیست (مسیر EXISTS) بدون خطا کار کند."""
        spec = {
            'entity': 'items',
            'joins': [{'to': 'count_tasks', 'type': 'left', 'as': 'count_tasks'}],
            'fields': ['fa_unic_code'],
            'filters': {'field': 'count_tasks.counter_name', 'operator': 'icontains', 'value': 'رضایی'},
            'warehouse_id': self.wh1.id,
        }
        res = ReportEngine(self.admin, spec).run()
        self.assertEqual(res.get('join_mode'), 'exists')
        self.assertEqual(res['count'], 2)
        codes = [r['fa_unic_code'] for r in res['rows']]
        self.assertIn('FA-001', codes)
        self.assertIn('FA-002', codes)

    def test_join_filter_only_exists_with_calculated_field_non_matching(self):
        """فیلتر EXISTS روی فیلد محاسباتی با مقدار ناموجود باید خروجی خالی بدهد."""
        spec = {
            'entity': 'items',
            'joins': [{'to': 'count_tasks', 'type': 'left', 'as': 'count_tasks'}],
            'fields': ['fa_unic_code'],
            'filters': {'field': 'count_tasks.counter_name', 'operator': 'icontains', 'value': 'شخص_ناموجود'},
            'warehouse_id': self.wh1.id,
        }
        res = ReportEngine(self.admin, spec).run()
        self.assertEqual(res.get('join_mode'), 'exists')
        self.assertEqual(res['count'], 0)
        self.assertEqual(res['rows'], [])

    def test_join_filter_only_exists_combined_and_filter(self):
        """ترکیب فیلتر پایه و فیلد محاسباتی پیوندی در مسیر EXISTS."""
        spec = {
            'entity': 'items',
            'joins': [{'to': 'count_tasks', 'type': 'left', 'as': 'count_tasks'}],
            'fields': ['fa_unic_code'],
            'filters': {
                'op': 'AND',
                'children': [
                    {'field': 'fa_unic_code', 'operator': 'eq', 'value': 'FA-001'},
                    {'field': 'count_tasks.counter_name', 'operator': 'icontains', 'value': 'رضایی'},
                ],
            },
            'warehouse_id': self.wh1.id,
        }
        res = ReportEngine(self.admin, spec).run()
        self.assertEqual(res['count'], 1)
        self.assertEqual(res['rows'][0]['fa_unic_code'], 'FA-001')

    def test_grouped_pagination_has_deterministic_order(self):
        """در حالت گروه‌بندی بدون مرتب‌سازی دستی، فیلدهای گروه‌بندی به عنوان tiebreak اضافه شوند."""
        spec = {
            'entity': 'items',
            'group_by': ['vendor'],
            'aggregations': [{'fn': 'count', 'field': 'id', 'alias': 'c'}],
            'warehouse_id': self.wh1.id,
        }
        eng = ReportEngine(self.admin, spec)
        qs, cols, mode = eng.build()
        self.assertIn('vendor', qs.query.order_by)

    def test_grouped_sum_on_base_field_with_filter_only_many_join_allowed(self):
        """جمع روی فیلد پایه وقتی JOIN چندمقداری فقط در فیلتر (EXISTS) است باید مجاز باشد و درست محاسبه شود."""
        spec = {
            'entity': 'items',
            'joins': [{'to': 'count_tasks', 'type': 'left', 'as': 'count_tasks'}],
            'group_by': ['vendor'],
            'aggregations': [{'fn': 'sum', 'field': 'inventory', 'alias': 'total_inv'}],
            'filters': {'field': 'count_tasks.status', 'operator': 'eq', 'value': 'COUNTED'},
            'warehouse_id': self.wh1.id,
        }
        res = ReportEngine(self.admin, spec).run()
        self.assertEqual(res.get('join_mode'), 'exists')
        self.assertEqual(res['count'], 1)
        self.assertEqual(res['rows'][0]['vendor'], 'زیمنس')
        self.assertEqual(Decimal(str(res['rows'][0]['total_inv'])), Decimal('10'))

    def test_grouped_sum_on_base_field_with_output_many_join_rejected(self):
        """جمع روی فیلد پایه وقتی JOIN چندمقداری در ستون‌های خروجی است باید رد شود (جلوگیری از ضرب دکارتی)."""
        spec = {
            'entity': 'items',
            'joins': [{'to': 'count_tasks', 'type': 'left', 'as': 'count_tasks'}],
            'group_by': ['count_tasks.status'],
            'aggregations': [{'fn': 'sum', 'field': 'inventory', 'alias': 'total_inv'}],
            'warehouse_id': self.wh1.id,
        }
        with self.assertRaises(ReportError) as ctx:
            ReportEngine(self.admin, spec).build()
        self.assertIn('در ترکیب با جدول چندمقداری', ctx.exception.message)

    def test_sensitive_field_bypass_view_sys_export_cannot_bypass(self):
        """کاربر دارای مجوز view_sys_export در حالت شمارش کور نباید فیلدهای حساس را ببیند."""
        from warehouses.services import clear_setting_cache
        SystemSetting.objects.filter(key='blind_counting').update(value='blind')
        clear_setting_cache('blind_counting', self.wh1.id)
        clear_setting_cache('blind_counting', None)
        export_user = make_user('exporter', perms=('view_sys_reports', 'view_wh_docs', 'view_sys_export'))
        cfg = get_registry()['items']
        allowed = cfg.allowed_fields(export_user, self.wh1.id)
        self.assertNotIn('inventory', allowed)

        # بازگرداندن تنظیم برای سایر تست‌ها
        SystemSetting.objects.filter(key='blind_counting').update(value='normal')
        clear_setting_cache('blind_counting', self.wh1.id)
        clear_setting_cache('blind_counting', None)

    def test_sensitive_field_bypass_manager_review_can_bypass(self):
        """کاربر دارای مجوز view_sys_manager_review در حالت شمارش کور به فیلدهای حساس دسترسی دارد."""
        from warehouses.services import clear_setting_cache
        SystemSetting.objects.filter(key='blind_counting').update(value='blind')
        clear_setting_cache('blind_counting', self.wh1.id)
        clear_setting_cache('blind_counting', None)
        manager_user = make_user('manager', perms=('view_sys_reports', 'view_wh_docs', 'view_sys_manager_review'))
        cfg = get_registry()['items']
        allowed = cfg.allowed_fields(manager_user, self.wh1.id)
        self.assertIn('inventory', allowed)

        # بازگرداندن تنظیم برای سایر تست‌ها
        SystemSetting.objects.filter(key='blind_counting').update(value='normal')
        clear_setting_cache('blind_counting', self.wh1.id)
        clear_setting_cache('blind_counting', None)

    def test_unhandled_exception_returns_500_generic_message(self):
        """خطای پیش‌بینی‌نشده سروری باید با کد ۵۰۰ و پیام امن فارسی برگردد بدون افشای متن داخلی استثنا."""
        client = APIClient()
        client.force_authenticate(self.admin)
        with patch('reports.views.ReportEngine.run', side_effect=RuntimeError('Database internal table fault')):
            r = client.post('/api/reports/run/', {'entity': 'items', 'fields': ['fa_unic_code']}, format='json')
            self.assertEqual(r.status_code, 500)
            self.assertIn('خطای داخلی در اجرای گزارش', r.json().get('error', ''))
            self.assertNotIn('Database internal table fault', r.json().get('error', ''))

    def test_export_queryset_handles_operational_error_timeout(self):
        """در صورت timeout خوردن محاسبه تعداد ردیف‌ها، خطای فارسی مناسب برگردانده شود."""
        from django.db import OperationalError
        eng = ReportEngine(self.admin, {'entity': 'items', 'fields': ['fa_unic_code']})
        with patch('django.db.models.query.QuerySet.count', side_effect=OperationalError('canceling statement due to statement timeout')):
            with self.assertRaises(ReportError) as ctx:
                eng.export_queryset()
            self.assertIn('بیش از حد طول کشید', ctx.exception.message)

    def test_export_job_detail_recovers_stale_running_job(self):
        """بررسی وضعیت job با نبض کهنه (بیش از ۵ دقیقه) باید آن را به failed تغییر دهد."""
        old_time = timezone.now() - timezone.timedelta(minutes=10)
        stale_job = ReportExportJob.objects.create(
            owner=self.admin, spec={'entity': 'items'}, report_name='stale',
            status='running', heartbeat_at=old_time,
        )
        client = APIClient()
        client.force_authenticate(self.admin)
        r = client.get(f'/api/reports/exports/{stale_job.pk}/')
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r.json().get('status'), 'failed')
        self.assertIn('پردازش متوقف شد', r.json().get('error_message', ''))
        stale_job.refresh_from_db()
        self.assertEqual(stale_job.status, 'failed')

    def test_multi_warehouse_blind_counting_masks_fields_when_any_warehouse_is_blind(self):
        """در کوئری بدون انتخاب انبار (چندانباری)، اگر حتی یک انبار blind باشد فیلدهای حساس باید ماسک شوند."""
        from warehouses.services import clear_setting_cache
        # تنظیم سراسری عادی، اما انبار ۲ روی blind
        SystemSetting.objects.filter(key='blind_counting', warehouse__isnull=True).update(value='normal')
        SystemSetting.objects.create(key='blind_counting', value='blind', warehouse=self.wh2)
        clear_setting_cache('blind_counting', self.wh2.id)
        clear_setting_cache('blind_counting', None)

        viewer_user = make_user('viewer_user', perms=('view_sys_reports', 'view_wh_docs'))
        cfg = get_registry()['items']
        # بدون انتخاب انبار (warehouse_id=None)
        allowed = cfg.allowed_fields(viewer_user, warehouse_id=None)
        self.assertNotIn('inventory', allowed)

        # پاک‌سازی رکورد تستی
        SystemSetting.objects.filter(key='blind_counting', warehouse=self.wh2).delete()
        clear_setting_cache('blind_counting', self.wh2.id)
        clear_setting_cache('blind_counting', None)

    def test_datetime_date_filter_includes_end_of_day(self):
        """فیلتر lte یا between روی فیلد datetime با ورودی تاریخ باید رکوردهای کل روز پایانی را شامل شود."""
        task = CountTask.objects.first()
        today = timezone.now().date()
        late_today = timezone.make_aware(timezone.datetime.combine(today, timezone.datetime.min.time().replace(hour=18, minute=30)))
        history = CountTaskHistory.objects.create(
            task=task,
            action_type='COUNTED',
            action_by=self.admin,
            counted_balance=10,
        )
        CountTaskHistory.objects.filter(pk=history.pk).update(created_at=late_today)

        # فیلتر lte با تاریخ امروز
        spec = {
            'entity': 'count_task_history',
            'fields': ['id', 'action_type', 'created_at'],
            'filters': {
                'field': 'created_at',
                'operator': 'lte',
                'value': str(today),
            }
        }
        res = ReportEngine(self.admin, spec).run()
        found = any(r.get('id') == history.id for r in res['rows'])
        self.assertTrue(found)

        # فیلتر between با تاریخ امروز
        spec_between = {
            'entity': 'count_task_history',
            'fields': ['id', 'action_type', 'created_at'],
            'filters': {
                'field': 'created_at',
                'operator': 'between',
                'value': [str(today), str(today)],
            }
        }
        res_between = ReportEngine(self.admin, spec_between).run()
        found_between = any(r.get('id') == history.id for r in res_between['rows'])
        self.assertTrue(found_between)

    def test_response_rows_do_not_leak_internal_join_id_columns(self):
        """ردیف‌های پاسخ فقط باید ستون‌های درخواستی را داشته باشند و کلیدهای داخلی ORM (مانند count_tasks_fr__id) نشت نکنند."""
        spec = {
            'entity': 'items',
            'fields': ['fa_unic_code', 'count_tasks.status'],
            'joins': [{'to': 'count_tasks'}],
        }
        res = ReportEngine(self.admin, spec).run()
        for row in res['rows']:
            self.assertIn('fa_unic_code', row)
            self.assertIn('count_tasks.status', row)
            self.assertNotIn('count_tasks_fr__id', row)
            self.assertNotIn('id', row)

    def test_export_content_disposition_utf8_filename(self):
        """هدر Content-Disposition باید با استاندارد RFC 5987 نام فارسی را با filename* انکود کند."""
        client = APIClient()
        client.force_authenticate(self.admin)
        r = client.post('/api/reports/export/', {
            'entity': 'items',
            'fields': ['fa_unic_code'],
            'report_name': 'گزارش موجودی کالا',
            'format': 'xlsx',
        }, format='json')
        self.assertEqual(r.status_code, 200)
        cd = r.headers.get('Content-Disposition', '')
        self.assertIn("filename*=UTF-8''", cd)
        from urllib.parse import quote
        self.assertIn(quote('گزارش موجودی کالا.xlsx'.encode('utf-8')), cd)

    def test_filter_on_join_computed_field_when_not_in_output(self):
        """فیلتر روی فیلد محاسباتی جدول پیوندی وقتی جدول پیوندی در خروجی نیست (مسیر EXISTS)."""
        spec = {
            'entity': 'items',
            'fields': ['fa_unic_code'],
            'joins': [{'to': 'count_tasks'}],
            'filters': {
                'field': 'count_tasks.counter_name',
                'operator': 'icontains',
                'value': 'علی',
            },
        }
        res = ReportEngine(self.admin, spec).run()
        codes = [r['fa_unic_code'] for r in res['rows']]
        self.assertIn('FA-001', codes)
        self.assertIn('FA-002', codes)
        self.assertNotIn('FA-003', codes)

    def test_grouped_pagination_stability(self):
        """صفحه‌بندی حالت گروه‌بندی‌شده باید دارای ترتیب پایدار و بدون خطا باشد."""
        spec = {
            'entity': 'items',
            'group_by': ['vendor'],
            'aggregations': [{'fn': 'count', 'field': 'id', 'alias': 'total_items'}],
            'page': 1,
            'page_size': 1,
        }
        res = ReportEngine(self.admin, spec).run()
        self.assertEqual(len(res['rows']), 1)
        self.assertGreaterEqual(res['count'], 2)

    def test_having_on_join_field_alias(self):
        """شرط HAVING روی تابع تجمیعی اعمال‌شده روی فیلد جدول پیوندی."""
        spec = {
            'entity': 'items',
            'joins': [{'to': 'count_tasks'}],
            'group_by': ['vendor'],
            'aggregations': [
                {'fn': 'sum', 'field': 'count_tasks.counted_balance', 'alias': 'sum_counted'}
            ],
            'having': [
                {'alias': 'sum_counted', 'operator': 'gte', 'value': 5}
            ]
        }
        res = ReportEngine(self.admin, spec).run()
        vendors = [r['vendor'] for r in res['rows']]
        self.assertIn('زیمنس', vendors)

    def test_export_queryset_operational_error_handling(self):
        """مسیر export_queryset در مواجهه با OperationalError (مانند timeout) باید خطای معنادار ReportError بدهد."""
        from django.db import OperationalError
        spec = {'entity': 'items', 'fields': ['fa_unic_code']}
        engine = ReportEngine(self.admin, spec)
        with patch('django.db.models.query.QuerySet.count', side_effect=OperationalError('canceling statement due to statement timeout')):
            with self.assertRaises(ReportError) as ctx:
                engine.export_queryset()
            self.assertIn('محاسبه تعداد ردیف‌های خروجی', str(ctx.exception))

    def test_export_job_list_view_ordering(self):
        """اندپوینت فهرست jobها باید به ترتیب نزولی زمان ایجاد (-created_at) مرتب باشد."""
        job1 = ReportExportJob.objects.create(
            owner=self.admin, report_name='گزارش اول',
            spec={'entity': 'items'}, total_rows=10,
        )
        job2 = ReportExportJob.objects.create(
            owner=self.admin, report_name='گزارش دوم',
            spec={'entity': 'items'}, total_rows=20,
        )
        client = APIClient()
        client.force_authenticate(self.admin)
        r = client.get('/api/reports/exports/')
        self.assertEqual(r.status_code, 200)
        items = r.json()
        self.assertGreaterEqual(len(items), 2)
        self.assertEqual(items[0]['id'], job2.pk)
        self.assertEqual(items[1]['id'], job1.pk)

    def test_aggregation_custom_display_label(self):
        """توابع تجمیعی باید از برچسب نمایشی دلخواه فارسی پشتیبانی کنند و آن را در columns برگردانند."""
        spec = {
            'entity': 'items',
            'group_by': ['vendor'],
            'aggregations': [
                {'fn': 'sum', 'field': 'inventory', 'alias': 'total_stock', 'label': 'مجموع کل موجودی انبار'}
            ],
        }
        res = ReportEngine(self.admin, spec).run()
        col = next((c for c in res['columns'] if c['key'] == 'total_stock'), None)
        self.assertIsNotNone(col)
        self.assertEqual(col['label'], 'مجموع کل موجودی انبار')

    def test_aggregation_label_max_length_validation(self):
        """عنوان نمایشی ستون تجمیعی نباید بیش از ۶۰ کاراکتر باشد."""
        spec = {
            'entity': 'items',
            'group_by': ['vendor'],
            'aggregations': [
                {'fn': 'sum', 'field': 'inventory', 'alias': 'total_stock', 'label': 'یک عنوان بسیار طولانی که بیش از شصت کاراکتر دارد و باید توسط سیستم اعتبارسنجی رد شود'}
            ],
        }
        with self.assertRaises(ReportError) as ctx:
            ReportEngine(self.admin, spec).run()
        self.assertIn('طول عنوان ستون', str(ctx.exception))





