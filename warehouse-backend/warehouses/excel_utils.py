"""
ماژول ابزارهای اکسل — انبارها (Warehouses)
تولید فایل اکسل خروجی، فایل قالب نمونه، و خوانش/اعتبارسنجی فایل آپلودی
"""
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse

from .models import Warehouse


# ── Column Definitions ──────────────────────────────────────────────
WAREHOUSES_COLUMNS = [
    {'header': 'کد (code)', 'field': 'code', 'width': 15},
    {'header': 'نام (name)', 'field': 'name', 'width': 25},
    {'header': 'نام پروژه (project_name)', 'field': 'project_name', 'width': 25},
    {'header': 'نوع (type)', 'field': 'type', 'width': 18},
    {'header': 'مکان (location)', 'field': 'location', 'width': 30},
    {'header': 'تلفن (phone_number)', 'field': 'phone_number', 'width': 18},
    {'header': 'ظرفیت (capacity)', 'field': 'capacity', 'width': 15},
    {'header': 'شرکت بهره‌بردار (operator_company)', 'field': 'operator_company', 'width': 25},
    {'header': 'رنگ (color)', 'field': 'color', 'width': 15},
    {'header': 'توضیحات (description)', 'field': 'description', 'width': 35},
]

MAX_IMPORT_ROWS = 500


def _apply_header_style(ws):
    """Apply professional styling to header row."""
    header_font = Font(name='B Nazanin', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='059669', end_color='059669', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_idx, col_def in enumerate(WAREHOUSES_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = col_def['width']


def generate_warehouses_excel(queryset):
    """
    Generate an Excel file from a Warehouse queryset.
    Returns an HttpResponse with the xlsx file attached.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'انبارها'
    ws.sheet_view.rightToLeft = True

    # Write headers
    for col_idx, col_def in enumerate(WAREHOUSES_COLUMNS, 1):
        ws.cell(row=1, column=col_idx, value=col_def['header'])
    _apply_header_style(ws)

    # Write data rows
    data_font = Font(name='B Nazanin', size=11)
    data_alignment = Alignment(horizontal='right', vertical='center')

    for row_idx, wh in enumerate(queryset, 2):
        row_data = [
            wh.code or '',
            wh.name or '',
            wh.project_name or '',
            wh.type or '',
            wh.location or '',
            wh.phone_number or '',
            wh.capacity if wh.capacity is not None else '',
            wh.operator_company or '',
            wh.color or '#6366f1',
            wh.description or '',
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_alignment

    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="warehouses_export.xlsx"'
    return response


def generate_warehouses_template():
    """
    Generate a template Excel file with headers and 2 sample rows.
    Returns an HttpResponse with the xlsx file attached.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'قالب انبارها'
    ws.sheet_view.rightToLeft = True

    # Write headers
    for col_idx, col_def in enumerate(WAREHOUSES_COLUMNS, 1):
        ws.cell(row=1, column=col_idx, value=col_def['header'])
    _apply_header_style(ws)

    # Sample rows
    sample_data = [
        ['WH-100', 'انبار مرکزی تهران', 'پروژه آلفا', 'اصلی', 'تهران - خیابان ولیعصر', '02112345678', 1000, 'شرکت فارس عالیش', '#6366f1', 'انبار مرکزی اصلی'],
        ['WH-101', 'انبار شعبه اصفهان', 'پروژه بتا', 'فرعی', 'اصفهان - خیابان چهارباغ', '03112345678', 500, 'شرکت پارس', '#22c55e', 'زیرمجموعه انبار مرکزی'],
    ]
    note_font = Font(name='B Nazanin', size=11, italic=True, color='6B7280')
    data_alignment = Alignment(horizontal='right', vertical='center')

    for row_idx, row in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = note_font
            cell.alignment = data_alignment

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="warehouses_template.xlsx"'
    return response


def parse_warehouses_excel(file):
    """
    Parse an uploaded Excel file and validate each row.
    Returns: { 'valid_rows': [...], 'errors': [...] }
    """
    try:
        wb = load_workbook(file, read_only=True, data_only=True)
    except Exception:
        return {'valid_rows': [], 'errors': [{'row': 0, 'field': 'file', 'message': 'فایل اکسل معتبر نیست یا قابل خواندن نمی‌باشد.'}]}

    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    if len(rows) > MAX_IMPORT_ROWS:
        return {'valid_rows': [], 'errors': [{'row': 0, 'field': 'file', 'message': f'حداکثر {MAX_IMPORT_ROWS} سطر قابل پردازش است. فایل شما {len(rows)} سطر دارد.'}]}

    # Pre-fetch existing warehouse codes for duplicate checking
    existing_codes = set(
        Warehouse.objects.exclude(code__isnull=True)
        .exclude(code='')
        .values_list('code', flat=True)
    )

    valid_rows = []
    errors = []
    seen_codes = set()

    for idx, row in enumerate(rows):
        row_num = idx + 2

        # Skip completely empty rows
        if not row or all(cell is None or str(cell).strip() == '' for cell in row):
            continue

        # Pad row
        row = list(row) + [None] * (len(WAREHOUSES_COLUMNS) - len(row))

        code = str(row[0] or '').strip() if row[0] else ''
        name = str(row[1] or '').strip()
        project_name = str(row[2] or '').strip() if row[2] else ''
        wh_type = str(row[3] or '').strip() if row[3] else ''
        location = str(row[4] or '').strip() if row[4] else ''
        phone_number = str(row[5] or '').strip() if row[5] else ''
        capacity_raw = row[6]
        operator_company = str(row[7] or '').strip() if row[7] else ''
        color = str(row[8] or '#6366f1').strip() if row[8] else '#6366f1'
        description = str(row[9] or '').strip() if row[9] else ''

        row_errors = []

        # ── Required fields ──
        if not name:
            row_errors.append({'row': row_num, 'field': 'name', 'message': 'نام انبار الزامی است.'})

        # ── Code uniqueness ──
        if code:
            if code in existing_codes or code in seen_codes:
                row_errors.append({'row': row_num, 'field': 'code', 'message': f'کد انبار «{code}» قبلاً ثبت شده است.'})
            else:
                seen_codes.add(code)

        # ── Capacity validation ──
        capacity = None
        if capacity_raw is not None and str(capacity_raw).strip() != '':
            try:
                capacity = int(float(str(capacity_raw).strip()))
                if capacity < 0:
                    row_errors.append({'row': row_num, 'field': 'capacity', 'message': 'ظرفیت باید عدد مثبت باشد.'})
            except (ValueError, TypeError):
                row_errors.append({'row': row_num, 'field': 'capacity', 'message': 'ظرفیت باید یک عدد صحیح باشد.'})

        # ── Color validation ──
        if color and not color.startswith('#'):
            color = '#6366f1'

        if row_errors:
            errors.extend(row_errors)
        else:
            valid_rows.append({
                'row_num': row_num,
                'data': {
                    'code': code or None,
                    'name': name,
                    'project_name': project_name or None,
                    'type': wh_type or None,
                    'location': location or None,
                    'phone_number': phone_number or None,
                    'capacity': capacity,
                    'operator_company': operator_company or None,
                    'color': color,
                    'description': description or None,
                }
            })

    wb.close()
    return {
        'valid_rows': valid_rows,
        'errors': errors
    }
